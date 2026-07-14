"""Write the wPPC .xlsx report: Report tab, Weights tab, and a Charts tab.

Charts are native Excel charts via openpyxl.chart (no plotting dependency) and
reference the data ranges in the other tabs, so they stay interactive/live in
the workbook.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .weights import FUNNEL_STATES, Weights

# Exact output column order (spec-mandated).
REPORT_COLUMNS = [
    "segment_id",
    "clicks",
    "conversions",
    "wPPC",
    "wPPC+",
    "wPPC_shrunk",
    "MAR",
    "stabilized",
    "closing_ratio",
]

_CURRENCY_FMT = "#,##0.00"
_INT_FMT = "#,##0"
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_report(output_path: str, results: pd.DataFrame, weights: Weights, platform: str,
                 *, decay: pd.DataFrame | None = None, run_meta: dict | None = None) -> None:
    """Render the report workbook to ``output_path``.

    ``decay`` is the optional two-period wPPC+ movement frame from
    ``model.build_decay`` (columns segment_id, wPPC+_prior, wPPC+_delta,
    delta_pct, trend). When None the output is byte-identical to the pre-decay
    path — no new columns, no new formatting. When supplied, four columns are
    appended at the right edge of the Report tab (positions 10-13); columns 1-9,
    the chart References and the existing conditional-formatting ranges are never
    touched.

    ``run_meta`` is the optional run-metadata block from ``model.build_run_meta``.
    When None (the default) no extra sheet is written, so existing callers/tests
    get the unchanged three-tab workbook. When supplied, a fourth "Run" sheet is
    appended carrying the run-level scalars (platform, baseline, replacement, k +
    k_source, segment/stabilization counts, self-check, telescope sum, generated,
    and the weights_version/drift/decay/incrementality statuses). The
    Report/Weights/Charts tabs are byte-unchanged either way.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    _write_report_tab(ws, results, platform, decay=decay)
    ws_weights = wb.create_sheet("Weights")
    _write_weights_tab(ws_weights, weights, platform)
    ws_charts = wb.create_sheet("Charts")
    _write_charts_tab(ws_charts, ws, ws_weights, results, weights)

    if run_meta is not None:
        ws_run = wb.create_sheet("Run")
        _write_run_tab(ws_run, run_meta)

    wb.save(output_path)


def _write_report_tab(ws, results: pd.DataFrame, platform: str, *, decay=None) -> None:
    headers = ["segment_id", "clicks", "conversions", "wPPC", "wPPC+",
               "wPPC_shrunk", "MAR", "stabilized (Y/N)", "closing_ratio"]

    # Decay is strictly additive: four columns appended at the RIGHT edge
    # (positions 10-13), joined by segment_id. Columns 1-9 are never renumbered.
    decay_lookup = None
    if decay is not None:
        headers = headers + ["wPPC+_prior", "wPPC+_delta", "delta_pct", "trend"]
        decay_lookup = {row["segment_id"]: row for _, row in decay.iterrows()}

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for _, row in results.iterrows():
        values = [
            row["segment_id"],
            int(row["clicks"]),
            int(row["conversions"]),
            round(float(row["wPPC"]), 2),
            round(float(row["wPPC+"]), 0),
            round(float(row["wPPC_shrunk"]), 2),
            round(float(row["MAR"]), 2),
            row["stabilized"],
            round(float(row["closing_ratio"]), 2),
        ]
        if decay is not None:
            d = decay_lookup.get(row["segment_id"])
            values.extend([
                _num_or_none(d["wPPC+_prior"] if d is not None else None, 0),
                _num_or_none(d["wPPC+_delta"] if d is not None else None, 0),
                _num_or_none(d["delta_pct"] if d is not None else None, 4),
                _text_or_none(d["trend"] if d is not None else None),
            ])
        ws.append(values)

    n = len(results)
    last = n + 1  # last data row (header is row 1)

    # Number formats.
    for r in range(2, last + 1):
        ws.cell(r, 2).number_format = _INT_FMT          # clicks
        ws.cell(r, 3).number_format = _INT_FMT          # conversions
        ws.cell(r, 4).number_format = _CURRENCY_FMT     # wPPC
        ws.cell(r, 5).number_format = "0"               # wPPC+
        ws.cell(r, 6).number_format = _CURRENCY_FMT     # wPPC_shrunk
        ws.cell(r, 7).number_format = _CURRENCY_FMT     # MAR
        ws.cell(r, 8).alignment = Alignment(horizontal="center")  # stabilized
        ws.cell(r, 9).number_format = "0.00"            # closing_ratio
        if decay is not None:
            ws.cell(r, 10).number_format = "0"              # wPPC+_prior
            ws.cell(r, 11).number_format = "+0;-0;0"        # wPPC+_delta
            ws.cell(r, 12).number_format = "0.0%"           # delta_pct
            ws.cell(r, 13).alignment = Alignment(horizontal="center")  # trend

    # Column widths.
    widths = [34, 9, 12, 10, 8, 12, 12, 14, 13]
    if decay is not None:
        widths = widths + [12, 12, 11, 10]  # prior, delta, pct, trend
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt

    ws.freeze_panes = "A2"

    if n == 0:
        return

    # Conditional formatting: wPPC+ around 100 (red<->green), MAR red<0 / green>0.
    plus_range = f"E2:E{last}"
    ws.conditional_formatting.add(
        plus_range,
        ColorScaleRule(
            start_type="num", start_value=60, start_color="F8696B",
            mid_type="num", mid_value=100, mid_color="FFEB84",
            end_type="num", end_value=160, end_color="63BE7B",
        ),
    )
    mar_range = f"G2:G{last}"
    ws.conditional_formatting.add(
        mar_range,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")),
    )
    ws.conditional_formatting.add(
        mar_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100")),
    )

    # Decay-only CF: flag fatiguing (Falling) segments red on the trend column.
    # Column 13 (M) — well clear of the existing E/G ranges, so no overlap.
    if decay is not None:
        trend_range = f"M2:M{last}"
        ws.conditional_formatting.add(
            trend_range,
            CellIsRule(operator="equal", formula=['"Falling"'],
                       fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")),
        )


def _num_or_none(value, ndigits):
    """Round a numeric to ``ndigits``, or None for None/NaN (leaves cell empty)."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return round(float(value), ndigits)


def _text_or_none(value):
    """Pass a trend string through, or None for None/NaN."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _write_weights_tab(ws, weights: Weights, platform: str) -> None:
    ws.append([f"Derived linear weights — platform: {platform}"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    ws.append(["State", "P(purchase|S)", "PE(S)", "w(S) incremental"])
    for cell in ws[3]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row = 4
    for s in FUNNEL_STATES:
        ws.append([s, round(weights.p[s], 6), round(weights.pe[s], 4), round(weights.w[s], 4)])
        ws.cell(row, 3).number_format = _CURRENCY_FMT
        ws.cell(row, 4).number_format = _CURRENCY_FMT
        row += 1
    # Repeat event: no P/PE, just the incremental weight.
    ws.append(["repeat", None, None, round(weights.w["repeat"], 4)])
    ws.cell(row, 4).number_format = _CURRENCY_FMT
    row += 1

    ws.append([])
    row += 1
    ws.append(["Self-check (telescope click..purchase)"])
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    ws.append(["Telescoped Σ w(click..purchase)", None, None, round(weights.telescope_sum, 4)])
    ws.cell(row, 4).number_format = _CURRENCY_FMT
    row += 1
    ws.append(["CM3_order (target)", None, None, round(weights.cm3_order, 4)])
    ws.cell(row, 4).number_format = _CURRENCY_FMT
    row += 1
    status = "PASS" if weights.self_check_pass else "FAIL"
    ws.append(["Self-check", None, None, status])
    ws.cell(row, 4).font = Font(bold=True, color="006100" if weights.self_check_pass else "9C0006")

    for i, wdt in enumerate([38, 14, 12, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt


def _write_charts_tab(ws_charts, ws_report, ws_weights, results: pd.DataFrame, weights: Weights) -> None:
    ws_charts["A1"] = "Visualizations (live — driven from the Report & Weights tabs)"
    ws_charts["A1"].font = Font(bold=True, size=12)

    n = len(results)
    if n == 0:
        ws_charts["A3"] = "No segments to chart."
        return
    last = n + 1

    # 1) MAR by segment — the decision sort key (already sorted desc; negatives
    #    extend the other way, so cut candidates read at a glance).
    mar_chart = BarChart()
    mar_chart.type = "bar"
    mar_chart.title = "MAR by segment (Margin Above Replacement)"
    mar_chart.y_axis.title = "MAR ($)"
    mar_chart.x_axis.title = "Segment"
    mar_data = Reference(ws_report, min_col=7, min_row=1, max_row=last)
    mar_cats = Reference(ws_report, min_col=1, min_row=2, max_row=last)
    mar_chart.add_data(mar_data, titles_from_data=True)
    mar_chart.set_categories(mar_cats)
    mar_chart.height = max(8, min(24, 1.2 * n))
    mar_chart.width = 18
    mar_chart.legend = None
    ws_charts.add_chart(mar_chart, "A3")

    # 2) wPPC+ by segment — 100 = account average.
    plus_chart = BarChart()
    plus_chart.type = "col"
    plus_chart.title = "wPPC+ by segment (100 = account average)"
    plus_chart.y_axis.title = "wPPC+"
    plus_chart.x_axis.title = "Segment"
    plus_data = Reference(ws_report, min_col=5, min_row=1, max_row=last)
    plus_chart.add_data(plus_data, titles_from_data=True)
    plus_chart.set_categories(mar_cats)
    plus_chart.height = 10
    plus_chart.width = 18
    plus_chart.legend = None
    ws_charts.add_chart(plus_chart, "L3")

    # 3) Derived incremental weights w(S) — the run-expectancy step.
    #    Weights tab: states in rows 4..(3+len), w(S) in column 4 (D).
    w_chart = BarChart()
    w_chart.type = "col"
    w_chart.title = "Derived incremental weights w(S)"
    w_chart.y_axis.title = "w(S) ($)"
    w_first = 4
    w_last = 3 + len(FUNNEL_STATES) + 1  # include the repeat row
    w_data = Reference(ws_weights, min_col=4, min_row=3, max_row=w_last)  # header row 3 for title
    w_cats = Reference(ws_weights, min_col=1, min_row=w_first, max_row=w_last)
    w_chart.add_data(w_data, titles_from_data=True)
    w_chart.set_categories(w_cats)
    w_chart.height = 10
    w_chart.width = 18
    w_chart.legend = None
    ws_charts.add_chart(w_chart, "A21")

    # 4) Closing ratio vs wPPC — leaky closers (<1) vs regression candidates (>1).
    sc = ScatterChart()
    sc.title = "Closing ratio vs wPPC"
    sc.x_axis.title = "wPPC ($)"
    sc.y_axis.title = "Closing ratio (realized CM3/click ÷ wPPC)"
    xref = Reference(ws_report, min_col=4, min_row=1, max_row=last)        # wPPC (title in row 1)
    yref = Reference(ws_report, min_col=9, min_row=1, max_row=last)        # closing_ratio
    series = Series(yref, xref, title_from_data=True)
    series.marker.symbol = "circle"
    series.marker.size = 7
    series.graphicalProperties.line.noFill = True
    sc.series.append(series)
    sc.height = 10
    sc.width = 18
    ws_charts.add_chart(sc, "L21")


def _status(block, default: str) -> str:
    """Reduce a run_meta status block (dict, or None) to a short label."""
    if block is None:
        return default
    if isinstance(block, dict):
        return str(block.get("status", default))
    return str(block)


def _drift_status(drift) -> str:
    """Human-readable weight-drift status from run_meta['drift']."""
    if drift is None:
        return "not-checked"
    if isinstance(drift, dict):
        return "flagged" if drift.get("flagged") else "no drift"
    return str(drift)


def _weights_version(wv):
    """Short label for the weights snapshot in the Run tab. The run-metadata
    contract carries the FULL snapshot dict (build_run_meta / test_drift); reduce
    it to its timestamp (the version identity) so it fits one cell — openpyxl
    cannot store a dict. Non-dict values pass through; None -> em dash. The full
    snapshot is preserved in the ``.weights.json`` sidecar."""
    if wv is None:
        return "—"
    if isinstance(wv, dict):
        return wv.get("timestamp") or "snapshot"
    return wv


def _write_run_tab(ws, run_meta: dict) -> None:
    """Write the run-metadata block as a labelled key/value sheet.

    Point-in-time methodology scalars up top, then the four pass-through statuses
    (weights_version / drift / decay / incrementality). Values are read straight
    from ``run_meta`` — nothing is recomputed here.
    """
    ws.append([f"Run metadata — platform: {run_meta.get('platform', '')}"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    ws.append(["Field", "Value"])
    for cell in ws[3]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    rows = [
        ("Platform", run_meta.get("platform")),
        ("Baseline wPPC", _round_or_blank(run_meta.get("baseline"), 4)),
        ("Replacement wPPC", _round_or_blank(run_meta.get("replacement"), 4)),
        ("k (stabilization)", _round_or_blank(run_meta.get("k"), 4)),
        ("k source", run_meta.get("k_source")),
        ("Segments", run_meta.get("n_segments")),
        ("Stabilized", run_meta.get("n_stabilized")),
        ("Self-check", "PASS" if run_meta.get("self_check_pass") else "FAIL"),
        ("Telescoped Σ w(click..purchase)", _round_or_blank(run_meta.get("telescope_sum"), 4)),
        ("Generated", run_meta.get("generated")),
        ("Weights version", _weights_version(run_meta.get("weights_version"))),
        ("Weight drift", _drift_status(run_meta.get("drift"))),
        ("Decay", _status(run_meta.get("decay"), "not-run")),
        ("Incrementality", _status(run_meta.get("incrementality"), "not-provided")),
    ]
    for label, value in rows:
        ws.append([label, value])

    ws.column_dimensions[get_column_letter(1)].width = 34
    ws.column_dimensions[get_column_letter(2)].width = 40
    ws.freeze_panes = "A4"


def _round_or_blank(value, ndigits):
    """Round a numeric for display, or pass non-numerics through unchanged."""
    if value is None:
        return None
    try:
        return round(float(value), ndigits)
    except (TypeError, ValueError):
        return value

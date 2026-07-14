"""CLI-405 — consolidated render-contract tests: white-label self-containment,
vendored-asset integrity, animate=False zero-GSAP, determinism (including a
rigorous "modulo timestamp" check), and 3-format agreement.

One model is built from the bundled sample with a PINNED ``generated`` (mirrors
tests/test_output.py's pattern and plugins/google-ads-audit's test_audit.py
discipline) so every render below is a pure function of known inputs. This file
overlaps deliberately with tests/test_charts.py (vendor SHA parity) and
tests/test_output.py (self-containment / animate=False / 3-format agreement) —
it is the single place those render-contract guarantees are asserted together,
with the determinism-modulo-timestamp check made rigorous rather than a loose
substring swap.
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from wppc import charts as C
from wppc import io as wio
from wppc import render_html as RH
from wppc.model import build_model, build_run_meta
from wppc.render_md import render_md
from wppc.report import REPORT_COLUMNS, write_report
from wppc.score import score
from wppc.weights import FUNNEL_STATES, derive_weights

_ROOT = Path(__file__).resolve().parents[1]
_SAMPLE_CSV = _ROOT / "sample_data" / "google_segments.sample.csv"
_SAMPLE_MAP = _ROOT / "config" / "mapping.google.sample.yaml"
PINNED_GENERATED = "2026-07-14T00:00:00+00:00"

# Two well-separated, non-overlapping timestamps for the modulo-timestamp
# determinism check — neither is a substring of the other.
_TS_A = "2020-01-01T00:00:00+00:00"
_TS_B = "2099-12-31T23:59:59+00:00"

_COL = {name: idx for idx, name in enumerate(REPORT_COLUMNS)}


def _build_sample_model(generated=PINNED_GENERATED):
    """Score the bundled sample CSV and assemble its model, end to end — the
    same path the CLI drives, with the clock pinned for reproducibility."""
    mapping = wio.load_mapping(str(_SAMPLE_MAP), "google")
    df, currency = wio.load_segments(str(_SAMPLE_CSV), mapping, "google")
    reach = {s: float(df[s].sum()) for s in FUNNEL_STATES}
    weights = derive_weights(
        reach,
        purchases_total=float(df["purchase"].sum()),
        cm3_order=float(currency["CM3_order"]),
        repeat_rate=float(currency["repeat_rate"]),
        cm3_repeat=float(currency["CM3_repeat"]),
    )
    results = score(df, weights)
    run_meta = build_run_meta(results, weights, "google", generated=generated)
    model = build_model(results, weights, run_meta)
    return model, results, weights, run_meta


@pytest.fixture(scope="module")
def sample():
    return _build_sample_model()


def _seg(model, seg_id):
    return next(s for s in model["segments"] if s["segment_id"] == seg_id)


def _md_row(md, seg_id):
    for line in md.splitlines():
        if line.startswith(f"| {seg_id} |"):
            return line
    raise AssertionError(f"segment row for {seg_id!r} not found in markdown")


def _html_data(html):
    """Parse the model back out of the HTML <script id="data"> block."""
    m = re.search(r'<script id="data" type="application/json">(.*?)</script>', html, re.S)
    assert m, "data script block not found"
    return json.loads(m.group(1))


def _strip_blobs(html):
    """Remove BOTH checksummed third-party blob regions (GSAP + Vega runtime)."""
    html = re.sub(re.escape(RH.GSAP_BEGIN) + ".*?" + re.escape(RH.GSAP_END), "", html, flags=re.S)
    html = re.sub(re.escape(C.VENDOR_BEGIN) + ".*?" + re.escape(C.VENDOR_END), "", html, flags=re.S)
    return html


# ---------------------------------------------------------------------------
# 1. Self-containment outside the two checksummed blob regions.
# ---------------------------------------------------------------------------
def test_self_contained_outside_blobs_animate_true(sample):
    model = sample[0]
    html = RH.render_html(model, animate=True)
    stripped = _strip_blobs(html)
    m = re.search(r"https?://|<link|src=|cdn", stripped)
    assert m is None, f"external ref outside blobs: {m.group(0)!r}"


def test_self_contained_outside_blobs_animate_false(sample):
    model = sample[0]
    html = RH.render_html(model, animate=False)
    stripped = _strip_blobs(html)
    m = re.search(r"https?://|<link|src=|cdn", stripped)
    assert m is None, f"external ref outside blobs: {m.group(0)!r}"


# ---------------------------------------------------------------------------
# 2. Vendored SHA-256 parity — every file listed in SHA256SUMS exists and
#    hashes match (both the chart runtime and GSAP).
# ---------------------------------------------------------------------------
def test_vendored_sha256_parity():
    sums_path = C.VENDOR_DIR / "SHA256SUMS"
    assert sums_path.exists(), "wppc/vendor/SHA256SUMS is missing"
    listed = {}
    for line in sums_path.read_text().strip().splitlines():
        digest, name = line.split()
        listed[name] = digest

    assert listed, "SHA256SUMS is empty"
    for name, digest in listed.items():
        fpath = C.VENDOR_DIR / name
        assert fpath.exists(), f"vendor file '{name}' listed in SHA256SUMS but missing on disk"
        got = hashlib.sha256(fpath.read_bytes()).hexdigest()
        assert got == digest, f"{name}: SHA256SUMS says {digest}, actual is {got}"


# ---------------------------------------------------------------------------
# 3. animate=False carries zero GSAP bytes (not even the token) and is still
#    self-contained.
# ---------------------------------------------------------------------------
def test_animate_false_zero_gsap_and_self_contained(sample):
    model = sample[0]
    html = RH.render_html(model, animate=False)
    assert "gsap" not in html.lower()
    assert RH.GSAP_BEGIN not in html
    assert RH.GSAP_END not in html
    # The chart runtime still ships — a static report still draws its charts.
    assert C.VENDOR_BEGIN in html and C.VENDOR_END in html
    stripped = _strip_blobs(html)
    assert re.search(r"https?://|<link|src=|cdn", stripped) is None


# ---------------------------------------------------------------------------
# 4a. Determinism: byte-identical output across repeated calls with the SAME
#     pinned ``generated`` — proves no clock/randomness/unstable ordering
#     leaks into the renderers.
# ---------------------------------------------------------------------------
def test_render_html_byte_identical_same_generated(sample):
    model = sample[0]
    for animate in (True, False):
        a = RH.render_html(model, animate=animate)
        b = RH.render_html(model, animate=animate)
        assert a == b, f"render_html not byte-identical (animate={animate})"


def test_render_md_byte_identical_same_generated(sample):
    model = sample[0]
    a = render_md(model)
    b = render_md(model)
    assert a == b


# ---------------------------------------------------------------------------
# 4b. Determinism modulo timestamp: build TWO full models from the same input
#     that differ ONLY in ``generated``, then prove the rendered outputs
#     differ ONLY at the timestamp's injection points.
#
#     Rigor beyond a loose ``.replace(TS1, TS).replace(TS2, TS)`` swap:
#       - the two renders must actually differ (guards against a no-op test)
#       - the timestamp must appear the SAME NUMBER of times in each render
#       - neither timestamp leaks into the other render
#       - substituting a placeholder for each run's own timestamp yields
#         byte-identical strings — the diff is confined to exactly those spans
# ---------------------------------------------------------------------------
def _assert_modulo_timestamp(text_a, ts_a, text_b, ts_b):
    assert text_a != text_b, "renders are identical before substitution — test would be vacuous"
    count_a = text_a.count(ts_a)
    count_b = text_b.count(ts_b)
    assert count_a > 0, f"timestamp {ts_a!r} does not appear in its own render"
    assert count_a == count_b, (
        f"timestamp appears {count_a} times in run A but {count_b} times in run B"
    )
    assert ts_b not in text_a, "run B's timestamp leaked into run A's render"
    assert ts_a not in text_b, "run A's timestamp leaked into run B's render"

    placeholder = "\x00__GENERATED__\x00"
    norm_a = text_a.replace(ts_a, placeholder)
    norm_b = text_b.replace(ts_b, placeholder)
    assert norm_a == norm_b, "renders differ at points other than the injected timestamp"


def test_render_html_deterministic_modulo_timestamp():
    model_a, *_ = _build_sample_model(generated=_TS_A)
    model_b, *_ = _build_sample_model(generated=_TS_B)
    # Sanity: the two independently-built models agree on everything BUT the
    # timestamp — otherwise the modulo-timestamp claim below would be moot.
    assert model_a["segments"] == model_b["segments"]
    assert model_a["weights_table"] == model_b["weights_table"]

    for animate in (True, False):
        html_a = RH.render_html(model_a, animate=animate)
        html_b = RH.render_html(model_b, animate=animate)
        _assert_modulo_timestamp(html_a, _TS_A, html_b, _TS_B)


def test_render_md_deterministic_modulo_timestamp():
    model_a, *_ = _build_sample_model(generated=_TS_A)
    model_b, *_ = _build_sample_model(generated=_TS_B)
    md_a = render_md(model_a)
    md_b = render_md(model_b)
    _assert_modulo_timestamp(md_a, _TS_A, md_b, _TS_B)


# ---------------------------------------------------------------------------
# 5. 3-format agreement: md, the HTML <script id="data"> block, and the xlsx
#    Report cells all carry the model's own segment values — no format
#    re-derives a number. Spot-checks two segments (a Scale and a Watch).
# ---------------------------------------------------------------------------
def test_three_format_agreement_spot_check(tmp_path, sample):
    model, results, weights, run_meta = sample
    md = render_md(model)
    html = RH.render_html(model, animate=True)
    html_model = _html_data(html)

    # The HTML embeds the model verbatim — proves it re-derives nothing.
    assert html_model["segments"] == model["segments"]

    xlsx = tmp_path / "report.xlsx"
    write_report(str(xlsx), results, weights, "google", run_meta=run_meta)
    ws = load_workbook(str(xlsx))["Report"]
    xlsx_by_seg = {row[_COL["segment_id"]]: row for row in ws.iter_rows(min_row=2, values_only=True)}

    for seg_id in ("kw_clearance", "kw_obscure_longtail"):
        seg = _seg(model, seg_id)
        xrow = xlsx_by_seg[seg_id]

        # model <-> xlsx (both computed in Python; equal to the last digit).
        assert xrow[_COL["wPPC+"]] == seg["wPPC+"]
        assert xrow[_COL["MAR"]] == seg["MAR"]
        assert xrow[_COL["stabilized"]] == seg["stabilized"]

        # model <-> html data block.
        html_seg = next(s for s in html_model["segments"] if s["segment_id"] == seg_id)
        assert html_seg["wPPC+"] == seg["wPPC+"]
        assert html_seg["MAR"] == seg["MAR"]
        assert html_seg["decision"] == seg["decision"]

        # model <-> md (the rendered cells carry the model's values).
        line = _md_row(md, seg_id)
        assert f"{seg['wPPC+']:,.0f}" in line
        assert f"{seg['MAR']:,.2f}" in line
        assert seg["decision"] in line

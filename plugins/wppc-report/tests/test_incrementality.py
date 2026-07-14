"""CLI-401 Layer-5 incrementality seam: inert-by-construction, v1.

The seam is additive and default-off, AND inert even when a table is
supplied: ``score()`` accepts an ``incrementality`` kwarg purely as the
future insertion point (see wppc/references/incrementality-seam.md) and
never reads it when computing any score. This suite proves that inertness,
the ``load_incrementality`` shape-validation, the ``build_incrementality_meta``
shape, the ``build_run_meta`` pass-through, and the CLI wiring end-to-end.
"""

import json
from pathlib import Path

import pandas as pd
import pytest
from click.testing import CliRunner
from openpyxl import load_workbook

from wppc.cli import cli
from wppc.model import build_incrementality_meta, build_run_meta, load_incrementality
from wppc.score import score

from conftest import make_segments

_REPO = Path(__file__).resolve().parents[1]
_GOOGLE_CSV = _REPO / "sample_data" / "google_segments.sample.csv"
_GOOGLE_MAPPING = _REPO / "config" / "mapping.google.sample.yaml"

_VALID_TABLE = {
    "tiers": [
        {
            "tier": "brand_search",
            "value": 0.35,
            "ci": [0.22, 0.48],
            "power": 0.82,
            "window": "2026-04-01/2026-06-30",
            "timestamp": "2026-07-01T00:00:00+00:00",
        },
        {
            "tier": "retargeting",
            "value": 0.15,
            "ci": [0.02, 0.30],
            "power": 0.55,
            "window": "2026-04-01/2026-06-30",
            "timestamp": "2026-07-01T00:00:00+00:00",
        },
    ]
}


def _write_table(tmp_path, table) -> Path:
    path = tmp_path / "incrementality.json"
    path.write_text(json.dumps(table, indent=2) + "\n", encoding="utf-8")
    return path


def _two_segment_df():
    return make_segments([
        {"segment_id": "a", "click": 100, "engagement": 60, "add_to_cart": 20,
         "initiate_checkout": 12, "purchase": 8, "repeats": 3},
        {"segment_id": "b", "click": 80, "engagement": 40, "add_to_cart": 10,
         "initiate_checkout": 6, "purchase": 4, "repeats": 1},
    ])


# ---------------------------------------------------------------------------
# Inertness proof
# ---------------------------------------------------------------------------

def test_score_with_incrementality_table_is_identical_to_bare(example_weights):
    df = _two_segment_df()

    bare = score(df, example_weights)
    with_table = score(df, example_weights, incrementality=_VALID_TABLE)

    # Additive attr aside, the frames must be pixel-identical.
    bare_cmp = bare.drop(columns=[]).copy()
    with_cmp = with_table.copy()
    pd.testing.assert_frame_equal(bare_cmp, with_cmp)

    bare_attrs = dict(bare.attrs)
    with_attrs = dict(with_table.attrs)
    assert bare_attrs.pop("incrementality_provided") is False
    assert with_attrs.pop("incrementality_provided") is True
    # Every other attr (baseline/replacement/k/k_source/n_segments/n_stabilized)
    # is unchanged.
    assert bare_attrs == with_attrs


def test_score_with_incrementality_none_matches_omitted_kwarg(example_weights):
    df = _two_segment_df()
    omitted = score(df, example_weights)
    explicit_none = score(df, example_weights, incrementality=None)
    pd.testing.assert_frame_equal(omitted, explicit_none)
    assert dict(omitted.attrs) == dict(explicit_none.attrs)


# ---------------------------------------------------------------------------
# load_incrementality shape validation
# ---------------------------------------------------------------------------

def test_load_incrementality_valid_file(tmp_path):
    path = _write_table(tmp_path, _VALID_TABLE)
    table = load_incrementality(str(path))
    assert table == _VALID_TABLE


def test_load_incrementality_missing_tiers_key(tmp_path):
    path = _write_table(tmp_path, {"not_tiers": []})
    with pytest.raises(ValueError, match="tiers"):
        load_incrementality(str(path))


def test_load_incrementality_missing_field_names_field_and_index(tmp_path):
    bad = json.loads(json.dumps(_VALID_TABLE))
    del bad["tiers"][1]["power"]
    path = _write_table(tmp_path, bad)
    with pytest.raises(ValueError) as excinfo:
        load_incrementality(str(path))
    assert "power" in str(excinfo.value)
    assert "1" in str(excinfo.value)  # tier index 1


def test_load_incrementality_non_numeric_value(tmp_path):
    bad = json.loads(json.dumps(_VALID_TABLE))
    bad["tiers"][0]["value"] = "high"
    path = _write_table(tmp_path, bad)
    with pytest.raises(ValueError, match="value"):
        load_incrementality(str(path))


def test_load_incrementality_bad_ci_shape(tmp_path):
    bad = json.loads(json.dumps(_VALID_TABLE))
    bad["tiers"][0]["ci"] = [0.1, 0.2, 0.3]
    path = _write_table(tmp_path, bad)
    with pytest.raises(ValueError, match="ci"):
        load_incrementality(str(path))


def test_load_incrementality_non_numeric_ci_entry(tmp_path):
    bad = json.loads(json.dumps(_VALID_TABLE))
    bad["tiers"][0]["ci"] = ["low", 0.2]
    path = _write_table(tmp_path, bad)
    with pytest.raises(ValueError, match="ci"):
        load_incrementality(str(path))


def test_load_incrementality_non_numeric_power(tmp_path):
    bad = json.loads(json.dumps(_VALID_TABLE))
    bad["tiers"][0]["power"] = None
    path = _write_table(tmp_path, bad)
    with pytest.raises(ValueError, match="power"):
        load_incrementality(str(path))


# ---------------------------------------------------------------------------
# build_incrementality_meta
# ---------------------------------------------------------------------------

def test_build_incrementality_meta_shape():
    meta = build_incrementality_meta("some/path.json", _VALID_TABLE)
    assert meta == {
        "status": "provided, not applied (v1)",
        "path": "some/path.json",
        "tiers": ["brand_search", "retargeting"],
    }


# ---------------------------------------------------------------------------
# build_run_meta pass-through
# ---------------------------------------------------------------------------

def test_build_run_meta_passes_through_incrementality(example_weights):
    df = _two_segment_df()
    results = score(df, example_weights)
    im_meta = build_incrementality_meta("some/path.json", _VALID_TABLE)

    meta = build_run_meta(results, example_weights, platform="google", incrementality=im_meta)
    assert meta["incrementality"] == im_meta
    json.dumps(meta["incrementality"])  # JSON-serializable

    bare = build_run_meta(results, example_weights, platform="google")
    assert bare["incrementality"] is None


# ---------------------------------------------------------------------------
# CLI wiring end-to-end
# ---------------------------------------------------------------------------

def _report_tab_values(path):
    ws = load_workbook(path)["Report"]
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return rows


def test_cli_with_valid_incrementality_file_echoes_and_matches_no_flag_output(tmp_path):
    runner = CliRunner()
    table_path = _write_table(tmp_path, _VALID_TABLE)

    out_bare = tmp_path / "bare.xlsx"
    bare_result = runner.invoke(cli, [
        "report", "--platform", "google",
        "--input", str(_GOOGLE_CSV),
        "--mapping", str(_GOOGLE_MAPPING),
        "--output", str(out_bare),
    ])
    assert bare_result.exit_code == 0, bare_result.output

    out_im = tmp_path / "with_im.xlsx"
    im_result = runner.invoke(cli, [
        "report", "--platform", "google",
        "--input", str(_GOOGLE_CSV),
        "--mapping", str(_GOOGLE_MAPPING),
        "--output", str(out_im),
        "--incrementality", str(table_path),
    ])
    assert im_result.exit_code == 0, im_result.output
    assert "Incrementality table provided (2 tiers)" in im_result.output
    assert "NOT applied" in im_result.output

    # The Report tab's cell values are byte-identical between the two runs —
    # the seam is fully inert.
    assert _report_tab_values(out_bare) == _report_tab_values(out_im)


def test_cli_with_malformed_incrementality_file_fails_precisely(tmp_path):
    runner = CliRunner()
    bad = json.loads(json.dumps(_VALID_TABLE))
    del bad["tiers"][0]["timestamp"]
    table_path = _write_table(tmp_path, bad)

    out = tmp_path / "out.xlsx"
    result = runner.invoke(cli, [
        "report", "--platform", "google",
        "--input", str(_GOOGLE_CSV),
        "--mapping", str(_GOOGLE_MAPPING),
        "--output", str(out),
        "--incrementality", str(table_path),
    ])
    assert result.exit_code != 0
    assert "timestamp" in result.output
    assert not out.exists()


# ---------------------------------------------------------------------------
# Hard-rule: no methodology literals in score.py
# ---------------------------------------------------------------------------

def test_no_incrementality_methodology_literals_in_score_kernel():
    """The staleness window / IM-blend math belongs in the contract doc, not
    the scoring kernel. score.py may reference the seam's naming (the
    mandated comment names ``IM_applied``) but must not carry the staleness
    day-count or the confidence-banding formula itself."""
    text = (_REPO / "wppc" / "score.py").read_text(encoding="utf-8")
    assert "90" not in text
    assert "confidence_weight" not in text
    # The multiplier itself is never computed/applied in v1.
    assert "IM_applied *" not in text
    assert "* IM_applied" not in text

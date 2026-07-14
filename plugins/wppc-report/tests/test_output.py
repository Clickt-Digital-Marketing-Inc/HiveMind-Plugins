"""CLI-403 — HTML-first 3-format output: model + md + html + Run tab.

One model is built ONCE from the bundled sample with a PINNED ``generated`` (so
renders are byte-reproducible), then the three formats are checked to agree with
it — the renderers TEMPLATE the model, they never re-derive a number. Also covers
HTML self-containment, the animate=False zero-GSAP guarantee, decision-lens
correctness, the derived-weights row-source contract, and white-labelling.
"""

import json
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from wppc import charts as C
from wppc import io as wio
from wppc import render_html as RH
from wppc.model import (
    build_decay,
    build_decay_meta,
    build_model,
    build_run_meta,
    build_weights_snapshot,
    classify_decision,
)
from wppc.render_md import _chart_svgs, render_md
from wppc.report import write_report
from wppc.score import score
from wppc.weights import FUNNEL_STATES, derive_weights

_ROOT = Path(__file__).resolve().parents[1]
_SAMPLE_CSV = _ROOT / "sample_data" / "google_segments.sample.csv"
_SAMPLE_MAP = _ROOT / "config" / "mapping.google.sample.yaml"
PINNED_GENERATED = "2026-07-14T00:00:00+00:00"

# Report-tab column positions (1-indexed) — see report.REPORT_COLUMNS.
_COL = {"segment_id": 1, "wPPC+": 5, "MAR": 7, "stabilized": 8}


def _build_sample_model(generated=PINNED_GENERATED):
    mapping = wio.load_mapping(str(_SAMPLE_MAP), "google")
    df, currency = wio.load_segments(str(_SAMPLE_CSV), mapping, "google")
    reach = {s: float(df[s].sum()) for s in FUNNEL_STATES}
    weights = derive_weights(
        reach,
        purchases_total=float(df["purchase"].sum()),
        cm3_order=float(currency["CM3_order"]),
        repeat_rate=float(currency["repeat_rate"]),
        cm3_repeat=float(currency["CM3_repeat"]),
    )
    results = score(df, weights)
    run_meta = build_run_meta(results, weights, "google", generated=generated)
    model = build_model(results, weights, run_meta)
    return model, results, weights, run_meta


@pytest.fixture(scope="module")
def sample():
    return _build_sample_model()


def _seg(model, seg_id):
    return next(s for s in model["segments"] if s["segment_id"] == seg_id)


def _md_row(md, seg_id):
    """The markdown segments-table line for one segment id."""
    for line in md.splitlines():
        if line.startswith(f"| {seg_id} |"):
            return line
    raise AssertionError(f"segment row for {seg_id!r} not found in markdown")


def _html_data(html):
    """Parse the model back out of the HTML <script id="data"> block."""
    m = re.search(r'<script id="data" type="application/json">(.*?)</script>', html, re.S)
    assert m, "data script block not found"
    return json.loads(m.group(1))


# ---------------------------------------------------------------------------
# 1. Three-format agreement / no re-derivation.
# ---------------------------------------------------------------------------
def test_three_format_agreement(tmp_path, sample):
    model, results, weights, run_meta = sample
    md = render_md(model)
    html = RH.render_html(model, animate=True)
    html_model = _html_data(html)

    xlsx = tmp_path / "report.xlsx"
    write_report(str(xlsx), results, weights, "google", run_meta=run_meta)
    ws = load_workbook(str(xlsx))["Report"]
    xlsx_by_seg = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        xlsx_by_seg[row[0]] = row

    # The HTML carries the model verbatim — proves it re-derives nothing.
    assert html_model["segments"] == model["segments"]

    # A Scale and a Cut segment must agree across model, md and xlsx.
    for seg_id in ("kw_brand_core", "kw_info_how_to"):
        seg = _seg(model, seg_id)
        xrow = xlsx_by_seg[seg_id]

        # model <-> xlsx (both computed in Python; equal to the last digit).
        assert xrow[_COL["wPPC+"] - 1] == seg["wPPC+"]
        assert xrow[_COL["MAR"] - 1] == seg["MAR"]
        assert xrow[_COL["stabilized"] - 1] == seg["stabilized"]

        # model <-> md (the rendered cells carry the model's values).
        line = _md_row(md, seg_id)
        assert f"{seg['wPPC+']:,.0f}" in line
        assert f"{seg['MAR']:,.2f}" in line
        assert seg["decision"] in line


# ---------------------------------------------------------------------------
# 2. HTML self-containment (outside the two checksummed blob regions).
# ---------------------------------------------------------------------------
def _strip_blobs(html):
    html = re.sub(re.escape(RH.GSAP_BEGIN) + ".*?" + re.escape(RH.GSAP_END), "", html, flags=re.S)
    html = re.sub(re.escape(C.VENDOR_BEGIN) + ".*?" + re.escape(C.VENDOR_END), "", html, flags=re.S)
    return html


def test_html_self_contained_outside_blobs(sample):
    model = sample[0]
    for animate in (True, False):
        html = RH.render_html(model, animate=animate)
        stripped = _strip_blobs(html)
        m = re.search(r"https?://|<link|src=|cdn", stripped)
        assert m is None, f"external ref outside blobs (animate={animate}): {m.group(0)!r}"


# ---------------------------------------------------------------------------
# 3. animate=False strips GSAP entirely (not even the token).
# ---------------------------------------------------------------------------
def test_animate_false_zero_gsap(sample):
    model = sample[0]
    html = RH.render_html(model, animate=False)
    assert "gsap" not in html.lower()
    assert RH.GSAP_BEGIN not in html
    assert RH.GSAP_END not in html
    # The chart runtime survives — a static report still draws its charts.
    assert C.VENDOR_BEGIN in html


def test_animate_true_has_gsap(sample):
    model = sample[0]
    html = RH.render_html(model, animate=True)
    assert RH.GSAP_BEGIN in html and RH.GSAP_END in html
    assert "gsap" in html.lower()


# ---------------------------------------------------------------------------
# 4. Decision lens correctness.
# ---------------------------------------------------------------------------
def test_classify_decision_all_branches():
    assert classify_decision("Y", 10.0) == "Scale"      # stabilized, MAR > 0
    assert classify_decision("Y", -10.0) == "Cut"       # stabilized, MAR < 0
    assert classify_decision("Y", 0.0) == "Watch"       # stabilized, MAR == 0
    assert classify_decision("N", 10.0) == "Watch"      # not stabilized (even MAR > 0)
    assert classify_decision("N", -10.0) == "Watch"     # not stabilized
    # bool stabilized accepted too.
    assert classify_decision(True, 5.0) == "Scale"
    assert classify_decision(False, 5.0) == "Watch"


def test_decision_lens_examples_and_counts(sample):
    model = sample[0]

    scale = _seg(model, "kw_brand_core")
    assert scale["stabilized"] == "Y" and scale["MAR"] > 0 and scale["decision"] == "Scale"

    cut = _seg(model, "kw_info_how_to")
    assert cut["stabilized"] == "Y" and cut["MAR"] < 0 and cut["decision"] == "Cut"

    watch = _seg(model, "kw_obscure_longtail")
    assert watch["stabilized"] == "N" and watch["decision"] == "Watch"

    lens = model["decision_lens"]
    assert lens["scale"] + lens["cut"] + lens["watch"] == model["provenance"]["n_segments"]
    assert lens["scale"] + lens["cut"] + lens["watch"] == len(model["segments"])


# ---------------------------------------------------------------------------
# 5. derived_weights renders from the weight rows, not the segment rows.
# ---------------------------------------------------------------------------
def test_derived_weights_uses_weight_rows(sample):
    model = sample[0]
    # The model maps this chart to the weights_table row source.
    assert model["charts"]["row_source"]["derived_weights"] == "weights_table"

    dw_title = next(c["title"] for c in model["charts"]["declarations"]
                    if c["id"] == "derived_weights")
    svgs = dict(_chart_svgs(model))
    # A funnel-state axis label only appears if the chart drew the weight rows;
    # the segment rows carry no "purchase" state.
    assert "purchase" in svgs[dw_title]

    # Guard: the segment rows genuinely lack the funnel-state field.
    assert all("state" not in s for s in model["segments"])


# ---------------------------------------------------------------------------
# 6. White-label — no vendor name in either rendered format.
# ---------------------------------------------------------------------------
def test_white_label_no_vendor_name(sample):
    model = sample[0]
    md = render_md(model)
    assert "Clickt" not in md
    for animate in (True, False):
        html = RH.render_html(model, animate=animate)
        assert "Clickt" not in html


# ---------------------------------------------------------------------------
# Run tab: added only when run_meta is passed; existing tabs unchanged.
# ---------------------------------------------------------------------------
def test_run_tab_optional_and_additive(tmp_path, sample):
    model, results, weights, run_meta = sample

    p_none = tmp_path / "none.xlsx"
    p_run = tmp_path / "run.xlsx"
    write_report(str(p_none), results, weights, "google")
    write_report(str(p_run), results, weights, "google", run_meta=run_meta)

    wb0 = load_workbook(str(p_none))
    wb1 = load_workbook(str(p_run))
    assert wb0.sheetnames == ["Report", "Weights", "Charts"]
    assert wb1.sheetnames == ["Report", "Weights", "Charts", "Run"]

    def cells(ws):
        return [[c.value for c in row] for row in ws.iter_rows()]

    # The Report/Weights tabs are byte-unchanged by the extra Run sheet.
    assert cells(wb0["Report"]) == cells(wb1["Report"])
    assert cells(wb0["Weights"]) == cells(wb1["Weights"])

    run_kv = {row[0]: row[1] for row in wb1["Run"].iter_rows(min_row=4, values_only=True)}
    assert run_kv["Platform"] == "google"
    assert run_kv["Self-check"] == "PASS"
    assert run_kv["Segments"] == model["provenance"]["n_segments"]
    assert run_kv["Generated"] == PINNED_GENERATED
    assert run_kv["Decay"] == "not-run"
    assert run_kv["Incrementality"] == "not-provided"


def _score_sample():
    """Score the bundled sample -> (weights, results). Used by the end-to-end
    regression tests below that exercise paths the module fixture doesn't."""
    mapping = wio.load_mapping(str(_SAMPLE_MAP), "google")
    df, currency = wio.load_segments(str(_SAMPLE_CSV), mapping, "google")
    reach = {s: float(df[s].sum()) for s in FUNNEL_STATES}
    weights = derive_weights(
        reach,
        purchases_total=float(df["purchase"].sum()),
        cm3_order=float(currency["CM3_order"]),
        repeat_rate=float(currency["repeat_rate"]),
        cm3_repeat=float(currency["CM3_repeat"]),
    )
    return weights, score(df, weights)


def test_build_model_with_decay_frame_serializes_trend():
    """Regression (surfaced by CLI-404): build_model(..., decay=<real frame>) must
    serialize the decay rows without a NameError. model._decay_rows() reduces the
    trend via a module-local _text_or_none; borrowing it only from report.py made
    the first real --prior-input run crash. Guards that model.py stays
    self-contained."""
    weights, results = _score_sample()
    _, prior_results = _score_sample()  # same period as prior -> Flat trends
    decay = build_decay(results, prior_results, band=5.0)
    run_meta = build_run_meta(
        results, weights, "google", generated=PINNED_GENERATED,
        decay=build_decay_meta("computed", prior_input="prior.csv", band=5.0),
    )
    model = build_model(results, weights, run_meta, decay=decay)
    assert model["decay"]["status"] == "computed"
    rows = model["decay"]["rows"]
    assert len(rows) == len(model["segments"])
    # trend serialized as text (proves _decay_rows ran, no NameError).
    assert all(r["trend"] in ("Rising", "Flat", "Falling") for r in rows)


def test_run_tab_accepts_dict_weights_version(tmp_path):
    """Regression (surfaced by CLI-404): the Run tab must accept a run_meta whose
    weights_version is the FULL snapshot dict (build_run_meta's contract, asserted
    in test_drift), not only a scalar. openpyxl cannot store a dict in a cell, so
    the writer reduces it to the snapshot timestamp. Guards the first real
    end-to-end xlsx run."""
    weights, results = _score_sample()
    snapshot = build_weights_snapshot(weights, "google", timestamp=PINNED_GENERATED)
    assert isinstance(snapshot, dict)  # the value that used to crash the cell writer
    run_meta = build_run_meta(
        results, weights, "google", generated=PINNED_GENERATED, weights_version=snapshot,
    )
    out = tmp_path / "run_wv.xlsx"
    write_report(str(out), results, weights, "google", run_meta=run_meta)  # must not raise
    wb = load_workbook(str(out))
    run_kv = {row[0]: row[1] for row in wb["Run"].iter_rows(min_row=4, values_only=True)}
    # dict reduced to its timestamp identity (full snapshot lives in the sidecar).
    assert run_kv["Weights version"] == snapshot["timestamp"]

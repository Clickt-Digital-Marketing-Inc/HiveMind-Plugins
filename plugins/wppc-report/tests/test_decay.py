"""W4 two-period decay: per-segment wPPC+ movement (current vs prior).

Decay is strictly additive and default-off: without ``--prior-input`` nothing is
computed and the xlsx path is unchanged. The trend band (in wPPC+ points) is
resolved as data at the CLI layer (CLI > mapping > default 5.0), never a literal
in the scoring kernels. Boundary semantics: ``delta == ±band`` classifies as
"Flat" (the band is inclusive; strict ``>``/``<`` gate Rising/Falling).
"""

import json
import re
from pathlib import Path

import pandas as pd
from click.testing import CliRunner
from openpyxl import load_workbook

from wppc.cli import DEFAULT_DECAY_BAND, cli, resolve_decay_band
from wppc.model import build_decay, build_decay_meta, build_run_meta
from wppc.report import write_report
from wppc.score import score

from conftest import make_segments

_REPO = Path(__file__).resolve().parents[1]
_GOOGLE_CSV = _REPO / "sample_data" / "google_segments.sample.csv"
_GOOGLE_MAPPING = _REPO / "config" / "mapping.google.sample.yaml"

_BASE_HEADERS = [
    "segment_id", "clicks", "conversions", "wPPC", "wPPC+",
    "wPPC_shrunk", "MAR", "stabilized (Y/N)", "closing_ratio",
]


def _plus_frame(pairs):
    """Minimal results-like frame: build_decay reads only segment_id + wPPC+."""
    return pd.DataFrame(
        {"segment_id": [p[0] for p in pairs], "wPPC+": [p[1] for p in pairs]}
    )


def test_trend_classification_and_boundary():
    band = 5.0
    # deltas vs a flat prior=100: A +10 Rising, B -10 Falling, C +3 Flat,
    # D +5 boundary -> Flat (inclusive), E -5 boundary -> Flat.
    current = _plus_frame([("A", 110), ("B", 90), ("C", 103), ("D", 105), ("E", 95)])
    prior = _plus_frame([("A", 100), ("B", 100), ("C", 100), ("D", 100), ("E", 100)])

    decay = build_decay(current, prior, band).set_index("segment_id")
    assert decay.loc["A", "trend"] == "Rising"
    assert decay.loc["B", "trend"] == "Falling"
    assert decay.loc["C", "trend"] == "Flat"
    assert decay.loc["D", "trend"] == "Flat"   # delta == +band -> Flat
    assert decay.loc["E", "trend"] == "Flat"   # delta == -band -> Flat

    assert decay.loc["A", "wPPC+_prior"] == 100
    assert decay.loc["A", "wPPC+_delta"] == 10
    assert abs(decay.loc["A", "delta_pct"] - 0.10) < 1e-9
    assert decay.loc["B", "wPPC+_delta"] == -10


def test_absent_from_prior_is_all_none():
    current = _plus_frame([("A", 110), ("Z", 120)])
    prior = _plus_frame([("A", 100)])

    decay = build_decay(current, prior, 5.0).set_index("segment_id")
    assert pd.isna(decay.loc["Z", "wPPC+_prior"])
    assert pd.isna(decay.loc["Z", "wPPC+_delta"])
    assert pd.isna(decay.loc["Z", "delta_pct"])
    assert pd.isna(decay.loc["Z", "trend"])   # trend None -> never fabricated
    # The present segment is unaffected.
    assert decay.loc["A", "trend"] == "Rising"


def test_prior_zero_guards_delta_pct_to_none():
    current = _plus_frame([("P", 50)])
    prior = _plus_frame([("P", 0)])

    decay = build_decay(current, prior, 5.0).set_index("segment_id")
    assert decay.loc["P", "wPPC+_prior"] == 0
    assert decay.loc["P", "wPPC+_delta"] == 50
    assert pd.isna(decay.loc["P", "delta_pct"])   # (cur-prior)/0 -> None
    assert decay.loc["P", "trend"] == "Rising"    # delta 50 > band, still classified


def test_resolve_decay_band_cli_over_mapping_over_default():
    # CLI value wins over everything.
    assert resolve_decay_band(10.0, {"decay_band": 3.0}) == 10.0
    # Mapping value wins when no CLI value.
    assert resolve_decay_band(None, {"decay_band": 3.0}) == 3.0
    # Falls through to the built-in default (5.0).
    assert resolve_decay_band(None, {}) == DEFAULT_DECAY_BAND
    assert resolve_decay_band(None, None) == DEFAULT_DECAY_BAND
    assert DEFAULT_DECAY_BAND == 5.0


def test_no_decay_band_literal_in_kernels():
    """The band literal must not appear in the scoring kernel — it lives at the
    CLI layer (default/mapping) and is passed to build_decay as data. (Word-
    boundary match so score.py's own 25.0/250.0 constants don't false-positive.)"""
    text = (_REPO / "wppc" / "score.py").read_text(encoding="utf-8")
    assert re.search(r"(?<![\d.])5\.0(?!\d)", text) is None, \
        "score.py must not carry the decay-band literal"
    # score.py owns no decay/band concept at all.
    assert "decay" not in text.lower()


def test_build_run_meta_passes_through_decay(example_weights):
    """The decay meta lands in run-metadata via build_run_meta's decay kwarg; the
    default keeps the prior stubbed (None) behaviour. incrementality stays None."""
    df = make_segments([
        {"segment_id": "solo", "click": 100, "engagement": 60, "add_to_cart": 20,
         "initiate_checkout": 12, "purchase": 8, "repeats": 3},
    ])
    results = score(df, example_weights)
    decay_meta = build_decay_meta("computed", prior_input="prior.csv", band=5.0)

    meta = build_run_meta(results, example_weights, platform="google", decay=decay_meta)
    assert meta["decay"] == decay_meta
    assert meta["incrementality"] is None
    # JSON-serializable.
    json.dumps(meta["decay"])

    # Omitting the kwarg preserves the stubbed None behaviour (default-off).
    bare = build_run_meta(results, example_weights, platform="google")
    assert bare["decay"] is None


def _two_segment_results(example_weights):
    df = make_segments([
        {"segment_id": "a", "click": 100, "engagement": 60, "add_to_cart": 20,
         "initiate_checkout": 12, "purchase": 8, "repeats": 3},
        {"segment_id": "b", "click": 80, "engagement": 40, "add_to_cart": 10,
         "initiate_checkout": 6, "purchase": 4, "repeats": 1},
    ])
    return score(df, example_weights)


def test_xlsx_appends_decay_columns_at_10_to_13(tmp_path, example_weights):
    results = _two_segment_results(example_weights)
    prior = _two_segment_results(example_weights)   # identical -> all Flat
    decay = build_decay(results, prior, 5.0)

    out = tmp_path / "with_decay.xlsx"
    write_report(str(out), results, example_weights, "google", decay=decay)

    ws = load_workbook(out)["Report"]
    # Columns 1-9 are EXACTLY the pre-existing headers (locks no mid-table insert).
    assert [ws.cell(1, c).value for c in range(1, 10)] == _BASE_HEADERS
    # New columns strictly at 10-13.
    assert ws.cell(1, 10).value == "wPPC+_prior"
    assert ws.cell(1, 11).value == "wPPC+_delta"
    assert ws.cell(1, 12).value == "delta_pct"
    assert ws.cell(1, 13).value == "trend"
    # Identical periods -> every trend cell is "Flat", delta 0.
    assert ws.cell(2, 13).value == "Flat"
    assert ws.cell(2, 11).value == 0


def test_xlsx_without_decay_has_exactly_nine_headers(tmp_path, example_weights):
    results = _two_segment_results(example_weights)

    out = tmp_path / "no_decay.xlsx"
    write_report(str(out), results, example_weights, "google")

    ws = load_workbook(out)["Report"]
    assert [ws.cell(1, c).value for c in range(1, 10)] == _BASE_HEADERS
    # Col 10 header is None -> no decay columns written (byte-path unchanged).
    assert ws.cell(1, 10).value is None


def test_cli_with_prior_input_echoes_decay_summary(tmp_path):
    runner = CliRunner()
    out = tmp_path / "report.xlsx"
    result = runner.invoke(cli, [
        "report", "--platform", "google",
        "--input", str(_GOOGLE_CSV),
        "--mapping", str(_GOOGLE_MAPPING),
        "--output", str(out),
        "--prior-input", str(_GOOGLE_CSV),   # same file -> deltas 0 -> all Flat
    ])
    assert result.exit_code == 0, result.output
    assert "Decay vs prior" in result.output
    assert "Flat" in result.output
    assert "absent-from-prior" in result.output
    # Decay columns landed in the workbook.
    ws = load_workbook(out)["Report"]
    assert ws.cell(1, 13).value == "trend"


def test_cli_without_prior_input_emits_no_decay(tmp_path):
    runner = CliRunner()
    out = tmp_path / "report.xlsx"
    result = runner.invoke(cli, [
        "report", "--platform", "google",
        "--input", str(_GOOGLE_CSV),
        "--mapping", str(_GOOGLE_MAPPING),
        "--output", str(out),
    ])
    assert result.exit_code == 0, result.output
    assert "Decay" not in result.output
    # No decay columns in the workbook.
    ws = load_workbook(out)["Report"]
    assert ws.cell(1, 10).value is None

#!/usr/bin/env python3
"""
build_audit_xlsx.py — Meta Ads Audit workbook generator.

Reads a single audit-payload.json (produced by the meta-ads-audit skill after it pulls data from
the Meta Ads MCP) and writes a deterministic, formula-driven .xlsx that is both an auditor working
tool and a client-facing report.

Usage
-----
  python3 build_audit_xlsx.py --input audit-payload.json --output meta-audit-acme-2026-06-24.xlsx
  python3 build_audit_xlsx.py --check meta-audit-acme-2026-06-24.xlsx

Design notes
------------
* The Health Score and per-category scores are REAL Excel formulas (SUMPRODUCT over per-check
  weighted-score / weighted-base helper columns), so if the auditor overrides a Flag cell the whole
  workbook recalculates on open. N/A checks are excluded from both numerator and denominator.
* Only openpyxl is required (>= 3.0). No network, no other deps. Output is deterministic.
* --check is a STRUCTURAL gate (tabs present, named ranges resolve, no literal #REF! in formulas).
  It cannot evaluate formula *values* — open in Excel/LibreOffice for that (it says so).

Payload schema (see SKILL.md for the authoritative version):
{
  "meta": {account_id, account_name, business_model, currency, windows{structure,creative,trend},
           generated_for_date, auditor, out_of_scope[]},
  "category_weights": {category_name: number, ...},   # optional; defaults applied
  "checks": [{id, category, name, severity, flag, observed, expected, recommendation}, ...],
  "sections": {section_key: {"columns": [...], "rows": [[...], ...]}},   # optional raw evidence
  "findings": [{id, title, category, severity, evidence, recommendation,
                impact, confidence, ease}, ...],                          # ICE fields optional
  "kpis": [{metric, value, unit?, benchmark?, flag?, notes?}, ...]        # optional scorecard
}

Optional values-only report tabs (12_Concentration / 13_Creative_Signals) are added when
build() receives the concentration= / creative_signals= blocks computed from raw pull files
by concentration.py / creative_signals.py. They are deliberately NOT in EXPECTED_TABS, so
--check stays green for workbooks built with or without them.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    sys.stderr.write("ERROR: openpyxl is required. Install with: python3 -m pip install 'openpyxl>=3.0'\n")
    sys.exit(2)

# Scoring constants, controlled vocabularies and the seven-lever section spine are
# single-sourced in audit_model.py so the xlsx formulas, the markdown record and the
# HTML explorer can never disagree (parity is asserted in tests/test_audit.py).
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from audit_model import (  # noqa: E402
    SEVERITY_WEIGHTS, FLAG_SCORES, SEVERITY_IMPACT, GRADE_CUTOFFS,
    ROADMAP_BUCKETS, DEFAULT_ICE, SEV_CANON, FLAG_CANON,
    SECTIONS, SECTION_WEIGHTS,
)

# ---------------------------------------------------------------------------- config

# Default category weights derive from the audit_model lever weights. CO/Competitive
# Landscape is qualitative (weight 0) and stays OUT of the scored default map, exactly
# as before. Integral floats collapse to int so cell values stay byte-identical.
DEFAULT_CATEGORY_WEIGHTS = {
    cat: (int(SECTION_WEIGHTS[code]) if float(SECTION_WEIGHTS[code]).is_integer()
          else SECTION_WEIGHTS[code])
    for code, cat, _tab, _key in SECTIONS if code != "CO"
}

# Controlled vocabularies (imported from audit_model). Drift here silently distorts the
# Health Score, so we normalize on build and flag (not error) anything we cannot map.
ALLOWED_SEVERITIES = set(SEV_CANON.values())
ALLOWED_FLAGS = set(FLAG_CANON.values())

# (sheet_title, category_name, code, section_key) — derived from audit_model.SECTIONS
# (the canonical order); the CO lever renders as the unscored 07_Competitive tab.
ANALYSIS_TABS = [(tab, cat, code, key)
                 for code, cat, tab, key in SECTIONS if code != "CO"]
COMPETITIVE_TAB = next((tab, cat, key)
                       for code, cat, tab, key in SECTIONS if code == "CO")

EXPECTED_TABS = [
    "00_Audit_Scope", "01_Executive_Summary",
    "02_Data_Infrastructure", "03_Account_Architecture", "04_Budget_Pacing",
    "05_Attribution", "06_Creative_Performance", "07_Competitive",
    "08_Future_Proofing", "09_Findings_Log", "10_ICE_Roadmap", "11_Reference",
]

# Values-only report tabs added by build(..., concentration=, creative_signals=).
# Deliberately NOT in EXPECTED_TABS: check() must stay green with or without them.
OPTIONAL_TABS = ["12_Concentration", "13_Creative_Signals"]

CHECK_HEADERS = ["ID", "Check", "Severity", "Sev.Wt", "Flag", "Score",
                 "W-Score", "W-Base", "Observed", "Expected", "Recommendation"]
CHECK_FIRST_ROW = 4  # header at row 3, data starts row 4

# ------------------------------------------------------------------------- styling

NAVY = "FF1E293B"
SLATE = "FFE2E8F0"
HILITE = "FFFFF7CD"
GREEN = "FFC6EFCE"
AMBER = "FFFFEB9C"
RED = "FFFFC7CE"
GREY = "FFD9D9D9"
BORDER_CLR = "FFCBD5E1"

TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="FFFFFFFF")
HEAD_FONT = Font(name="Calibri", size=10, bold=True, color="FF0F172A")
BODY_FONT = Font(name="Calibri", size=10, color="FF0F172A")
MUTED_FONT = Font(name="Calibri", size=9, italic=True, color="FF64748B")
BIG_FONT = Font(name="Calibri", size=28, bold=True, color="FF0F172A")

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FILL = PatternFill("solid", fgColor=SLATE)
HILITE_FILL = PatternFill("solid", fgColor=HILITE)
_thin = Side(style="thin", color=BORDER_CLR)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

FMT_MONEY = "#,##0.00"
FMT_PCT = "0.0%"
FMT_CTR = "0.000%"   # all-click CTR is a small fraction; 0.1% granularity is too coarse
FMT_SCORE = "0.00"   # fatigue / saturation / frequency scores

TAB_COLORS = {  # color spine: setup / deliverable / analysis / post-audit / reference
    "00_Audit_Scope": "1E40AF", "01_Executive_Summary": "B91C1C",
    "02_Data_Infrastructure": "CA8A04", "03_Account_Architecture": "CA8A04",
    "04_Budget_Pacing": "CA8A04", "05_Attribution": "CA8A04",
    "06_Creative_Performance": "CA8A04", "07_Competitive": "CA8A04",
    "08_Future_Proofing": "CA8A04", "09_Findings_Log": "15803D",
    "10_ICE_Roadmap": "15803D", "11_Reference": "475569",
    "12_Concentration": "475569", "13_Creative_Signals": "475569",
}


def q(sheet):
    """Quote a sheet name for cross-sheet references (always safe)."""
    return "'%s'" % sheet.replace("'", "''")


# Excel formula strings are DERIVED from the audit_model constants so the workbook
# can never drift from the model/JS scoring. The rendered strings are byte-identical
# to the previous hand-written formulas (verified against a pre-change build).

def _sev_wt_formula(r):
    """Nested-IF mirror of SEVERITY_WEIGHTS over the Severity cell (column C)."""
    expr = "0"
    for sev, w in reversed(list(SEVERITY_WEIGHTS.items())):
        expr = 'IF(C%d="%s",%g,%s)' % (r, sev, w, expr)
    return "=" + expr


def _flag_score_formula(r):
    """Nested-IF mirror of FLAG_SCORES over the Flag cell (column E); N/A/blank
    display as "" (the W-Score/W-Base columns exclude them from the score)."""
    expr = '""'
    for flag, s in reversed(list(FLAG_SCORES.items())):
        expr = 'IF(E%d="%s",%g,%s)' % (r, flag, s, expr)
    return "=" + expr


def _grade_formula(cell):
    """Nested-IF mirror of GRADE_CUTOFFS over the Health Score cell."""
    expr = '"%s"' % GRADE_CUTOFFS[-1][1]
    for cutoff, letter in reversed(GRADE_CUTOFFS[:-1]):
        expr = 'IF(%s>=%g,"%s",%s)' % (cell, cutoff, letter, expr)
    return "=" + expr


def _bucket_formula(r):
    """Nested-IF mirror of ROADMAP_BUCKETS over the ICE Priority cell (column H)."""
    expr = '"%s"' % ROADMAP_BUCKETS[-1][1]
    for cutoff, label in reversed(ROADMAP_BUCKETS[:-1]):
        expr = 'IF(H%d>=%g,"%s",%s)' % (r, cutoff, label, expr)
    return "=" + expr


def add_name(wb, name, ref):
    dn = DefinedName(name, attr_text=ref)
    try:
        wb.defined_names[name] = dn          # openpyxl >= 3.1 (dict-like)
    except (TypeError, AttributeError):
        wb.defined_names.add(dn)             # openpyxl 3.0 (list-like)


def defined_name_set(wb):
    try:
        return set(wb.defined_names.keys())                       # >= 3.1
    except AttributeError:
        return {dn.name for dn in wb.defined_names.definedName}   # 3.0


def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = CTR
        c.border = BORDER


def set_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


def _section_caption(ws, row, text, ncols):
    """Merged section caption band (used by the values-only report tabs)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    return row + 1


def _kv(ws, row, label, value, fmt=None):
    """Label/value line (values only). Value overflows to the right — neighbors empty."""
    a = ws.cell(row=row, column=1, value=label)
    a.font = HEAD_FONT
    b = ws.cell(row=row, column=2, value=value)
    b.font = BODY_FONT
    b.alignment = LEFT
    if fmt:
        b.number_format = fmt
    return row + 1


# --------------------------------------------------------------------- tab builders

def build_scope(wb, meta):
    ws = wb.create_sheet("00_Audit_Scope")
    title_row(ws, "Meta Ads Account Audit — Scope & Provenance", 4)
    set_widths(ws, [26, 60])
    rows = [
        ("Business model", meta.get("business_model", "")),
        ("Account ID", meta.get("account_id", "")),
        ("Account name", meta.get("account_name", "")),
        ("Currency", meta.get("currency", "")),
        ("Structure window", meta.get("windows", {}).get("structure", "last_30d")),
        ("Creative window", meta.get("windows", {}).get("creative", "last_90d")),
        ("Trend window", meta.get("windows", {}).get("trend", "")),
        ("Generated for", meta.get("generated_for_date", "")),
        ("Auditor", meta.get("auditor", "")),
        ("Framework", "Comprehensive Meta Ads account audit"),
        ("Data source", "Meta Ads MCP (live pull) — see 11_Reference for tool provenance"),
    ]
    r = 3
    for label, val in rows:
        a = ws.cell(row=r, column=1, value=label); a.font = HEAD_FONT; a.border = BORDER
        b = ws.cell(row=r, column=2, value=val); b.font = BODY_FONT; b.alignment = WRAP; b.border = BORDER
        r += 1
    add_name(wb, "business_model", "%s!$B$3" % q("00_Audit_Scope"))

    r += 1
    h = ws.cell(row=r, column=1, value="Out of scope (not measurable via the Meta MCP — assess separately):")
    h.font = HEAD_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    oos = meta.get("out_of_scope") or [
        "Landing-page / CRO (load speed, headline match, friction, heatmaps, post-click CVR/bounce)",
        "Testing discipline (calendar, 10%/10x mix, sample-size & duration guardrails)",
        "Business/goal (ICP, customer value tiers, margin by SKU, payback, seasonality/promo calendar)",
        "Incrementality economics (MER, NC-ROAS, new-customer revenue, 1DV-vs-7DC conversion splits)",
        "Thumb-stop (3s) rate and relevance/quality rankings (not exposed by this MCP)",
    ]
    for item in oos:
        c = ws.cell(row=r, column=1, value="•  " + item); c.font = MUTED_FONT; c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
    ws.sheet_view.showGridLines = False
    return ws


def build_analysis_tab(wb, sheet_title, category, code, checks, section):
    ws = wb.create_sheet(sheet_title)
    title_row(ws, "%s" % category, len(CHECK_HEADERS))
    set_widths(ws, [10, 34, 10, 8, 9, 7, 9, 9, 40, 34, 44])
    header_row(ws, 3, CHECK_HEADERS)

    r = CHECK_FIRST_ROW
    if not checks:
        ws.cell(row=r, column=1, value="(no checks supplied)").font = MUTED_FONT
        last = r
    else:
        for chk in checks:
            ws.cell(row=r, column=1, value=chk.get("id", "")).font = BODY_FONT
            ws.cell(row=r, column=2, value=chk.get("name", "")).font = BODY_FONT
            ws.cell(row=r, column=3, value=chk.get("severity", "")).font = BODY_FONT
            # D Sev.Wt (formula from severity, derived from SEVERITY_WEIGHTS)
            ws.cell(row=r, column=4, value=_sev_wt_formula(r))
            ws.cell(row=r, column=5, value=chk.get("flag", "")).font = BODY_FONT  # E Flag
            # F Score (display, derived from FLAG_SCORES)
            ws.cell(row=r, column=6, value=_flag_score_formula(r))
            # G W-Score, H W-Base (numeric, N/A & blank excluded)
            ws.cell(row=r, column=7, value='=IF(OR(E{r}="N/A",E{r}=""),0,F{r}*D{r})'.format(r=r))
            ws.cell(row=r, column=8, value='=IF(OR(E{r}="N/A",E{r}=""),0,D{r})'.format(r=r))
            ws.cell(row=r, column=9, value=chk.get("observed", "")).font = BODY_FONT
            ws.cell(row=r, column=10, value=chk.get("expected", "")).font = BODY_FONT
            ws.cell(row=r, column=11, value=chk.get("recommendation", "")).font = BODY_FONT
            for col in range(1, len(CHECK_HEADERS) + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER
                cell.alignment = WRAP if col in (2, 9, 10, 11) else CTR
            r += 1
        last = r - 1

    # named ranges over the W-Score / W-Base columns (exact check span)
    span_last = max(last, CHECK_FIRST_ROW)
    add_name(wb, "wscore_%s" % code, "%s!$G$%d:$G$%d" % (q(sheet_title), CHECK_FIRST_ROW, span_last))
    add_name(wb, "wbase_%s" % code, "%s!$H$%d:$H$%d" % (q(sheet_title), CHECK_FIRST_ROW, span_last))

    # Flag dropdown + conditional formatting
    flag_range = "E%d:E%d" % (CHECK_FIRST_ROW, span_last)
    dv = DataValidation(type="list", formula1='"PASS,FLAG,FAIL,N/A"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(flag_range)
    ws.conditional_formatting.add(flag_range, CellIsRule(operator="equal", formula=['"PASS"'],
                                  fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(flag_range, CellIsRule(operator="equal", formula=['"FLAG"'],
                                  fill=PatternFill("solid", fgColor=AMBER)))
    ws.conditional_formatting.add(flag_range, CellIsRule(operator="equal", formula=['"FAIL"'],
                                  fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(flag_range, CellIsRule(operator="equal", formula=['"N/A"'],
                                  fill=PatternFill("solid", fgColor=GREY)))

    # optional raw evidence table to the RIGHT (col M=13) so it never collides with G/H named ranges
    _render_raw_block(ws, section, start_col=13)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def _render_raw_block(ws, section, start_col):
    if not section or not section.get("columns"):
        return
    cap = ws.cell(row=3, column=start_col, value="Evidence / raw pull")
    cap.font = HEAD_FONT
    cols = section["columns"]
    header_row(ws, 4, cols, start_col=start_col)  # put header at row 4 (below caption)
    for i in range(len(cols)):
        ws.column_dimensions[get_column_letter(start_col + i)].width = 20
    rr = 5
    for row in section.get("rows", []):
        for i, val in enumerate(row):
            c = ws.cell(row=rr, column=start_col + i, value=val)
            c.font = BODY_FONT; c.border = BORDER; c.alignment = LEFT
        rr += 1


def build_competitive(wb, section):
    sheet_title, category, _ = COMPETITIVE_TAB
    ws = wb.create_sheet(sheet_title)
    title_row(ws, "%s (qualitative — not scored)" % category, 6)
    if section and section.get("columns"):
        cols = section["columns"]
        header_row(ws, 3, cols)
        set_widths(ws, [max(16, len(c)) for c in cols])
        r = 4
        for row in section.get("rows", []):
            for i, val in enumerate(row):
                c = ws.cell(row=r, column=1 + i, value=val)
                c.font = BODY_FONT; c.border = BORDER; c.alignment = WRAP
            r += 1
    else:
        ws.cell(row=3, column=1,
                value="No competitor set supplied. Provide competitor page names/search terms + "
                      "country and re-run ads_library_search.").font = MUTED_FONT
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def build_exec_summary(wb, weights, findings, kpis=None):
    ws = wb.create_sheet("01_Executive_Summary")
    title_row(ws, "Executive Summary — Account Health", 6)
    set_widths(ws, [30, 12, 10, 12, 11, 30])

    ws.cell(row=3, column=1, value="Health Score").font = HEAD_FONT
    ws.cell(row=4, column=1, value="Grade").font = HEAD_FONT

    # scorecard table — category rows derive from audit_model.SECTIONS (CO unscored)
    sc_head = ["Category", "Weight", "Base", "Score(0-100)", "Included"]
    header_row(ws, 6, sc_head)
    cats = [(cat, code) for code, cat, _tab, _key in SECTIONS if code != "CO"]
    first, r = 7, 7
    for name, code in cats:
        ws.cell(row=r, column=1, value=name).font = BODY_FONT
        ws.cell(row=r, column=2, value=weights.get(name, DEFAULT_CATEGORY_WEIGHTS.get(name, 0)))
        ws.cell(row=r, column=3, value="=IFERROR(SUM(wbase_%s),0)" % code)
        ws.cell(row=r, column=4, value="=IFERROR(SUM(wscore_%s)/C%d*100,0)" % (code, r))
        ws.cell(row=r, column=5, value="=IF(C%d>0,1,0)" % r)
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
            if col != 1:
                ws.cell(row=r, column=col).alignment = CTR
        r += 1
    last = r - 1
    add_name(wb, "category_weights", "%s!$B$%d:$B$%d" % (q("01_Executive_Summary"), first, last))

    # health score + grade formulas
    hs = ws.cell(row=3, column=2,
                 value="=IFERROR(SUMPRODUCT($D${f}:$D${l},$B${f}:$B${l},$E${f}:$E${l})/"
                       "SUMPRODUCT($B${f}:$B${l},$E${f}:$E${l}),0)".format(f=first, l=last))
    hs.font = BIG_FONT; hs.alignment = CTR; hs.fill = HILITE_FILL; hs.number_format = "0.0"
    gr = ws.cell(row=4, column=2, value=_grade_formula("B3"))  # from GRADE_CUTOFFS
    gr.font = BIG_FONT; gr.alignment = CTR; gr.fill = HILITE_FILL

    # optional KPI scorecard (values only, informational — same row schema as the
    # md/html renderers: metric/value/unit/benchmark/flag/notes)
    r = last + 2
    if kpis:
        cap = ws.cell(row=r, column=1, value="KPI scorecard (informational)")
        cap.font = HEAD_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        header_row(ws, r, ["Metric", "Value", "Unit", "Benchmark", "Flag", "Notes"])
        r += 1
        for k in kpis:
            ws.cell(row=r, column=1, value=k.get("metric", ""))
            ws.cell(row=r, column=2, value=k.get("value", ""))
            ws.cell(row=r, column=3, value=k.get("unit", ""))
            ws.cell(row=r, column=4, value=k.get("benchmark", ""))
            ws.cell(row=r, column=5, value=k.get("flag", ""))
            ws.cell(row=r, column=6, value=k.get("notes", ""))
            for col in range(1, 7):
                c = ws.cell(row=r, column=col)
                c.font = BODY_FONT
                c.border = BORDER
                c.alignment = WRAP if col in (1, 6) else CTR
            r += 1
        r += 1

    # top quick wins (static, computed in Python by ICE priority)
    qw_row = r
    cap = ws.cell(row=qw_row, column=1, value="Top priorities (by ICE — see 10_ICE_Roadmap)")
    cap.font = HEAD_FONT
    ws.merge_cells(start_row=qw_row, start_column=1, end_row=qw_row, end_column=6)
    ranked = sorted(
        findings or [],
        key=lambda f: _ice_priority(f), reverse=True,
    )[:5]
    rr = qw_row + 1
    for f in ranked:
        line = "[%s] %s — %s (ICE %d)" % (
            f.get("severity", ""), f.get("title", ""),
            f.get("recommendation", ""), _ice_priority(f))
        c = ws.cell(row=rr, column=1, value=line); c.font = BODY_FONT; c.alignment = WRAP
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
        rr += 1
    if not ranked:
        ws.cell(row=rr, column=1, value="(no findings logged)").font = MUTED_FONT

    ws.sheet_view.showGridLines = False
    return ws


def _ice_num(v):
    """Numeric ICE value or None (missing/blank/non-numeric). Mirrors
    audit_model._ice_num: integral floats collapse to int so cells stay clean."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        num = v
    else:
        s = str(v).strip() if v is not None else ""
        if not s:
            return None
        try:
            num = float(s)
        except ValueError:
            return None
    if isinstance(num, float) and num.is_integer():
        return int(num)
    return num


def _ice_fields(f):
    """(impact, confidence, ease) as NUMBERS. Missing/blank/non-numeric fields take
    the audit_model defaults (impact <- SEVERITY_IMPACT[severity], confidence/ease <-
    DEFAULT_ICE) so the =E*F*G Priority formulas on 10_ICE_Roadmap never hit #VALUE!."""
    impact = _ice_num(f.get("impact"))
    if impact is None:
        sev = SEV_CANON.get(str(f.get("severity", "") or "").strip().lower())
        impact = SEVERITY_IMPACT.get(sev, DEFAULT_ICE)
    confidence = _ice_num(f.get("confidence"))
    if confidence is None:
        confidence = DEFAULT_ICE
    ease = _ice_num(f.get("ease"))
    if ease is None:
        ease = DEFAULT_ICE
    return impact, confidence, ease


def _ice_priority(f):
    impact, confidence, ease = _ice_fields(f)
    return impact * confidence * ease


def build_findings(wb, findings):
    ws = wb.create_sheet("09_Findings_Log")
    title_row(ws, "Findings Log", 6)
    heads = ["ID", "Title", "Category", "Severity", "Evidence", "Recommendation"]
    header_row(ws, 3, heads)
    set_widths(ws, [10, 32, 24, 10, 46, 46])
    r = 4
    for f in findings or []:
        ws.cell(row=r, column=1, value=f.get("id", ""))
        ws.cell(row=r, column=2, value=f.get("title", ""))
        ws.cell(row=r, column=3, value=f.get("category", ""))
        ws.cell(row=r, column=4, value=f.get("severity", ""))
        ws.cell(row=r, column=5, value=f.get("evidence", ""))
        ws.cell(row=r, column=6, value=f.get("recommendation", ""))
        for col in range(1, 7):
            c = ws.cell(row=r, column=col); c.font = BODY_FONT; c.border = BORDER
            c.alignment = WRAP if col in (2, 5, 6) else CTR
        r += 1
    if not findings:
        ws.cell(row=4, column=1, value="(no findings logged)").font = MUTED_FONT
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def build_ice(wb, findings):
    ws = wb.create_sheet("10_ICE_Roadmap")
    title_row(ws, "ICE Prioritization & Roadmap", 9)
    heads = ["ID", "Title", "Category", "Severity", "Impact", "Confidence",
             "Ease", "Priority", "Bucket"]
    header_row(ws, 3, heads)
    set_widths(ws, [10, 34, 22, 10, 9, 11, 8, 10, 12])
    r = 4
    for f in findings or []:
        impact, confidence, ease = _ice_fields(f)  # numeric seeds — never #VALUE!
        ws.cell(row=r, column=1, value=f.get("id", ""))
        ws.cell(row=r, column=2, value=f.get("title", ""))
        ws.cell(row=r, column=3, value=f.get("category", ""))
        ws.cell(row=r, column=4, value=f.get("severity", ""))
        ws.cell(row=r, column=5, value=impact)
        ws.cell(row=r, column=6, value=confidence)
        ws.cell(row=r, column=7, value=ease)
        ws.cell(row=r, column=8, value="=E{r}*F{r}*G{r}".format(r=r))
        ws.cell(row=r, column=9, value=_bucket_formula(r))  # from ROADMAP_BUCKETS
        for col in range(1, 10):
            c = ws.cell(row=r, column=col); c.font = BODY_FONT; c.border = BORDER
            c.alignment = WRAP if col in (2, 3) else CTR
        r += 1
    if not findings:
        ws.cell(row=4, column=1, value="(no findings logged)").font = MUTED_FONT
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def build_reference(wb):
    ws = wb.create_sheet("11_Reference")
    title_row(ws, "Reference — Thresholds, Glossary & Provenance", 2)
    set_widths(ws, [40, 70])
    lines = [
        ("Health Score", "Σ(category_score×weight×included) / Σ(weight×included). N/A checks excluded."),
        ("Flag → numeric", "PASS=1.0, FLAG=0.5, FAIL=0.0, N/A=excluded."),
        ("Severity weights", "Critical 5.0 · High 3.0 · Medium 1.5 · Low 0.5."),
        ("Grade bands", "A ≥90 · B 75-89 · C 60-74 · D 40-59 · F <40."),
        ("ICE", "Priority = Impact×Confidence×Ease (1-1000). Quick win: Ease≥8 & Impact≥7."),
        ("Roadmap buckets", "30-day ≥500 · 60-day 250-499 · 90-day 100-249 · Parking lot <100."),
        ("Top-3 spend", "Healthy ≥60% of spend in top 3 campaigns (fragmentation guard)."),
        ("Learning", "≥25 optimization events / ad set / 30d (results). <25 = starved."),
        ("Prospecting:Retargeting", "≈80:20; flag if retargeting >50% of spend."),
        ("CTR-Link", "Ecom prospecting ≥0.8-1.0%. Derived from cost_per_action_type[link_click]."),
        ("ThruPlay / hold-through", "Thumb-stop(3s) unavailable; ThruPlay(15s)/impr + P100/P25 substituted."),
        ("Frequency", "Prospecting <3 healthy; >5 fatigue."),
        ("EMQ", "Purchase EMQ ≥8 healthy; <6 fail."),
        ("CAPI/dedup", "SERVER_ONLY event volume present alongside WEB_ONLY (ads_get_dataset_stats)."),
        ("Attribution", "Prefer 7d_click; flag heavy 1d_view reliance (attribution_setting)."),
        ("Provenance", "ads_get_ad_entities, ads_get_dataset_quality/stats, ads_get_creatives, "
                       "ads_get_ad_account_custom_audiences, ads_insights_*, ads_library_search, "
                       "ads_get_opportunity_score."),
        ("Recalculation note", "Scores are live formulas. Override any Flag cell and the workbook "
                               "recomputes on open in Excel/LibreOffice."),
    ]
    r = 3
    for k, v in lines:
        a = ws.cell(row=r, column=1, value=k); a.font = HEAD_FONT; a.border = BORDER; a.alignment = WRAP
        b = ws.cell(row=r, column=2, value=v); b.font = BODY_FONT; b.border = BORDER; b.alignment = WRAP
        r += 1
    ws.sheet_view.showGridLines = False
    return ws


def build_concentration_tab(wb, block):
    """Values-only concentration report (HHI / Effective-N / Gini / Pareto-ABC).

    Layout ported from google-ads-audit generate_workbook.build_concentration_tab.
    Informational — never scored, never in EXPECTED_TABS (check() must stay green
    for workbooks built without raw pull files). No formulas: every cell is a value
    computed by concentration.compute_concentration."""
    ws = wb.create_sheet("12_Concentration")
    title_row(ws, "Concentration — spend vs conversions (HHI)", 6)
    set_widths(ws, [44, 14, 14, 12, 12, 8])
    r = 3
    c = ws.cell(row=r, column=1, value=(
        "HHI bands (merger-guideline cutoffs): <1,500 unconcentrated · 1,500-2,500 "
        "moderate · >2,500 high. Small dimensions: read Effective-N instead."))
    c.font = MUTED_FONT
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2
    block = block or {}
    for dim in block.get("dimensions", []) or []:
        label = dim.get("label", "")
        if dim.get("window"):
            label = "%s (%s)" % (label, dim["window"])
        r = _section_caption(ws, r, label, 6)
        r = _kv(ws, r, "Verdict", dim.get("verdict", ""))
        for side, side_label in (("spend", "Spend"), ("conv", "Conversions")):
            m = dim.get(side)
            if m:
                r = _kv(ws, r, side_label,
                        "HHI {hhi:,.1f} ({band}) · Effective-N {eff_n} · Gini {gini}".format(
                            hhi=m.get("hhi", 0.0), band=m.get("band", ""),
                            eff_n=m.get("eff_n", ""), gini=m.get("gini", "")))
            else:
                r = _kv(ws, r, side_label, "no signal in this window")
        r = _kv(ws, r, "Entities", "%d (from %d raw rows)"
                % (dim.get("n_entities", 0), dim.get("n_rows_raw", 0)))
        if dim.get("caveat"):
            r = _kv(ws, r, "Caveat", dim["caveat"])
        r += 1
        header_row(ws, r, ["Entity", "Spend", "Conv", "Spend %", "Conv %", "ABC"])
        r += 1
        for t in dim.get("top", []) or []:
            ws.cell(row=r, column=1, value=t.get("name", "")).alignment = WRAP
            ws.cell(row=r, column=2, value=t.get("spend")).number_format = FMT_MONEY
            ws.cell(row=r, column=3, value=t.get("conv"))
            ws.cell(row=r, column=4, value=t.get("spend_share")).number_format = FMT_PCT
            ws.cell(row=r, column=5, value=t.get("conv_share")).number_format = FMT_PCT
            ws.cell(row=r, column=6, value=t.get("abc"))
            for col in range(1, 7):
                cell = ws.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.border = BORDER
            r += 1
        tail = dim.get("tail")
        if tail:
            c = ws.cell(row=r, column=1, value="… plus %d more (tail)" % tail.get("n", 0))
            c.font = MUTED_FONT
            ws.cell(row=r, column=2, value=tail.get("spend")).number_format = FMT_MONEY
            ws.cell(row=r, column=3, value=tail.get("conv"))
            ws.cell(row=r, column=4, value=tail.get("spend_share")).number_format = FMT_PCT
            r += 1
        r += 1
    for note in block.get("notes", []) or []:
        c = ws.cell(row=r, column=1, value="Note: %s" % note)
        c.font = MUTED_FONT
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    ws.sheet_view.showGridLines = False
    return ws


def build_creative_signals_tab(wb, block):
    """Values-only Creative Signals report (fatigue / reach saturation / effective-
    frequency zones / ranking decomposition) from creative_signals.compute_creative_signals.

    Informational — never scored, never in EXPECTED_TABS (check() must stay green for
    workbooks built without raw pull files). Tail aggregates carry NO reach/frequency
    (non-additive across ads). No formulas: every cell is a precomputed value."""
    ws = wb.create_sheet("13_Creative_Signals")
    ncols = 12
    title_row(ws, "Creative Signals — fatigue, saturation & effective frequency", ncols)
    set_widths(ws, [36, 11, 12, 11, 8, 9, 10, 10, 18, 9, 10, 11])
    block = block or {}
    r = 3

    # account baselines + fatigue-band summary
    label = "Account baselines"
    if block.get("window"):
        label = "%s (%s)" % (label, block["window"])
    r = _section_caption(ws, r, label, ncols)
    base = block.get("baselines") or {}
    r = _kv(ws, r, "CTR (all-click)", base.get("ctr"), fmt=FMT_CTR)
    r = _kv(ws, r, "CPM", base.get("cpm"), fmt=FMT_MONEY)
    r = _kv(ws, r, "Ads in baseline", base.get("n_ads"))
    summary = block.get("summary") or {}
    if summary:
        r = _kv(ws, r, "Fatigue bands (ads)",
                "saturated %d · watch %d · fresh %d · below floor %d · high saturation %d"
                % (summary.get("saturated", 0), summary.get("watch", 0),
                   summary.get("fresh", 0), summary.get("below_floor", 0),
                   summary.get("high_saturation", 0)))
    r += 1

    # per-ad fatigue / saturation table (top spenders; tail aggregated below)
    r = _section_caption(ws, r, "Ads — fatigue & reach saturation (top spenders)", ncols)
    header_row(ws, r, ["Ad", "Spend", "Impressions", "Reach", "Freq", "CTR", "CPM",
                       "Results", "Indicator", "Fatigue", "Band", "Saturation"])
    r += 1
    for a in block.get("ads", []) or []:
        ws.cell(row=r, column=1, value=a.get("name", "")).alignment = WRAP
        ws.cell(row=r, column=2, value=a.get("spend")).number_format = FMT_MONEY
        ws.cell(row=r, column=3, value=a.get("impressions"))
        ws.cell(row=r, column=4, value=a.get("reach"))
        ws.cell(row=r, column=5, value=a.get("frequency")).number_format = FMT_SCORE
        ws.cell(row=r, column=6, value=a.get("ctr")).number_format = FMT_CTR
        ws.cell(row=r, column=7, value=a.get("cpm")).number_format = FMT_MONEY
        ws.cell(row=r, column=8, value=a.get("results"))
        ws.cell(row=r, column=9, value=a.get("results_indicator"))
        ws.cell(row=r, column=10, value=a.get("fatigue")).number_format = FMT_SCORE
        ws.cell(row=r, column=11, value=a.get("fatigue_band"))
        ws.cell(row=r, column=12, value=a.get("saturation")).number_format = FMT_SCORE
        for col in range(1, ncols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = BODY_FONT
            cell.border = BORDER
        r += 1
    tail = block.get("tail")
    if tail:
        share = tail.get("spend_share")
        note = "… plus %d more (tail) — reach/frequency omitted (non-additive)" % tail.get("n", 0)
        if share is not None:
            note = ("… plus %d more (tail; %.1f%% of spend) — reach/frequency omitted "
                    "(non-additive)" % (tail.get("n", 0), share * 100.0))
        c = ws.cell(row=r, column=1, value=note)
        c.font = MUTED_FONT
        ws.cell(row=r, column=2, value=tail.get("spend")).number_format = FMT_MONEY
        ws.cell(row=r, column=3, value=tail.get("impressions"))
        r += 1
    r += 1

    # effective-frequency zones — per AD SET (frequency is level-native there)
    zones = block.get("zones") or {}
    r = _section_caption(
        ws, r, "Effective-frequency zones — per ad set (<3 under · 3-7 effective · >7 oversaturated)",
        ncols)
    r = _kv(ws, r, "Zone counts", "under %d · effective %d · oversaturated %d"
            % (zones.get("under", 0), zones.get("effective", 0),
               zones.get("oversaturated", 0)))
    zrows = zones.get("rows") or []
    if zrows:
        header_row(ws, r, ["Ad set", "Frequency", "Zone", "Spend"])
        r += 1
        for z in zrows:
            ws.cell(row=r, column=1, value=z.get("name", "")).alignment = WRAP
            ws.cell(row=r, column=2, value=z.get("frequency")).number_format = FMT_SCORE
            ws.cell(row=r, column=3, value=z.get("zone"))
            ws.cell(row=r, column=4, value=z.get("spend")).number_format = FMT_MONEY
            for col in range(1, 5):
                cell = ws.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.border = BORDER
            r += 1
    r += 1

    # ranking decomposition — available on the manual CSV path only
    rankings = block.get("rankings") or {}
    r = _section_caption(ws, r, "Ranking decomposition (quality / engagement / conversion)", ncols)
    if rankings.get("available"):
        rsum = rankings.get("summary") or {}
        weak = rsum.get("weakest") or {}
        r = _kv(ws, r, "Weakest-lever counts",
                "quality %d · engagement %d · conversion %d (of %d ranked; %d all-unknown)"
                % (weak.get("quality", 0), weak.get("engagement", 0),
                   weak.get("conversion", 0), rsum.get("n_ranked", 0),
                   rsum.get("all_unknown", 0)))
        header_row(ws, r, ["Ad", "Spend", "Quality", "Engagement", "Conversion",
                           "Weakest", "Known", "Fix order"])
        r += 1
        for row in rankings.get("rows", []) or []:
            ws.cell(row=r, column=1, value=row.get("name", "")).alignment = WRAP
            ws.cell(row=r, column=2, value=row.get("spend")).number_format = FMT_MONEY
            ws.cell(row=r, column=3, value=row.get("quality"))
            ws.cell(row=r, column=4, value=row.get("engagement"))
            ws.cell(row=r, column=5, value=row.get("conversion"))
            ws.cell(row=r, column=6, value=row.get("weakest"))
            ws.cell(row=r, column=7, value=row.get("known_count"))
            ws.cell(row=r, column=8, value=", ".join(row.get("priority") or []))
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.border = BORDER
            r += 1
    else:
        c = ws.cell(row=r, column=1,
                    value="Rankings not available in this input — quality/engagement/"
                          "conversion rankings unlock on the manual CSV export path.")
        c.font = MUTED_FONT
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1
    r += 1

    for note in block.get("notes", []) or []:
        c = ws.cell(row=r, column=1, value="Note: %s" % note)
        c.font = MUTED_FONT
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1
    ws.sheet_view.showGridLines = False
    return ws


# ------------------------------------------------------------------------- assembly

def validate_and_normalize(checks, weights):
    """Normalize severity/flag casing in place; return human-readable warnings for vocab drift.

    Drift is reported, not fatal — but it WILL change the score (unknown severity → weight 0;
    unknown/empty flag → excluded; category that matches no scored bucket → dropped from all tabs).
    """
    warns = []
    weight_keys = set(weights)
    for chk in checks:
        cid = chk.get("id", "?")
        sev = str(chk.get("severity", "")).strip()
        canon = SEV_CANON.get(sev.lower())
        if canon:
            chk["severity"] = canon
        elif sev:
            warns.append("check %s: unknown severity %r → Sev.Wt 0 (not scored)" % (cid, sev))
        else:
            warns.append("check %s: missing severity → Sev.Wt 0 (not scored)" % cid)
        flag = str(chk.get("flag", "")).strip()
        fcanon = FLAG_CANON.get(flag.lower())
        if fcanon:
            chk["flag"] = fcanon
        elif flag:
            warns.append("check %s: unknown flag %r → excluded from score" % (cid, flag))
        else:
            warns.append("check %s: empty flag → excluded from score" % cid)
        cat = chk.get("category", "")
        if cat not in weight_keys:
            warns.append("check %s: category %r matches no scored category → "
                         "will not appear on any analysis tab" % (cid, cat))
    return warns


def build(payload, *, concentration=None, creative_signals=None):
    """Payload dict -> Workbook. concentration / creative_signals are the optional
    values-only report blocks (from concentration.py / creative_signals.py, computed
    from raw pull files — never from the payload); each adds one OPTIONAL_TABS sheet."""
    meta = payload.get("meta", {})
    weights = {**DEFAULT_CATEGORY_WEIGHTS, **(payload.get("category_weights") or {})}
    checks = payload.get("checks", []) or []
    sections = payload.get("sections", {}) or {}
    findings = payload.get("findings", []) or []

    for w in validate_and_normalize(checks, weights):
        print("WARN: %s" % w)

    by_cat = {}
    for chk in checks:
        by_cat.setdefault(chk.get("category", ""), []).append(chk)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet; we create all explicitly in order

    build_scope(wb, meta)
    build_exec_summary(wb, weights, findings, payload.get("kpis") or [])
    for sheet_title, category, code, section_key in ANALYSIS_TABS:
        build_analysis_tab(wb, sheet_title, category, code,
                           by_cat.get(category, []), sections.get(section_key))
        if sheet_title == "06_Creative_Performance":
            build_competitive(wb, sections.get(COMPETITIVE_TAB[2]))
    build_findings(wb, findings)
    build_ice(wb, findings)
    build_reference(wb)
    if concentration:
        build_concentration_tab(wb, concentration)
    if creative_signals:
        build_creative_signals_tab(wb, creative_signals)

    # numeric order + tab colors
    order = {name: i for i, name in enumerate(EXPECTED_TABS + OPTIONAL_TABS)}
    wb._sheets.sort(key=lambda s: order.get(s.title, 99))
    for ws in wb.worksheets:
        if ws.title in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS[ws.title]
    wb.active = wb.sheetnames.index("01_Executive_Summary")
    return wb


# ---------------------------------------------------------------------------- check

def check(path):
    errors, warns = [], []
    try:
        wb = load_workbook(path)
    except Exception as e:  # noqa: BLE001
        print("FAIL: could not open workbook: %s" % e)
        return 1

    for t in EXPECTED_TABS:
        if t not in wb.sheetnames:
            errors.append("missing tab: %s" % t)

    names = defined_name_set(wb)
    required_names = {"business_model", "category_weights"}
    for _, _, code, _ in ANALYSIS_TABS:
        required_names.add("wscore_%s" % code)
        required_names.add("wbase_%s" % code)
    for n in sorted(required_names):
        if n not in names:
            errors.append("missing named range: %s" % n)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and "#REF!" in v:
                    errors.append("#REF! in %s!%s" % (ws.title, cell.coordinate))

    # Semantic vocab scan on the analysis tabs (col C = Severity, col E = Flag).
    # Out-of-vocab values silently shift the Health Score, so surface them as warnings.
    analysis_titles = {t[0] for t in ANALYSIS_TABS}
    for ws in wb.worksheets:
        if ws.title not in analysis_titles:
            continue
        r = CHECK_FIRST_ROW
        while True:
            cid = ws.cell(row=r, column=1).value
            if cid in (None, ""):
                break
            sev = ws.cell(row=r, column=3).value
            flag = ws.cell(row=r, column=5).value
            if sev and str(sev) not in ALLOWED_SEVERITIES:
                warns.append("%s!C%d severity %r not in %s" % (ws.title, r, sev, sorted(ALLOWED_SEVERITIES)))
            if flag and str(flag) not in ALLOWED_FLAGS:
                warns.append("%s!E%d flag %r not in %s" % (ws.title, r, flag, sorted(ALLOWED_FLAGS)))
            r += 1

    print("Workbook: %s" % path)
    print("Tabs: %d/%d present" % (len([t for t in EXPECTED_TABS if t in wb.sheetnames]), len(EXPECTED_TABS)))
    print("Named ranges: %d/%d present" % (len([n for n in required_names if n in names]), len(required_names)))
    if warns:
        for w in warns:
            print("WARN: %s" % w)
    if errors:
        for e in errors:
            print("ERROR: %s" % e)
        print("RESULT: FAIL (%d errors)" % len(errors))
        return 1
    print("RESULT: PASS (structural). Note: open in Excel/LibreOffice to verify computed values.")
    return 0


# ----------------------------------------------------------------------------- main

def main(argv=None):
    ap = argparse.ArgumentParser(description="Meta Ads Audit workbook generator.")
    ap.add_argument("--input", help="audit-payload.json")
    ap.add_argument("--output", help="output .xlsx path")
    ap.add_argument("--check", dest="check_path", help="structural-check an existing .xlsx")
    args = ap.parse_args(argv)

    if args.check_path:
        return check(args.check_path)

    if not args.input or not args.output:
        ap.error("provide --input and --output (or --check <file>)")

    with open(args.input, "r", encoding="utf-8") as fh:
        payload = json.load(fh)
    wb = build(payload)
    wb.save(args.output)
    print("Wrote %s" % args.output)
    return 0


if __name__ == "__main__":
    sys.exit(main())

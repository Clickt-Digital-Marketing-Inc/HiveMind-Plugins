#!/usr/bin/env python3
# Copyright (c) 2026 Clickt Digital Marketing Inc. All rights reserved.
"""Conformance tests for the meta-ads-audit pipeline. Stdlib only
(openpyxl needed for the xlsx sections only — they skip gracefully without it).

Run: python3 tests/test_audit.py

Guards: lever-WEIGHTED Health-Score correctness (hand oracles), N/A exclusion,
JS<->Python<->Excel constant parity, self-containment (GSAP sentinel +
checksum), determinism, concentration metrics, bounded embeds, creative
signals (MediaMetrics mirrors), meta_rows raw-shape parsing, the manual UI-CSV
path, the deterministic pre-scorer, merge semantics, and — section 13 — a
DIFFERENTIAL parity sweep that recalculates the workbook through LibreOffice
and asserts its health+grade equal the model's. Constant parity proves the
kernels share inputs; only the differential sweep proves they agree on output.
Section 13 skips (never fails) when soffice is unavailable.
"""
from __future__ import annotations

import copy
import hashlib
import json
import math
import random
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent / "scripts"
FIX = HERE / "fixtures"
sys.path.insert(0, str(SCRIPTS))

import audit_model
import audit_html
import audit_md
import concentration as conc
import creative_signals as cs
import manual_csv as mc
import meta_rows as mr
import prescore as ps

EXAMPLE = SCRIPTS / "example_payload.json"
PAYLOAD = json.loads(EXAMPLE.read_text())
MICRO = json.loads((FIX / "micro_payload.json").read_text())
GROUPS = json.loads((FIX / "prescore_rows.json").read_text())

FAILS = []
SKIPS = []
N_CHECKS = 0


def check(name, cond, detail=""):
    global N_CHECKS
    N_CHECKS += 1
    print(("  ok   " if cond else "FAIL   ") + name + ("" if cond else "   :: " + detail))
    if not cond:
        FAILS.append(name)


def skip_note(msg):
    print("  skip " + msg)
    SKIPS.append(msg)


def approx(a, b, tol=1e-9):
    return a is not None and abs(a - b) <= tol


# ============================================================================
# 1: model oracle — example_payload.json (full 40-check-ID coverage)
#
# Hand arithmetic (SEVERITY_WEIGHTS C5/H3/M1.5/L.5; PASS 1 / FLAG .5 / FAIL 0):
#   DI: P(C5)+F(C0)+P(H3)+FL(C2.5)+P(M1.5)+FL(M.75)+P(H3) = 15.75 / 24  -> 65.625
#   AR: FL(H1.5)+F(C0)+P(H3)+P(C5)+FL(H1.5)+P(M1.5)+FL(M.75)+FL(L.25)
#                                                       = 13.50 / 22.5 -> 60.0
#   BP: FL(M.75)+F(H0)+P(M1.5)+P(M1.5)+FL(M.75)         =  4.50 /  9   -> 50.0
#   AT: P(M1.5)+FL(H1.5)+P(M1.5)+FL(M.75)+N/A           =  5.25 /  7.5 -> 70.0
#   CR: N/A+FL(M.75)+P(M1.5)+FL(H1.5)+F(H0)+F(H0)+FL(H1.5)+P(L.5)+FL(M.75)
#                                                       =  6.50 / 17   -> 38.235294
#   CO: two N/A -> possible 0 -> EXCLUDED (weight 0 anyway)
#   FP: FL(M.75)+P(H3)+P(M1.5)+N/A                      =  5.25 /  6   -> 87.5
#   health = (65.625*20 + 60*20 + 50*15 + 70*10 + 38.235294*25 + 87.5*10) / 100
#          = 5793.38235 / 100 = 57.9338 -> 57.9 (grade D)
#   totals: earned 50.75, possible 86.0; counts 15 pass / 15 flag / 5 fail / 5 n/a
# ============================================================================
print("--- 1: model oracle (example payload, weighted health) ---")
model = audit_model.compute_model(PAYLOAD, generated="2026-07-10T00:00:00")
check("health score == 57.9 (hand oracle)", model["health"]["score"] == 57.9,
      str(model["health"]))
check("grade == D", model["health"]["grade"] == "D")
check("health.weighted is True", model["health"]["weighted"] is True)
check("earned/possible == 50.75 / 86.0",
      model["health"]["earned"] == 50.75 and model["health"]["possible"] == 86.0,
      str(model["health"]))
check("no-loss: 40 checks", model["meta"]["n_checks"] == 40, str(model["meta"]["n_checks"]))
check("no-loss: 5 findings", model["meta"]["n_findings"] == 5)
check("counts 15/15/5/5",
      (model["summary"]["n_pass"], model["summary"]["n_flag"],
       model["summary"]["n_fail"], model["summary"]["n_na"]) == (15, 15, 5, 5),
      str(model["summary"]))

ids = [c["id"] for s in model["sections"] for c in s["checks"]]
check("40 unique check IDs bucketed", len(ids) == 40 and len(set(ids)) == 40)
prefix_counts = {}
for i in ids:
    prefix_counts[i.split("-")[0]] = prefix_counts.get(i.split("-")[0], 0) + 1
check("lever ID coverage DI7/AR8/BP5/AT5/CR9/CO2/FP4",
      prefix_counts == {"DI": 7, "AR": 8, "BP": 5, "AT": 5, "CR": 9, "CO": 2, "FP": 4},
      str(prefix_counts))
check("no unknown-category warnings", model["warnings"] == [], str(model["warnings"]))

secs = {s["code"]: s for s in model["sections"]}
check("sections in canonical order",
      [s["code"] for s in model["sections"]] == ["DI", "AR", "BP", "AT", "CR", "CO", "FP"])
check("section scores 65.6/60.0/50.0/70.0/38.2/87.5",
      (secs["DI"]["score_pct"], secs["AR"]["score_pct"], secs["BP"]["score_pct"],
       secs["AT"]["score_pct"], secs["CR"]["score_pct"], secs["FP"]["score_pct"])
      == (65.6, 60.0, 50.0, 70.0, 38.2, 87.5),
      str({k: s["score_pct"] for k, s in secs.items()}))
na = [c for s in model["sections"] for c in s["checks"] if c["result"] == "N/A"]
check("5 N/A checks excluded (earned & possible None)",
      len(na) == 5 and all(c["earned"] is None and c["possible"] is None for c in na),
      f"{len(na)} n/a rows")
check("whole-lever-N/A CO excluded: possible 0, not included, score None",
      secs["CO"]["possible"] == 0.0 and secs["CO"]["included"] is False
      and secs["CO"]["score_pct"] is None, str(secs["CO"]["possible"]))
check("CO weight is 0", secs["CO"]["weight"] == 0.0)
check("evidence passthrough on architecture section",
      secs["AR"]["evidence"] and secs["AR"]["evidence"]["columns"][0] == "Campaign")
check("evidence None when not supplied", secs["AT"]["evidence"] is None)

# CO weight-0 exercise: scoring the CO checks must NOT move the health score
# (weight 0 contributes nothing to numerator or denominator).
co_scored = copy.deepcopy(PAYLOAD)
for c in co_scored["checks"]:
    if c["id"].startswith("CO-"):
        c["flag"] = "PASS"
m_co = audit_model.compute_model(co_scored, generated="2026-07-10T00:00:00")
check("CO scored PASS at weight 0 leaves health at 57.9",
      m_co["health"]["score"] == 57.9, str(m_co["health"]))

# category_weights override: zero out Creative Performance ->
# health = (65.625*20 + 60*20 + 50*15 + 70*10 + 87.5*10) / 75 = 4837.5/75 = 64.5
ovr = copy.deepcopy(PAYLOAD)
ovr["category_weights"] = {"Creative Performance": 0}
m_ovr = audit_model.compute_model(ovr, generated="2026-07-10T00:00:00")
check("category_weights override flows through (CR->0 gives 64.5 / C)",
      m_ovr["health"]["score"] == 64.5 and m_ovr["health"]["grade"] == "C",
      str(m_ovr["health"]))
check("unknown category_weights key warned",
      any("unknown category" in w for w in audit_model.compute_model(
          {**copy.deepcopy(PAYLOAD), "category_weights": {"Nope": 5}},
          generated="x")["warnings"]))

# findings: ICE defaults + priority + roadmap bucket
# F-101 9*9*6=486 -> 60-day; F-102 Critical defaults 9*5*5=225 -> 90-day;
# F-103 10*10*10=1000 -> 30-day; F-104 High 7*5*5=175 -> 90-day;
# F-105 Low 3*5*5=75 -> Parking lot
fnd = {f["id"]: f for f in model["findings"]}
check("ICE explicit F-101: 486 / 60-day",
      fnd["F-101"]["priority"] == 486 and fnd["F-101"]["bucket"] == "60-day",
      str(fnd["F-101"]))
check("ICE defaults F-102 (Critical): 9*5*5=225 / 90-day",
      (fnd["F-102"]["impact"], fnd["F-102"]["confidence"], fnd["F-102"]["ease"],
       fnd["F-102"]["priority"], fnd["F-102"]["bucket"]) == (9, 5, 5, 225, "90-day"),
      str(fnd["F-102"]))
check("ICE F-103: 1000 / 30-day", fnd["F-103"]["priority"] == 1000
      and fnd["F-103"]["bucket"] == "30-day")
check("ICE defaults F-104 (High): 175 / 90-day",
      fnd["F-104"]["priority"] == 175 and fnd["F-104"]["bucket"] == "90-day")
check("ICE defaults F-105 (Low): 75 / Parking lot",
      fnd["F-105"]["priority"] == 75 and fnd["F-105"]["bucket"] == "Parking lot")
check("findings_by_bucket tallies",
      model["summary"]["findings_by_bucket"]
      == {"30-day": 1, "60-day": 1, "90-day": 2, "Parking lot": 1},
      str(model["summary"]["findings_by_bucket"]))
check("stem() example", audit_model.stem(PAYLOAD)
      == "meta-audit_example-outfitters-sample_2026-07-10",
      audit_model.stem(PAYLOAD))

# ============================================================================
# 2: micro-oracle — the weighted formula demonstrably differs from flat
#
# Micro fixture: DI-01 Critical PASS (5/5 -> 100, w20); AT-05 High N/A
# (lever excluded); CR-03 Medium FLAG (.75/1.5) + CR-06 High FAIL (0/3)
# -> CR 0.75/4.5 = 16.6667, w25.
#   weighted = (100*20 + 16.6667*25) / 45 = 2416.667/45 = 53.7037 -> 53.7 (D)
#   flat     = 5.75 / 9.5 * 100 = 60.526 -> 60.5  (differs!)
# ============================================================================
print("--- 2: micro-oracle (weighted vs flat) ---")
m_micro = audit_model.compute_model(MICRO, generated="2026-07-10T00:00:00")
check("micro weighted health == 53.7 / D",
      m_micro["health"]["score"] == 53.7 and m_micro["health"]["grade"] == "D",
      str(m_micro["health"]))
flat = round(m_micro["health"]["earned"] / m_micro["health"]["possible"] * 100, 1)
check("flat formula would say 60.5 (earned 5.75 / possible 9.5)",
      flat == 60.5 and m_micro["health"]["earned"] == 5.75
      and m_micro["health"]["possible"] == 9.5, str(m_micro["health"]))
check("weighted differs from flat", m_micro["health"]["score"] != flat)
at_sec = [s for s in m_micro["sections"] if s["code"] == "AT"][0]
check("whole-lever-N/A AT excluded (possible 0, included False)",
      at_sec["possible"] == 0.0 and at_sec["included"] is False)
check("stem() micro", audit_model.stem(MICRO)
      == "meta-audit_micro-oracle-co-sample_2026-07-10", audit_model.stem(MICRO))

# ============================================================================
# 3: three-way parity — JS kernel <-> audit_model <-> xlsx constants
# ============================================================================
print("--- 3: three-way constant parity ---")
tpl = audit_html._TEMPLATE


def js_map(name):
    m = re.search(name + r"\s*=\s*\{([^}]*)\}", tpl)
    out = {}
    for part in m.group(1).split(","):
        k, v = part.split(":")
        out[k.strip()] = float(v.strip())
    return out


# The HTML mirrors ONLY what it recomputes live (the ICE re-rank). Everything
# else is displayed from the Python model, so there is no second implementation
# to assert. Mirroring a constant the browser never reads is worse than not
# mirroring it: the old SEV_W/FLAG/SECT_W/GRADES assertions all passed while
# guarding a healthOf() no caller invoked, which is why a live rounding
# divergence in sectionPct went unnoticed. Assert the live ones, and assert
# that the dead kernel stays dead.
check("JS IMPACT == audit_model.SEVERITY_IMPACT", js_map("IMPACT") == audit_model.SEVERITY_IMPACT)
js_buckets = re.search(r"var BUCKETS\s*=\s*\[(.*?)\];", tpl)
js_bkt = [(int(n), lbl) for n, lbl in
          re.findall(r"\[(\d+),'([^']+)'\]", js_buckets.group(1))]
check("JS BUCKETS == audit_model.ROADMAP_BUCKETS",
      js_bkt == [tuple(x) for x in audit_model.ROADMAP_BUCKETS], str(js_bkt))
check("HTML recomputes NO score: the dead JS health kernel is gone",
      not any(sym in tpl for sym in ("healthOf", "scoreCheck", "gradeOf",
                                     "SECT_W", "GRADES", "SEV_W")))
check("sectionPct displays the model's score_pct (no JS re-rounding)",
      "s.score_pct" in tpl and "Math.round(e/p" not in tpl)
check("Findings Horizon recomputes from the LIVE ICE (bucketVal), not f.bucket",
      "bucketVal(r)" in tpl and "esc(r.f.bucket" not in tpl)
# The ICE seed must not re-default what Python already filled: `||5` treats a
# legitimate confidence/ease of 0 as falsy and substitutes 5, so the row showed
# an ICE (and, once Horizon followed the live value, a bucket) that Python's
# priority/bucket contradicted. Guard the falsy-zero form specifically.
check("ICE seeds from the model without a second, disagreeing default",
      "iceSeed(f.confidence)" in tpl and "iceSeed(f.ease)" in tpl
      and "+f.confidence||5" not in tpl and "+f.ease||5" not in tpl
      and "(r.f.impact||0)" not in tpl)
check("no xmlns on inline SVG (template)", "xmlns" not in tpl)
# Competitive Landscape is structurally weight-0: the workbook builds no CO row,
# so a nonzero CO weight is a number ONLY compute_model could honour — health
# would move in md/html while the xlsx silently disagreed. (Found by the
# section-13 differential sweep, not by constant parity.)
_co = audit_model.compute_model({
    "meta": {"account_name": "CO"},
    "category_weights": {"Competitive Landscape": 40.0, "Creative Performance": 60.0},
    "checks": [{"id": "CO-01", "category": "Competitive Landscape", "name": "a",
                "severity": "Critical", "flag": "FAIL"},
               {"id": "CR-02", "category": "Creative Performance", "name": "b",
                "severity": "Critical", "flag": "PASS"}]})
check("a Competitive Landscape weight override is ignored (pinned to 0), not honoured",
      _co["health"]["score"] == 100.0
      and [s for s in _co["sections"] if s["code"] == "CO"][0]["weight"] == 0.0,
      str(_co["health"]))
check("...and the ignored override is warned about, not silently dropped",
      any("is qualitative" in w for w in _co["warnings"]), str(_co["warnings"]))
# "CO is special" was spelled out in six places across two files; it now has ONE
# declaration that both the model and the workbook read.
check("UNSCORED_SECTIONS is the single source: SECTION_WEIGHTS agrees with it",
      all(audit_model.SECTION_WEIGHTS[c] == 0.0
          for c in audit_model.UNSCORED_SECTIONS)
      and audit_model.UNSCORED_SECTIONS == frozenset({"CO"}))
# The set is plural because compute_model's weight rule is; the workbook is not
# — it renders exactly ONE unscored lever as the Competitive tab. A second would
# vanish from the workbook while md/html still listed it, so the mismatch has to
# fail loudly rather than let next() pick a winner.
check("the workbook refuses to import if UNSCORED_SECTIONS stops being exactly one",
      len([1 for code, _c, _t, _k in audit_model.SECTIONS
           if code in audit_model.UNSCORED_SECTIONS]) == 1)

# round1 — the kernel adopts Excel's half-away rule. Python's round() is
# half-even, so a .x5 boundary printed one number in the model and another in
# the workbook. Ported from google-ads-audit (same defect, one decimal over).
check("round1 rounds half UP (Excel/JS rule), unlike Python's banker's round()",
      audit_model.round1(62.25) == 62.3 and round(62.25, 1) == 62.2
      and audit_model.round1(6.25) == 6.3 and audit_model.round1(11.25) == 11.3,
      str(audit_model.round1(62.25)))
check("round1 leaves non-boundary values alone",
      audit_model.round1(59.9901) == 60.0 and audit_model.round1(0.0) == 0.0
      and audit_model.round1(43.74) == 43.7)
_tie = audit_model.compute_model({
    "meta": {"account_name": "Tie"},
    "category_weights": {"Creative Performance": 62.25, "Attribution": 37.75},
    "checks": [{"id": "CR-02", "category": "Creative Performance", "name": "a",
                "severity": "Critical", "flag": "PASS"},
               {"id": "AT-02", "category": "Attribution", "name": "b",
                "severity": "Critical", "flag": "FAIL"}]})
check("health on an exact .x5 tie uses Excel's rule (62.3, not banker's 62.2)",
      _tie["health"]["score"] == 62.3, str(_tie["health"]["score"]))
# earned = H FLAG(3×.5=1.5) + M FLAG(1.5×.5=.75) = 2.25
# possible = 3 + 1.5 + 5+5+5 (C FAIL ×3) + 0.5 (L FAIL) = 20  ->  11.25 exactly
_pct = audit_model.compute_model({
    "meta": {"account_name": "Pct"},
    "checks": [{"id": "CR-05", "category": "Creative Performance", "name": "b",
                "severity": "High", "flag": "FLAG"},
               {"id": "CR-02", "category": "Creative Performance", "name": "a",
                "severity": "Medium", "flag": "FLAG"},
               {"id": "CR-03", "category": "Creative Performance", "name": "c",
                "severity": "Critical", "flag": "FAIL"},
               {"id": "CR-06", "category": "Creative Performance", "name": "d",
                "severity": "Critical", "flag": "FAIL"},
               {"id": "CR-07", "category": "Creative Performance", "name": "e",
                "severity": "Critical", "flag": "FAIL"},
               {"id": "CR-08", "category": "Creative Performance", "name": "f",
                "severity": "Low", "flag": "FAIL"}]})
_cr = [s for s in _pct["sections"] if s["code"] == "CR"][0]
check("lever score_pct on a .x5 tie (2.25/20 = 11.25) is 11.3, matching Excel",
      _cr["earned"] == 2.25 and _cr["possible"] == 20.0
      and _cr["score_pct"] == 11.3, str(_cr["score_pct"]))

try:
    import build_audit_xlsx as bax
    check("xlsx SEVERITY_WEIGHTS IS audit_model's (identity)",
          bax.SEVERITY_WEIGHTS is audit_model.SEVERITY_WEIGHTS)
    check("xlsx FLAG_SCORES IS audit_model's", bax.FLAG_SCORES is audit_model.FLAG_SCORES)
    check("xlsx SEVERITY_IMPACT IS audit_model's", bax.SEVERITY_IMPACT is audit_model.SEVERITY_IMPACT)
    check("xlsx GRADE_CUTOFFS IS audit_model's", bax.GRADE_CUTOFFS is audit_model.GRADE_CUTOFFS)
    check("xlsx SECTIONS IS audit_model's", bax.SECTIONS is audit_model.SECTIONS)
    check("xlsx ANALYSIS_TABS derived from SECTIONS (CO excluded)",
          [t[0] for t in bax.ANALYSIS_TABS]
          == [tab for code, _c, tab, _k in audit_model.SECTIONS if code != "CO"]
          and bax.COMPETITIVE_TAB[0] == "07_Competitive")
    check("xlsx DEFAULT_CATEGORY_WEIGHTS derived from SECTION_WEIGHTS",
          bax.DEFAULT_CATEGORY_WEIGHTS
          == {cat: int(audit_model.SECTION_WEIGHTS[code])
              for code, cat, _t, _k in audit_model.SECTIONS if code != "CO"},
          str(bax.DEFAULT_CATEGORY_WEIGHTS))
except SystemExit:
    bax = None
    skip_note("xlsx constant-parity checks (openpyxl unavailable)")

# ============================================================================
# 4: self-containment — GSAP sentinel excise-by-checksum, then grep
# ============================================================================
print("--- 4: self-containment ---")
gsap_bytes = (SCRIPTS / "vendor" / "gsap.min.js").read_bytes()
want = (SCRIPTS / "vendor" / "SHA256SUMS").read_text().split()[0]
check("vendored GSAP matches SHA256SUMS",
      hashlib.sha256(gsap_bytes).hexdigest() == want)

html = audit_html.render_html(model, animate=True)
blob = audit_html.gsap_blob()
check("GSAP blob embedded between sentinels", blob in html)
stripped = html.replace(blob, "")
hits = re.findall(r"https?://|<link|src=|cdn", stripped)
check("no external refs outside the GSAP sentinels", not hits, str(hits[:4]))

h_noanim = audit_html.render_html(model, animate=False)
check("animate=False carries zero GSAP bytes",
      audit_html.GSAP_BEGIN not in h_noanim and "GreenSock" not in h_noanim)
check("animate=False still self-contained",
      not re.findall(r"https?://|<link|src=|cdn", h_noanim))
check("html without blocks embeds nulls",
      '"concentration":null' in html and '"prescore":null' in html
      and '"creative_signals":null' in html)

# ============================================================================
# 5: determinism — pure function of the payload except meta.generated
# ============================================================================
print("--- 5: determinism ---")
a = audit_html.render_html(audit_model.compute_model(PAYLOAD, generated="TS1"))
b = audit_html.render_html(audit_model.compute_model(PAYLOAD, generated="TS2"))
check("HTML deterministic modulo generated", a.replace("TS1", "TS") == b.replace("TS2", "TS"))
ma = audit_md.render_md(audit_model.compute_model(PAYLOAD, generated="TS1"))
mb = audit_md.render_md(audit_model.compute_model(PAYLOAD, generated="TS2"))
check("markdown deterministic modulo generated", ma.replace("TS1", "TS") == mb.replace("TS2", "TS"))
j1 = json.dumps(audit_model.compute_model(PAYLOAD, generated="TS"), sort_keys=True)
j2 = json.dumps(audit_model.compute_model(PAYLOAD, generated="TS"), sort_keys=True)
check("model JSON deterministic (same generated)", j1 == j2)

# ============================================================================
# 6: concentration — metric hand-oracles + assembly
#
# spend6 [50,20,10,10,5,5]: HHI (.5²+.2²+.1²+.1²+.05²+.05²)*10000 = 3150;
#   eff_n 1/.315 = 3.17; gini (2*485)/(6*100) - 7/6 = 0.45; ABC AAABBC.
# conv6 [30,0,5,0,1,0]: HHI ((30/36)²+(5/36)²+(1/36)²)*10000 = 7145.1;
#   eff_n 1.40; gini 2*209/(6*36) - 7/6 = 0.769; ABC crossing-inclusive ACBCCC.
# ============================================================================
print("--- 6: concentration ---")
spend6 = [50, 20, 10, 10, 5, 5]
conv6 = [30, 0, 5, 0, 1, 0]
check("hhi(spend6) == 3150.0", approx(conc.hhi(spend6), 3150.0), str(conc.hhi(spend6)))
check("effective_n(spend6) == 3.17", round(conc.effective_n(spend6), 2) == 3.17)
check("gini(spend6) == 0.45", round(conc.gini(spend6), 3) == 0.45, str(conc.gini(spend6)))
check("abc(spend6) == AAABBC", conc.pareto_abc(spend6) == ["A", "A", "A", "B", "B", "C"],
      str(conc.pareto_abc(spend6)))
check("hhi(conv6) == 7145.1", round(conc.hhi(conv6), 1) == 7145.1, str(conc.hhi(conv6)))
check("effective_n(conv6) == 1.40", round(conc.effective_n(conv6), 2) == 1.40)
check("gini(conv6) == 0.769", round(conc.gini(conv6), 3) == 0.769, str(conc.gini(conv6)))
check("abc(conv6) crossing-inclusive == ACBCCC",
      conc.pareto_abc(conv6) == ["A", "C", "B", "C", "C", "C"], str(conc.pareto_abc(conv6)))

check("verdict fragility", conc.verdict(3000, 3000)[0] == "fragility")
check("verdict consolidate", conc.verdict(1000, 3000)[0] == "consolidate")
check("verdict review_bidding", conc.verdict(3000, 1000)[0] == "review_bidding")
check("verdict diversified", conc.verdict(1000, 1000)[0] == "diversified")
check("verdict no_conv_signal", conc.verdict(3000, None)[0] == "no_conv_signal")
check("verdict insufficient", conc.verdict(None, None)[0] == "insufficient")
check("equal k=4: hhi 2500 / band moderate / eff_n 4 / gini 0",
      conc.hhi([1, 1, 1, 1]) == 2500.0 and conc.hhi_band(2500.0) == "moderate"
      and conc.effective_n([1, 1, 1, 1]) == 4.0 and conc.gini([1, 1, 1, 1]) == 0.0)
check("band boundary 1500 -> moderate", conc.hhi_band(1500.0) == "moderate")
check("single entity: hhi 10000 / eff_n 1 / gini 0 / lorenz / abc A",
      conc.hhi([7]) == 10000.0 and conc.effective_n([7]) == 1.0 and conc.gini([7]) == 0.0
      and conc.lorenz_points([7]) == [[0.0, 0.0], [1.0, 1.0]]
      and conc.pareto_abc([7]) == ["A"])

# Entity identity: Meta names are NOT unique (reusing "Broad"/"LAL 1%" across
# campaigns, or one creative across ad sets, is the standard workflow). Keying
# by name merged distinct entities and inflated EVERY downstream metric — 4
# evenly-split ads read as 2, doubling HHI (2500 -> 5000), halving Effective-N
# (4 -> 2) and flipping the verdict (diversified -> fragility).
_dupe_ads = [{"id": str(i), "name": ("Broad", "Broad", "LAL 1%", "LAL 1%")[i - 1],
              "spend": 25.0, "conv_results": 5.0,
              "date_start": "2026-06-01", "date_stop": "2026-06-30"}
             for i in (1, 2, 3, 4)]
_d_ads = [d for d in conc.compute_concentration(ad_rows=_dupe_ads)["dimensions"]
          if d["key"] == "ads"][0]
check("duplicate NAMES with distinct ids stay 4 entities (keyed by id)",
      _d_ads["n_entities"] == 4 and _d_ads["n_rows_raw"] == 4, str(_d_ads["n_entities"]))
check("dupe-name oracle: HHI 2500 / eff-N 4 / diversified (not 5000 / 2 / fragility)",
      _d_ads["spend"]["hhi"] == 2500.0 and _d_ads["spend"]["eff_n"] == 4.0
      and _d_ads["verdict_key"] == "diversified", str(_d_ads["spend"]))
check("shared names are called out in the caveat",
      "shared by more than one entity" in (_d_ads["caveat"] or ""), str(_d_ads["caveat"]))
# The caveat describes what the TABLE shows: entities dropped by the all-zero
# filter must not be described (they are not on screen to explain).
_drop = [{"id": "1", "name": "Broad", "spend": 0.0, "conv_results": 0.0},
         {"id": "2", "name": "Broad", "spend": 0.0, "conv_results": 0.0},
         {"id": "3", "name": "Solo", "spend": 50.0, "conv_results": 10.0}]
_d_drop = [d for d in conc.compute_concentration(ad_rows=_drop)["dimensions"]
           if d["key"] == "ads"][0]
check("a shared name whose entities were ALL dropped (all-zero) is not described",
      _d_drop["n_entities"] == 1
      and "shared by more than one entity" not in (_d_drop["caveat"] or ""),
      str(_d_drop["caveat"]))
# split/merged are properties of a NAME, not of the dimension: one id-less row
# used to flip the whole report to "rows were merged" when none had been, and
# hid the genuine shared-name case behind it.
_mix = [{"id": "1", "name": "Broad", "spend": 25.0, "conv_results": 5.0},
        {"id": "2", "name": "Broad", "spend": 25.0, "conv_results": 5.0},
        {"name": "Solo", "spend": 50.0, "conv_results": 10.0}]
_d_mix = [d for d in conc.compute_concentration(ad_rows=_mix)["dimensions"]
          if d["key"] == "ads"][0]
check("mixed id presence: nothing merged, so the caveat reports the SPLIT only",
      _d_mix["n_entities"] == 3
      and "shared by more than one entity" in (_d_mix["caveat"] or "")
      and "carry no id" not in (_d_mix["caveat"] or ""), str(_d_mix["caveat"]))
check("entity label is the NAME, never the opaque id",
      {t["name"] for t in _d_ads["top"]} == {"Broad", "LAL 1%"},
      str([t["name"] for t in _d_ads["top"]]))
# id-less rows (the UI-export path) must still merge — that IS the limit of
# that data — but must SAY so rather than pretend precision it lacks.
_idless = [{"name": "Broad", "spend": 25.0, "conv_results": 5.0},
           {"name": "Broad", "spend": 25.0, "conv_results": 5.0},
           {"name": "LAL 1%", "spend": 50.0, "conv_results": 10.0}]
_d_idless = [d for d in conc.compute_concentration(ad_rows=_idless)["dimensions"]
             if d["key"] == "ads"][0]
check("id-less rows merge by name but admit it in the caveat",
      _d_idless["n_entities"] == 2
      and "carry no id" in (_d_idless["caveat"] or ""), str(_d_idless["caveat"]))
# Objectives are a genuine many-rows-to-one-bucket dimension: keyed by VALUE.
_obj_rows = [{"id": "1", "name": "c1", "objective": "OUTCOME_LEADS", "spend": 10.0,
              "conv_results": 1.0},
             {"id": "2", "name": "c2", "objective": "OUTCOME_LEADS", "spend": 10.0,
              "conv_results": 1.0}]
_d_obj = [d for d in conc.compute_concentration(campaign_rows=_obj_rows)["dimensions"]
          if d["key"] == "objectives"][0]
check("objectives still group BY VALUE (2 campaigns, 1 objective)",
      _d_obj["n_entities"] == 1, str(_d_obj["n_entities"]))

rows6 = json.loads((FIX / "conc_rows_6.json").read_text())
ads31 = json.loads((FIX / "conc_ads_31.json").read_text())
block = conc.compute_concentration(
    campaign_rows=rows6, ad_rows=ads31,
    windows={"structure": "2026-06-11 – 2026-07-10", "creative": "2026-04-12 – 2026-07-10"})
dims = {d["key"]: d for d in block["dimensions"]}
check("dimensions: campaigns + ads + objectives",
      set(dims) == {"campaigns", "ads", "objectives"}, str(set(dims)))
dc = dims["campaigns"]
check("campaigns spend hhi 3150.0 / band high",
      dc["spend"]["hhi"] == 3150.0 and dc["spend"]["band"] == "high", str(dc["spend"]))
check("campaigns conv hhi 7145.1 (conv_results only)",
      dc["conv"]["hhi"] == 7145.1, str(dc["conv"]))
check("campaigns verdict fragility", dc["verdict_key"] == "fragility")
check("campaigns top sorted spend desc, name asc tie-break",
      [t["name"] for t in dc["top"]][:4] == ["camp-a", "camp-b", "camp-c", "camp-d"])
check("small-N caveat leans on Effective-N", "Effective-N" in (dc["caveat"] or ""))
check("conversion-indicator exclusion note fires",
      any("non-conversion result indicators excluded" in n
          and "Reach" in n and "video_continuous_2_sec_watched_actions" in n
          for n in block["notes"]), str(block["notes"]))
dobj = dims["objectives"]
check("objectives dimension groups campaigns by objective (65/35)",
      dobj["n_entities"] == 2 and dobj["top"][0]["name"] == "OUTCOME_LEADS"
      and dobj["top"][0]["spend"] == 65.0 and dobj["top"][1]["spend"] == 35.0,
      str(dobj["top"]))
check("windows labeled per dimension (structure vs creative)",
      dc["window"] == "2026-06-11 – 2026-07-10"
      and dims["ads"]["window"] == "2026-04-12 – 2026-07-10")

# bounded embeds: 31 ads -> top-25 + tail 6; math on the FULL universe
da = dims["ads"]
vals31 = [float(31 - i) for i in range(31)]
check("embed cap: top 25 of 31, tail n 6, tail spend 21.0",
      len(da["top"]) == 25 and da["tail"]["n"] == 6 and da["tail"]["spend"] == 21.0,
      str(da["tail"]))
check("cap does not distort math: hhi on the full universe",
      da["spend"]["hhi"] == round(conc.hhi(vals31), 1), str(da["spend"]["hhi"]))
check("lorenz downsampled <= 101 pts", len(da["lorenz"]["spend"]) <= 101)

no_obj = [{k: v for k, v in r.items() if k != "objective"} for r in rows6]
block2 = conc.compute_concentration(campaign_rows=no_obj)
check("objectives omitted with note when objective absent",
      "objectives" not in {d["key"] for d in block2["dimensions"]}
      and any("objectives dimension omitted" in n for n in block2["notes"]),
      str(block2["notes"]))
check("row-derived window fallback (ads dates)",
      {d["key"]: d for d in conc.compute_concentration(ad_rows=ads31)["dimensions"]}
      ["ads"]["window"] == "2026-04-12 – 2026-07-10")
check("absent inputs -> None", conc.compute_concentration() is None)
check("all-empty inputs -> None", conc.compute_concentration(campaign_rows=[]) is None)

# ============================================================================
# 7: creative signals — MediaMetrics-mirror hand-oracles
#
# fatigue(f=4, ctr=.01, cpm=0, ctr_b=.02, cpm_b=0):
#   freq_comp = 1-exp(-(4-1)/3) = 1-exp(-1) = 0.6321206 (w .5)
#   ctr erosion = 1-.01/.02 = 0.5 (w .3); cpm term dropped (baseline 0)
#   score = (.5*.6321206 + .3*.5)/.8 = .4660603/.8 = 0.5825754 -> 0.5826
# ============================================================================
print("--- 7: creative signals ---")
f_oracle = cs.creative_fatigue_score(4, 0.01, 0, 0.02, 0)
check("fatigue oracle 0.5826", round(f_oracle, 4) == 0.5826, str(f_oracle))
f_hi = cs.creative_fatigue_score(100, 0, 1000, 0.01, 1)
check("fatigue clamps: floor at 0, ceiling toward 1, components capped",
      cs.creative_fatigue_score(0, 5, 0, 1, 1) == 0.0     # all terms clamp to 0
      and 0.999 < f_hi <= 1.0                              # saturates near 1
      and approx(cs.creative_fatigue_score(1, 1, 50, 1, 1), 0.2),  # CPM-inflation
      str(f_hi))  # term min(1, 49) -> 1, weight .2 of total 1.0 -> exactly 0.2
check("reach_saturation(700, 1000) == 0.30",
      round(cs.reach_saturation(700, 1000), 4) == 0.3)
check("reach_saturation guards zero impressions", cs.reach_saturation(5, 0) == 0.0)
zf = [cs.effective_frequency(f)["zone"] for f in (2.9, 3.0, 7.0, 7.01)]
check("frequency zones at 2.9/3.0/7.0/7.01",
      zf == ["under", "effective", "effective", "oversaturated"], str(zf))
check("effective flag boundary at 3.0",
      cs.effective_frequency(2.9)["effective"] is False
      and cs.effective_frequency(3.0)["effective"] is True)
rd = cs.ranking_decomposition("BELOW_AVERAGE", "AVERAGE", "UNKNOWN")
check("ranking weakest quality / known 2 / priority [quality, engagement]",
      rd["weakest"] == "quality" and rd["known_count"] == 2
      and rd["priority"] == ["quality", "engagement"] and rd["all_unknown"] is False,
      str(rd))
rd2 = cs.ranking_decomposition("UNKNOWN", None, "")
check("all-unknown ranking never scored",
      rd2["all_unknown"] is True and rd2["weakest"] is None and rd2["priority"] == [])
base = cs.account_baselines([
    {"clicks": 30, "impressions": 2000, "spend": 20},
    {"clicks": 10, "impressions": 2000, "spend": 10}])
check("totals-based baselines ctr .01 / cpm 7.5",
      approx(base["ctr"], 0.01) and approx(base["cpm"], 7.5) and base["n_ads"] == 2,
      str(base))
# Absent is NOT zero. On the Meta path a missing metric is "Not available (…)"
# and meta_rows omits the key; _f() maps that to 0.0, which is right for a
# scalar and WRONG for a ratio — an ad with impressions but no clicks used to
# add its impressions to the denominator and nothing to the numerator, halving
# the baseline that every ad's CTR-erosion term is measured against.
base_gap = cs.account_baselines([{"name": "a", "spend": 10.0, "impressions": 1000.0,
                                  "clicks": 9.0},
                                 {"name": "b", "spend": 10.0, "impressions": 1000.0}])
check("baseline CTR pairs its terms: 0.009 over the 1 ad that has clicks (not 0.0045)",
      approx(base_gap["ctr"], 0.009) and base_gap["n_ctr_rows"] == 1
      and base_gap["n_cpm_rows"] == 2 and base_gap["n_ads"] == 2, str(base_gap))
_blk_gap = cs.compute_creative_signals(
    [{"name": "a", "spend": 10.0, "impressions": 1000.0, "clicks": 9.0},
     {"name": "b", "spend": 10.0, "impressions": 1000.0}], [])
check("partial baseline coverage is stated in the notes",
      any("Baseline CTR computed over 1 of 2 ads" in n for n in _blk_gap["notes"]),
      str(_blk_gap["notes"]))
check("baselines block keeps its public shape (ctr/cpm/n_ads only)",
      set(_blk_gap["baselines"]) == {"ctr", "cpm", "n_ads"},
      str(_blk_gap["baselines"]))

cs_block = cs.compute_creative_signals(ads31)
check("cs bounded embed: 25 ads + tail 6", len(cs_block["ads"]) == 25
      and cs_block["tail"]["n"] == 6, str(cs_block["tail"]))
check("tail has NO reach/frequency keys (non-additive)",
      set(cs_block["tail"]) == {"n", "spend", "impressions", "spend_share"},
      str(set(cs_block["tail"])))
check("summary counts span the FULL universe (31 fresh)",
      cs_block["summary"]["fresh"] == 31 and cs_block["summary"]["below_floor"] == 0,
      str(cs_block["summary"]))
check("cs baselines from totals (ctr .01, cpm 1.0, 31 ads)",
      cs_block["baselines"] == {"ctr": 0.01, "cpm": 1.0, "n_ads": 31},
      str(cs_block["baselines"]))
# ad-01: freq 1.43 vs equal-to-baseline ctr/cpm -> score = .5*(1-exp(-0.43/3))
want_f = round(0.5 * (1.0 - math.exp(-0.43 / 3.0)), 4)
check("per-ad fatigue matches the formula (band fresh)",
      cs_block["ads"][0]["fatigue"] == want_f
      and cs_block["ads"][0]["fatigue_band"] == "fresh",
      str(cs_block["ads"][0]))
check("per-ad saturation 0.3 (reach = 70% of impressions)",
      cs_block["ads"][0]["saturation"] == 0.3)
check("cs window from row dates", cs_block["window"] == "2026-04-12 – 2026-07-10")
check("rankings unavailable on raw-path rows",
      cs_block["rankings"] == {"available": False}
      and any("rankings" in n and "manual CSV" in n for n in cs_block["notes"]),
      str(cs_block["rankings"]))

floor_block = cs.compute_creative_signals([
    {"name": "tiny", "spend": 1.0, "impressions": 500.0, "clicks": 5.0,
     "frequency": 2.0, "ctr": 0.01, "cpm": 2.0}])
check("below min-impressions floor: fatigue null + below_floor counted",
      floor_block["ads"][0]["fatigue"] is None
      and floor_block["ads"][0]["fatigue_band"] is None
      and floor_block["summary"]["below_floor"] == 1,
      str(floor_block["summary"]))
check("floor note fires", any("impression floor" in n for n in floor_block["notes"]),
      str(floor_block["notes"]))

zone_sets = [{"name": "z-a", "spend": 40.0, "frequency": 2.9},
             {"name": "z-b", "spend": 30.0, "frequency": 3.0},
             {"name": "z-c", "spend": 20.0, "frequency": 7.0},
             {"name": "z-d", "spend": 10.0, "frequency": 7.01}]
zb = cs.compute_creative_signals(ads31, adset_rows=zone_sets)
check("zones per AD SET: under 1 / effective 2 / oversaturated 1",
      (zb["zones"]["under"], zb["zones"]["effective"], zb["zones"]["oversaturated"])
      == (1, 2, 1), str(zb["zones"]))
check("zone rows spend-desc with zone labels",
      [r["zone"] for r in zb["zones"]["rows"]]
      == ["under", "effective", "effective", "oversaturated"],
      str(zb["zones"]["rows"]))
check("no ad rows -> None", cs.compute_creative_signals([]) is None)

# ============================================================================
# 8: meta_rows — raw MCP envelope + human-formatted value parsing
# (raw_shape_*.json are sanitized captures: exact quirks, generic names)
# ============================================================================
print("--- 8: meta_rows (raw shapes) ---")
check("num_or_none money with NBSP + ISO suffix",
      mr.num_or_none("CA$1,023.31 CAD") == 1023.31)
check("num_or_none count / decimal strings",
      mr.num_or_none("583,301") == 583301.0 and mr.num_or_none("3.21") == 3.21)
check("num_or_none percent -> fraction", approx(mr.num_or_none("0.0658%"), 0.000658))
check("num_or_none missing markers -> None",
      mr.num_or_none("Not available (Uses ad set daily budget.)") is None
      and mr.num_or_none("") is None and mr.num_or_none("--") is None
      and mr.num_or_none("—") is None and mr.num_or_none(None) is None)
check("parse_human_date explicit month map",
      mr.parse_human_date("11 June 2026") == "2026-06-11"
      and mr.parse_human_date("4 July 2026") == "2026-07-04"
      and mr.parse_human_date("2026-07-10") == "2026-07-10"
      and mr.parse_human_date("Someday soon") is None)
check("parse_result string form", mr.parse_result("181,893 (Reach)") == (181893.0, "Reach"))
check("parse_result keeps nested parens",
      mr.parse_result({"value": "107 (Leads (form))"}) == (107.0, "Leads (form)"))
check("parse_result list form",
      mr.parse_result({"value": [{"indicator": "video_continuous_2_sec_watched_actions",
                                  "values": [{"value": 245042}]}]})
      == (245042.0, "video_continuous_2_sec_watched_actions"))
check("parse_result missing -> None",
      mr.parse_result("Not available") is None and mr.parse_result(None) is None)
check("is_conversion_indicator token rule",
      mr.is_conversion_indicator("Leads (form)") and mr.is_conversion_indicator("Purchases")
      and not mr.is_conversion_indicator("Reach")
      and not mr.is_conversion_indicator("video_continuous_2_sec_watched_actions")
      and not mr.is_conversion_indicator("Landing page views")
      and not mr.is_conversion_indicator(""))

camps = mr.load_rows(str(FIX / "raw_shape_campaigns.json"), level="campaign")
check("string envelope DOUBLE PARSE -> 4 campaign rows, spend desc",
      len(camps) == 4 and camps[0]["name"] == "Campaign A"
      and camps[0]["spend"] == 1023.31, str([c["name"] for c in camps]))
ca = {c["name"]: c for c in camps}
check("'Not available (...)' budget omitted; real budget parsed",
      "daily_budget" not in ca["Campaign A"] and ca["Campaign B"]["daily_budget"] == 30.0)
check("dual results: Reach string (NO conv_results)",
      ca["Campaign A"]["results"] == 181780.0
      and ca["Campaign A"]["results_indicator"] == "Reach"
      and "conv_results" not in ca["Campaign A"])
check("dual results: list form (video indicator, NO conv_results)",
      ca["Campaign B"]["results"] == 245042.0
      and ca["Campaign B"]["results_indicator"] == "video_continuous_2_sec_watched_actions"
      and "conv_results" not in ca["Campaign B"])
check("nested-paren indicator -> conv_results set",
      ca["Campaign C"]["results_indicator"] == "Leads (form)"
      and ca["Campaign C"]["conv_results"] == 107.0)
check("human dates -> ISO", ca["Campaign A"]["created_time"] == "2026-06-11"
      and ca["Campaign A"]["date_stop"] == "2026-07-10")
check("ctr recomputed from counts (not the returned percent)",
      approx(ca["Campaign A"]["ctr"], 384.0 / 583301.0))
check("percent-ctr fallback when counts absent",
      approx(mr.normalize({"name": "x", "ctr": "0.0658%"}, level="ad")["ctr"], 0.000658)
      and approx(mr.normalize({"name": "x", "ctr": 2.85}, level="ad")["ctr"], 0.0285))

asets = mr.load_rows(str(FIX / "raw_shape_adsets.json"), level="adset")
aa = {r["name"]: r for r in asets}
check("adset rows carry campaign_id linkage",
      all("campaign_id" in r for r in asets)
      and aa["Ad Set A"]["campaign_id"] == "111000000000001")
check("adset attribution + frequency + reach survive",
      aa["Ad Set A"]["attribution_setting"] == "1d_view_7d_click"
      and aa["Ad Set A"]["frequency"] == 3.21 and aa["Ad Set A"]["reach"] == 181893.0)
check("CAMPAIGN_PAUSED status + OFFSITE_CONVERSIONS goal survive",
      aa["Ad Set D"]["effective_status"] == "CAMPAIGN_PAUSED"
      and aa["Ad Set D"]["optimization_goal"] == "OFFSITE_CONVERSIONS")

ads = mr.load_rows(str(FIX / "raw_shape_ads.json"), level="ad")
ad = {r["name"]: r for r in ads}
check("video quartiles + thruplay aliased to canonical keys",
      ad["Ad A"]["video_p25"] == 48091.0 and ad["Ad A"]["thruplay"] == 4772.0)
check("static ad: 'Not available' video fields omitted",
      "video_p25" not in ad["Ad C"] and "thruplay" not in ad["Ad C"])
check("unknown source keys ignored gracefully",
      "cost_per_video_view" not in ad["Ad A"] and "cost_per_action_type" not in ad["Ad A"])

a7 = mr.load_rows(str(FIX / "raw_shape_adsets_7d.json"), level="adset")
check("window_label: 7-day pull -> inclusive 7 days",
      mr.window_label(a7) == ("2026-07-04 – 2026-07-10", 7), str(mr.window_label(a7)))
check("window_label: 30-day pull -> 30 days",
      mr.window_label(camps) == ("2026-06-11 – 2026-07-10", 30))

dsets = mr.load_datasets(str(FIX / "raw_shape_datasets.json"))
check("datasets deduped by dataset_id (2 unique from 4 rows, order kept)",
      len(dsets) == 2 and [d["name"] for d in dsets] == ["Dataset One", "Dataset Two"],
      str([d.get("name") for d in dsets]))
check("epoch-zero last_fired preserved verbatim",
      dsets[0]["last_fired_time"].startswith("1969-12-31"))
dq = mr.load_dataset_quality(str(FIX / "raw_shape_dataset_quality.json"))
check("dataset quality: channel -> events with numeric composite",
      [e["event_name"] for e in dq["web"]] == ["Form", "Lead", "PageView"]
      and dq["web"][1]["event_match_quality"]["composite_score"] == 7)

stamp = mr.file_stamp(str(FIX / "raw_shape_campaigns.json"))
check("file_stamp shape + real sha256",
      set(stamp) == {"file", "sha256", "bytes"}
      and stamp["sha256"] == hashlib.sha256(
          (FIX / "raw_shape_campaigns.json").read_bytes()).hexdigest())
try:
    mr.load_rows(str(FIX / "raw_shape_datasets.json"), level="campaign")
    check("wrong-pull guard raises RawResultError", False)
except mr.RawResultError:
    check("wrong-pull guard raises RawResultError", True)
with tempfile.TemporaryDirectory() as td:
    p = Path(td) / "decoded.json"
    p.write_text(json.dumps({"ad_entities": [{"name": "Camp Z", "amount_spent": "12.50"}],
                             "summary": {"total_count": 1}}))
    check("already-decoded ad_entities list accepted (future-proofing)",
          mr.load_rows(str(p), level="campaign")[0]["spend"] == 12.5)

# ============================================================================
# 9: manual UI-CSV path (NEEDS-REAL-EXPORT-VALIDATION format)
# ============================================================================
print("--- 9: manual CSV ---")
uc_rows, uc_meta = mc.campaigns_rows(str(FIX / "ui_campaigns.csv"))
check("ui campaigns: 3 rows (summary 'Results from…' dropped)", len(uc_rows) == 3,
      str(len(uc_rows)))
check("ui campaigns: spend-desc order, quoted thousands parsed",
      uc_rows[0]["name"] == "Campaign Beta" and uc_rows[0]["spend"] == 2500.50,
      str([r["name"] for r in uc_rows]))
check("ui campaigns: multiline quoted name survives",
      any("\n" in r["name"] and "Alpha" in r["name"] for r in uc_rows))
byname = {r["name"].replace("\n", " "): r for r in uc_rows}
check("ui campaigns: '--' Results -> key omitted",
      "results" not in byname["Campaign Beta"])
check("ui campaigns: conversion-like indicator -> conv_results",
      byname["Campaign Alpha"]["results"] == 120.0
      and byname["Campaign Alpha"]["conv_results"] == 120.0
      and byname["Campaign Alpha"]["results_indicator"] == "Leads (form)")
check("ui campaigns: budget routed by type column (daily/lifetime/omitted)",
      byname["Campaign Alpha"]["daily_budget"] == 50.0
      and byname["Campaign Gamma"]["lifetime_budget"] == 100.0
      and "daily_budget" not in byname["Campaign Beta"]
      and "lifetime_budget" not in byname["Campaign Beta"])
check("ui campaigns: ctr/cpm recomputed from counts",
      approx(byname["Campaign Alpha"]["ctr"], 600.0 / 50000.0)
      and approx(byname["Campaign Alpha"]["cpm"], 1000.0 / 50000.0 * 1000.0))
check("ui campaigns meta: window + inclusive days + provenance",
      uc_meta["window"] == "2026-06-11 – 2026-07-10" and uc_meta["window_days"] == 30
      and uc_meta["n_rows_raw"] == 3
      and set(uc_meta["stamp"]) == {"file", "sha256", "bytes"}, str(uc_meta))

# The 'CTR (all)' fallback fires only when 'Clicks (all)' is absent — which the
# committed fixtures never are, so this path shipped unexecuted. Ads Manager
# emits CTR on the PERCENT scale ("0.87" = 0.87%), and Meta all-click CTR is
# typically 0.5-1.5%: the old "fraction if <= 1" heuristic therefore guessed
# wrong for the COMMON case and rendered 0.87% as 87.00%.
with tempfile.TemporaryDirectory() as td:
    _p = Path(td) / "ctr_only.csv"
    _p.write_text(
        "Ad name,Amount spent (CAD),Impressions,CTR (all),Results,Result indicator,"
        "Reporting starts,Reporting ends\n"
        "Sub-one,100.00,10000,0.87,5,Purchases,2026-07-04,2026-07-10\n"
        "With-sign,100.00,10000,1.25%,5,Purchases,2026-07-04,2026-07-10\n"
        "Over-one,100.00,10000,2.50,5,Purchases,2026-07-04,2026-07-10\n",
        encoding="utf-8")
    _c = {r["name"]: r for r in mc.ads_rows(str(_p))[0]}
    check("CSV 'CTR (all)' is percent-scale: 0.87 -> 0.0087 (NOT 0.87)",
          approx(_c["Sub-one"]["ctr"], 0.0087), str(_c["Sub-one"].get("ctr")))
    check("CSV 'CTR (all)' with a % sign parses identically (1.25% -> 0.0125)",
          approx(_c["With-sign"]["ctr"], 0.0125), str(_c["With-sign"].get("ctr")))
    check("CSV 'CTR (all)' > 1 still percent-scale (2.50 -> 0.025)",
          approx(_c["Over-one"]["ctr"], 0.025), str(_c["Over-one"].get("ctr")))

ua_rows, ua_meta = mc.ads_rows(str(FIX / "ui_ads.csv"))
u1 = {r["name"]: r for r in ua_rows}
check("ui ads: rankings canonicalised",
      u1["Ad One"]["quality_ranking"] == "ABOVE_AVERAGE"
      and u1["Ad One"]["engagement_rate_ranking"] == "AVERAGE"
      and u1["Ad One"]["conversion_rate_ranking"] == "BELOW_AVERAGE")
check("ui ads: '--' and 'Not enough data' rankings omitted",
      "quality_ranking" not in u1["Ad Two"]
      and "engagement_rate_ranking" not in u1["Ad Two"]
      and u1["Ad Two"]["conversion_rate_ranking"] == "AVERAGE")
check("ui ads: link clicks parsed (CSV-only unlock)",
      u1["Ad One"]["link_clicks"] == 350.0)
check("ui ads meta: 7-day window (CR-07 true-band unlock)",
      ua_meta["window_days"] == 7, str(ua_meta))

us_rows, us_meta = mc.adsets_rows(str(FIX / "ui_adsets.csv"))
check("ui adsets: frequency + attribution + campaign_name kept",
      us_rows[0]["frequency"] == 2.0
      and us_rows[0]["attribution_setting"] == "7-day click"
      and us_rows[0]["campaign_name"] == "Campaign Alpha", str(us_rows[0]))

try:
    mc.campaigns_rows(str(FIX / "ui_ads.csv"))
    check("ui wrong-report guard raises", False)
except mc.ManualCsvError:
    check("ui wrong-report guard raises", True)
try:
    mc.load_ui_csv(str(FIX / "ui_ads.csv"), level="nope")
    check("unknown level raises", False)
except mc.ManualCsvError:
    check("unknown level raises", True)

csv_cs = cs.compute_creative_signals(ua_rows)
check("rankings available on CSV-shaped rows",
      csv_cs["rankings"]["available"] is True
      and csv_cs["rankings"]["summary"]["n_ranked"] == 3)
check("weakest-lever tally: conversion 2 / quality 1",
      csv_cs["rankings"]["summary"]["weakest"]
      == {"quality": 1, "engagement": 0, "conversion": 2},
      str(csv_cs["rankings"]["summary"]))
check("top ranked row is the top spender with its weakest lever",
      csv_cs["rankings"]["rows"][0]["name"] == "Ad One"
      and csv_cs["rankings"]["rows"][0]["weakest"] == "conversion")

# ============================================================================
# 10: deterministic pre-scorer — per-rule hand oracles
# ============================================================================
print("--- 10: prescore ---")
G = GROUPS

# AR-01: top-3 = 40+25+15 = 80 of 100 -> 80.0% PASS
p = ps.compute_prescore(campaign_rows=G["campaigns_ar01"], business_model="Lead Gen")
check("AR-01 80.0% PASS", p["checks"]["AR-01"]["result"] == "PASS"
      and "80.0%" in p["checks"]["AR-01"]["observed"], str(p["checks"].get("AR-01")))

# AR-02: starved 3 of 5 = 60% > 50% -> FAIL (counts in observed)
p = ps.compute_prescore(adset_rows=G["adsets_ar02"], business_model="Lead Gen")
check("AR-02 3-of-5 starved 60% FAIL", p["checks"]["AR-02"]["result"] == "FAIL"
      and p["checks"]["AR-02"]["observed"].startswith("3 of 5"),
      str(p["checks"].get("AR-02")))
p = ps.compute_prescore(adset_rows=G["adsets_ar02_flag"], business_model="Lead Gen")
check("AR-02 count>=3 at 25% share -> FLAG", p["checks"]["AR-02"]["result"] == "FLAG"
      and p["checks"]["AR-02"]["observed"].startswith("3 of 12"),
      str(p["checks"].get("AR-02")))

# AR-03/AR-04: OUTCOME_LEADS -> OFFSITE_CONVERSIONS aligned; -> REACH hard FAIL
p = ps.compute_prescore(campaign_rows=G["campaigns_ar04"],
                        adset_rows=G["adsets_ar04"], business_model="Lead Gen")
check("AR-03 PASS (one goal per campaign)", p["checks"]["AR-03"]["result"] == "PASS")
check("AR-04 hard mismatch FAIL, OFFSITE_CONVERSIONS aligned for LEADS",
      p["checks"]["AR-04"]["result"] == "FAIL"
      and "1 HARD" in p["checks"]["AR-04"]["observed"]
      and "1 aligned" in p["checks"]["AR-04"]["observed"]
      and "Ad Set B" in p["checks"]["AR-04"]["observed"],
      str(p["checks"].get("AR-04")))
check("AR-04 severity Critical (benchmarks §2)",
      p["checks"]["AR-04"]["severity"] == "Critical")

# BP-02 homogeneous: Campaign D = 12% spend / 3% results -> FLAG
p = ps.compute_prescore(campaign_rows=G["campaigns_bp02"], business_model="Lead Gen")
check("BP-02 offender 12.0%/3.0% -> FLAG", p["checks"]["BP-02"]["result"] == "FLAG"
      and "12.0% spend / 3.0% results" in p["checks"]["BP-02"]["observed"],
      str(p["checks"].get("BP-02")))
# BP-02 heterogeneous indicators: never scored — evidence only
p = ps.compute_prescore(campaign_rows=G["campaigns_bp02_mixed"], business_model="Lead Gen")
check("BP-02 heterogeneous -> skipped with reason",
      "BP-02" not in p["checks"]
      and any(s["id"] == "BP-02" and "heterogeneous" in s["reason"]
              for s in p["skipped"]), str(p["skipped"]))
check("BP-02 heterogeneous -> spend-mix evidence",
      "spend by results indicator" in p["evidence"]["BP-02"]["observed"],
      str(p["evidence"].get("BP-02")))

# AT-02 40% FLAG / AT-03 60% PASS + AT-01 evidence
p = ps.compute_prescore(adset_rows=G["adsets_at"], business_model="Lead Gen")
check("AT-02 40.0% 1d_view -> FLAG", p["checks"]["AT-02"]["result"] == "FLAG"
      and "40.0%" in p["checks"]["AT-02"]["observed"], str(p["checks"].get("AT-02")))
check("AT-03 60.0% exactly-7d_click -> PASS", p["checks"]["AT-03"]["result"] == "PASS"
      and "60.0%" in p["checks"]["AT-03"]["observed"], str(p["checks"].get("AT-03")))
check("AT-01 evidence: attribution spend mix",
      "Attribution-setting spend mix" in p["evidence"]["AT-01"]["observed"])

# CR-03 30% FLAG / CR-06 78% FAIL / CR-04 evidence-only on Lead Gen raw path
pres_ads = ps.compute_prescore(ad_rows=G["ads_cr"], business_model="Lead Gen")
check("CR-03 hold-through 30.0% -> FLAG",
      pres_ads["checks"]["CR-03"]["result"] == "FLAG"
      and "Hold-through 30.0%" in pres_ads["checks"]["CR-03"]["observed"],
      str(pres_ads["checks"].get("CR-03")))
check("CR-06 top-5 78.0% -> FAIL", pres_ads["checks"]["CR-06"]["result"] == "FAIL"
      and "78.0%" in pres_ads["checks"]["CR-06"]["observed"],
      str(pres_ads["checks"].get("CR-06")))
check("CR-04 not scored on Lead Gen; all-click CTR evidence labeled",
      "CR-04" not in pres_ads["checks"]
      and "All-click CTR 1.000%" in pres_ads["evidence"]["CR-04"]["observed"]
      and "not scored" in pres_ads["evidence"]["CR-04"]["observed"],
      str(pres_ads["evidence"].get("CR-04")))
check("CR-08 skipped without created_time (manual-CSV degradation)",
      any(s["id"] == "CR-08" for s in pres_ads["skipped"]))

# CR-04 scored on the CSV path + Ecommerce: 590/79000 = 0.75% -> FLAG
p = ps.compute_prescore(ad_rows=ua_rows, business_model="Ecommerce")
check("CR-04 CSV+Ecommerce: CTR-Link 0.75% FLAG",
      p["checks"]["CR-04"]["result"] == "FLAG"
      and "CTR-Link 0.75%" in p["checks"]["CR-04"]["observed"],
      str(p["checks"].get("CR-04")))

# CR-07: 7-day window -> true bands (6.0 > 5 FAIL); 30-day -> PASS-only FLAG
p = ps.compute_prescore(adset7_rows=G["adsets7_cr07"], business_model="Lead Gen")
check("CR-07 true bands on 7d window: 6.00 -> FAIL",
      p["checks"]["CR-07"]["result"] == "FAIL"
      and "7-day window" in p["checks"]["CR-07"]["observed"],
      str(p["checks"].get("CR-07")))
p = ps.compute_prescore(adset_rows=G["adsets30_cr07"], business_model="Lead Gen")
check("CR-07 PASS-only mode on 30d window: 6.00 -> FLAG (never FAIL)",
      p["checks"]["CR-07"]["result"] == "FLAG"
      and "PASS-only mode" in p["checks"]["CR-07"]["observed"],
      str(p["checks"].get("CR-07")))
p = ps.compute_prescore(adset_rows=us_rows, business_model="Lead Gen")
check("CR-07 true bands via 7d CSV: 2.19 -> PASS",
      p["checks"]["CR-07"]["result"] == "PASS"
      and "2.19" in p["checks"]["CR-07"]["observed"], str(p["checks"].get("CR-07")))

# CR-08 age bands vs generated_for_date (newest ACTIVE ad = 2026-06-10)
for gd, expect, days in (("2026-07-10", "PASS", 30), ("2026-07-11", "FLAG", 31),
                         ("2026-08-09", "FLAG", 60), ("2026-08-10", "FAIL", 61)):
    p = ps.compute_prescore(ad_rows=G["ads_cr08"], generated_for_date=gd,
                            business_model="Lead Gen")
    check("CR-08 %d days -> %s" % (days, expect),
          p["checks"]["CR-08"]["result"] == expect
          and ("%d days before" % days) in p["checks"]["CR-08"]["observed"],
          str(p["checks"].get("CR-08")))

# DI-01 / DI-04 from the raw-shape optional pulls
p = ps.compute_prescore(datasets=dsets, dataset_quality=dq, business_model="Lead Gen")
check("DI-01 PASS: 2 unique active datasets, never-fired flagged",
      p["checks"]["DI-01"]["result"] == "PASS"
      and "2 of 2 unique datasets active" in p["checks"]["DI-01"]["observed"]
      and "never fired" in p["checks"]["DI-01"]["observed"]
      and "Dataset One" in p["checks"]["DI-01"]["observed"],
      str(p["checks"].get("DI-01")))
check("DI-04 Lead EMQ 7.0 -> FLAG (quality-shaped fixture)",
      p["checks"]["DI-04"]["result"] == "FLAG"
      and "EMQ (Lead) composite 7.0" in p["checks"]["DI-04"]["observed"],
      str(p["checks"].get("DI-04")))
p = ps.compute_prescore(dataset_quality=dq, business_model="Ecommerce")
check("DI-04 skip-with-note when the primary event is absent",
      any(s["id"] == "DI-04" and "Purchase" in s["reason"] and "Lead" in s["reason"]
          for s in p["skipped"]), str(p["skipped"]))

# graceful skips + determinism + None
bare = ps.compute_prescore(campaign_rows=[{"name": "solo", "spend": 1.0}])
skipped_ids = {s["id"] for s in bare["skipped"]}
check("graceful skips when inputs missing",
      {"AR-02", "AR-03", "AR-04", "BP-02", "AT-02", "AT-03", "CR-03", "CR-04",
       "CR-06", "CR-07", "CR-08", "DI-01", "DI-04"} <= skipped_ids
      and list(bare["checks"]) == ["AR-01"], str(sorted(skipped_ids)))
check("every skip carries a reason", all(s["reason"] for s in bare["skipped"]))
check("business_model default note", any("Lead Gen assumed" in n for n in bare["notes"]))
check("prescore None on no input at all", ps.compute_prescore() is None)
p1 = ps.compute_prescore(ad_rows=G["ads_cr"], business_model="Lead Gen")
check("prescore deterministic (same args -> identical JSON)",
      json.dumps(pres_ads, sort_keys=True) == json.dumps(p1, sort_keys=True))

# KPI rows name the window; creative-signals evidence + Fatigued Ads KPI
kpis = {k["metric"]: k for k in pres_ads["kpis"]}
check("KPI set from ads input (Spend/CTR/CPM/Top-5)",
      {"Spend", "CTR (all-click)", "CPM", "Top-5 Ad Spend Share"} <= set(kpis),
      str(sorted(kpis)))
check("KPI values: Spend 100.0 / CTR 1.0 / CPM 1.0 / Top-5 78.0 FAIL",
      kpis["Spend"]["value"] == 100.0 and kpis["CTR (all-click)"]["value"] == 1.0
      and kpis["CPM"]["value"] == 1.0 and kpis["Top-5 Ad Spend Share"]["value"] == 78.0
      and kpis["Top-5 Ad Spend Share"]["flag"] == "FAIL", str(kpis))
check("KPI notes name the window",
      "2026-04-12 – 2026-07-10" in kpis["Spend"]["notes"], str(kpis["Spend"]))
cs_small = cs.compute_creative_signals(G["ads_cr"])
p = ps.compute_prescore(ad_rows=G["ads_cr"], business_model="Lead Gen",
                        creative_signals=cs_small)
check("CR-01 evidence from creative signals",
      "Thumb-stop (3s) is not exposed" in p["evidence"]["CR-01"]["observed"])
check("Fatigued Ads KPI present (informational, machine-scored)",
      any(k["metric"] == "Fatigued Ads" and k["flag"] == "N/A" for k in p["kpis"]),
      str([k["metric"] for k in p["kpis"]]))

# Results KPI headline honesty: mixed indicators must not sum Reach with Leads
mix = [
    {"name": "aw", "spend": 100.0, "results": 100000.0,
     "results_indicator": "Reach"},
    {"name": "lg", "spend": 100.0, "results": 50.0,
     "results_indicator": "Leads (form)", "conv_results": 50.0},
]
pmix = ps.compute_prescore(campaign_rows=mix, business_model="Lead Gen")
kmix = {k["metric"]: k for k in pmix["kpis"]}
check("Results KPI: mixed indicators -> conversion-like only (50, not 100050)",
      kmix["Results"]["value"] == 50.0
      and "conversion-like results only" in kmix["Results"]["notes"]
      and "Reach" in kmix["Results"]["notes"], str(kmix.get("Results")))
check("Cost per Result blends over conversion-like results (200/50 = 4.0)",
      kmix["Cost per Result"]["value"] == 4.0
      and "conversion-like results only" in kmix["Cost per Result"]["notes"],
      str(kmix.get("Cost per Result")))
hom = [{"name": "a", "spend": 10.0, "results": 5.0,
        "results_indicator": "Leads", "conv_results": 5.0},
       {"name": "b", "spend": 10.0, "results": 15.0,
        "results_indicator": "Leads", "conv_results": 15.0}]
phom = ps.compute_prescore(campaign_rows=hom, business_model="Lead Gen")
khom = {k["metric"]: k for k in phom["kpis"]}
check("Results KPI: homogeneous indicator sums raw with indicator note",
      khom["Results"]["value"] == 20.0
      and "indicator: Leads" in khom["Results"]["notes"], str(khom.get("Results")))

# ============================================================================
# 11: merge — purity, corrections, injection, KPIs, health flip
#
# Pre-merge micro payload (CR-03/CR-06 drafted PASS): DI 100 (w20) + CR 100
# (w25) -> health 100.0. Machine truth: CR-03 FLAG + CR-06 FAIL ->
# CR = 0.75/4.5 = 16.6667 -> health (100*20 + 16.6667*25)/45 = 53.7.
# ============================================================================
print("--- 11: merge ---")
pre = copy.deepcopy(MICRO)
for c in pre["checks"]:
    if c["id"] in ("CR-03", "CR-06"):
        c["flag"] = "PASS"
pre["kpis"] = [{"metric": "CPM", "value": 99.0, "unit": "", "benchmark": "",
                "flag": "N/A", "notes": "drafted"},
               {"metric": "Conv rate", "value": 2.0, "unit": "%", "benchmark": "",
                "flag": "N/A", "notes": "drafted"}]
check("pre-merge micro health == 100.0 (drafted PASS)",
      audit_model.compute_model(pre, generated="T")["health"]["score"] == 100.0)

merged, blockp, plog = ps.merge_into_findings(pre, pres_ads)
check("merge is PURE (input payload unmutated)",
      all(c["flag"] == "PASS" for c in pre["checks"] if c["id"].startswith("CR-"))
      and pre["kpis"][0]["value"] == 99.0)
check("corrections recorded PASS->FLAG / PASS->FAIL",
      blockp["corrected"] == [{"id": "CR-03", "from": "PASS", "to": "FLAG"},
                              {"id": "CR-06", "from": "PASS", "to": "FAIL"}],
      str(blockp["corrected"]))
check("correction log line format",
      any(l.startswith("prescore: CR-06 PASS->FAIL (") and "78.0%" in l for l in plog),
      str(plog))
mrow = next(c for c in merged["checks"] if c["id"] == "CR-06")
check("recommendation preserved on the corrected check",
      mrow["recommendation"] == "keep monitoring" and mrow["flag"] == "FAIL"
      and "78.0%" in mrow["observed"])
# The score is machine-enforced; findings[] is not — so drift runs BOTH ways.
# Direction 1 (missing): machine scores FAIL/FLAG, no finding covers it.
check("corrections with no matching finding are reported as unreconciled/missing",
      blockp["unreconciled"] == [{"id": "CR-03", "result": "FLAG", "reason": "missing"},
                                 {"id": "CR-06", "result": "FAIL", "reason": "missing"}],
      str(blockp["unreconciled"]))
check("...and each gets a WARNING log line naming the drift",
      sum(1 for l in plog if l.startswith("prescore: WARNING CR-")
          and "the narrative did not" in l) == 2, str(plog))
_rec = copy.deepcopy(pre)
_rec["findings"] = [{"id": "F-001", "title": "Top-5 ad spend concentration",
                     "category": "Creative Performance", "severity": "High",
                     "evidence": "CR-06: top-5 ads hold 78.0% of spend",
                     "recommendation": "diversify"},
                    {"id": "F-002", "title": "CR-03 hold-through", "category":
                     "Creative Performance", "severity": "Medium",
                     "evidence": "x", "recommendation": "y"}]
_m2, _b2, _l2 = ps.merge_into_findings(_rec, pres_ads)
check("a finding that cites the check id reconciles it (no false alarm)",
      _b2["unreconciled"] == [], str(_b2["unreconciled"]))

# Direction 2 (cleared): the machine CLEARS a check a finding still argues —
# the score rose while the roadmap still tells the client to fix it. Only
# reporting direction 1 made this half a feature; a live run corrects
# `AR-02 FAIL->PASS`, so the uncovered direction is the routine one.
_clr = copy.deepcopy(pre)
for _c in _clr["checks"]:
    if _c["id"] == "CR-06":
        _c["flag"] = "FAIL"          # the model drafted a failure...
_clr["findings"] = [{"id": "F-001", "title": "Top-5 ad concentration",
                     "category": "Creative Performance", "severity": "High",
                     "evidence": "CR-06 shows top-5 ads dominating",
                     "recommendation": "diversify"}]
_pass6 = copy.deepcopy(pres_ads)
_pass6["checks"]["CR-06"] = dict(_pass6["checks"]["CR-06"],
                                 result="PASS", observed="top-5 ads 41.0% <= 70%")
_m3, _b3, _l3 = ps.merge_into_findings(_clr, _pass6)
check("a CLEARED check that a finding still argues is reported (reason=cleared)",
      {"id": "CR-06", "result": "PASS", "reason": "cleared"} in _b3["unreconciled"],
      str(_b3["unreconciled"]))
check("...with a log line telling the auditor to drop or amend the finding",
      any("CR-06 machine-scored PASS but a finding still argues it" in l
          for l in _l3), str(_l3))
check("an untouched check is never reported as drift (only corrected/injected)",
      all(u["id"] in {c["id"] for c in _b3["corrected"]} | set(_b3["injected"])
          for u in _b3["unreconciled"]), str(_b3["unreconciled"]))
m_merged = audit_model.compute_model(merged, generated="2026-07-10T00:00:00",
                                     prescore=blockp)
check("correction flips health 100.0 -> 53.7 / D (hand oracle)",
      m_merged["health"]["score"] == 53.7 and m_merged["health"]["grade"] == "D",
      str(m_merged["health"]))
kpi_metrics = [k["metric"] for k in merged["kpis"]]
check("KPI CPM replaced IN PLACE (machine value), Conv rate survives",
      kpi_metrics[0] == "CPM" and merged["kpis"][0]["value"] == 1.0
      and "Conv rate" in kpi_metrics, str(kpi_metrics))
check("unmatched machine KPIs appended",
      "Spend" in kpi_metrics and "Top-5 Ad Spend Share" in kpi_metrics)
check("kpis_replaced records CPM", blockp["kpis_replaced"] == ["CPM"])
check("merge with None prescore is a no-op",
      ps.merge_into_findings(pre, None) == (pre, None, []))

# injection: payload missing CR-03 -> a full check row is created
pre2 = copy.deepcopy(pre)
pre2["checks"] = [c for c in pre2["checks"] if c["id"] != "CR-03"]
merged2, block2, _log2 = ps.merge_into_findings(pre2, pres_ads)
irow = next(c for c in merged2["checks"] if c["id"] == "CR-03")
check("injected check row: right category + framework name + machine flag",
      irow["category"] == "Creative Performance"
      and irow["name"] == "Hold-through (P100/P25)" and irow["flag"] == "FLAG"
      and irow["expected"] and irow["recommendation"] == "", str(irow))
check("injection recorded", block2["injected"] == ["CR-03"], str(block2["injected"]))

# prescore threading through the renderers
html_p = audit_html.render_html(m_merged)
check("html shows the machine-scored overview line", "machine-scored" in html_p)
stripped_p = html_p.replace(audit_html.gsap_blob(), "")
check("html with prescore still self-contained",
      not re.findall(r"https?://|<link|src=|cdn", stripped_p))
md_p = audit_md.render_md(m_merged)
check("md prescore footer names the corrected ids",
      "checks machine-scored" in md_p
      and "CR-06" in md_p.rsplit("checks machine-scored", 1)[1][:250],
      repr(md_p.rsplit("checks machine-scored", 1)[-1][:120]))
# The old guard here re-scanned the WHOLE template for [NN,'X'] grade literals,
# so any stray JS literal of that shape could corrupt it (hence "pollution").
# The BUCKETS check in section 3 is scoped to the `var BUCKETS = [...]`
# declaration, so it cannot be polluted by data or markup. What is worth
# asserting after a data-heavy render is that the rendered document still
# carries the live kernel and no scoring math sneaked back in (html_p above).
check("rendered html keeps the live ICE kernel and recomputes no score",
      "var BUCKETS" in html_p and "function bucketOf" in html_p
      and "healthOf" not in html_p and "scoreCheck" not in html_p)
# Where machine-vs-narrative drift surfaces is a deliberate split: the md
# record and the xlsx working copy warn the AUDITOR; the HTML is the
# client-shareable deliverable and must not carry "our findings disagree with
# our score" — that is a note to resolve before sending, not a disclosure.
check("md record surfaces unreconciled drift to the auditor",
      "Findings not reconciled with the machine results" in md_p
      and "CR-06" in md_p.split("Findings not reconciled")[1][:400], md_p[-400:])
check("the CLIENT-facing html does NOT carry the drift warning",
      "not reconciled" not in html_p.lower())
# Not rendering it is not enough: the whole model is embedded as JSON, so
# view-source handed the client the exact list of checks whose findings we know
# are stale. It must be absent from the PAYLOAD, not just the UI.
check("...and does not embed unreconciled in its data block either",
      "unreconciled" not in html_p and '"corrected"' in html_p,
      "unreconciled leaked into the client deliverable")
check("stripping it does not mutate the caller's model (md still sees it)",
      (m_merged.get("prescore") or {}).get("unreconciled") is not None)

# renderers with concentration + creative signals blocks
m_full = audit_model.compute_model(PAYLOAD, generated="2026-07-10T00:00:00",
                                   concentration=block, creative_signals=cs_block)
html_f = audit_html.render_html(m_full)
check("html embeds concentration + creative signals data",
      '"verdict_key"' in html_f and '"fatigue_band"' in html_f)
check("html with all blocks self-contained",
      not re.findall(r"https?://|<link|src=|cdn",
                     html_f.replace(audit_html.gsap_blob(), "")))
md_f = audit_md.render_md(m_full)
check("md has Concentration + Creative Signals sections",
      "## Concentration" in md_f and "## Creative Signals" in md_f)
check("md omits the sections when blocks absent",
      "## Concentration" not in ma and "## Creative Signals" not in ma)
f1 = audit_html.render_html(audit_model.compute_model(
    PAYLOAD, generated="TS1", concentration=block, creative_signals=cs_block))
f2 = audit_html.render_html(audit_model.compute_model(
    PAYLOAD, generated="TS2", concentration=block, creative_signals=cs_block))
check("HTML with all blocks deterministic modulo generated",
      f1.replace("TS1", "TS") == f2.replace("TS2", "TS"))

# Adversarial client-controlled strings. Account names are client data and flow
# into BOTH renderers' headers; the md's YAML frontmatter is the fragile one (a
# bare quote closes the scalar and the whole block stops parsing, defeating the
# "Obsidian-ingestible" promise).
_adv = copy.deepcopy(PAYLOAD)
_adv["meta"]["account_name"] = 'Acme "Prime" Ltd </script><script>x()</script>'
_m_adv = audit_model.compute_model(_adv, generated="2026-07-10T00:00:00")
_md_adv = audit_md.render_md(_m_adv)
_fm = _md_adv.split("---")[1]
check("md frontmatter escapes quotes in the account name (block stays parseable)",
      '\\"Prime\\"' in _fm and _fm.count('title: "') == 1
      and all(l.count('"') % 2 == 0 for l in _fm.strip().splitlines()
              if ": " in l and '"' in l),
      repr([l for l in _fm.strip().splitlines() if l.startswith("title:")]))
_html_adv = audit_html.render_html(_m_adv)
check("html never emits a literal </script> from client data",
      "</script><script>x()" not in _html_adv)
check("html data block survives the account name (still valid JSON)",
      json.loads(re.search(r'<script id="data" type="application/json">(.*?)</script>',
                           _html_adv, re.S).group(1).replace("<\\/", "</")
                 )["meta"]["account_name"] == _adv["meta"]["account_name"])

# ============================================================================
# 12: workbook — tabs, named ranges, optional tabs, corrected flags, ICE seeds
# ============================================================================
print("--- 12: workbook ---")
if bax is None:
    skip_note("workbook section entirely (openpyxl unavailable)")
else:
    from openpyxl import load_workbook as _lw

    def _cells(ws):
        for row in ws.iter_rows(values_only=True):
            for v in row:
                yield v

    with tempfile.TemporaryDirectory() as td:
        x1 = Path(td) / "plain.xlsx"
        bax.build(json.loads(EXAMPLE.read_text())).save(x1)
        wb1 = _lw(x1)
        check("EXPECTED_TABS all present",
              all(t in wb1.sheetnames for t in bax.EXPECTED_TABS),
              str(wb1.sheetnames))
        check("optional tabs absent when blocks not provided",
              all(t not in wb1.sheetnames for t in bax.OPTIONAL_TABS))
        names = bax.defined_name_set(wb1)
        wanted = {"business_model", "category_weights"}
        for _t, _c, code, _k in bax.ANALYSIS_TABS:
            wanted |= {"wscore_%s" % code, "wbase_%s" % code}
        check("named ranges resolve (wscore/wbase per lever + globals)",
              wanted <= names, str(sorted(wanted - names)))
        check("check() green without optional tabs", bax.check(str(x1)) == 0)

        ice = wb1["10_ICE_Roadmap"]
        seeds = {}
        r = 4
        while ice.cell(row=r, column=1).value:
            seeds[ice.cell(row=r, column=1).value] = (
                ice.cell(row=r, column=5).value, ice.cell(row=r, column=6).value,
                ice.cell(row=r, column=7).value)
            r += 1
        check("ICE numeric seeds for findings missing ICE fields",
              seeds.get("F-102") == (9, 5, 5) and seeds.get("F-104") == (7, 5, 5)
              and seeds.get("F-105") == (3, 5, 5), str(seeds))
        check("explicit ICE fields kept", seeds.get("F-101") == (9, 9, 6))
        check("no #VALUE! anywhere in the workbook",
              not any(isinstance(v, str) and "#VALUE!" in v
                      for ws in wb1.worksheets for v in _cells(ws)))
        check("KPI block renders on the exec summary",
              any(v == "KPI scorecard (informational)"
                  for v in _cells(wb1["01_Executive_Summary"])))

        x2 = Path(td) / "blocks.xlsx"
        bax.build(json.loads(EXAMPLE.read_text()), concentration=block,
                  creative_signals=cs_block).save(x2)
        wb2 = _lw(x2)
        check("12_Concentration + 13_Creative_Signals appear with blocks",
              "12_Concentration" in wb2.sheetnames
              and "13_Creative_Signals" in wb2.sheetnames)
        check("check() green WITH optional tabs", bax.check(str(x2)) == 0)
        check("creative-signals tab carries the band summary",
              any(v == "Fatigue bands (ads)" for v in _cells(wb2["13_Creative_Signals"])))
        check("concentration tab carries the verdict rows",
              any(v == "Verdict" for v in _cells(wb2["12_Concentration"])))

        x3 = Path(td) / "merged.xlsx"
        bax.build(copy.deepcopy(merged)).save(x3)
        wb3 = _lw(x3)
        cp = wb3["06_Creative_Performance"]
        rows_cr = {}
        r = 4
        while cp.cell(row=r, column=1).value:
            rows_cr[cp.cell(row=r, column=1).value] = cp.cell(row=r, column=5).value
            r += 1
        check("corrected FAIL lands in the check cell (CR-06)",
              rows_cr.get("CR-06") == "FAIL" and rows_cr.get("CR-03") == "FLAG",
              str(rows_cr))
        check("check() green on the merged workbook", bax.check(str(x3)) == 0)


# ============================================================================
print("--- 13: differential parity (Python vs RECALCULATED xlsx) ---")
#
# Everything above asserts the three kernels share the same INPUTS (constants).
# This asserts they produce the same OUTPUT for the same payload — which is a
# different claim, and the only one a client can actually see. It exists
# because two real defects lived exactly in that gap while 234 checks stayed
# green:
#   * the Score cell wasn't ROUNDed, but its number_format was, so the grade
#     formula graded a raw 59.99 as D while md/html displayed "60.0 / C";
#   * a flag outside the vocabulary (validate_and_normalize tolerates it with
#     a WARN) made F*D -> #VALUE!, which SUM propagated and the Score cell's
#     IFERROR then rendered as a plausible 0 — a whole lever silently lost.
# Neither is visible to openpyxl, which reads FORMULAS; both need a real
# recalculation. soffice is the gate: skip (never fail) when it is absent, so
# the suite still runs anywhere the plugin installs.
_SOFFICE = next((p for p in (
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice", "/usr/local/bin/soffice",
) if Path(p).exists()), None) or shutil.which("soffice")


def _recalc(xlsx_path, outdir):
    """Round-trip a workbook through LibreOffice so formulas hold VALUES."""
    subprocess.run([_SOFFICE, "--headless", "--convert-to", "xlsx",
                    "--outdir", str(outdir), str(xlsx_path)],
                   check=True, capture_output=True, timeout=180)
    return Path(outdir) / Path(xlsx_path).name


def _rand_payload(rng, n=14):
    """A deterministic pseudo-random payload. Weights are floats so raw health
    lands on cutoff boundaries often enough to catch rounding drift."""
    cats = [c for _code, c, _t, _k in audit_model.SECTIONS]
    checks = []
    for i in range(n):
        checks.append({
            "id": "X-%02d" % i, "category": rng.choice(cats), "name": "c%d" % i,
            "severity": rng.choice(["Critical", "High", "Medium", "Low"]),
            "flag": rng.choice(["PASS", "FLAG", "FAIL", "N/A"]),
            "observed": "", "expected": "", "recommendation": "",
        })
    return {
        "meta": {"account_id": "1", "account_name": "Diff", "currency": "USD",
                 "business_model": "Lead Gen", "generated_for_date": "2026-07-15",
                 "windows": {"structure": "", "creative": "", "trend": ""},
                 "auditor": "t", "out_of_scope": []},
        "category_weights": {c: round(rng.uniform(5, 40), 4) for c in cats},
        "checks": checks, "findings": [],
    }


if _SOFFICE is None:
    skip_note("differential recalc parity (soffice not found)")
else:
    try:
        from openpyxl import load_workbook as _lw2
    except ImportError:
        skip_note("differential recalc parity (openpyxl unavailable)")
    else:
        rng = random.Random(20260715)  # seeded: deterministic, no wall clock
        cases = [_rand_payload(rng) for _ in range(12)]
        # Pin the two regressions explicitly, so they fail loudly by name even
        # if the random sweep drifts. Health lands at exactly 59.9901 raw.
        cases.append({
            "meta": {"account_id": "1", "account_name": "Edge", "currency": "USD",
                     "business_model": "Lead Gen", "generated_for_date": "2026-07-15",
                     "windows": {"structure": "", "creative": "", "trend": ""},
                     "auditor": "t", "out_of_scope": []},
            "category_weights": {"Creative Performance": 59.9901,
                                 "Attribution": 40.0099},
            "checks": [
                {"id": "CR-02", "category": "Creative Performance", "name": "a",
                 "severity": "Critical", "flag": "PASS", "observed": "",
                 "expected": "", "recommendation": ""},
                {"id": "AT-02", "category": "Attribution", "name": "b",
                 "severity": "Critical", "flag": "FAIL", "observed": "",
                 "expected": "", "recommendation": ""}],
            "findings": [],
        })
        cases.append({  # exact .x5 tie: Python round() is half-EVEN (62.2),
            # Excel ROUND is half-AWAY (62.3). round1 makes the kernel share
            # Excel's rule. Weights 62.25/37.75 with a 100/0 split land the raw
            # quotient exactly on the boundary.
            "meta": {"account_id": "1", "account_name": "Tie", "currency": "USD",
                     "business_model": "Lead Gen", "generated_for_date": "2026-07-15",
                     "windows": {"structure": "", "creative": "", "trend": ""},
                     "auditor": "t", "out_of_scope": []},
            "category_weights": {"Creative Performance": 62.25,
                                 "Attribution": 37.75},
            "checks": [
                {"id": "CR-02", "category": "Creative Performance", "name": "a",
                 "severity": "Critical", "flag": "PASS", "observed": "",
                 "expected": "", "recommendation": ""},
                {"id": "AT-02", "category": "Attribution", "name": "b",
                 "severity": "Critical", "flag": "FAIL", "observed": "",
                 "expected": "", "recommendation": ""}],
            "findings": [],
        })
        cases.append({  # drifted flag vocab: excluded from BOTH sums, not #VALUE!
            "meta": {"account_id": "1", "account_name": "Drift", "currency": "USD",
                     "business_model": "Lead Gen", "generated_for_date": "2026-07-15",
                     "windows": {"structure": "", "creative": "", "trend": ""},
                     "auditor": "t", "out_of_scope": []},
            "checks": [
                {"id": "CR-02", "category": "Creative Performance", "name": "a",
                 "severity": "Critical", "flag": "PASS", "observed": "",
                 "expected": "", "recommendation": ""},
                {"id": "CR-05", "category": "Creative Performance", "name": "b",
                 "severity": "Critical", "flag": "PASS", "observed": "",
                 "expected": "", "recommendation": ""},
                {"id": "CR-03", "category": "Creative Performance", "name": "c",
                 "severity": "Critical", "flag": "PARTIAL", "observed": "",
                 "expected": "", "recommendation": ""}],
            "findings": [],
        })
        mismatches = []
        with tempfile.TemporaryDirectory() as td:
            for i, p in enumerate(cases):
                model = audit_model.compute_model(copy.deepcopy(p))
                x = Path(td) / ("d%02d.xlsx" % i)
                bax.build(copy.deepcopy(p)).save(x)
                out = Path(td) / ("r%02d" % i)
                out.mkdir()
                ws = _lw2(_recalc(x, out), data_only=True)["01_Executive_Summary"]
                xl_score, xl_grade = ws["B3"].value, ws["B4"].value
                py_score = model["health"]["score"]
                py_grade = model["health"]["grade"]
                # `float(xl_score or 0)` would turn an UNEVALUATED cell (None)
                # into 0.0 and pass vacuously against an all-N/A payload whose
                # model health is also 0.0 — the sweep must fail, not shrug,
                # when the workbook computed nothing. And the tolerance is exact
                # (1e-9, not 0.05): both sides now round with audit_model.round1
                # / Excel ROUND, so ANY difference is a real divergence. A loose
                # tolerance here would have hidden exactly the .x5 tie this
                # section exists to catch.
                if xl_score is None:
                    mismatches.append("case %d: xlsx B3 did not evaluate" % i)
                elif not approx(float(xl_score), py_score, 1e-9) \
                        or xl_grade != py_grade:
                    mismatches.append("case %d: python %r/%s vs xlsx %r/%s"
                                      % (i, py_score, py_grade, xl_score, xl_grade))
        check("recalculated xlsx health+grade == audit_model for all %d payloads"
              % len(cases), not mismatches, "; ".join(mismatches[:4]))

        # The sweep above compares the HEADLINE only, which is how a lever score
        # displaying 11.25 beside the md's 11.3 slipped through: D's value is
        # deliberately unrounded (health reads it), so parity there is about the
        # DISPLAY rule. Assert the number_format that carries it.
        tie = {"meta": {"account_id": "1", "account_name": "Disp", "currency": "USD",
                        "business_model": "Lead Gen", "generated_for_date": "2026-07-15",
                        "windows": {"structure": "", "creative": "", "trend": ""},
                        "auditor": "t", "out_of_scope": []},
               "checks": [
                   {"id": "CR-05", "category": "Creative Performance", "name": "b",
                    "severity": "High", "flag": "FLAG"},
                   {"id": "CR-02", "category": "Creative Performance", "name": "a",
                    "severity": "Medium", "flag": "FLAG"},
                   {"id": "CR-03", "category": "Creative Performance", "name": "c",
                    "severity": "Critical", "flag": "FAIL"},
                   {"id": "CR-06", "category": "Creative Performance", "name": "d",
                    "severity": "Critical", "flag": "FAIL"},
                   {"id": "CR-07", "category": "Creative Performance", "name": "e",
                    "severity": "Critical", "flag": "FAIL"},
                   {"id": "CR-08", "category": "Creative Performance", "name": "f",
                    "severity": "Low", "flag": "FAIL"}],
               "findings": []}
        with tempfile.TemporaryDirectory() as td2:
            xt = Path(td2) / "disp.xlsx"
            bax.build(copy.deepcopy(tie)).save(xt)
            ws_f = _lw2(xt)["01_Executive_Summary"]
            fmts, raw = {}, {}
            for rr2 in range(7, 14):
                nm = ws_f.cell(row=rr2, column=1).value
                if nm:
                    fmts[nm] = ws_f.cell(row=rr2, column=4).number_format
            out2 = Path(td2) / "rc"
            out2.mkdir()
            ws_v = _lw2(_recalc(xt, out2), data_only=True)["01_Executive_Summary"]
            for rr2 in range(7, 14):
                nm = ws_v.cell(row=rr2, column=1).value
                if nm:
                    raw[nm] = ws_v.cell(row=rr2, column=4).value
            model_t = audit_model.compute_model(copy.deepcopy(tie))
            cr_t = [s for s in model_t["sections"] if s["code"] == "CR"][0]
            check("lever Score cell keeps the UNROUNDED value (health SUMPRODUCT reads it)",
                  approx(raw.get("Creative Performance"), 11.25, 1e-9),
                  str(raw.get("Creative Performance")))
            check("lever Score cell rounds for DISPLAY to match the model's score_pct",
                  fmts.get("Creative Performance") == "0.0" and cr_t["score_pct"] == 11.3,
                  "fmt=%r score_pct=%r" % (fmts.get("Creative Performance"),
                                           cr_t["score_pct"]))


# ============================================================================
print("--- 14: build_audit end-to-end (the orchestrator seam) ---")
#
# Every section above tests a leaf module. build_audit.py — which uniquely owns
# window plumbing, raw/csv exclusion, the prescore->model wiring and the final
# JSON line — had NO coverage, and that is exactly where the honest-window
# violation lived: it seeded concentration's window labels from the payload's
# REQUESTED preset ("last_30d"), so a 7-day pull could be labelled "last_30d".
# Unit fixtures could never catch it because they hand compute_concentration
# the correct labels the real caller never produced.
#
# The fixtures make a clean discriminator: sample-payload's meta.windows say
# "last_30d"/"last_90d" while the captured pulls carry real dates.
_BUILD = SCRIPTS / "build_audit.py"
with tempfile.TemporaryDirectory() as td:
    raw = Path(td) / "raw"
    raw.mkdir()
    for key, fname in (("campaigns", "raw_shape_campaigns.json"),
                       ("adsets", "raw_shape_adsets.json"),
                       ("ads", "raw_shape_ads.json")):
        shutil.copy(FIX / fname, raw / (key + ".json"))
    out = Path(td) / "out"
    proc = subprocess.run(
        [sys.executable, str(_BUILD), "--input", str(HERE / "sample-payload.json"),
         "--outdir", str(out), "--brand", "E2E", "--formats", "md,html",
         "--raw-dir", str(raw), "--business-model", "Lead Gen",
         "--no-downloads"],  # never write to a real ~/Downloads from tests
        capture_output=True, text=True, timeout=180)
    check("build_audit exits 0 with raw pulls + payload", proc.returncode == 0,
          proc.stderr[-400:])
    tail = json.loads(proc.stdout.strip().splitlines()[-1])
    check("final stdout line is the JSON contract "
          "(outputs/health/grade/checks/findings/prescore_corrections)",
          {"outputs", "health", "grade", "checks", "findings",
           "prescore_corrections"} <= set(tail), proc.stdout[-200:])
    html_e2e = Path(tail["outputs"]["html"]).read_text(encoding="utf-8")
    model_e2e = json.loads(re.search(
        r'<script id="data" type="application/json">(.*?)</script>',
        html_e2e, re.S).group(1).replace("<\\/", "</"))
    wins = {d["key"]: d["window"] for d in model_e2e["concentration"]["dimensions"]}
    check("concentration windows are DATA-derived, never the requested preset",
          all(re.match(r"^\d{4}-\d{2}-\d{2} – \d{4}-\d{2}-\d{2}$", w)
              for w in wins.values()), str(wins))
    check("no dimension is labelled with a preset string from meta.windows",
          not any(w in ("last_30d", "last_90d", "last_7d") for w in wins.values()),
          str(wins))
    check("an unknown --formats value fails loudly instead of writing nothing "
          "and exiting 0",
          subprocess.run(
              [sys.executable, str(_BUILD), "--input", str(HERE / "sample-payload.json"),
               "--outdir", str(out), "--formats", "pdf", "--no-downloads"],
              capture_output=True, text=True, timeout=120).returncode == 1)
    check("the ads window reports the pull's OWN captured span",
          wins.get("ads") == "2026-04-12 – 2026-07-10", str(wins.get("ads")))
    check("--no-downloads honoured: bundle written only to --outdir",
          sorted(p.suffix for p in out.iterdir()) == [".html", ".md"],
          str(sorted(p.name for p in out.iterdir())))
    # The raw/csv exclusion is a guard the orchestrator alone enforces.
    bad = subprocess.run(
        [sys.executable, str(_BUILD), "--input", str(HERE / "sample-payload.json"),
         "--outdir", str(out), "--formats", "md", "--raw-dir", str(raw),
         "--csv-dir", str(raw), "--no-downloads"],
        capture_output=True, text=True, timeout=120)
    check("--raw-* and --csv-* are mutually exclusive (exit 1, named reason)",
          bad.returncode == 1 and "not both" in bad.stderr, bad.stderr[-200:])

print()
if SKIPS:
    print(f"skipped: {len(SKIPS)} block(s) — " + "; ".join(SKIPS))
if FAILS:
    print(f"FAILED ({len(FAILS)} of {N_CHECKS}): " + ", ".join(FAILS))
    sys.exit(1)
print(f"All {N_CHECKS} meta-ads-audit conformance tests passed.")
sys.exit(0)

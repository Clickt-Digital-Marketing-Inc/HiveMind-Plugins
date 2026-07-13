#!/usr/bin/env python3
"""
build_cro_workbook.py — Shopify CRO Audit workbook generator (11-step framework).

Reads a single cro-payload.json (assembled by the shopify-cro-audit skill after it ingests the
Shopify/GA4 CSVs, mines reviews/surveys, and WebFetches the storefront + competitors) and writes a
deterministic, formula-driven .xlsx that is both an auditor working tool and a client-facing report.

Usage
-----
  python3 build_cro_workbook.py --input cro-payload.json --output shopify-cro-audit-acme-2026-06-25.xlsx
  python3 build_cro_workbook.py --check shopify-cro-audit-acme-2026-06-25.xlsx

Design notes
------------
* The payload is the single authoritative input. The skill TRANSCRIBES the analytics figures that
  GA4/Shopify already report (it does not re-derive rates), and does the qualitative analysis
  (review mining, heuristics, surveys). This script renders those into the workbook and expresses all
  GRADING + PRIORITIZATION as live Excel formulas, so an auditor can override an input and the
  workbook recomputes on open:
    - Funnel Health Score grades the funnel rates against the named benchmark cells.
    - Triangulation count is a live formula over each finding's comma-separated step sources.
    - Roadmap Priority is the framework's (Impact × 2) + Ease — NOT ICE (the framework dropped
      Confidence on purpose).
* Steps with no data are rendered as structured templates and stamped "Not run — data not provided"
  on the Scope tab and the step tab. Nothing is fabricated.
* Only openpyxl (>= 3.0) is required. No network, no other deps. Output is deterministic.
* Optional values-only report tabs 15_Concentration / 16_CVR_Signals are added when build()
  receives the concentration= / cvr_signals= blocks (computed by concentration.py /
  cvr_signals.py from raw pull files or CSV exports — never from the payload). They are
  deliberately NOT in EXPECTED_TABS, so --check stays green with or without them.
* --check is a STRUCTURAL gate (tabs present, named ranges resolve, no literal #REF!, vocab sane).
  It cannot evaluate formula *values* — open in Excel/LibreOffice for that (it says so).
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.stderr.write("ERROR: openpyxl is required. Install with: python3 -m pip install 'openpyxl>=3.1'\n")
    sys.exit(2)

# Scoring constants and controlled vocabularies are single-sourced in audit_model.py so the
# xlsx cell formulas, the markdown record and the HTML explorer can never disagree (parity is
# asserted in tests/test_audit.py). This module imports them back and keeps NO local copies.
# SEVERITY_IMPACT is re-exported for the same parity assertions (the workbook itself renders
# the payload's Impact values verbatim).
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from audit_model import (  # noqa: E402,F401
    BENCH, SEV_CANON, CHANGE_CANON, STATUS_CANON, SEVERITY_IMPACT,
)

# ---------------------------------------------------------------------------- config

ALLOWED_SEVERITIES = set(SEV_CANON.values())
ALLOWED_CHANGE = set(CHANGE_CANON.values())

# The 11 analysis methods → step tab title. Order matters (tabs are sorted by EXPECTED_TABS).
STEP_TABS = [
    (1, "02_Analytics", "Step 1 — GA4 & Shopify Analytics"),
    (2, "03_Heuristic_LIFT", "Step 2 — Heuristic Analysis (LIFT Model)"),
    (3, "04_Review_Mining", "Step 3 — Review Mining"),
    (4, "05_Customer_Support", "Step 4 — Customer Support Analysis"),
    (5, "06_Heatmaps", "Step 5 — Heatmap & Scrollmap Analysis"),
    (6, "07_PostPurchase_Survey", "Step 6 — Post-Purchase Survey Analysis"),
    (7, "08_Email_Survey", "Step 7 — Email Long Survey Analysis"),
    (8, "09_User_Testing", "Step 8 — User Testing Analysis"),
    (9, "10_Marketing_Match", "Step 9 — Marketing Strategy Analysis"),
    (10, "11_Competitor", "Step 10 — Competitor Analysis"),
    (11, "13_Roadmap", "Step 11 — Testing Roadmap"),
]

EXPECTED_TABS = [
    "00_Audit_Scope", "01_Executive_Summary", "02_Analytics", "03_Heuristic_LIFT",
    "04_Review_Mining", "05_Customer_Support", "06_Heatmaps", "07_PostPurchase_Survey",
    "08_Email_Survey", "09_User_Testing", "10_Marketing_Match", "11_Competitor",
    "12_Findings_Log", "13_Roadmap", "14_Reference",
]

# Values-only report tabs added by build(..., concentration=, cvr_signals=).
# Deliberately NOT in EXPECTED_TABS: check() must stay green with or without them,
# and they append strictly AFTER 14_Reference (no layout change to any existing tab —
# rate_atc/rate_checkout/rate_cvr stay pinned to 02_Analytics C6/C7/C8).
OPTIONAL_TABS = ["15_Concentration", "16_CVR_Signals"]

REQUIRED_NAMES = {
    "bench_atc", "bench_checkout", "bench_cvr", "bench_mobile", "bench_desktop",
    "rate_atc", "rate_checkout", "rate_cvr", "funnel_health", "roadmap_priority",
}

# ------------------------------------------------------------------------- styling

NAVY = "FF1E293B"
SLATE = "FFE2E8F0"
HILITE = "FFFFF7CD"
GREEN = "FFC6EFCE"
AMBER = "FFFFEB9C"
RED = "FFFFC7CE"
GREY = "FFE5E7EB"
BORDER_CLR = "FFCBD5E1"

TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="FFFFFFFF")
HEAD_FONT = Font(name="Calibri", size=10, bold=True, color="FF0F172A")
BODY_FONT = Font(name="Calibri", size=10, color="FF0F172A")
MUTED_FONT = Font(name="Calibri", size=9, italic=True, color="FF64748B")
BIG_FONT = Font(name="Calibri", size=28, bold=True, color="FF0F172A")
SECT_FONT = Font(name="Calibri", size=11, bold=True, color="FF0F172A")

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FILL = PatternFill("solid", fgColor=SLATE)
HILITE_FILL = PatternFill("solid", fgColor=HILITE)
SECT_FILL = PatternFill("solid", fgColor="FFF1F5F9")
NOTRUN_FILL = PatternFill("solid", fgColor=GREY)
GREEN_FILL = PatternFill("solid", fgColor=GREEN)
AMBER_FILL = PatternFill("solid", fgColor=AMBER)
RED_FILL = PatternFill("solid", fgColor=RED)
_thin = Side(style="thin", color=BORDER_CLR)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

TAB_COLORS = {
    "00_Audit_Scope": "1E40AF", "01_Executive_Summary": "B91C1C",
    "02_Analytics": "CA8A04", "03_Heuristic_LIFT": "CA8A04", "04_Review_Mining": "CA8A04",
    "05_Customer_Support": "CA8A04", "06_Heatmaps": "CA8A04", "07_PostPurchase_Survey": "CA8A04",
    "08_Email_Survey": "CA8A04", "09_User_Testing": "CA8A04", "10_Marketing_Match": "CA8A04",
    "11_Competitor": "CA8A04", "12_Findings_Log": "15803D", "13_Roadmap": "15803D",
    "14_Reference": "475569",
    "15_Concentration": "475569", "16_CVR_Signals": "475569",
}

# Number formats for the values-only report tabs. Rates arriving from cvr_signals /
# concentration are FRACTIONS (SHAPE-NOTES pin) — FMT_RATE lets Excel render
# 0.019804 as 1.98% without the value ever being multiplied by 100.
FMT_MONEY = "#,##0.00"
FMT_INT = "#,##0"
FMT_SHARE = "0.0%"
FMT_RATE = "0.00%"
FMT_Z = "0.00"


# ------------------------------------------------------------------------- helpers

def q(sheet):
    return "'%s'" % sheet.replace("'", "''")


def add_name(wb, name, ref):
    dn = DefinedName(name, attr_text=ref)
    try:
        wb.defined_names[name] = dn          # openpyxl >= 3.1 (dict-like)
    except (TypeError, AttributeError):
        wb.defined_names.add(dn)             # openpyxl 3.0 (list-like)


def defined_name_set(wb):
    try:
        return set(wb.defined_names.keys())
    except AttributeError:
        return {dn.name for dn in wb.defined_names.definedName}


def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = CTR
        c.border = BORDER


def set_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


def section_label(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECT_FONT
    c.fill = SECT_FILL


def banner(ws, row, text, ncols, fill=NOTRUN_FILL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, bold=True, color="FF334155")
    c.fill = fill
    c.alignment = WRAP


def write_table(ws, row, columns, rows, widths=None, wrap_cols=None):
    """Generic evidence table. Returns the next free row."""
    header_row(ws, row, columns)
    if widths:
        set_widths(ws, widths)
    wrap_cols = wrap_cols or set(range(1, len(columns) + 1))
    r = row + 1
    if not rows:
        ws.cell(row=r, column=1, value="(not provided)").font = MUTED_FONT
        return r + 1
    for record in rows:
        for i, val in enumerate(record):
            c = ws.cell(row=r, column=1 + i, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = WRAP if (i + 1) in wrap_cols else CTR
        r += 1
    return r


def step_status(meta, step_no):
    for s in meta.get("steps", []) or []:
        if int(s.get("step", -1)) == step_no:
            st = STATUS_CANON.get(str(s.get("status", "")).strip().lower(), s.get("status", ""))
            return st, s.get("reason", "")
    return "", ""


def notrun_banner_if_needed(ws, meta, step_no, ncols):
    """Stamp a status banner at row 2 when a step is not_run/partial/absent. Returns True if blocked."""
    st, reason = step_status(meta, step_no)
    if st == "run":
        return False
    if st == "partial":
        banner(ws, 2, "PARTIAL — limited data provided. %s" % (reason or ""), ncols, fill=AMBER_FILL)
        return False
    msg = "NOT RUN — data not provided." + ((" " + reason) if reason else "")
    banner(ws, 2, msg + "  Template retained below for when the data arrives.", ncols)
    return True


# --------------------------------------------------------------------- tab builders

def build_scope(wb, meta):
    ws = wb.create_sheet("00_Audit_Scope")
    title_row(ws, "Shopify CRO Audit — Scope & Provenance", 4)
    set_widths(ws, [26, 34, 16, 46])

    rows = [
        ("Store name", meta.get("store_name", "")),
        ("Store URL", meta.get("store_url", "")),
        ("Currency", meta.get("currency", "")),
        ("Date range", meta.get("date_range", "")),
        ("Generated for", meta.get("generated_for_date", "")),
        ("Auditor", meta.get("auditor", "")),
        ("Framework", "11-step CRO audit checklist for ecommerce"),
        ("Prioritization", "Priority = (Impact × 2) + Ease  (1–10 each; ICE deliberately dropped)"),
        ("Method", "Organized by analysis method, not by page type. Triangulated across sources."),
    ]
    r = 3
    for label, val in rows:
        a = ws.cell(row=r, column=1, value=label); a.font = HEAD_FONT; a.border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        b = ws.cell(row=r, column=2, value=val); b.font = BODY_FONT; b.alignment = WRAP; b.border = BORDER
        r += 1
    add_name(wb, "store_url", "%s!$B$4" % q("00_Audit_Scope"))

    r += 1
    section_label(ws, r, "Data inventory (what was provided vs. missing)", 4); r += 1
    header_row(ws, r, ["Dataset", "Status", "", "Notes"]); r += 1
    inv = meta.get("data_inventory") or []
    if not inv:
        ws.cell(row=r, column=1, value="(no data inventory supplied)").font = MUTED_FONT; r += 1
    for item in inv:
        ws.cell(row=r, column=1, value=item.get("dataset", "")).font = BODY_FONT
        st = str(item.get("status", "")).strip().lower()
        sc = ws.cell(row=r, column=2, value="Provided" if st == "provided" else "Missing")
        sc.font = BODY_FONT
        sc.fill = PatternFill("solid", fgColor=GREEN if st == "provided" else RED)
        sc.alignment = CTR
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=4)
        nt = ws.cell(row=r, column=4, value=item.get("notes", "")); nt.font = BODY_FONT; nt.alignment = WRAP
        for col in (1, 2, 4):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    r += 1
    section_label(ws, r, "11-step coverage (run / partial / not run)", 4); r += 1
    header_row(ws, r, ["Step", "Analysis method", "Status", "Reason / note"]); r += 1
    steps = meta.get("steps") or []
    if not steps:
        ws.cell(row=r, column=1, value="(no step coverage supplied)").font = MUTED_FONT; r += 1
    for s in steps:
        ws.cell(row=r, column=1, value=s.get("step", "")).font = BODY_FONT
        ws.cell(row=r, column=2, value=s.get("name", "")).font = BODY_FONT
        st = STATUS_CANON.get(str(s.get("status", "")).strip().lower(), s.get("status", ""))
        label = {"run": "Run", "partial": "Partial", "not_run": "Not run"}.get(st, st or "—")
        sc = ws.cell(row=r, column=3, value=label); sc.font = BODY_FONT; sc.alignment = CTR
        sc.fill = PatternFill("solid", fgColor={"run": GREEN, "partial": AMBER, "not_run": GREY}.get(st, "FFFFFFFF"))
        rc = ws.cell(row=r, column=4, value=s.get("reason", "")); rc.font = BODY_FONT; rc.alignment = WRAP
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1

    ws.sheet_view.showGridLines = False
    return ws


def build_analytics(wb, analytics):
    ws = wb.create_sheet("02_Analytics")
    title_row(ws, "Step 1 — GA4 & Shopify Analytics", 6)
    set_widths(ws, [34, 14, 16, 16, 16, 30])
    a = analytics or {}
    funnel = a.get("funnel", {}) or {}
    blocked = not funnel and not a

    # Conversion funnel (rate cells become named ranges rate_atc/checkout/cvr at fixed rows 6/7/8)
    section_label(ws, 3, "Conversion funnel (vs FY2025 DTC benchmarks)", 6)
    header_row(ws, 4, ["Stage", "Count", "% of sessions", "Benchmark %", "Index (100 = bench)", "Read"])
    stages = [
        ("Sessions", funnel.get("sessions", ""), None, None),
        ("Added to cart", funnel.get("atc", ""), funnel.get("atc_rate", ""), "bench_atc"),
        ("Reached checkout", funnel.get("checkout", ""), funnel.get("checkout_rate", ""), "bench_checkout"),
        ("Completed purchase", funnel.get("purchases", ""), funnel.get("cvr", ""), "bench_cvr"),
    ]
    r = 5
    for name, count, rate, bench_name in stages:
        ws.cell(row=r, column=1, value=name).font = BODY_FONT
        cc = ws.cell(row=r, column=2, value=count); cc.font = BODY_FONT; cc.alignment = CTR
        if bench_name is None:  # Sessions row = the 100% base
            rc = ws.cell(row=r, column=3, value=100 if count != "" else "")
            rc.number_format = "0.0"; rc.alignment = CTR; rc.font = BODY_FONT
            ws.cell(row=r, column=6, value="Funnel base").font = MUTED_FONT
        else:
            rc = ws.cell(row=r, column=3, value=rate); rc.font = BODY_FONT; rc.alignment = CTR
            rc.fill = HILITE_FILL; rc.number_format = "0.00"
            bc = ws.cell(row=r, column=4, value="=%s" % bench_name); bc.alignment = CTR; bc.number_format = "0.00"
            ic = ws.cell(row=r, column=5, value="=IFERROR(C{r}/D{r}*100,\"\")".format(r=r))
            ic.alignment = CTR; ic.number_format = "0"
            ws.cell(row=r, column=6,
                    value='=IF(C{r}="","",IF(C{r}>=D{r},"At / above benchmark",'
                          'IF(C{r}>=D{r}*0.7,"Below benchmark","Well below benchmark")))'.format(r=r)).font = BODY_FONT
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER
        r += 1
    add_name(wb, "rate_atc", "%s!$C$6" % q("02_Analytics"))
    add_name(wb, "rate_checkout", "%s!$C$7" % q("02_Analytics"))
    add_name(wb, "rate_cvr", "%s!$C$8" % q("02_Analytics"))

    r = 10
    section_label(ws, r, "Device segmentation (frame the mobile gap as intent, not broken UX)", 6); r += 1
    r = write_table(ws, r, ["Device", "Sessions", "CVR %", "Benchmark %", "Index", ""],
                    [[d.get("device", ""), d.get("sessions", ""), d.get("cvr", ""),
                      ("=bench_mobile" if str(d.get("device", "")).lower().startswith("mob")
                       else ("=bench_desktop" if str(d.get("device", "")).lower().startswith("desk") else "")),
                      ""]
                     for d in (a.get("device") or [])],
                    wrap_cols={6})

    r += 1
    section_label(ws, r, "Acquisition — conversion rate by channel", 6); r += 1
    r = write_table(ws, r, ["Channel", "Sessions", "CVR %", "Revenue", "", ""],
                    [[c.get("channel", ""), c.get("sessions", ""), c.get("cvr", ""), c.get("revenue", ""), "", ""]
                     for c in (a.get("channels") or [])], wrap_cols={6})

    r += 1
    section_label(ws, r, "Top landing pages (traffic concentration → where to focus)", 6); r += 1
    r = write_table(ws, r, ["Landing page", "Sessions", "Share %", "", "", ""],
                    [[p.get("page", ""), p.get("sessions", ""), p.get("share_pct", ""), "", "", ""]
                     for p in (a.get("landing_pages") or [])], wrap_cols={1})

    r += 1
    section_label(ws, r, "Revenue concentration (hero SKUs)", 6); r += 1
    r = write_table(ws, r, ["Product", "Revenue", "Share %", "", "", ""],
                    [[p.get("product", ""), p.get("revenue", ""), p.get("share_pct", ""), "", "", ""]
                     for p in (a.get("revenue_concentration") or [])], wrap_cols={1})

    r += 1
    nvr = a.get("new_vs_returning", {}) or {}
    section_label(ws, r, "New vs. returning & AOV", 6); r += 1
    pairs = [
        ("New-visitor CVR %", nvr.get("new_cvr", "")),
        ("Returning-visitor CVR %", nvr.get("returning_cvr", "")),
        ("Average order value", a.get("aov", "")),
    ]
    for label, val in pairs:
        ws.cell(row=r, column=1, value=label).font = HEAD_FONT
        vc = ws.cell(row=r, column=2, value=val); vc.font = BODY_FONT; vc.alignment = CTR
        ws.cell(row=r, column=1).border = BORDER; vc.border = BORDER
        r += 1
    ws.cell(row=r, column=1, value="AOV CVR band").font = HEAD_FONT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2,
            value='=IF(B{a}="","",IF(B{a}<60,"Sub-$60 band — peers median CVR ~4.63%",'
                  'IF(B{a}<=200,"$60–$200 band","Over-$200 band — peers median CVR ~0.95%")))'.format(a=r - 1)).font = BODY_FONT

    if blocked:
        banner(ws, 2, "NOT RUN — no analytics provided. Funnel/benchmark formulas stay live; "
                      "fill the highlighted cells from GA4 + Shopify and the workbook grades on open.", 6)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return ws


def build_heuristic(wb, detail, meta):
    ws = wb.create_sheet("03_Heuristic_LIFT")
    title_row(ws, "Step 2 — Heuristic Analysis (LIFT Model)", 5)
    notrun_banner_if_needed(ws, meta, 2, 5)
    section_label(ws, 3, "LIFT factors: Value Proposition · Relevance · Clarity · Urgency · Anxiety · Distraction", 5)
    rows = [[f.get("page", ""), f.get("lift_factor", ""), f.get("severity", ""),
             f.get("observed", ""), f.get("recommendation", "")]
            for f in ((detail or {}).get("findings") or [])]
    write_table(ws, 4, ["Page / section", "LIFT factor", "Severity", "Observed", "Test recommendation"],
                rows, widths=[26, 18, 11, 46, 46], wrap_cols={1, 4, 5})
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def build_review_mining(wb, detail, meta):
    ws = wb.create_sheet("04_Review_Mining")
    title_row(ws, "Step 3 — Review Mining", 4)
    notrun_banner_if_needed(ws, meta, 3, 4)
    set_widths(ws, [40, 14, 40, 40])
    d = detail or {}
    section_label(ws, 3, "Theme distribution (% of reviews mentioning — may exceed 100%)", 4)
    r = write_table(ws, 4, ["Theme", "% of reviews", "", ""],
                    [[t.get("theme", ""), t.get("pct", ""), "", ""] for t in (d.get("themes") or [])],
                    wrap_cols={1})
    r += 1
    section_label(ws, r, "Objections (what almost stopped them)", 4); r += 1
    r = write_table(ws, r, ["Objection", "", "", ""], [[o, "", "", ""] for o in (d.get("objections") or [])],
                    wrap_cols={1})
    r += 1
    section_label(ws, r, "Purchase drivers (what convinced them)", 4); r += 1
    r = write_table(ws, r, ["Driver", "", "", ""], [[o, "", "", ""] for o in (d.get("drivers") or [])],
                    wrap_cols={1})
    r += 1
    section_label(ws, r, "Customer voice (verbatim phrases to reuse in copy)", 4); r += 1
    write_table(ws, r, ["Phrase", "", "", ""], [[o, "", "", ""] for o in (d.get("voice") or [])], wrap_cols={1})
    ws.sheet_view.showGridLines = False
    return ws


def build_support(wb, detail, meta):
    ws = wb.create_sheet("05_Customer_Support")
    title_row(ws, "Step 4 — Customer Support Analysis", 3)
    notrun_banner_if_needed(ws, meta, 4, 3)
    rows = [[r.get("question_or_complaint", ""), r.get("category", ""), r.get("site_gap", "")]
            for r in ((detail or {}).get("rows") or [])]
    write_table(ws, 3, ["Question / complaint", "Category", "Site gap → proactive fix"], rows,
                widths=[50, 22, 50], wrap_cols={1, 3})
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return ws


def build_heatmaps(wb, detail, meta):
    ws = wb.create_sheet("06_Heatmaps")
    title_row(ws, "Step 5 — Heatmap & Scrollmap Analysis", 4)
    notrun_banner_if_needed(ws, meta, 5, 4)
    rows = [[r.get("page", ""), r.get("device", ""), r.get("metric", ""), r.get("observation", "")]
            for r in ((detail or {}).get("rows") or [])]
    write_table(ws, 3, ["Page", "Device", "Metric (scroll / click / dead zone)", "Observation → action"],
                rows, widths=[26, 12, 30, 52], wrap_cols={1, 3, 4})
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return ws


def build_postpurchase(wb, detail, meta):
    ws = wb.create_sheet("07_PostPurchase_Survey")
    title_row(ws, "Step 6 — Post-Purchase Survey Analysis", 4)
    notrun_banner_if_needed(ws, meta, 6, 4)
    set_widths(ws, [40, 14, 16, 40])
    d = detail or {}
    section_label(ws, 3, "What nearly stopped you from ordering (active conversion killers)", 4)
    r = write_table(ws, 4, ["Near-abandonment factor", "", "", ""],
                    [[x, "", "", ""] for x in (d.get("near_abandonment") or [])], wrap_cols={1})
    r += 1
    section_label(ws, r, "What convinced you to order (triggers to amplify)", 4); r += 1
    r = write_table(ws, r, ["Purchase trigger", "", "", ""],
                    [[x, "", "", ""] for x in (d.get("triggers") or [])], wrap_cols={1})
    r += 1
    section_label(ws, r, "Attribution check — survey 'how did you hear' vs GA4", 4); r += 1
    write_table(ws, r, ["Channel", "Survey %", "GA4 %", "Gap / note"],
                [[a.get("channel", ""), a.get("survey_pct", ""), a.get("ga4_pct", ""), a.get("note", "")]
                 for a in (d.get("attribution") or [])], wrap_cols={4})
    ws.sheet_view.showGridLines = False
    return ws


def build_email_survey(wb, detail, meta):
    ws = wb.create_sheet("08_Email_Survey")
    title_row(ws, "Step 7 — Email Long Survey Analysis", 3)
    notrun_banner_if_needed(ws, meta, 7, 3)
    rows = [[r.get("insight_type", ""), r.get("finding", ""), r.get("pct_or_n", "")]
            for r in ((detail or {}).get("rows") or [])]
    write_table(ws, 3, ["Insight type", "Finding", "% / n"], rows,
                widths=[26, 64, 14], wrap_cols={1, 2})
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return ws


def build_user_testing(wb, detail, meta):
    ws = wb.create_sheet("09_User_Testing")
    title_row(ws, "Step 8 — User Testing Analysis", 3)
    notrun_banner_if_needed(ws, meta, 8, 3)
    rows = [[r.get("tester_or_theme", ""), r.get("quote", ""), r.get("issue", "")]
            for r in ((detail or {}).get("rows") or [])]
    write_table(ws, 3, ["Tester / theme", "Verbatim quote", "Friction → fix"], rows,
                widths=[22, 56, 46], wrap_cols={2, 3})
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return ws


def build_marketing(wb, detail, meta):
    ws = wb.create_sheet("10_Marketing_Match")
    title_row(ws, "Step 9 — Marketing Strategy Analysis", 3)
    notrun_banner_if_needed(ws, meta, 9, 3)
    rows = [[r.get("area", ""), r.get("observed", ""), r.get("gap", "")]
            for r in ((detail or {}).get("rows") or [])]
    write_table(ws, 3, ["Area (ad-match / promo / channel gap)", "Observed", "Gap → action"], rows,
                widths=[30, 50, 46], wrap_cols={1, 2, 3})
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return ws


def build_competitor(wb, detail, meta):
    ws = wb.create_sheet("11_Competitor")
    title_row(ws, "Step 10 — Competitor Analysis", 6)
    notrun_banner_if_needed(ws, meta, 10, 6)
    d = detail or {}
    r = 3
    offer = d.get("offer_table") or {}
    section_label(ws, r, "Offer & pricing comparison", 6); r += 1
    cols = offer.get("columns") or ["Brand", "Price", "Subscription discount", "Bundles", "Guarantee", "Shipping"]
    r = write_table(ws, r, cols[:6] + [""] * (6 - len(cols[:6])), offer.get("rows") or [], wrap_cols=set(range(1, 7)))
    r += 1
    atf = d.get("atf") or {}
    section_label(ws, r, "Above-the-fold comparison (mobile)", 6); r += 1
    cols = atf.get("columns") or ["Brand", "Headline", "Star rating", "Trust badges", "Primary CTA", "Notes"]
    r = write_table(ws, r, cols[:6] + [""] * (6 - len(cols[:6])), atf.get("rows") or [], wrap_cols=set(range(1, 7)))
    r += 1
    section_label(ws, r, "Messaging gaps / untapped angles", 6); r += 1
    write_table(ws, r, ["Gap or angle no competitor is using", "", "", "", "", ""],
                [[x, "", "", "", "", ""] for x in (d.get("messaging_gaps") or [])], wrap_cols={1})
    ws.sheet_view.showGridLines = False
    return ws


def build_findings(wb, findings):
    ws = wb.create_sheet("12_Findings_Log")
    title_row(ws, "Findings Log — triangulated across the 11 analyses", 9)
    heads = ["ID", "Title", "Step sources", "# sources", "Severity", "Page", "Evidence", "Recommendation", "Change"]
    header_row(ws, 3, heads)
    set_widths(ws, [9, 30, 24, 9, 10, 18, 40, 44, 9])
    r = 4
    for f in findings or []:
        ws.cell(row=r, column=1, value=f.get("id", ""))
        ws.cell(row=r, column=2, value=f.get("title", ""))
        ws.cell(row=r, column=3, value=", ".join(f.get("step_sources", []) or []))
        # live triangulation count over the comma-separated sources string
        ws.cell(row=r, column=4,
                value='=IF(C{r}="",0,LEN(C{r})-LEN(SUBSTITUTE(C{r},",",""))+1)'.format(r=r))
        ws.cell(row=r, column=5, value=f.get("severity", ""))
        ws.cell(row=r, column=6, value=f.get("page", ""))
        ws.cell(row=r, column=7, value=f.get("evidence", ""))
        ws.cell(row=r, column=8, value=f.get("recommendation", ""))
        ws.cell(row=r, column=9, value=f.get("change_type", ""))
        for col in range(1, 10):
            c = ws.cell(row=r, column=col); c.font = BODY_FONT; c.border = BORDER
            c.alignment = WRAP if col in (2, 3, 6, 7, 8) else CTR
        r += 1
    if not findings:
        ws.cell(row=4, column=1, value="(no findings logged)").font = MUTED_FONT
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def _priority(f):
    try:
        return int(f.get("impact", 0)) * 2 + int(f.get("ease", 0))
    except (TypeError, ValueError):
        return 0


def build_roadmap(wb, findings):
    ws = wb.create_sheet("13_Roadmap")
    title_row(ws, "Step 11 — Testing Roadmap   (Priority = Impact × 2 + Ease)", 9)
    heads = ["ID", "Title", "Severity", "Impact (1-10)", "Ease (1-10)", "Priority",
             "Bucket", "Test / Ship", "Expected lever"]
    header_row(ws, 3, heads)
    set_widths(ws, [9, 34, 10, 13, 12, 10, 16, 11, 30])
    ranked = sorted(findings or [], key=_priority, reverse=True)
    first = 4
    r = first
    for f in ranked:
        ws.cell(row=r, column=1, value=f.get("id", ""))
        ws.cell(row=r, column=2, value=f.get("title", ""))
        ws.cell(row=r, column=3, value=f.get("severity", ""))
        ic = ws.cell(row=r, column=4, value=f.get("impact", "")); ic.fill = HILITE_FILL
        ec = ws.cell(row=r, column=5, value=f.get("ease", "")); ec.fill = HILITE_FILL
        ws.cell(row=r, column=6, value='=IF(OR(D{r}="",E{r}=""),"",D{r}*2+E{r})'.format(r=r))
        ws.cell(row=r, column=7,
                value='=IF(F{r}="","",IF(F{r}>=24,"Now",IF(F{r}>=20,"Next",IF(F{r}>=15,"Soon","Later"))))'.format(r=r))
        ws.cell(row=r, column=8, value=f.get("change_type", ""))
        ws.cell(row=r, column=9, value=f.get("expected_lever", ""))
        for col in range(1, 10):
            c = ws.cell(row=r, column=col); c.font = BODY_FONT; c.border = BORDER
            c.alignment = WRAP if col in (2, 9) else CTR
        r += 1
    last = max(r - 1, first)
    add_name(wb, "roadmap_priority", "%s!$F$%d:$F$%d" % (q("13_Roadmap"), first, last))

    if not ranked:
        ws.cell(row=first, column=1, value="(no findings logged)").font = MUTED_FONT

    # dropdowns
    dv_change = DataValidation(type="list", formula1='"Test,Ship"', allow_blank=True)
    ws.add_data_validation(dv_change); dv_change.add("H%d:H%d" % (first, last))
    dv_sev = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_sev); dv_sev.add("C%d:C%d" % (first, last))
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ws


def build_exec(wb, analytics, findings, meta):
    ws = wb.create_sheet("01_Executive_Summary")
    title_row(ws, "Executive Summary — CRO Opportunity", 6)
    set_widths(ws, [34, 14, 12, 14, 12, 30])

    ws.cell(row=3, column=1, value="Funnel Health (100 = at benchmark)").font = HEAD_FONT
    # Mean of (rate / benchmark) across ONLY the funnel stages that have a numeric rate, ×100.
    # Blank/unmeasurable stages are excluded (not treated as 0), so a store with reliable CVR but
    # un-tracked ATC/checkout still grades honestly on what is measured.
    hs = ws.cell(row=3, column=2,
                 value='=IFERROR(ROUND(MIN(150,100*('
                       'N(ISNUMBER(rate_atc))*IFERROR(rate_atc/bench_atc,0)'
                       '+N(ISNUMBER(rate_checkout))*IFERROR(rate_checkout/bench_checkout,0)'
                       '+N(ISNUMBER(rate_cvr))*IFERROR(rate_cvr/bench_cvr,0))'
                       '/(N(ISNUMBER(rate_atc))+N(ISNUMBER(rate_checkout))+N(ISNUMBER(rate_cvr)))),0),"N/A")')
    hs.font = BIG_FONT; hs.alignment = CTR; hs.fill = HILITE_FILL; hs.number_format = "0"
    add_name(wb, "funnel_health", "%s!$B$3" % q("01_Executive_Summary"))
    ws.cell(row=4, column=1, value="Grade").font = HEAD_FONT
    gr = ws.cell(row=4, column=2,
                 value='=IF(NOT(ISNUMBER(B3)),"—",IF(B3>=110,"A",IF(B3>=90,"B",IF(B3>=70,"C",IF(B3>=50,"D","F")))))')
    gr.font = BIG_FONT; gr.alignment = CTR; gr.fill = HILITE_FILL

    # funnel snapshot mirrors the analytics rate cells (live)
    section_label(ws, 6, "Funnel snapshot (live from 02_Analytics)", 6)
    header_row(ws, 7, ["Stage", "Rate %", "Benchmark %", "Index", "", ""])
    snap = [("Added to cart", "rate_atc", "bench_atc"),
            ("Reached checkout", "rate_checkout", "bench_checkout"),
            ("Completed purchase", "rate_cvr", "bench_cvr")]
    r = 8
    for name, rate_name, bench_name in snap:
        ws.cell(row=r, column=1, value=name).font = BODY_FONT
        ws.cell(row=r, column=2, value="=IFERROR(%s,\"\")" % rate_name).number_format = "0.00"
        ws.cell(row=r, column=3, value="=%s" % bench_name).number_format = "0.00"
        ws.cell(row=r, column=4, value="=IFERROR(%s/%s*100,\"\")" % (rate_name, bench_name)).number_format = "0"
        for col in range(1, 5):
            c = ws.cell(row=r, column=col); c.border = BORDER
            if col != 1:
                c.alignment = CTR
        r += 1

    # findings-by-severity counts (static, from payload)
    r += 1
    section_label(ws, r, "Findings by severity", 6); r += 1
    counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for f in findings or []:
        sev = SEV_CANON.get(str(f.get("severity", "")).strip().lower())
        if sev:
            counts[sev] += 1
    header_row(ws, r, ["Critical", "High", "Medium", "Low", "", ""]); r += 1
    for i, sev in enumerate(["Critical", "High", "Medium", "Low"]):
        c = ws.cell(row=r, column=1 + i, value=counts[sev]); c.font = BODY_FONT; c.alignment = CTR; c.border = BORDER
    r += 2

    # top opportunities by (Impact*2)+Ease
    section_label(ws, r, "Top opportunities — Priority = (Impact × 2) + Ease  (see 13_Roadmap)", 6); r += 1
    ranked = sorted(findings or [], key=_priority, reverse=True)[:7]
    for f in ranked:
        line = "[%s · %s] %s — %s (Priority %d)" % (
            f.get("severity", ""), f.get("change_type", "") or "Test",
            f.get("title", ""), f.get("recommendation", ""), _priority(f))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=line); c.font = BODY_FONT; c.alignment = WRAP
        r += 1
    if not ranked:
        ws.cell(row=r, column=1, value="(no findings logged)").font = MUTED_FONT

    ws.sheet_view.showGridLines = False
    return ws


def build_reference(wb):
    ws = wb.create_sheet("14_Reference")
    title_row(ws, "Reference — LIFT, Benchmarks & Prioritization", 2)
    set_widths(ws, [44, 66])

    section_label(ws, 3, "Funnel benchmarks — aggregated across 21 Shopify stores, FY2025 (named cells, editable)", 2)
    bench_rows = [
        ("Added to cart (% of sessions)", BENCH["atc"], "bench_atc"),
        ("Reached checkout (% of sessions)", BENCH["checkout"], "bench_checkout"),
        ("Completed purchase / CVR (% of sessions)", BENCH["cvr"], "bench_cvr"),
        ("Mobile CVR (%)", BENCH["mobile"], "bench_mobile"),
        ("Desktop CVR (%)", BENCH["desktop"], "bench_desktop"),
    ]
    r = 4
    for label, val, name in bench_rows:
        a = ws.cell(row=r, column=1, value=label); a.font = HEAD_FONT; a.border = BORDER; a.alignment = WRAP
        b = ws.cell(row=r, column=2, value=val); b.font = BODY_FONT; b.border = BORDER; b.alignment = CTR
        b.number_format = "0.00"
        add_name(wb, name, "%s!$B$%d" % (q("14_Reference"), r))
        r += 1

    r += 1
    section_label(ws, r, "Method & glossary", 2); r += 1
    lines = [
        ("Prioritization", "Priority = (Impact × 2) + Ease. Impact & Ease 1–10. ICE's Confidence is "
                           "deliberately dropped — triangulation already encodes confidence."),
        ("Roadmap buckets", "Now ≥24 · Next 20–23 · Soon 15–19 · Later <15 (of max 30)."),
        ("Impact rule", "Weight impact by the traffic of the page being changed (Step 1). A big win on a "
                        "500-session page is low impact; a modest win on a 50k-session page is high."),
        ("Triangulation", "A finding seen in multiple sources (analytics + heatmap + customer voice) is "
                          "higher-confidence and should score higher Impact. # sources is live in 12_Findings_Log."),
        ("Test vs Ship", "Test = uncertain/risky (layout, pricing, messaging). Ship = low-risk obvious fix "
                         "(broken link, missing shipping info, clear UX bug). Not every fix needs an A/B test."),
        ("LIFT Model", "Value Proposition (core) · Relevance · Clarity · Urgency (drivers) · Anxiety · "
                       "Distraction (inhibitors)."),
        ("Funnel Health", "Mean of (rate / benchmark) across the funnel stages that HAVE data (ATC, "
                          "checkout, CVR), ×100. Blank/unmeasurable stages are excluded, not scored as 0. "
                          "100 = at benchmark. Grade: A ≥110 · B 90–109 · C 70–89 · D 50–69 · F <50."),
        ("Mobile gap", "Mobile CVR is typically 40–60% below desktop. This reflects purchase INTENT, not "
                       "broken UX — do not recommend cloning desktop onto mobile."),
        ("AOV CVR bands", "Stores <$60 → median CVR ~4.63%. Stores >$200 → median CVR ~0.95%. Read CVR "
                          "relative to price point."),
        ("Honesty", "Steps without data are marked 'Not run' on 00_Audit_Scope and the step tab. "
                    "Nothing is fabricated."),
        ("Framework", "11-step CRO audit checklist for ecommerce. "
                      "11 analyses: analytics, heuristics (LIFT), review mining, support, heatmaps, "
                      "post-purchase survey, email survey, user testing, marketing match, competitor, roadmap."),
        ("Recalculation", "Grades, triangulation counts and priorities are live formulas. Override an input "
                          "(rate, Impact, Ease, benchmark) and the workbook recomputes on open."),
    ]
    for k, v in lines:
        a = ws.cell(row=r, column=1, value=k); a.font = HEAD_FONT; a.border = BORDER; a.alignment = WRAP
        b = ws.cell(row=r, column=2, value=v); b.font = BODY_FONT; b.border = BORDER; b.alignment = WRAP
        r += 1
    ws.sheet_view.showGridLines = False
    return ws


# ------------------------------------------- values-only report tabs (OPTIONAL_TABS)

def _section_caption(ws, row, text, ncols):
    """Merged section caption band (used by the values-only report tabs)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    return row + 1


def _kv(ws, row, label, value, fmt=None):
    """Label/value line (values only). Value overflows to the right — neighbors empty."""
    a = ws.cell(row=row, column=1, value=label)
    a.font = HEAD_FONT
    b = ws.cell(row=row, column=2, value=value)
    b.font = BODY_FONT
    b.alignment = LEFT
    if fmt:
        b.number_format = fmt
    return row + 1


# Per-dimension entity/weight/outcome labels + number formats for 15_Concentration.
# The block keeps the meta `spend`/`conv` JSON keys (math ported verbatim) — only the
# nouns are CRO: products weigh revenue vs orders, landing pages / channels weigh
# sessions vs conversions (channels: revenue when the export carries it — the block's
# notes say which; hence the neutral "Outcome" label).
_CONC_STYLE = {
    "products": ("Product", "Revenue", FMT_MONEY, "Orders", FMT_INT),
    "landing_pages": ("Landing page", "Sessions", FMT_INT, "Conversions", FMT_INT),
    "channels": ("Channel", "Sessions", FMT_INT, "Outcome", FMT_MONEY),
}
_CONC_STYLE_DEFAULT = ("Entity", "Weight", FMT_MONEY, "Outcome", FMT_MONEY)


def build_concentration_tab(wb, block):
    """Values-only concentration report (HHI / Effective-N / Gini / Pareto-ABC).

    Layout ported from meta-ads-audit build_audit_xlsx.build_concentration_tab.
    Informational — never scored, never in EXPECTED_TABS (check() must stay green
    for workbooks built without raw pulls or CSV exports). No formulas: every cell
    is a value computed by concentration.compute_concentration."""
    ws = wb.create_sheet("15_Concentration")
    title_row(ws, "Concentration — where revenue, traffic & conversions pile up", 6)
    set_widths(ws, [44, 14, 14, 12, 12, 8])
    r = 3
    c = ws.cell(row=r, column=1, value=(
        "HHI bands (merger-guideline cutoffs): <1,500 unconcentrated · 1,500-2,500 "
        "moderate · >2,500 high. Small dimensions: read Effective-N instead."))
    c.font = MUTED_FONT
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2
    block = block or {}
    for dim in block.get("dimensions", []) or []:
        ent_label, w_label, w_fmt, o_label, o_fmt = _CONC_STYLE.get(
            dim.get("key", ""), _CONC_STYLE_DEFAULT)
        label = dim.get("label", "")
        if dim.get("window"):
            label = "%s (%s)" % (label, dim["window"])
        r = _section_caption(ws, r, label, 6)
        r = _kv(ws, r, "Verdict", dim.get("verdict", ""))
        for side, side_label in (("spend", w_label), ("conv", o_label)):
            m = dim.get(side)
            if m:
                r = _kv(ws, r, side_label,
                        "HHI {hhi:,.1f} ({band}) · Effective-N {eff_n} · Gini {gini}".format(
                            hhi=m.get("hhi", 0.0), band=m.get("band", ""),
                            eff_n=m.get("eff_n", ""), gini=m.get("gini", "")))
            else:
                r = _kv(ws, r, side_label, "no signal in this window")
        r = _kv(ws, r, "Entities", "%d (from %d raw rows)"
                % (dim.get("n_entities", 0), dim.get("n_rows_raw", 0)))
        if dim.get("caveat"):
            r = _kv(ws, r, "Caveat", dim["caveat"])
        r += 1
        header_row(ws, r, [ent_label, w_label, o_label,
                           "%s %%" % w_label, "%s %%" % o_label, "ABC"])
        r += 1
        for t in dim.get("top", []) or []:
            ws.cell(row=r, column=1, value=t.get("name", "")).alignment = WRAP
            ws.cell(row=r, column=2, value=t.get("spend")).number_format = w_fmt
            ws.cell(row=r, column=3, value=t.get("conv")).number_format = o_fmt
            ws.cell(row=r, column=4, value=t.get("spend_share")).number_format = FMT_SHARE
            ws.cell(row=r, column=5, value=t.get("conv_share")).number_format = FMT_SHARE
            ws.cell(row=r, column=6, value=t.get("abc"))
            for col in range(1, 7):
                cell = ws.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.border = BORDER
            r += 1
        tail = dim.get("tail")
        if tail:
            c = ws.cell(row=r, column=1, value="… plus %d more (tail)" % tail.get("n", 0))
            c.font = MUTED_FONT
            ws.cell(row=r, column=2, value=tail.get("spend")).number_format = w_fmt
            ws.cell(row=r, column=3, value=tail.get("conv")).number_format = o_fmt
            ws.cell(row=r, column=4, value=tail.get("spend_share")).number_format = FMT_SHARE
            r += 1
        r += 1
    for note in block.get("notes", []) or []:
        c = ws.cell(row=r, column=1, value="Note: %s" % note)
        c.font = MUTED_FONT
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    ws.sheet_view.showGridLines = False
    return ws


def build_cvr_signals_tab(wb, block):
    """Values-only CVR Signals report (Wilson CIs / two-proportion z / significance
    gate / empirical-Bayes shrinkage) from cvr_signals.compute_cvr_signals.

    Every rate in the block is a FRACTION (SHAPE-NOTES pin) — cells keep the raw
    fraction and FMT_RATE renders it as a percent, so nothing is multiplied by 100.
    Informational — never scored, never in EXPECTED_TABS (check() must stay green
    without it). No formulas: every cell is a precomputed value."""
    ws = wb.create_sheet("16_CVR_Signals")
    ncols = 8
    title_row(ws, "CVR Signals — significance, gates & shrinkage", ncols)
    set_widths(ws, [40, 11, 12, 10, 9, 11, 11, 9])
    block = block or {}
    r = 3

    label = "Site conversion rate"
    if block.get("window"):
        label = "%s (%s)" % (label, block["window"])
    r = _section_caption(ws, r, label, ncols)
    site = block.get("site") or {}
    r = _kv(ws, r, "Sessions", site.get("sessions"), fmt=FMT_INT)
    r = _kv(ws, r, "Conversions", site.get("conversions"), fmt=FMT_INT)
    r = _kv(ws, r, "CVR", site.get("cvr"), fmt=FMT_RATE)
    ci = site.get("ci") or []
    if len(ci) == 2:
        r = _kv(ws, r, "Wilson 95% CI",
                "%.2f%% – %.2f%%" % (ci[0] * 100.0, ci[1] * 100.0))
    r = _kv(ws, r, "Significance gate n*", block.get("min_sessions"), fmt=FMT_INT)
    prior = block.get("prior") or {}
    if prior:
        r = _kv(ws, r, "Shrinkage prior", "rate %.2f%% · k=%s sessions (%s)"
                % ((prior.get("rate") or 0.0) * 100.0, prior.get("k", ""),
                   prior.get("basis", "")))
    r += 1

    # segment splits — sibling-complement two-proportion z per row
    segments = block.get("segments") or {}
    hz = segments.get("headline_device_z")
    for seg_key, seg_label in (("device", "Device split"),
                               ("channels", "Channel split"),
                               ("new_vs_returning", "New vs returning")):
        rows = segments.get(seg_key) or []
        if not rows:
            continue
        r = _section_caption(
            ws, r, "%s — two-proportion z vs sibling complement (|z| ≥ 1.96 significant)"
            % seg_label, ncols)
        if seg_key == "device" and hz:
            r = _kv(ws, r, "Mobile vs desktop z", "%.2f (%s)"
                    % (hz.get("z"), "significant" if hz.get("significant")
                       else "not significant"))
        header_row(ws, r, ["Segment", "Sessions", "Conversions", "CVR", "z",
                           "Significant", "Derived"])
        r += 1
        for s in rows:
            ws.cell(row=r, column=1, value=s.get("name", "")).alignment = WRAP
            ws.cell(row=r, column=2, value=s.get("sessions")).number_format = FMT_INT
            ws.cell(row=r, column=3, value=s.get("conversions")).number_format = FMT_INT
            ws.cell(row=r, column=4, value=s.get("cvr")).number_format = FMT_RATE
            ws.cell(row=r, column=5, value=s.get("z")).number_format = FMT_Z
            ws.cell(row=r, column=6,
                    value=("yes" if s.get("significant")
                           else ("no" if s.get("z") is not None else "")))
            ws.cell(row=r, column=7, value=("yes" if s.get("derived") else ""))
            for col in range(1, 8):
                cell = ws.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.border = BORDER
            r += 1
        r += 1

    # landing pages — raw vs shrunk CVR, Wilson lower bound, significance gate
    pages = block.get("pages") or []
    if pages:
        r = _section_caption(
            ws, r, "Landing pages — top %d by sessions (full-universe math ran first)"
            % len(pages), ncols)
        uni = block.get("pages_universe") or {}
        if uni:
            r = _kv(ws, r, "Page universe", "%d pages · %s sessions · %d below the n*=%s gate"
                    % (uni.get("n", 0), format(uni.get("sessions", 0), ","),
                       uni.get("gated_n", 0), block.get("min_sessions", "")))
        header_row(ws, r, ["Landing page", "Sessions", "Conversions", "Derived",
                           "CVR raw", "CVR shrunk", "Wilson LB", "Gated"])
        r += 1
        for p in pages:
            ws.cell(row=r, column=1, value=p.get("page", "")).alignment = WRAP
            ws.cell(row=r, column=2, value=p.get("sessions")).number_format = FMT_INT
            ws.cell(row=r, column=3, value=p.get("conversions")).number_format = FMT_INT
            ws.cell(row=r, column=4, value=("yes" if p.get("derived") else ""))
            ws.cell(row=r, column=5, value=p.get("cvr_raw")).number_format = FMT_RATE
            ws.cell(row=r, column=6, value=p.get("cvr_shrunk")).number_format = FMT_RATE
            ws.cell(row=r, column=7, value=p.get("wilson_lb")).number_format = FMT_RATE
            ws.cell(row=r, column=8, value=("yes" if p.get("gated") else ""))
            for col in range(1, ncols + 1):
                cell = ws.cell(row=r, column=col)
                cell.font = BODY_FONT
                cell.border = BORDER
            r += 1
        r += 1

    for note in block.get("notes", []) or []:
        c = ws.cell(row=r, column=1, value="Note: %s" % note)
        c.font = MUTED_FONT
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1
    ws.sheet_view.showGridLines = False
    return ws


# ------------------------------------------------------------------------- assembly

def normalize(payload):
    """Normalize severity/change casing in findings; return warnings for vocab drift."""
    warns = []
    for f in payload.get("findings", []) or []:
        fid = f.get("id", "?")
        sev = str(f.get("severity", "")).strip()
        canon = SEV_CANON.get(sev.lower())
        if canon:
            f["severity"] = canon
        elif sev:
            warns.append("finding %s: unknown severity %r" % (fid, sev))
        ct = str(f.get("change_type", "")).strip()
        cc = CHANGE_CANON.get(ct.lower())
        if cc:
            f["change_type"] = cc
        elif ct:
            warns.append("finding %s: change_type %r not Test/Ship" % (fid, ct))
    return warns


def build(payload, *, concentration=None, cvr_signals=None):
    """Payload dict -> Workbook. concentration / cvr_signals are the optional
    values-only report blocks (from concentration.py / cvr_signals.py, computed from
    raw pull files or CSV exports — never from the payload); each adds one
    OPTIONAL_TABS sheet after 14_Reference. Existing tabs are untouched either way."""
    meta = payload.get("meta", {}) or {}
    analytics = payload.get("analytics", {}) or {}
    detail = payload.get("steps_detail", {}) or {}
    findings = payload.get("findings", []) or []

    for w in normalize(payload):
        print("WARN: %s" % w)

    wb = Workbook()
    wb.remove(wb.active)

    build_scope(wb, meta)
    build_exec(wb, analytics, findings, meta)
    build_analytics(wb, analytics)
    build_heuristic(wb, detail.get("heuristic"), meta)
    build_review_mining(wb, detail.get("review_mining"), meta)
    build_support(wb, detail.get("support"), meta)
    build_heatmaps(wb, detail.get("heatmaps"), meta)
    build_postpurchase(wb, detail.get("post_purchase_survey"), meta)
    build_email_survey(wb, detail.get("email_survey"), meta)
    build_user_testing(wb, detail.get("user_testing"), meta)
    build_marketing(wb, detail.get("marketing"), meta)
    build_competitor(wb, detail.get("competitor"), meta)
    build_findings(wb, findings)
    build_roadmap(wb, findings)
    build_reference(wb)
    if concentration:
        build_concentration_tab(wb, concentration)
    if cvr_signals:
        build_cvr_signals_tab(wb, cvr_signals)

    order = {name: i for i, name in enumerate(EXPECTED_TABS + OPTIONAL_TABS)}
    wb._sheets.sort(key=lambda s: order.get(s.title, 99))
    for ws in wb.worksheets:
        if ws.title in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS[ws.title]
    wb.active = wb.sheetnames.index("01_Executive_Summary")
    return wb


# ---------------------------------------------------------------------------- check

def check(path):
    errors, warns = [], []
    try:
        wb = load_workbook(path)
    except Exception as e:  # noqa: BLE001
        print("FAIL: could not open workbook: %s" % e)
        return 1

    for t in EXPECTED_TABS:
        if t not in wb.sheetnames:
            errors.append("missing tab: %s" % t)

    names = defined_name_set(wb)
    for n in sorted(REQUIRED_NAMES):
        if n not in names:
            errors.append("missing named range: %s" % n)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and ("#REF!" in v or "#NAME?" in v):
                    errors.append("%s in %s!%s" % ("#REF!" if "#REF!" in v else "#NAME?", ws.title, cell.coordinate))

    # vocab scan on Findings (col E severity, col I change) and Roadmap (col C severity, col H change)
    for title, sev_col, chg_col, first in (("12_Findings_Log", 5, 9, 4), ("13_Roadmap", 3, 8, 4)):
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        r = first
        while True:
            fid = ws.cell(row=r, column=1).value
            if fid in (None, ""):
                break
            sev = ws.cell(row=r, column=sev_col).value
            chg = ws.cell(row=r, column=chg_col).value
            if sev and str(sev) not in ALLOWED_SEVERITIES and not str(sev).startswith("="):
                warns.append("%s!%s%d severity %r not in %s" %
                             (title, get_column_letter(sev_col), r, sev, sorted(ALLOWED_SEVERITIES)))
            if chg and str(chg) not in ALLOWED_CHANGE and not str(chg).startswith("="):
                warns.append("%s!%s%d change %r not Test/Ship" % (title, get_column_letter(chg_col), r, chg))
            r += 1

    print("Workbook: %s" % path)
    print("Tabs: %d/%d present" % (len([t for t in EXPECTED_TABS if t in wb.sheetnames]), len(EXPECTED_TABS)))
    print("Named ranges: %d/%d present" % (len([n for n in REQUIRED_NAMES if n in names]), len(REQUIRED_NAMES)))
    for w in warns:
        print("WARN: %s" % w)
    if errors:
        for e in errors:
            print("ERROR: %s" % e)
        print("RESULT: FAIL (%d errors)" % len(errors))
        return 1
    print("RESULT: PASS (structural). Note: open in Excel/LibreOffice to verify computed values.")
    return 0


# ----------------------------------------------------------------------------- main

def main(argv=None):
    ap = argparse.ArgumentParser(description="Shopify CRO Audit workbook generator (11-step).")
    ap.add_argument("--input", help="cro-payload.json")
    ap.add_argument("--output", help="output .xlsx path")
    ap.add_argument("--check", dest="check_path", help="structural-check an existing .xlsx")
    args = ap.parse_args(argv)

    if args.check_path:
        return check(args.check_path)
    if not args.input or not args.output:
        ap.error("provide --input and --output (or --check <file>)")

    with open(args.input, "r", encoding="utf-8") as fh:
        payload = json.load(fh)
    wb = build(payload)
    wb.save(args.output)
    print("Wrote %s" % args.output)
    return 0


if __name__ == "__main__":
    sys.exit(main())

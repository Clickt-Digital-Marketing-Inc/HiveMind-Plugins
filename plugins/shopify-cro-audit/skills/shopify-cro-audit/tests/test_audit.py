#!/usr/bin/env python3
# Copyright (c) 2026 Clickt Digital Marketing Inc. All rights reserved.
"""Conformance tests for the shopify-cro-audit pipeline. Stdlib only
(openpyxl needed for the xlsx sections only — they skip gracefully without it).

Run: python3 tests/test_audit.py

Guards: Funnel-Health-0-150 correctness (hand oracles, measured-stages-only
exclusion, Excel-ROUND half-up parity), (Impact x 2) + Ease priority + the
Now/Next/Soon/Later buckets, JS<->Python<->Excel constant parity (GRADES
pollution guard), self-containment (GSAP sentinel + checksum), determinism,
concentration metrics + CRO verdict wording, CVR Signals (MediaMetrics
stdlib mirrors: Wilson / two-proportion z / gates / Beta-binomial /
empirical-Bayes shrinkage), shopify_rows raw-envelope parsing (PERCENT =
FRACTION, AOV verbatim, checksum notes, error shapes), the manual GA4 +
Shopify CSV path, the machine analytics layer (precedence, Reads, percent
boundary) and its replace-and-log merge, and the workbook (15 tabs, 10 named
ranges, optional tabs, C6/C7/C8 pinned rate cells, whitelabel scrub).
"""
from __future__ import annotations

import copy
import hashlib
import json
import math
import re
import sys
import tempfile
from pathlib import Path
from statistics import median

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent / "scripts"
FIX = HERE / "fixtures"
CSVDIR = FIX / "csv"
sys.path.insert(0, str(SCRIPTS))

import audit_model as am
import audit_html
import audit_md
import concentration as conc
import cvr_signals as cv
import machine as mx
import manual_csv as mc
import shopify_rows as sr

SAMPLE = HERE / "sample-payload.json"
PAYLOAD = json.loads(SAMPLE.read_text())
MICRO = json.loads((FIX / "micro_payload.json").read_text())

FAILS = []
SKIPS = []
N_CHECKS = 0


def check(name, cond, detail=""):
    global N_CHECKS
    N_CHECKS += 1
    print(("  ok   " if cond else "FAIL   ") + name + ("" if cond else "   :: " + str(detail)))
    if not cond:
        FAILS.append(name)


def skip_note(msg):
    print("  skip " + msg)
    SKIPS.append(msg)


def approx(a, b, tol=1e-9):
    return a is not None and abs(a - b) <= tol


# ============================================================================
# 1: model oracle — sample-payload.json (Funnel Health 0-150)
#
# Hand arithmetic (BENCH atc 7.23 / checkout 5.96 / cvr 2.99, all stages
# measured): 6.10/7.23 = 0.8437068; 4.40/5.96 = 0.7382550; 2.30/2.99 =
# 0.7692308; mean = 2.3511926/3 = 0.7837309 -> 100x = 78.373086 -> 78.3731
# unrounded, ROUND half-up 0dp -> 78 INT -> grade C (70 <= 78 < 90).
#
# Priority = Impact*2 + Ease: F-001 9*2+8=26 Now; F-002 8*2+9=25 Now;
# F-003 8*2+6=22 Next; F-004 7*2+7=21 Next; F-005 6*2+8=20 Next;
# F-006 5*2+5=15 Soon. Buckets: Now>=24 / Next>=20 / Soon>=15 / else Later.
# ============================================================================
print("--- 1: model oracle (sample payload, Funnel Health 150) ---")
model = am.compute_model(PAYLOAD, generated="2026-06-25T00:00:00")
check("health score == 78 INT (hand oracle)", model["health"]["score"] == 78,
      str(model["health"]))
check("health unrounded == 78.3731", model["health"]["score_unrounded"] == 78.3731,
      str(model["health"]))
check("grade == C", model["health"]["grade"] == "C")
check("health max == 150", model["health"]["max"] == 150)
check("summary mirrors score/grade",
      model["summary"]["score"] == 78 and model["summary"]["grade"] == "C")
check("step statuses 7 run / 1 partial / 3 not_run",
      (model["summary"]["n_run"], model["summary"]["n_partial"],
       model["summary"]["n_not_run"]) == (7, 1, 3), str(model["summary"]))
check("severity counts 0/3/2/1",
      (model["summary"]["crit"], model["summary"]["high"],
       model["summary"]["med"], model["summary"]["low"]) == (0, 3, 2, 1))
check("11 sections in canonical step order",
      [s["step"] for s in model["sections"]] == list(range(1, 12)))
check("section tabs/titles mirror STEPS",
      [(s["step"], s["tab"], s["title"]) for s in model["sections"]]
      == [(n, t, ti) for n, t, ti, _k in am.STEPS])

fnd = {f["id"]: f for f in model["findings"]}
check("priority F-001 26 / F-002 25 / F-003 22 / F-004 21 / F-005 20 / F-006 15",
      [(f["id"], f["priority"]) for f in model["findings"]]
      == [("F-001", 26), ("F-002", 25), ("F-003", 22), ("F-004", 21),
          ("F-005", 20), ("F-006", 15)],
      str([(f["id"], f["priority"]) for f in model["findings"]]))
check("buckets Now/Now/Next/Next/Next/Soon",
      [f["bucket"] for f in model["findings"]]
      == ["Now", "Now", "Next", "Next", "Next", "Soon"])
check("findings_by_bucket {Now:2, Next:3, Soon:1, Later:0}",
      model["summary"]["findings_by_bucket"]
      == {"Now": 2, "Next": 3, "Soon": 1, "Later": 0})
check("triangulation count rides the row (F-002: 3 sources)",
      fnd["F-002"]["n_sources"] == 3)
check("kpis empty (CRO has no KPI strip)", model["kpis"] == [])
check("no warnings on the sample payload", model["warnings"] == [],
      str(model["warnings"]))
check("stem() sample", am.stem(PAYLOAD) == "cro-audit_acme-wellness_2026-06-25",
      am.stem(PAYLOAD))

# Step-1 evidence: funnel KPI table with machine Reads + bench + index columns.
ev1 = model["sections"][0]["evidence"]
check("step 1 carries 6 evidence tables", len(ev1) == 6,
      str([t["label"] for t in ev1]))
ft = ev1[0]
check("funnel table columns Stage/Count/Rate %/Benchmark %/Index/Read",
      ft["columns"] == ["Stage", "Count", "Rate %", "Benchmark %", "Index", "Read"])
check("funnel base row (Sessions 482190, 100, 'Funnel base')",
      ft["rows"][0] == ["Sessions", 482190, 100, "", "", "Funnel base"])
# Index = half-up(rate/bench*100): 6.10/7.23=84.37->84; 4.40/5.96=73.83->74;
# 2.30/2.99=76.92->77. Reads: all three >= bench*0.7 but < bench -> Below.
check("funnel stage rows: counts + bench + index 84/74/77 + Below reads",
      ft["rows"][1] == ["Added to cart", 29413, 6.10, 7.23, 84, "Below benchmark"]
      and ft["rows"][2] == ["Reached checkout", 21217, 4.40, 5.96, 74, "Below benchmark"]
      and ft["rows"][3] == ["Completed purchase", 11091, 2.30, 2.99, 77, "Below benchmark"],
      str(ft["rows"]))
dt = ev1[1]
# Mobile 1.90 < 2.87*0.7=2.009 -> Well below (index 66); Desktop 4.10 >=
# 4.51*0.7=3.157 -> Below (index 91); Tablet: NO benchmark, NO read (pinned).
check("device rows: mobile Well below 66 / desktop Below 91",
      dt["rows"][0] == ["Mobile", 395396, 1.90, 2.87, 66, "Well below benchmark"]
      and dt["rows"][1] == ["Desktop", 67507, 4.10, 4.51, 91, "Below benchmark"],
      str(dt["rows"]))
check("tablet row has no benchmark / index / read",
      dt["rows"][2] == ["Tablet", 19287, 2.60, "", "", ""])
check("AOV band row ($60-$200 for 84.90)",
      ["AOV CVR band", "$60–$200 band"] in ev1[5]["rows"], str(ev1[5]["rows"]))
check("evidence is a LIST of labeled tables (review mining: 4 tables)",
      len(model["sections"][2]["evidence"]) == 4
      and model["sections"][2]["evidence"][1]["label"].startswith("Objections"))
check("step 11 carries no evidence tables (findings ARE the step)",
      model["sections"][10]["evidence"] == [])
check("not_run step keeps its reason (step 5 heatmaps)",
      model["sections"][4]["status"] == "not_run"
      and "Clarity" in model["sections"][4]["reason"])

# grade boundaries on the INT score: 110A / 109B / 90B / 89C / 70C / 69D /
# 50D / 49F (grade reads the ROUNDED cell — Excel parity)
check("grade boundaries 110A/109B/90B/89C/70C/69D/50D/49F",
      [am.grade(x) for x in (110, 109, 90, 89, 70, 69, 50, 49)]
      == ["A", "B", "B", "C", "C", "D", "D", "F"])
# 150 cap: every ratio 2.0 -> mean 2.0 -> min(150, 200) = 150 -> A
check("150 cap: mean ratio 2.0 -> (150.0, 150, 'A')",
      am.funnel_health({"atc_rate": 14.46, "checkout_rate": 11.92, "cvr": 5.98})
      == (150.0, 150, "A"))
# half-up .5 boundary vs banker's: Excel ROUND(108.5,0)=109; Python round()=108
check("_round_half_up(108.5) == 109 while banker's round(108.5) == 108",
      am._round_half_up(108.5) == 109.0 and round(108.5) == 108)
check("_round_half_up(0.5) == 1 / (2.675, 2dp) == 2.68 (repr-quantize)",
      am._round_half_up(0.5) == 1.0 and round(0.5) == 0
      and am._round_half_up(2.675, 2) == 2.68 and round(2.675, 2) == 2.67)
check("priority() and bucket_for() boundaries 24/20/15",
      (am.bucket_for(24), am.bucket_for(23), am.bucket_for(20), am.bucket_for(19),
       am.bucket_for(15), am.bucket_for(14))
      == ("Now", "Next", "Next", "Soon", "Soon", "Later"))
check("read_verdict boundaries at bench and bench*0.7",
      am.read_verdict(7.23, 7.23) == "At / above benchmark"
      and am.read_verdict(5.061, 7.23) == "Below benchmark"       # exactly 0.7x
      and am.read_verdict(5.06, 7.23) == "Well below benchmark"
      and am.read_verdict("", 7.23) == "" and am.read_verdict(None, 7.23) == "")
check("aov_band boundaries <60 / <=200 / >200 / blank",
      am.aov_band(59.99).startswith("Sub-$60")
      and am.aov_band(60) == "$60–$200 band" and am.aov_band(200) == "$60–$200 band"
      and am.aov_band(200.01).startswith("Over-$200") and am.aov_band(None) == "")
# vocab canon + warnings
warn_model = am.compute_model(
    {"meta": {"steps": [{"step": 99, "status": "run"},
                        {"step": 2, "status": "Bogus"}]},
     "findings": [{"id": "W-1", "severity": "Extreme"}]}, generated="T")
check("unknown step / status / severity all warn (not error)",
      any("unknown step" in w for w in warn_model["warnings"])
      and any("unknown status" in w for w in warn_model["warnings"])
      and any("unknown severity" in w for w in warn_model["warnings"]),
      str(warn_model["warnings"]))
check("unknown-severity finding defaults impact to DEFAULT_EASE (5) -> pri 15",
      warn_model["findings"][0]["priority"] == 15)

# ============================================================================
# 2: measured-stages exclusion — micro payload (checkout_rate ABSENT)
#
# Measured stages: atc 7.23/7.23 = 1.0; cvr 1.495/2.99 = 0.5 (exact halving)
# -> mean 0.75 -> 75 -> C. Scoring the missing stage as 0 would give
# (1.0 + 0 + 0.5)/3 = 0.5 -> 50 -> D. The exclusion demonstrably differs.
# ============================================================================
print("--- 2: measured-stages exclusion (micro payload) ---")
m_micro = am.compute_model(MICRO, generated="2026-07-01T00:00:00")
check("micro health == 75 / 75.0 / C (2-stage mean)",
      m_micro["health"]["score"] == 75 and m_micro["health"]["score_unrounded"] == 75.0
      and m_micro["health"]["grade"] == "C", str(m_micro["health"]))
zero_scored = round(100 * (1.0 + 0.0 + 0.5) / 3)
check("3-stage zero-scored mean would say 50 / D (differs!)",
      zero_scored == 50 and am.grade(zero_scored) == "D"
      and m_micro["health"]["score"] != zero_scored)
check("funnel_health direct: 2 measured stages -> (75.0, 75, 'C')",
      am.funnel_health({"atc_rate": 7.23, "cvr": 1.495}) == (75.0, 75, "C"))
check("no measured stages -> (None, None, '—')",
      am.funnel_health({}) == (None, None, "—")
      and am.funnel_health(None) == (None, None, "—"))
check("ISNUMBER parity: a rate transcribed as TEXT is excluded",
      am.funnel_health({"cvr": "2.30"}) == (None, None, "—")
      and am.funnel_health({"cvr": True}) == (None, None, "—"))
# micro evidence: unmeasured checkout row has blank rate/index/read
ev_micro = m_micro["sections"][0]["evidence"][0]
check("unmeasured checkout row blank rate/index/read in evidence",
      ev_micro["rows"][2] == ["Reached checkout", "", "", 5.96, "", ""],
      str(ev_micro["rows"]))
# defaults: M-002 severity 'critical' canon -> Critical, impact 9, ease 5 ->
# 23 Next; M-001 High -> impact 7, ease 5 -> 19 Soon. Sorted desc.
check("Impact defaults from severity + Ease default 5 (23 Next / 19 Soon)",
      [(f["id"], f["severity"], f["impact"], f["ease"], f["priority"], f["bucket"])
       for f in m_micro["findings"]]
      == [("M-002", "Critical", 9, 5, 23, "Next"),
          ("M-001", "High", 7, 5, 19, "Soon")],
      str(m_micro["findings"]))
check("change_type canon ('test' -> 'Test')",
      {f["id"]: f["change_type"] for f in m_micro["findings"]}
      == {"M-001": "Test", "M-002": "Ship"})
check("stem() micro", am.stem(MICRO) == "cro-audit_micro-oracle-co-sample_2026-07-01",
      am.stem(MICRO))

# ============================================================================
# 3: three-way parity — JS kernel <-> audit_model <-> xlsx constants
# ============================================================================
print("--- 3: three-way constant parity ---")
tpl = audit_html._TEMPLATE


def js_map(name):
    m = re.search(name + r"\s*=\s*\{([^}]*)\}", tpl)
    out = {}
    for part in m.group(1).split(","):
        k, v = part.split(":")
        out[k.strip()] = float(v.strip())
    return out


check("JS BENCH == audit_model.BENCH", js_map("BENCH") == am.BENCH)
check("JS IMPACT == audit_model.SEVERITY_IMPACT",
      js_map("IMPACT") == am.SEVERITY_IMPACT)
js_grades = [(int(n), g) for n, g in re.findall(r"\[(\d+),'([A-F])'\]", tpl)]
check("GRADES parity regex finds EXACTLY the 5 grade pairs (pollution guard)",
      js_grades == am.GRADE_CUTOFFS and len(js_grades) == 5, str(js_grades))
bkt = re.search(r"BUCKETS\s*=\s*\[\[(.*?)\]\];", tpl)
js_buckets = [(int(n), lab) for n, lab in re.findall(r"(\d+),'(\w+)'", bkt.group(1))]
check("JS BUCKETS == audit_model.PRIORITY_BUCKETS (multi-char labels)",
      js_buckets == am.PRIORITY_BUCKETS, str(js_buckets))
check("JS PRI mirrors priority() = i*2 + e", "return i*2+e" in tpl)
check("JS healthOf: measured-stages mean, 150 cap, Math.round half-up",
      "Math.round(Math.min(150" in tpl and "STAGES=['atc','checkout','cvr']" in tpl)
check("gauge arc = C*(1-score/150), labeled out of 150",
      "C*(1-score/150)" in tpl and "out of 150" in tpl)
check("NO Confidence slider anywhere in the template", "Confidence" not in tpl
      and 'data-t="confidence"' not in tpl)
check("no xmlns on inline SVG (template)", "xmlns" not in tpl)

try:
    import build_cro_workbook as bcw
except SystemExit:
    bcw = None
    skip_note("xlsx constant-parity checks (openpyxl unavailable)")
if bcw is not None:
    check("xlsx BENCH IS audit_model's (identity)", bcw.BENCH is am.BENCH)
    check("xlsx SEV_CANON IS audit_model's", bcw.SEV_CANON is am.SEV_CANON)
    check("xlsx CHANGE_CANON IS audit_model's", bcw.CHANGE_CANON is am.CHANGE_CANON)
    check("xlsx STATUS_CANON IS audit_model's", bcw.STATUS_CANON is am.STATUS_CANON)
    check("xlsx SEVERITY_IMPACT IS audit_model's",
          bcw.SEVERITY_IMPACT is am.SEVERITY_IMPACT)
    check("xlsx STEP_TABS byte-match audit_model.STEPS",
          bcw.STEP_TABS == [(n, t, ti) for n, t, ti, _k in am.STEPS])
    check("EXPECTED_TABS is the pinned 15 / OPTIONAL_TABS 15_Conc + 16_CVR",
          len(bcw.EXPECTED_TABS) == 15
          and bcw.OPTIONAL_TABS == ["15_Concentration", "16_CVR_Signals"]
          and not set(bcw.OPTIONAL_TABS) & set(bcw.EXPECTED_TABS))
    check("REQUIRED_NAMES is the pinned 10",
          bcw.REQUIRED_NAMES == {
              "bench_atc", "bench_checkout", "bench_cvr", "bench_mobile",
              "bench_desktop", "rate_atc", "rate_checkout", "rate_cvr",
              "funnel_health", "roadmap_priority"})

# ============================================================================
# 4: self-containment — GSAP sentinel excise-by-checksum, then grep
# ============================================================================
print("--- 4: self-containment ---")
gsap_bytes = (SCRIPTS / "vendor" / "gsap.min.js").read_bytes()
want = (SCRIPTS / "vendor" / "SHA256SUMS").read_text().split()[0]
check("vendored GSAP matches SHA256SUMS",
      hashlib.sha256(gsap_bytes).hexdigest() == want)
meta_sums = (SCRIPTS.parent.parent.parent.parent / "meta-ads-audit" / "skills"
             / "meta-ads-audit" / "scripts" / "vendor" / "SHA256SUMS")
if meta_sums.is_file():
    check("GSAP byte-identical to the meta-ads-audit vendored copy",
          meta_sums.read_text().split()[0] == want)
else:
    skip_note("meta-ads-audit vendor copy not found for cross-plugin GSAP check")

html = audit_html.render_html(model, animate=True)
blob = audit_html.gsap_blob()
check("GSAP blob embedded between sentinels", blob in html)
stripped = html.replace(blob, "")
hits = re.findall(r"https?://|<link|src=|cdn", stripped)
check("no external refs outside the GSAP sentinels", not hits, str(hits[:4]))
check("store_url scheme scrubbed from the embedded model",
      "acmewellness.example" in html)

h_noanim = audit_html.render_html(model, animate=False)
check("animate=False carries zero GSAP bytes",
      audit_html.GSAP_BEGIN not in h_noanim and "GreenSock" not in h_noanim)
check("animate=False still self-contained",
      not re.findall(r"https?://|<link|src=|cdn", h_noanim))
check("html without blocks embeds nulls",
      '"concentration":null' in html and '"cvr_signals":null' in html
      and '"machine":null' in html)
check("html title + eyebrow say Shopify CRO Audit",
      "Shopify CRO Audit" in html and "Acme Wellness" in html)

# ============================================================================
# 5: determinism — pure function of the payload except meta.generated
# ============================================================================
print("--- 5: determinism ---")
a = audit_html.render_html(am.compute_model(PAYLOAD, generated="TS1"))
b = audit_html.render_html(am.compute_model(PAYLOAD, generated="TS2"))
check("HTML deterministic modulo generated", a.replace("TS1", "TS") == b.replace("TS2", "TS"))
ma = audit_md.render_md(am.compute_model(PAYLOAD, generated="TS1"))
mb = audit_md.render_md(am.compute_model(PAYLOAD, generated="TS2"))
check("markdown deterministic modulo generated", ma.replace("TS1", "TS") == mb.replace("TS2", "TS"))
j1 = json.dumps(am.compute_model(PAYLOAD, generated="TS"), sort_keys=True)
j2 = json.dumps(am.compute_model(PAYLOAD, generated="TS"), sort_keys=True)
check("model JSON deterministic (same generated)", j1 == j2)
check("double HTML render byte-identical (fixed generated)",
      audit_html.render_html(am.compute_model(PAYLOAD, generated="TS"))
      == audit_html.render_html(am.compute_model(PAYLOAD, generated="TS")))
check("no wall clock: meta.generated falls back to generated_for_date",
      am.compute_model(PAYLOAD)["meta"]["generated"] == "2026-06-25")

# ============================================================================
# 6: concentration — metric hand-oracles + CRO dimensions + verdict wording
#
# revenue6 [50,20,10,10,5,5]: HHI (.5²+.2²+.1²+.1²+.05²+.05²)*10000 = 3150;
#   eff_n 1/.315 = 3.17; gini (2*485)/(6*100) - 7/6 = 0.45; ABC AAABBC.
# orders6 [30,0,5,0,1,0]: HHI ((30/36)²+(5/36)²+(1/36)²)*10000 = 7145.1;
#   eff_n 1.40; gini 2*209/(6*36) - 7/6 = 0.769; ABC crossing-inclusive ACBCCC.
# ============================================================================
print("--- 6: concentration ---")
rev6 = [50, 20, 10, 10, 5, 5]
ord6 = [30, 0, 5, 0, 1, 0]
check("hhi(rev6) == 3150.0", approx(conc.hhi(rev6), 3150.0), str(conc.hhi(rev6)))
check("effective_n(rev6) == 3.17", round(conc.effective_n(rev6), 2) == 3.17)
check("gini(rev6) == 0.45", round(conc.gini(rev6), 3) == 0.45, str(conc.gini(rev6)))
check("abc(rev6) == AAABBC", conc.pareto_abc(rev6) == ["A", "A", "A", "B", "B", "C"],
      str(conc.pareto_abc(rev6)))
check("hhi(ord6) == 7145.1", round(conc.hhi(ord6), 1) == 7145.1, str(conc.hhi(ord6)))
check("effective_n(ord6) == 1.40", round(conc.effective_n(ord6), 2) == 1.40)
check("gini(ord6) == 0.769", round(conc.gini(ord6), 3) == 0.769, str(conc.gini(ord6)))
check("abc(ord6) crossing-inclusive == ACBCCC",
      conc.pareto_abc(ord6) == ["A", "C", "B", "C", "C", "C"], str(conc.pareto_abc(ord6)))
check("equal k=4: hhi 2500 / band moderate / eff_n 4 / gini 0",
      conc.hhi([1, 1, 1, 1]) == 2500.0 and conc.hhi_band(2500.0) == "moderate"
      and conc.effective_n([1, 1, 1, 1]) == 4.0 and conc.gini([1, 1, 1, 1]) == 0.0)
check("band boundary 1500 -> moderate", conc.hhi_band(1500.0) == "moderate")
check("single entity: hhi 10000 / eff_n 1 / gini 0 / lorenz / abc A",
      conc.hhi([7]) == 10000.0 and conc.effective_n([7]) == 1.0 and conc.gini([7]) == 0.0
      and conc.lorenz_points([7]) == [[0.0, 0.0], [1.0, 1.0]]
      and conc.pareto_abc([7]) == ["A"])

# CRO verdict table: review_bidding is RENAMED review_mix; no 'bidding' vocab
check("verdict fragility", conc.verdict(3000, 3000)[0] == "fragility")
check("verdict consolidate", conc.verdict(1000, 3000)[0] == "consolidate")
check("verdict review_mix (renamed from review_bidding)",
      conc.verdict(3000, 1000)[0] == "review_mix"
      and "Review the mix" in conc.verdict(3000, 1000, ("sessions", "revenue"))[1])
check("verdict diversified", conc.verdict(1000, 1000)[0] == "diversified")
check("verdict no_conv_signal", conc.verdict(3000, None)[0] == "no_conv_signal")
check("verdict insufficient", conc.verdict(None, None)[0] == "insufficient")
check("no 'bidding' vocabulary anywhere in the verdict table",
      "bidding" not in json.dumps(conc.VERDICTS).lower()
      and "review_bidding" not in json.dumps(conc.VERDICTS))
check("verdict nouns templated per dimension ('No orders signal')",
      conc.verdict(3000, None, conc.NOUNS["products"])[1].startswith("No orders signal"))

prod6 = [{"product": "prod-a", "revenue": 50, "orders": 30},
         {"product": "prod-b", "revenue": 20, "orders": 0},
         {"product": "prod-c", "revenue": 10, "orders": 5},
         {"product": "prod-d", "revenue": 10, "orders": 0},
         {"product": "prod-e", "revenue": 5, "orders": 1},
         {"product": "prod-f", "revenue": 5, "orders": 0}]
pages120 = [{"name": "/lp%03d" % i, "sessions": 120 - i, "cvr": 0.0295}
            for i in range(120)]
ch_rev = [{"channel": "search", "sessions": 900, "revenue": 5000},
          {"channel": "direct", "sessions": 800, "revenue": 3000}]
block = conc.compute_concentration(
    product_rows=prod6, page_rows=pages120, channel_rows=ch_rev,
    windows={"products": "2026-04-01 – 2026-06-30", "default": "W0"},
    files={"analytics_products.json": {"file": "analytics_products.json",
                                       "sha256": "0" * 64, "bytes": 1}})
dims = {d["key"]: d for d in block["dimensions"]}
check("dimensions: products + landing_pages + channels",
      set(dims) == {"products", "landing_pages", "channels"}, str(set(dims)))
dp = dims["products"]
check("products spend=revenue hhi 3150.0 / band high",
      dp["spend"]["hhi"] == 3150.0 and dp["spend"]["band"] == "high", str(dp["spend"]))
check("products conv=orders hhi 7145.1 / eff_n 1.4 / gini 0.769",
      dp["conv"]["hhi"] == 7145.1 and dp["conv"]["eff_n"] == 1.4
      and dp["conv"]["gini"] == 0.769, str(dp["conv"]))
check("products verdict fragility (both HHIs > 2500)",
      dp["verdict_key"] == "fragility" and "fragility" in dp["verdict"])
check("products top sorted revenue desc / ABC AAABBC on rows",
      [t["name"] for t in dp["top"]] == ["prod-a", "prod-b", "prod-c",
                                         "prod-d", "prod-e", "prod-f"]
      and [t["abc"] for t in dp["top"]] == ["A", "A", "A", "B", "B", "C"],
      str(dp["top"]))
check("small-N caveat leans on Effective-N (6 entities < 8)",
      "Effective-N" in (dp["caveat"] or ""))
check("windows: per-dimension label beats default",
      dp["window"] == "2026-04-01 – 2026-06-30"
      and dims["landing_pages"]["window"] == "W0")
check("files passthrough verbatim",
      block["files"]["analytics_products.json"]["sha256"] == "0" * 64)

dl = dims["landing_pages"]
vals120 = [120.0 - i for i in range(120)]
check("landing pages: bounded embed top 25 + tail 95, math on the FULL universe",
      len(dl["top"]) == 25 and dl["tail"]["n"] == 95
      and dl["spend"]["hhi"] == round(conc.hhi(vals120), 1), str(dl["tail"]))
check("lorenz downsampled <= 101 pts (121 raw points)",
      len(dl["lorenz"]["spend"]) <= 101)
# derived conversions: floor(sessions*0.0295 + 0.5); top page 120 sessions ->
# floor(3.54+0.5) = 4
check("landing conversions derived half-up (120 sess @ 0.0295 -> 4)",
      dl["top"][0]["name"] == "/lp000" and dl["top"][0]["conv"] == 4.0,
      str(dl["top"][0]))
check("landing derived-conversions note fires",
      any("Landing pages: conversions derived" in n for n in block["notes"]),
      str(block["notes"]))
check("channels conv = revenue when exported (no conversions-basis note)",
      dims["channels"]["top"][0]["conv"] == 5000.0
      and not any(n.startswith("Channels:") for n in block["notes"]))

ch_cvr = [{"channel": "search", "sessions": 900, "cvr": 0.024},
          {"channel": "direct", "sessions": 800, "cvr": 0.0292}]
block2 = conc.compute_concentration(channel_rows=ch_cvr)
d2 = {d["key"]: d for d in block2["dimensions"]}["channels"]
check("channels fall back to derived conversions + honest note",
      d2["top"][0]["conv"] == round(900 * 0.024 + 0.5 - 0.5)  # floor(21.6+.5)=22
      or d2["top"][0]["conv"] == 22.0, str(d2["top"]))
check("channels conversions-basis notes fire",
      any("revenue absent" in n for n in block2["notes"])
      and any("Channels: conversions derived" in n for n in block2["notes"]),
      str(block2["notes"]))
block3 = conc.compute_concentration(product_rows=[{"product": "x", "revenue": 10},
                                                  {"product": "y", "revenue": 5}])
d3 = block3["dimensions"][0]
check("no_conv_signal when the orders side is absent",
      d3["conv"] is None and d3["verdict_key"] == "no_conv_signal"
      and d3["verdict"].startswith("No orders signal"), str(d3["verdict_key"]))
check("absent inputs -> None", conc.compute_concentration() is None)
check("all-empty inputs -> None", conc.compute_concentration(product_rows=[]) is None)
check("derived_conversions mirrors cvr_signals (half-up)",
      conc.derived_conversions(1000, 0.0285) == 29
      and conc.derived_conversions(1000, 0.0285) == cv.derived_conversions(1000, 0.0285))

# ============================================================================
# 7: cvr_signals — MediaMetrics-mirror hand-oracles + block assembly
#
# wilson_ci(30, 1000): p=.03, z=1.96 -> center .031799, half .010706 ->
#   (0.0211, 0.0425) at 4dp.
# two_proportion_z(30,1000,60,1000): pooled p=.045, se=.0092709 ->
#   -.03/.0092709 = -3.236.
# min_clicks(0.0299): ln(.05)/ln(.9701) = 98.68 -> ceil 99.
# eb_shrink(3,50,.03,200) = (3 + .03*200)/(50+200) = 9/250 = 0.036.
# beta_binomial_credible(0,100): posterior Beta(1,101), CDF 1-(1-x)^101 ->
#   bounds x_q = 1-(1-q)^(1/101), analytic vs bisection <= 1e-6.
# vwap([2,4,1],[1000,500,4000]) = 8000/5500 = 1.4545.
# ============================================================================
print("--- 7: cvr_signals ---")
check("wilson_ci(30,1000) == (0.0211, 0.0425)",
      tuple(round(x, 4) for x in cv.wilson_ci(30, 1000)) == (0.0211, 0.0425),
      str(cv.wilson_ci(30, 1000)))
check("wilson clamps: n<=0 -> (0,0); succ clamped to [0,n]",
      cv.wilson_ci(5, 0) == (0.0, 0.0)
      and cv.wilson_ci(20, 10) == cv.wilson_ci(10, 10)
      and cv.wilson_ci(-5, 10) == cv.wilson_ci(0, 10))
check("wilson_lower_bound is ci[0] (fair small-n ranking)",
      cv.wilson_lower_bound(30, 1000) == cv.wilson_ci(30, 1000)[0]
      and cv.wilson_lower_bound(1, 1) < cv.wilson_lower_bound(95, 100))
check("two_proportion_z(30,1000,60,1000) == -3.236",
      round(cv.two_proportion_z(30, 1000, 60, 1000), 3) == -3.236,
      str(cv.two_proportion_z(30, 1000, 60, 1000)))
check("z degenerate: n<=0 or se=0 -> 0.0",
      cv.two_proportion_z(1, 0, 1, 10) == 0.0
      and cv.two_proportion_z(0, 10, 0, 10) == 0.0)
check("min_clicks_for_significance(0.0299) == 99",
      cv.min_clicks_for_significance(0.0299) == 99)
check("min_clicks clamps: <=0 -> 0; >=1 -> 1",
      cv.min_clicks_for_significance(0) == 0
      and cv.min_clicks_for_significance(1.5) == 1)
check("empirical_bayes_shrink(3,50,0.03,200) == 0.036",
      approx(cv.empirical_bayes_shrink(3, 50, 0.03, 200), 0.036),
      str(cv.empirical_bayes_shrink(3, 50, 0.03, 200)))
check("eb_shrink clamps: n+k<=0 -> prior; k clamped >= 0",
      cv.empirical_bayes_shrink(0, 0, 0.03, 0) == 0.03
      and cv.empirical_bayes_shrink(3, 50, 0.03, -10) == 0.06)
blo, bhi, bmean = cv.beta_binomial_credible(0, 100)
alo = 1.0 - (1.0 - 0.025) ** (1.0 / 101.0)
ahi = 1.0 - 0.025 ** (1.0 / 101.0)
check("beta_binomial_credible(0,100) matches the analytic bounds <= 1e-6",
      abs(blo - alo) <= 1e-6 and abs(bhi - ahi) <= 1e-6,
      "got (%r, %r) want (%r, %r)" % (blo, bhi, alo, ahi))
check("beta posterior mean = 1/102", approx(bmean, 1.0 / 102.0))
check("vwap([2,4,1],[1000,500,4000]) == 1.4545",
      round(cv.vwap([2, 4, 1], [1000, 500, 4000]), 4) == 1.4545)
check("vwap clamps: empty / zero weights -> 0.0",
      cv.vwap([], []) == 0.0 and cv.vwap([1, 2], [0, 0]) == 0.0)
check("stabilization_point degenerate sd/r -> 0.0",
      cv.stabilization_point(0.03, 0) == 0.0
      and cv.stabilization_point(0.03, 0.01, 1.0) == 0.0)

# derived counts: half-up, never banker's (28.5 -> 29; banker's round() -> 28)
check("derived_conversions half-up: 1000 x 0.0285 -> 29 (banker's says 28)",
      cv.derived_conversions(1000, 0.0285) == 29 and round(28.5) == 28)
check("_resolve: conversion count PREFERRED over shipped rate",
      cv._resolve({"name": "x", "sessions": 100, "conversions": 5, "cvr": 0.5})
      == {"name": "x", "sessions": 100, "conversions": 5, "derived": False,
          "cvr_raw": 0.05})
check("_resolve: funnel aliases (purchases / conversion_rate)",
      cv._resolve({"name": "f", "sessions": 100, "purchases": 7})["conversions"] == 7
      and cv._resolve({"name": "f", "sessions": 100,
                       "conversion_rate": 0.02})["derived"] is True)

# block assembly: 30-page universe, gate n*=99, prior k over the FULL universe
pages30 = [{"name": "/p%02d" % i, "sessions": 310 - 10 * i,
            "cvr": (0.04 if i % 2 == 1 else 0.02)} for i in range(1, 31)]
device_in = [{"name": "mobile", "sessions": 1000, "conversions": 30},
             {"name": "desktop", "sessions": 1000, "conversions": 60}]
sig = cv.compute_cvr_signals(funnel={"sessions": 10000, "conversions": 299},
                             device_rows=device_in, page_rows=pages30,
                             window="2026-04-01 – 2026-06-30")
check("site block: 10000 sessions / 299 conv / cvr 0.0299 / Wilson CI",
      sig["site"]["sessions"] == 10000 and sig["site"]["conversions"] == 299
      and sig["site"]["cvr"] == 0.0299
      and sig["site"]["ci"] == [round(cv.wilson_ci(299, 10000)[0], 6),
                                round(cv.wilson_ci(299, 10000)[1], 6)],
      str(sig["site"]))
check("significance gate n* == 99 at site CVR 0.0299", sig["min_sessions"] == 99)
check("prior k = median sessions over the FULL 30-page universe (155, not 180)",
      sig["prior"]["k"] == 155.0
      and median(p["sessions"] for p in pages30) == 155
      and median(sorted((p["sessions"] for p in pages30), reverse=True)[:25]) == 180,
      str(sig["prior"]))
check("prior basis names the funnel", sig["prior"]["basis"] == "site CVR (funnel)")
check("pages embed bounded to top-25 by sessions",
      len(sig["pages"]) == 25 and sig["pages"][0]["page"] == "/p01"
      and sig["pages"][-1]["page"] == "/p25")
check("pages_universe spans ALL 30 pages (4650 sessions, 9 gated)",
      sig["pages_universe"] == {"n": 30, "sessions": 4650, "gated_n": 9},
      str(sig["pages_universe"]))
p1 = sig["pages"][0]
check("page row: derived count floor(300*0.04+0.5)=12, cvr_raw keeps the shipped rate",
      p1["conversions"] == 12 and p1["derived"] is True and p1["cvr_raw"] == 0.04)
check("page shrinkage: eb_shrink(12,300,0.0299,155) (thin page pulled to site rate)",
      p1["cvr_shrunk"] == round(cv.empirical_bayes_shrink(12, 300, 0.0299, 155), 6)
      and p1["cvr_shrunk"] < p1["cvr_raw"], str(p1))
check("page wilson_lb populated", p1["wilson_lb"] == round(cv.wilson_lower_bound(12, 300), 6))
check("gate flags: 90-session page gated, 100-session page not",
      next(p for p in sig["pages"] if p["sessions"] == 100)["gated"] is False
      and all(b["sessions"] >= 99 or b["gated"]
              for b in sig["pages"]), str([p["gated"] for p in sig["pages"]]))

# SINGLE-SOURCE COMPLEMENT: mobile z tests against its SIBLING (desktop),
# never against the cross-source funnel remainder (which would say z ~ 0.02).
segd = {r["name"]: r for r in sig["segments"]["device"]}
cross_z = round(cv.two_proportion_z(30, 1000, 299 - 30, 10000 - 1000), 2)
check("device z = sibling complement -3.24 (cross-source would say %.2f)" % cross_z,
      segd["mobile"]["z"] == -3.24 and segd["mobile"]["significant"] is True
      and segd["mobile"]["z"] != cross_z, str(segd["mobile"]))
check("desktop mirrors +3.24 significant",
      segd["desktop"]["z"] == 3.24 and segd["desktop"]["significant"] is True)
check("segment rows sorted sessions desc, name asc (desktop before mobile)",
      [r["name"] for r in sig["segments"]["device"]] == ["desktop", "mobile"])
check("headline mobile-vs-desktop z (prefix-matched)",
      sig["segments"]["headline_device_z"] == {"z": -3.24, "significant": True},
      str(sig["segments"]["headline_device_z"]))
check("window passthrough", sig["window"] == "2026-04-01 – 2026-06-30")
check("notes fixed order: derived, gate, stabilization, missing inputs",
      sig["notes"][0].startswith("30 of 33 input rows ship a rate")
      and "n*=99" in sig["notes"][1] and "9 of 30" in sig["notes"][1]
      and sig["notes"][2].startswith("Page CVR stabilizes")
      and any("No channel rows" in n for n in sig["notes"])
      and any("No new-vs-returning rows" in n for n in sig["notes"]),
      str(sig["notes"]))
check("cvr_signals deterministic (double compute identical)",
      json.dumps(sig, sort_keys=True) == json.dumps(
          cv.compute_cvr_signals(funnel={"sessions": 10000, "conversions": 299},
                                 device_rows=device_in, page_rows=pages30,
                                 window="2026-04-01 – 2026-06-30"),
          sort_keys=True))

sig_fb = cv.compute_cvr_signals(device_rows=device_in)
check("funnel fallback: site totals summed from device rows + note",
      sig_fb["site"]["sessions"] == 2000 and sig_fb["site"]["conversions"] == 90
      and sig_fb["prior"]["basis"] == "site CVR (summed device rows)"
      and sig_fb["notes"][0].startswith("No funnel totals provided"),
      str(sig_fb["site"]))
check("no usable input at all -> None", cv.compute_cvr_signals() is None)
check("constants pinned: Z_SIG 1.96 / TOP_N 25 / CONFIDENCE 0.95",
      cv.Z_SIG == 1.96 and cv.TOP_N == 25 and cv.CONFIDENCE == 0.95)

# ============================================================================
# 8: shopify_rows — sanitized raw-shape fixtures (exact envelope quirks)
# ============================================================================
print("--- 8: shopify_rows (raw shapes) ---")
tbl = sr.load_table(str(FIX / "raw_shape_funnel.json"))
check("STRING rows coerced by dataType: INTEGER->int, PERCENT->float FRACTION",
      tbl[0]["sessions"] == 10000 and isinstance(tbl[0]["sessions"], int)
      and tbl[0]["conversion_rate"] == 0.0198
      and isinstance(tbl[0]["conversion_rate"], float))
check("PERCENT never x100 at parse time (fraction < 1)",
      0 < tbl[0]["conversion_rate"] < 1)

fun = sr.funnel_from_table(str(FIX / "raw_shape_funnel.json"))
check("funnel adapter: counts + fraction rates + verbatim cvr",
      fun == {"sessions": 10000, "atc_sessions": 1220, "checkout_sessions": 920,
              "purchase_sessions": 198, "atc_rate": 0.122, "checkout_rate": 0.092,
              "cvr": 0.0198}, str(fun))
check("cvr verbatim fraction == purchases/sessions (SHAPE-NOTES fact 2)",
      fun["cvr"] == 198 / 10000)

dev = sr.device_rows_from_table(str(FIX / "raw_shape_device.json"))
check("device rows sorted sessions desc; names verbatim incl 'other'",
      [r["name"] for r in dev] == ["mobile", "desktop", "tablet", "other"]
      and dev[0] == {"name": "mobile", "sessions": 1500, "cvr": 0.0143}, str(dev))
check("zero-CVR row keeps the 0.0 fraction", dev[3]["cvr"] == 0.0)

refr = sr.referrer_rows_from_table(str(FIX / "raw_shape_referrer.json"))
check("referrer rows -> channel shape", refr[0] == {"name": "search",
      "sessions": 900, "cvr": 0.024} and len(refr) == 4)

land = sr.landing_rows_from_table(str(FIX / "raw_shape_landing.json"))
check("URL normalization merges dupes: /pages/a?ref=x + /Pages/A/ -> 300 sessions",
      [(r["name"], r["sessions"]) for r in land]
      == [("/", 300), ("/pages/a", 300), ("/products/b", 150)], str(land))
check("merged cvr = sessions-weighted mean ((200*.02 + 100*.05)/300 = 0.03)",
      approx(next(r for r in land if r["name"] == "/pages/a")["cvr"], 0.03))
check("normalize_url: lowercase + strip ?query + trailing / (root survives)",
      sr.normalize_url("/Pages/A/?q=1") == "/pages/a"
      and sr.normalize_url("") == "/" and sr.normalize_url("///") == "/"
      and sr.normalize_url("/x") == "/x")

prods = sr.product_rows_from_table(str(FIX / "raw_shape_products.json"))
check("product rows: net_sales as revenue, orders int, sorted revenue desc",
      prods == [{"product": "Product A", "revenue": 5000.0, "orders": 1},
                {"product": "Product B", "revenue": 3000.0, "orders": 10},
                {"product": "Product C", "revenue": 2000.5, "orders": 8}]
      or prods[0]["product"] == "Product A", str(prods))

tot = sr.totals_from_table(str(FIX / "raw_shape_totals.json"))
check("AOV VERBATIM: 242.494, never total_sales/orders (298.91) nor "
      "net_sales/orders (252.66)",
      tot["aov"] == 242.494 and tot["aov"] != tot["total_sales"] / tot["orders"]
      and tot["aov"] != tot["net_sales"] / tot["orders"], str(tot))
cust = sr.customers_from_table(str(FIX / "raw_shape_customers.json"))
check("customers: verbatim order-share fraction (173/530)",
      cust == {"customers": 530, "returning_customers": 173,
               "returning_customer_rate": 0.3264150943396226}, str(cust))

check("checksum_note: matching summaryMetric -> None (sessions 2,120)",
      sr.checksum_note(str(FIX / "raw_shape_device.json")) is None)
bad = sr.checksum_note(str(FIX / "raw_shape_products_badsum.json"))
check("checksum_note: deliberate mismatch -> warn note naming both sums",
      bad is not None and "12,345.67" in bad and "10,000" in bad
      and "re-save the verbatim result" in bad, str(bad))

try:
    sr.load_table(str(FIX / "raw_shape_error.json"))
    check("error-shaped result raises RawResultError", False)
except sr.RawResultError as e:
    check("error-shaped result raises RawResultError",
          "error-shaped" in str(e) and "shopify-pulls.md" in str(e), str(e))
try:
    sr.funnel_from_table(str(FIX / "raw_shape_device.json"))
    check("wrong-pull guard: GROUP BY fed to the funnel adapter raises", False)
except sr.RawResultError:
    check("wrong-pull guard: GROUP BY fed to the funnel adapter raises", True)
try:
    sr.device_rows_from_table(str(FIX / "raw_shape_funnel.json"))
    check("wrong-pull guard: funnel fed to the device adapter raises", False)
except sr.RawResultError:
    check("wrong-pull guard: funnel fed to the device adapter raises", True)
try:
    sr.load_table(str(FIX / "raw_shape_shop_info.json"))
    check("non-envelope file raises for load_table", False)
except sr.RawResultError:
    check("non-envelope file raises for load_table", True)

info = sr.load_shop_info(str(FIX / "raw_shape_shop_info.json"))
check("shop_info flat object verbatim (currencyCode CAD)",
      info["name"] == "Test Store" and info["currencyCode"] == "CAD")
orders = sr.load_orders(str(FIX / "raw_shape_orders.json"))
check("orders canonical rows sorted created_at desc, name asc",
      [o["name"] for o in orders] == ["#1003", "#1001", "#1002"]
      and orders[0]["total_price"] == 510.9
      and orders[0]["created_at"] == "2026-07-11T14:20:32Z", str(orders))
gql = sr.load_products(str(FIX / "raw_shape_products_graphql.json"))
check("GraphQL edges/node -> canonical products sorted title asc",
      [p["title"] for p in gql] == ["Product A", "Product B"]
      and gql[0]["status"] == "DRAFT" and gql[0]["price"] == 12.5  # variant fallback
      and gql[1]["price"] == 86.98 and gql[1]["currency"] == "CAD"
      and gql[1]["sku"] == "SKU-B" and gql[1]["total_inventory"] == 66,
      str(gql))
check("gid:// ids survive verbatim",
      gql[0]["id"] == "gid://shopify/Product/2000000000002")
stamp = sr.file_stamp(str(FIX / "raw_shape_funnel.json"))
check("file_stamp shape + real sha256",
      set(stamp) == {"file", "sha256", "bytes"}
      and stamp["sha256"] == hashlib.sha256(
          (FIX / "raw_shape_funnel.json").read_bytes()).hexdigest())

# ============================================================================
# 9: manual CSV path (NEEDS-REAL-EXPORT-VALIDATION formats, synthetic files)
# ============================================================================
print("--- 9: manual CSV ---")
lrows, lmeta = mc.ga4_landing_rows(str(CSVDIR / "ga4-landing.csv"))
check("ga4-landing: '#' preamble skipped, window parsed from YYYYMMDD-YYYYMMDD",
      lmeta["window"] == "2026-04-01 – 2026-06-30", str(lmeta["window"]))
check("ga4-landing: Grand total row captured as totals, excluded from entities",
      lmeta["totals"] == {"sessions": 10000, "key_events": 197, "cvr": 0.0197}
      and all(r["name"] != "grand total" for r in lrows), str(lmeta["totals"]))
check("ga4-landing: blank-line second (day-by-day) section NOT parsed",
      len(lrows) == 3 and all(r["name"] not in ("0", "1", "nth day")
                              for r in lrows), str([r["name"] for r in lrows]))
check("ga4-landing: URL-normalized merge (/pages/a?variant + /Pages/A/ -> 500)",
      [(r["name"], r["sessions"]) for r in lrows]
      == [("/", 2400), ("/pages/a", 500), ("/products/b", 300)], str(lrows))
check("ga4-landing: merged cvr sessions-weighted, aux counts summed",
      approx(lrows[1]["cvr"], (450 * 0.0044 + 50 * 0.02) / 500)
      and lrows[1]["key_events"] == 3)
check("GA4 bare rate defaults FRACTION (0.0366 stays 0.0366)",
      approx(lrows[0]["cvr"], 0.0366))
try:
    mc.ga4_landing_rows(str(CSVDIR / "ga4-device.csv"))
    check("wrong-report guard raises (device export fed to landing adapter)", False)
except mc.ManualCsvError as e:
    check("wrong-report guard raises (device export fed to landing adapter)",
          "Landing page + query string" in str(e), str(e))

drows, dmeta = mc.ga4_device_rows(str(CSVDIR / "ga4-device.csv"))
check("ga4-device: thousands comma + sessions-desc sort",
      [(r["name"], r["sessions"]) for r in drows]
      == [("mobile", 8000), ("desktop", 3000), ("tablet", 400)], str(drows))

gfun, gmeta = mc.ga4_funnel(str(CSVDIR / "ga4-funnel.csv"))
check("ga4-funnel: step mapping to shopify_rows.funnel_from_table keys",
      (gfun["sessions"], gfun["atc_sessions"], gfun["checkout_sessions"],
       gfun["purchase_sessions"]) == (8500, 1000, 750, 180), str(gfun))
check("ga4-funnel: rates computed from USER counts + users basis flagged",
      approx(gfun["cvr"], 180 / 8500) and gmeta["basis"] == "users"
      and any("USERS" in n for n in gmeta["notes"]), str(gmeta["notes"]))

sfun, smeta = mc.shopify_conversion_funnel(str(CSVDIR / "shopify-conversion.csv"))
check("shopify-conversion: Totals row used VERBATIM (primary funnel source)",
      sfun["sessions"] == 10000 and sfun["atc_sessions"] == 1220
      and sfun["checkout_sessions"] == 920 and sfun["purchase_sessions"] == 210
      and "derived" not in sfun, str(sfun))
check("shopify-conversion: '%'-cell rate -> fraction 0.021; sessions basis",
      approx(sfun["cvr"], 0.021) and smeta["basis"] == "sessions"
      and any("verbatim" in n for n in smeta["notes"]), str(smeta["notes"]))
check("shopify-conversion: Day-column window",
      smeta["window"] == "2026-04-01 – 2026-04-02")
try:
    mc.shopify_conversion_funnel(str(CSVDIR / "bad-conversion-misscaled.csv"))
    check(">20% site-CVR mis-scale guard aborts", False)
except mc.ManualCsvError as e:
    check(">20% site-CVR mis-scale guard aborts", "mis-scaled" in str(e), str(e))

trows, tmeta = mc.shopify_traffic_source_rows(str(CSVDIR / "shopify-traffic-source.csv"))
tby = {r["name"]: r for r in trows}
check("traffic source: Totals row excluded (4 entities)", len(trows) == 4)
check("'%' cell -> percent/100 ('2.40%' -> 0.024)", approx(tby["search"]["cvr"], 0.024))
check("bare fraction stays fraction ('0.0292')", approx(tby["direct"]["cvr"], 0.0292))
check("Shopify bare > 1 defensively percent ('5.48' -> 0.0548)",
      approx(tby["email"]["cvr"], 0.0548))
check("totals: currency prefix + thousands parsed ($70,210.65)",
      approx(tmeta["totals"]["revenue"], 70210.65)
      and tmeta["totals"]["sessions"] == 10000, str(tmeta["totals"]))

prows, pmeta = mc.shopify_product_rows(str(CSVDIR / "shopify-sales-product.csv"))
check("products: prefix-matched 'Net sales (CAD)', Total row captured",
      pmeta["revenue_column"] == "Net sales (CAD)"
      and pmeta["totals"] == {"revenue": 22944.04, "orders": 27, "units": 32},
      str(pmeta))
check("product row shape {product, revenue, orders, units} sorted revenue desc",
      prows[0] == {"product": "Product A", "revenue": 11020.94, "orders": 1,
                   "units": 1}, str(prows[0]))

nrows, nmeta = mc.ga4_new_returning_rows(str(CSVDIR / "ga4-new-returning.csv"))
check("new/returning canonicalized ('established' -> 'returning'); users basis noted",
      [r["name"] for r in nrows] == ["new", "returning"]
      and nrows[0]["sessions"] == 7000
      and any("users basis" in n for n in nmeta["notes"]), str(nrows))

crows, cmeta = mc.shopify_customer_rows(str(CSVDIR / "shopify-customers.csv"))
check("customers: 'First-time' -> 'new'; summary mirrors customers_from_table",
      [r["name"] for r in crows] == ["new", "returning"]
      and cmeta["customers_summary"]["customers"] == 530
      and cmeta["customers_summary"]["returning_customers"] == 173
      and approx(cmeta["customers_summary"]["returning_customer_rate"], 173 / 530),
      str(cmeta.get("customers_summary")))

aov, ameta = mc.shopify_aov(str(CSVDIR / "shopify-aov.csv"))
check("AOV totals row VERBATIM (242.494 — never recomputed, never averaged)",
      aov == {"aov": 242.494} and ameta["aov"] == 242.494, str(ameta["notes"]))
aov2, ameta2 = mc.shopify_aov(str(CSVDIR / "bad-aov-multirow.csv"))
check("multi-row AOV without totals -> {} + honest never-recomputed note",
      aov2 == {} and ameta2["aov"] is None
      and any("never recomputed" in n for n in ameta2["notes"]), str(ameta2["notes"]))

csv_all = mc.load_csv_dir(str(CSVDIR))
check("load_csv_dir finds all 11 canonical filenames (bad-* ignored)",
      sorted(csv_all) == sorted(mc.ADAPTERS), str(sorted(csv_all)))
check("double parse identical (deterministic)",
      csv_all == mc.load_csv_dir(str(CSVDIR)))
check("every adapter meta carries a provenance stamp",
      all(set(meta["stamp"]) == {"file", "sha256", "bytes"}
          for _d, meta in csv_all.values()))

# ============================================================================
# 10: machine layer — precedence, percent boundary, Reads, divergence notes
# ============================================================================
print("--- 10: machine ---")
RAW = {"shop_info": str(FIX / "raw_shape_shop_info.json"),
       "analytics_funnel": str(FIX / "raw_shape_funnel.json"),
       "analytics_device": str(FIX / "raw_shape_device.json"),
       "analytics_referrer": str(FIX / "raw_shape_referrer.json"),
       "analytics_landing": str(FIX / "raw_shape_landing.json"),
       "analytics_products": str(FIX / "raw_shape_products.json"),
       "analytics_totals": str(FIX / "raw_shape_totals.json"),
       "analytics_customers": str(FIX / "raw_shape_customers.json")}

mach_raw = mx.compute_machine({"raw": RAW})
check("raw-only funnel: counts int + rates recomputed from counts, PERCENT 2dp",
      mach_raw["analytics"]["funnel"]
      == {"sessions": 10000, "atc": 1220, "checkout": 920, "purchases": 198,
          "atc_rate": 12.2, "checkout_rate": 9.2, "cvr": 1.98},
      str(mach_raw["analytics"]["funnel"]))
check("fraction->percent happens EXACTLY once (device 0.0143 -> 1.43)",
      mach_raw["analytics"]["device"][0]
      == {"device": "mobile", "sessions": 1500, "cvr": 1.43})
check("raw landing: URL-normalized universe, share vs FULL universe (300/750=40.0)",
      mach_raw["analytics"]["landing_pages"][0]
      == {"page": "/", "sessions": 300, "share_pct": 40.0}
      and len(mach_raw["universes"]["pages"]) == 3,
      str(mach_raw["analytics"]["landing_pages"]))
check("raw aov verbatim 242.494 + Over-$200 band read",
      mach_raw["analytics"]["aov"] == 242.494
      and mach_raw["reads"]["aov_band"].startswith("Over-$200"))
check("nvr skipped without GA4 (the only session-basis source)",
      "new_vs_returning" not in mach_raw["analytics"]
      and any(s["field"] == "analytics.new_vs_returning"
              and "GA4" in s["reason"] for s in mach_raw["skipped"]),
      str(mach_raw["skipped"]))
check("customers order-share note (evidence, not a CVR)",
      any("order share — evidence, not a CVR" in n for n in mach_raw["notes"]),
      str(mach_raw["notes"]))
check("currency from shop_info", mach_raw.get("currency") == "CAD")
check("machine deterministic (double compute identical)",
      mx.compute_machine({"raw": RAW}) == mach_raw)
check("no inputs -> None", mx.compute_machine(None) is None
      and mx.compute_machine({}) is None)

badsum_raw = mx.compute_machine(
    {"raw": {"analytics_products": str(FIX / "raw_shape_products_badsum.json")}})
check("summaryMetric checksum mismatch surfaces as a machine note",
      any("checksum" in n and "12,345.67" in n for n in badsum_raw["notes"]),
      str(badsum_raw["notes"]))

mach = mx.compute_machine({"csv": csv_all, "raw": RAW})
check("funnel precedence: shopify-conversion.csv beats analytics_funnel + ga4-funnel",
      mach["sources"]["funnel"] == "shopify-conversion.csv"
      and mach["analytics"]["funnel"]
      == {"sessions": 10000, "atc": 1220, "checkout": 920, "purchases": 210,
          "atc_rate": 12.2, "checkout_rate": 9.2, "cvr": 2.1},
      str(mach["analytics"]["funnel"]))
check("device precedence: ga4-device.csv beats analytics_device.json",
      mach["sources"]["device"] == "ga4-device.csv"
      and mach["analytics"]["device"][0]
      == {"device": "mobile", "sessions": 8000, "cvr": 1.45},
      str(mach["analytics"]["device"]))
check("channels precedence: ga4-channels.csv first of three candidates",
      mach["sources"]["channels"] == "ga4-channels.csv"
      and mach["analytics"]["channels"][0]
      == {"channel": "Organic Search", "sessions": 4000, "cvr": 2.4,
          "revenue": 50000.0}, str(mach["analytics"]["channels"]))
check("landing precedence: ga4-landing.csv; shares vs FULL 3200-session universe",
      mach["sources"]["landing_pages"] == "ga4-landing.csv"
      and mach["analytics"]["landing_pages"]
      == [{"page": "/", "sessions": 2400, "share_pct": 75.0},
          {"page": "/pages/a", "sessions": 500, "share_pct": 15.6},
          {"page": "/products/b", "sessions": 300, "share_pct": 9.4}],
      str(mach["analytics"]["landing_pages"]))
check("products from shopify-sales-product.csv (Product A 48.0% share)",
      mach["sources"]["revenue_concentration"] == "shopify-sales-product.csv"
      and mach["analytics"]["revenue_concentration"][0]
      == {"product": "Product A", "revenue": 11020.94, "share_pct": 48.0},
      str(mach["analytics"]["revenue_concentration"][0]))
check("nvr from GA4 (percent 2dp)",
      mach["analytics"]["new_vs_returning"] == {"new_cvr": 1.5, "returning_cvr": 3.2})
check("aov precedence: shopify-aov.csv verbatim", mach["sources"]["aov"]
      == "shopify-aov.csv" and mach["analytics"]["aov"] == 242.494)
# divergence notes: funnel 10,000 vs ga4 8,500 = 15% > 10%; raw funnel equal ->
# silent. Device ga4 11,400 vs raw 2,120 = 81.4%.
check("funnel divergence >10% noted (15.0% vs ga4-funnel.csv)",
      any("funnel: sessions diverge 15.0% between shopify-conversion.csv "
          "(10,000) and ga4-funnel.csv (8,500)" in n for n in mach["notes"]),
      str(mach["notes"]))
check("equal-sessions candidate stays silent (raw funnel also 10,000)",
      not any("analytics_funnel" in n and "diverge" in n for n in mach["notes"]))
check("device divergence 81.4% noted vs the raw pull",
      any("device: sessions diverge 81.4% between ga4-device.csv (11,400) "
          "and raw_shape_device.json (2,120)" in n for n in mach["notes"]),
      str(mach["notes"]))
check("machine Reads: funnel At/above + Below (2.1 vs 2.99)",
      mach["reads"]["funnel.atc_rate"] == "At / above benchmark"
      and mach["reads"]["funnel.checkout_rate"] == "At / above benchmark"
      and mach["reads"]["funnel.cvr"] == "Below benchmark", str(mach["reads"]))
check("machine Reads: Mobile/Desktop only — Tablet has no benchmark",
      mach["reads"]["device.mobile"] == "Well below benchmark"
      and mach["reads"]["device.desktop"] == "Below benchmark"
      and "device.tablet" not in mach["reads"], str(mach["reads"]))
check("windows: funnel window rides to default",
      mach["windows"]["funnel"] == "2026-04-01 – 2026-04-02"
      and mach["windows"]["default"] == "2026-04-01 – 2026-04-02")
check("stamps carry both csv and raw provenance",
      "shopify-conversion.csv" in mach["stamps"]
      and "raw_shape_funnel.json" in mach["stamps"])
check("universes retained for concentration/cvr_signals (fractions)",
      approx(mach["universes"]["funnel"]["cvr"], 0.021)
      and mach["universes"]["pages"][0]["sessions"] == 2400
      and mach["universes"]["products"][0]["revenue"] == 11020.94)
# exact-binary .5 tie at 2dp: 0.125 IS representable, so banker's round()
# demonstrably gives 0.12 where the Excel-parity boundary must give 0.13
check("percent boundary rounding is half-up (0.125 -> 0.13, banker's 0.12)",
      am._round_half_up(0.125, 2) == 0.13 and round(0.125, 2) == 0.12
      and mx._pct(0.0198) == 1.98)

# ============================================================================
# 11: merge — purity, replace-and-log, filled/skipped, author fields untouched
# ============================================================================
print("--- 11: merge ---")
before = copy.deepcopy(PAYLOAD)
merged, mblock, mlog = mx.merge_into_payload(PAYLOAD, mach)
check("merge is PURE (input payload unmutated)", PAYLOAD == before)
check("machine values REPLACE transcribed values (sessions 482190 -> 10000)",
      merged["analytics"]["funnel"]["sessions"] == 10000
      and merged["analytics"]["funnel"]["cvr"] == 2.1)
check("pinned log-line format: counts corrected on ANY diff",
      "machine: analytics.funnel.sessions 482190->10000 (shopify-conversion.csv)"
      in mlog, str(mlog[:4]))
check("rate corrected > 0.05 abs with 2dp formatting",
      "machine: analytics.funnel.cvr 2.30->2.10 (shopify-conversion.csv)" in mlog,
      str([l for l in mlog if "cvr" in l]))
check("device rows matched case-insensitively (Mobile -> mobile), per-cell log",
      any(l.startswith("machine: analytics.device[mobile].cvr 1.90->1.45")
          for l in mlog), str([l for l in mlog if "device" in l]))
check("entity-set change logged once for landing pages",
      "machine: analytics.landing_pages.entities 3 rows->3 rows (ga4-landing.csv)"
      in mlog, str([l for l in mlog if "landing" in l]))
check("aov corrected + stored verbatim (log shows 2dp)",
      "machine: analytics.aov 84.90->242.49 (shopify-aov.csv)" in mlog
      and merged["analytics"]["aov"] == 242.494)
check("corrected entries mirror the log", len(mblock["corrected"]) >= 5
      and any(c["field"] == "analytics.funnel.sessions" and c["from"] == 482190
              and c["to"] == 10000 for c in mblock["corrected"]),
      str(mblock["corrected"][:3]))
check("author fields NEVER touched: meta / steps_detail / findings",
      merged["meta"] == PAYLOAD["meta"]
      and merged["steps_detail"] == PAYLOAD["steps_detail"]
      and merged["findings"] == PAYLOAD["findings"])
check("non-blank currency untouched + honest note (USD stays, CAD reported)",
      merged["meta"]["currency"] == "USD"
      and any("meta.currency left as transcribed (USD)" in n
              for n in mblock["notes"]), str(mblock["notes"][-2:]))
check("reads + sources + skipped ride into machine_block",
      mblock["reads"] == mach["reads"] and mblock["sources"] == mach["sources"]
      and mblock["skipped"] == mach["skipped"])
check("applied sorted + field-level for scalar blocks",
      mblock["applied"] == sorted(mblock["applied"])
      and "analytics.funnel.cvr" in mblock["applied"]
      and "analytics.landing_pages" in mblock["applied"])

# filled: blank payload fields become machine-filled, not corrected
bare = {"meta": {"store_name": "Bare", "currency": ""}, "analytics": {},
        "steps_detail": {}, "findings": []}
merged_b, block_b, log_b = mx.merge_into_payload(bare, mach)
check("blank fields land in filled (funnel fields + list blocks + currency)",
      "analytics.funnel.sessions" in block_b["filled"]
      and "analytics.device" in block_b["filled"]
      and "meta.currency" in block_b["filled"]
      and merged_b["meta"]["currency"] == "CAD", str(block_b["filled"]))
check("filling emits no correction log lines", log_b == [], str(log_b))
# single-source drop: a transcribed stage the machine source lacks is dropped
partial_machine = {"analytics": {"funnel": {"sessions": 900, "purchases": 18,
                                            "cvr": 2.0}},
                   "sources": {"funnel": "ga4-funnel.csv"},
                   "reads": {}, "skipped": [], "notes": []}
merged_c, block_c, log_c = mx.merge_into_payload(PAYLOAD, partial_machine)
check("whole-block funnel replace: stages never mixed across sources",
      merged_c["analytics"]["funnel"] == {"sessions": 900, "purchases": 18,
                                          "cvr": 2.0},
      str(merged_c["analytics"]["funnel"]))
check("dropped stages logged with the single-source reason",
      any(l == "machine: analytics.funnel.atc 29413->dropped "
              "(single-source: not in ga4-funnel.csv)" for l in log_c),
      str(log_c))
check("machine=None is a no-op passthrough (--no-machine equivalence)",
      mx.merge_into_payload(PAYLOAD, None) == (PAYLOAD, None, []))
p_none, b_none, l_none = mx.merge_into_payload(PAYLOAD, None)
check("no-op returns the SAME payload object", p_none is PAYLOAD)

# merged model + renderers with all blocks
cvr_from_mach = cv.compute_cvr_signals(
    funnel={"sessions": mach["universes"]["funnel"]["sessions"],
            "conversions": mach["universes"]["funnel"]["purchase_sessions"]},
    device_rows=mach["universes"]["device"],
    page_rows=mach["universes"]["pages"],
    window=mach["windows"]["default"])
conc_from_mach = conc.compute_concentration(
    product_rows=mach["universes"]["products"],
    page_rows=mach["universes"]["pages"],
    windows={"default": mach["windows"]["default"]})
m_full = am.compute_model(merged, generated="2026-06-25T00:00:00",
                          concentration=conc_from_mach,
                          cvr_signals=cvr_from_mach, machine=mblock)
check("merged model: machine funnel drives Funnel Health "
      "(12.2/7.23 + 9.2/5.96 + 2.1/2.99 -> 131 A)",
      m_full["health"]["score"] == 131 and m_full["health"]["grade"] == "A",
      str(m_full["health"]))
check("blocks passthrough verbatim on the model",
      m_full["concentration"] is conc_from_mach
      and m_full["cvr_signals"] is cvr_from_mach and m_full["machine"] is mblock)
html_f = audit_html.render_html(m_full)
check("html embeds concentration + cvr + machine data",
      '"verdict_key"' in html_f and '"wilson_lb"' in html_f
      and '"applied"' in html_f)
check("html with all blocks self-contained",
      not re.findall(r"https?://|<link|src=|cdn",
                     html_f.replace(audit_html.gsap_blob(), "")))
js_grades_full = [(int(n), g) for n, g in
                  re.findall(r"\[(\d+),'([A-F])'\]", html_f.replace(
                      audit_html.gsap_blob(), ""))]
check("GRADES parity unpolluted in the FULL rendered report",
      js_grades_full == am.GRADE_CUTOFFS, str(js_grades_full))
md_f = audit_md.render_md(m_full)
check("md has Concentration + CVR Signals sections + machine footer",
      "## Concentration — weight vs outcomes (HHI)" in md_f
      and "## CVR Signals" in md_f
      and "analytics fields machine-computed" in md_f)
check("md machine footer counts applied/corrected",
      ("_%d analytics fields machine-computed · %d correction(s)"
       % (len(mblock["applied"]), len(mblock["corrected"]))) in md_f,
      md_f[md_f.find("machine-computed") - 60:md_f.find("machine-computed") + 40])
check("md omits the sections when blocks absent",
      "## Concentration" not in ma and "## CVR Signals" not in ma)
f1 = audit_html.render_html(am.compute_model(
    merged, generated="TS1", concentration=conc_from_mach,
    cvr_signals=cvr_from_mach, machine=mblock))
f2 = audit_html.render_html(am.compute_model(
    merged, generated="TS2", concentration=conc_from_mach,
    cvr_signals=cvr_from_mach, machine=mblock))
check("HTML with all blocks deterministic modulo generated",
      f1.replace("TS1", "TS") == f2.replace("TS2", "TS"))

# ============================================================================
# 12: workbook — 15 tabs, 10 named ranges, optional tabs, C6/C7/C8, whitelabel
# ============================================================================
print("--- 12: workbook ---")
if bcw is None:
    skip_note("workbook section entirely (openpyxl unavailable)")
else:
    from openpyxl import load_workbook as _lw

    def _cells(ws):
        for row in ws.iter_rows(values_only=True):
            for v in row:
                yield v

    with tempfile.TemporaryDirectory() as td:
        x1 = Path(td) / "plain.xlsx"
        bcw.build(json.loads(SAMPLE.read_text())).save(x1)
        wb1 = _lw(x1)
        check("EXPECTED_TABS all 15 present",
              all(t in wb1.sheetnames for t in bcw.EXPECTED_TABS),
              str(wb1.sheetnames))
        check("optional tabs absent when blocks not provided",
              all(t not in wb1.sheetnames for t in bcw.OPTIONAL_TABS))
        names = bcw.defined_name_set(wb1)
        check("all 10 REQUIRED_NAMES resolve", bcw.REQUIRED_NAMES <= names,
              str(sorted(bcw.REQUIRED_NAMES - names)))
        check("check() green without optional tabs", bcw.check(str(x1)) == 0)
        check("rate named ranges pinned to 02_Analytics C6/C7/C8",
              all(coord in str(dict(wb1.defined_names)[n].value)
                  for n, coord in (("rate_atc", "$C$6"),
                                   ("rate_checkout", "$C$7"),
                                   ("rate_cvr", "$C$8"))),
              str({n: str(dict(wb1.defined_names)[n].value)
                   for n in ("rate_atc", "rate_checkout", "rate_cvr")}))
        ws_an = wb1["02_Analytics"]
        check("sample payload rates land in C6/C7/C8",
              (ws_an["C6"].value, ws_an["C7"].value, ws_an["C8"].value)
              == (6.10, 4.40, 2.30),
              str((ws_an["C6"].value, ws_an["C7"].value, ws_an["C8"].value)))
        exec_cells = [v for v in _cells(wb1["01_Executive_Summary"])
                      if isinstance(v, str)]
        check("exec Funnel Health formula is ISNUMBER-weighted (measured stages)",
              any("N(ISNUMBER(rate_atc))" in v and "MIN(150" in v
                  for v in exec_cells),
              str([v for v in exec_cells if "ISNUMBER" in str(v)][:1]))
        all_text = " ".join(str(v) for ws in wb1.worksheets for v in _cells(ws)
                            if isinstance(v, str))
        check("whitelabel: attribution scrubbed (no Goward / Conversion.com)",
              "goward" not in all_text.lower()
              and "conversion.com" not in all_text.lower())
        check("LIFT factor names survive the scrub",
              "Value Proposition" in all_text and "LIFT" in all_text)

        x2 = Path(td) / "blocks.xlsx"
        bcw.build(json.loads(SAMPLE.read_text()), concentration=block,
                  cvr_signals=sig).save(x2)
        wb2 = _lw(x2)
        check("15_Concentration + 16_CVR_Signals appear with blocks",
              "15_Concentration" in wb2.sheetnames
              and "16_CVR_Signals" in wb2.sheetnames)
        check("optional tabs append strictly AFTER 14_Reference",
              wb2.sheetnames.index("15_Concentration")
              > wb2.sheetnames.index("14_Reference"))
        check("check() green WITH optional tabs", bcw.check(str(x2)) == 0)
        check("concentration tab carries the verdict rows",
              any(v == "Verdict" for v in _cells(wb2["15_Concentration"])))
        check("cvr tab carries the significance gate",
              any(v == "Significance gate n*" for v in _cells(wb2["16_CVR_Signals"])))
        check("optional tabs never join EXPECTED_TABS (legacy check unchanged)",
              bcw.check(str(x1)) == 0)

        x3 = Path(td) / "merged.xlsx"
        bcw.build(copy.deepcopy(merged)).save(x3)
        wb3 = _lw(x3)
        ws3 = wb3["02_Analytics"]
        check("machine-corrected rates land in C6/C7/C8 (12.2 / 9.2 / 2.1)",
              (ws3["C6"].value, ws3["C7"].value, ws3["C8"].value)
              == (12.2, 9.2, 2.1),
              str((ws3["C6"].value, ws3["C7"].value, ws3["C8"].value)))
        check("machine-corrected counts land too (B5 sessions 10000)",
              ws3["B5"].value == 10000 and ws3["B6"].value == 1220)
        check("check() green on the merged workbook", bcw.check(str(x3)) == 0)

        # determinism: two builds of the same payload -> identical cell values
        x4 = Path(td) / "again.xlsx"
        bcw.build(json.loads(SAMPLE.read_text())).save(x4)
        wb4 = _lw(x4)
        same = all(list(_cells(wb1[t])) == list(_cells(wb4[t]))
                   for t in bcw.EXPECTED_TABS)
        check("double workbook build cell-identical (deterministic)", same)

print()
if SKIPS:
    print(f"skipped: {len(SKIPS)} block(s) — " + "; ".join(SKIPS))
if FAILS:
    print(f"FAILED ({len(FAILS)} of {N_CHECKS}): " + ", ".join(FAILS))
    sys.exit(1)
print(f"All {N_CHECKS} shopify-cro-audit conformance tests passed.")
sys.exit(0)

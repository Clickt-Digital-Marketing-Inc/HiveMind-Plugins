#!/usr/bin/env python3
"""Tests for the PMax listing-group waste-filter core (stdlib only; run directly).

    python3 tests/test_pmax_listing.py

Asserts the documented fixture result for BOTH universes (partitions + products),
no-row-loss, dedupe, the strict-inequality block boundaries, empty-universe edges,
fractional conversions, and md/html bundle parity + lazy openpyxl import.
Exit 0 = all pass, 1 = a failure.
"""
import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent / "scripts"))
sys.path.insert(0, str(HERE.parents[2] / "_shared"))  # the shared render toolkit

import pmax_listing_core as core  # noqa: E402

FIXTURE = HERE / "sample-pmax-findings.json"
EMPTY_RETAIL_FIXTURE = HERE / "sample-pmax-findings-empty-retail.json"
_failures = []


def check(name, cond, detail=""):
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  — {detail}" if detail and not cond else ""))
    if not cond:
        _failures.append(name)


def chart_args():
    """`[]` when the pinned chart renderer is importable, `["--no-charts"]` when it is not.

    This skill's spec DECLARES charts, and `_shared/render/bundle.py` deliberately
    hard-fails (exit 2) when charts are declared while `vl-convert-python` is absent,
    rather than silently shipping a chartless report. That is correct product
    behavior — but the builder test below asserts on md/html/xlsx EMISSION, on the
    md root-cause sentence and on the campaign-benchmark no-row-loss table, never on
    chart SVGs, so a missing optional native wheel must not turn it red (HM-803).
    Opting out only when the renderer is genuinely unavailable keeps the full chart
    path under test on any machine that has the dependency `requirements.txt` pins.

    The probe mirrors the guard it compensates for (`_shared/render/charts.py`'s
    `render_chart_svg`, and `_has_vl_convert()` in the shared toolkit's own tests):
    a REAL import, never `importlib.util.find_spec`. A dist that is locatable but
    not importable (arch-mismatched wheel, half-extracted native extension)
    resolves under `find_spec` while the builder still exits 2 — the exact red this
    helper exists to remove. The selected mode is printed so a green log can be told
    apart from a chart-blind one.
    """
    try:
        import vl_convert  # noqa: F401
        mode = []
    except Exception:      # absent, or present-but-unimportable
        mode = ["--no-charts"]
    print("    chart mode: " + ("charts rendered" if not mode
                                else "charts skipped (vl_convert unimportable)"))
    return mode


def test_fixture_counts():
    print("test_fixture_counts")
    model = core.compute_model(core.load_findings(str(FIXTURE)))
    s = model["summary"]
    check("partitions block1 == 1", s["block1"] == 1, f"got {s['block1']}")
    check("partitions block2 == 1", s["block2"] == 1, f"got {s['block2']}")
    check("partitions no_benchmark == 1", s["no_benchmark"] == 1, f"got {s['no_benchmark']}")
    check("partitions universe == 7", s["universe"] == 7, f"got {s['universe']}")
    check("partitions scored == 6", s["scored"] == 6, f"got {s['scored']}")
    check("partitions flagged == 260.0", abs(s["flagged_spend"] - 260.0) < 1e-6, f"got {s['flagged_spend']}")
    i = s["item"]
    check("products block1 == 1", i["block1"] == 1, f"got {i['block1']}")
    check("products block2 == 1", i["block2"] == 1, f"got {i['block2']}")
    check("products no_benchmark == 1", i["no_benchmark"] == 1, f"got {i['no_benchmark']}")
    check("products universe == 6", i["universe"] == 6, f"got {i['universe']}")
    check("products flagged == 230.0", abs(i["flagged_spend"] - 230.0) < 1e-6, f"got {i['flagged_spend']}")
    f = core.load_findings(str(FIXTURE))
    check("partition rows preserved == input listing_groups",
          len(model["rows"]) == len(f["listing_groups"]), f"{len(model['rows'])} vs {len(f['listing_groups'])}")
    check("product rows preserved == input products",
          len(model["items"]) == len(f["products"]), f"{len(model['items'])} vs {len(f['products'])}")
    # tier concentration + signal (HM-539): the fixture is a clean 0-signal result at
    # the default thresholds — an honest result, not an error (mirrors the 0/0 story).
    check("fixture partitions tier_signals == 0 (clean)", s["tier_signals"] == 0, f"got {s['tier_signals']}")
    check("fixture products tier_signals == 0 (clean)", i["tier_signals"] == 0, f"got {i['tier_signals']}")
    check("fixture provenance source defaults to mcp",
          model["provenance"]["source"] == "mcp", model["provenance"]["source"])


def test_no_benchmark_rows_present():
    print("test_no_benchmark_rows_present")
    model = core.compute_model(core.load_findings(str(FIXTURE)))
    for key, word in (("rows", "partition"), ("items", "product")):
        nb = [r for r in model[key] if r["status"] == "no_benchmark"]
        check(f"{word} no_benchmark row kept with status", len(nb) == 1, f"got {len(nb)}")
        check(f"{word} no_benchmark row has empty block", bool(nb) and nb[0]["block"] == "")


def test_block_boundaries_strict():
    print("test_block_boundaries_strict")
    # campaign: cpa = 1000/50 = 20 ; clicks/conv = 500/50 = 10 ; F = 1.5 -> bars at 30 and 15
    base_bench = [{"campaign_id": 1, "campaign": "X", "clicks": 500, "cost": 1000.0, "conversions": 50.0}]

    def lg(cost, clicks, conv):
        return {"campaign_id": 1, "campaign": "X", "asset_group": "AG", "listing_group_id": "1~1",
                "listing_group": "L", "dimension": "", "impressions": 1000, "clicks": clicks,
                "cost": cost, "conversions": conv, "conversions_value": 0.0}

    # Block 1: lg_cpa exactly at the bar (30.0) must NOT flag (strict >)
    s = core.compute_model({"meta": {}, "benchmarks": base_bench,
                            "listing_groups": [lg(60.0, 5, 2)]})["summary"]  # 60/2 = 30.0
    check("B1 at-bar (cpa==30) not flagged", s["block1"] == 0, f"got {s['block1']}")
    s = core.compute_model({"meta": {}, "benchmarks": base_bench,
                            "listing_groups": [lg(61.0, 5, 2)]})["summary"]  # 30.5 > 30
    check("B1 above-bar (cpa==30.5) flagged", s["block1"] == 1, f"got {s['block1']}")

    # Block 2: clicks exactly at the bar (15) must NOT flag; 16 flags
    s = core.compute_model({"meta": {}, "benchmarks": base_bench,
                            "listing_groups": [lg(50.0, 15, 0)]})["summary"]
    check("B2 at-bar (clicks==15) not flagged", s["block2"] == 0, f"got {s['block2']}")
    s = core.compute_model({"meta": {}, "benchmarks": base_bench,
                            "listing_groups": [lg(50.0, 16, 0)]})["summary"]
    check("B2 above-bar (clicks==16) flagged", s["block2"] == 1, f"got {s['block2']}")

    # conv==0 but campaign has conv: eligible for B2 only; conv>0 cheap: no flag
    s = core.compute_model({"meta": {}, "benchmarks": base_bench,
                            "listing_groups": [lg(10.0, 4, 3)]})["summary"]  # cpa 3.3, cheap
    check("cheap converter not flagged", s["block1"] == 0 and s["block2"] == 0)


def test_no_benchmark_campaign_blocks_both():
    print("test_no_benchmark_campaign_blocks_both")
    # campaign with 0 conversions -> undefined benchmark -> never classify, never drop
    bench = [{"campaign_id": 9, "campaign": "Zero", "clicks": 80, "cost": 500.0, "conversions": 0.0}]
    lgs = [{"campaign_id": 9, "campaign": "Zero", "asset_group": "AG", "listing_group_id": "9~1",
            "listing_group": "L", "dimension": "", "impressions": 1000, "clicks": 50, "cost": 300.0,
            "conversions": 0.0, "conversions_value": 0.0}]
    m = core.compute_model({"meta": {}, "benchmarks": bench, "listing_groups": lgs})
    s = m["summary"]
    check("no-benchmark campaign -> 0 flags", s["block1"] == 0 and s["block2"] == 0)
    check("no-benchmark row kept", s["universe"] == 1 and s["no_benchmark"] == 1)
    check("row carries no_benchmark status", m["rows"][0]["status"] == "no_benchmark")


def test_empty_universes():
    print("test_empty_universes")
    bench = [{"campaign_id": 1, "campaign": "X", "clicks": 100, "cost": 1000.0, "conversions": 50.0}]
    # products omitted entirely
    m = core.compute_model({"meta": {}, "benchmarks": bench,
                            "listing_groups": [{"campaign_id": 1, "campaign": "X", "asset_group": "AG",
                              "listing_group_id": "1~1", "listing_group": "All", "dimension": "",
                              "impressions": 0, "clicks": 0, "cost": 0.0, "conversions": 0.0,
                              "conversions_value": 0.0}]})
    check("absent products -> empty items", m["items"] == [] and m["summary"]["item"]["universe"] == 0)
    check("item sensitivity computed without crash", len(m["item_sensitivity"]) == len(core.FACTOR_LADDER))
    # listing_groups omitted entirely (products-only run)
    m2 = core.compute_model({"meta": {}, "benchmarks": bench,
                             "products": [{"campaign_id": 1, "campaign": "X", "item_id": "A",
                               "title": "A", "impressions": 10, "clicks": 1, "cost": 1.0,
                               "conversions": 0.0, "conversions_value": 0.0}]})
    check("absent partitions -> empty rows", m2["rows"] == [] and m2["summary"]["universe"] == 0)
    check("partition sensitivity computed without crash", len(m2["sensitivity"]) == len(core.FACTOR_LADDER))


def test_empty_retail_universe_present_but_empty():
    """HM-599: `listing_groups: []` and `products: []` (both PRESENT, both empty)
    is a legitimate empty universe — a feedless lead-gen account — not a missing
    source. It must load and compute cleanly, never raise FindingsError."""
    print("test_empty_retail_universe_present_but_empty")
    findings = core.load_findings(str(EMPTY_RETAIL_FIXTURE))
    check("present-but-empty listing_groups/products load without error", True)
    model = core.compute_model(findings)
    s = model["summary"]
    check("partitions universe == 0", s["universe"] == 0, f"got {s['universe']}")
    check("products universe == 0", s["item"]["universe"] == 0, f"got {s['item']['universe']}")
    check("no rows dropped/fabricated (rows == [])", model["rows"] == [])
    check("no items dropped/fabricated (items == [])", model["items"] == [])
    check("campaign benchmarks still present (2 rows, independent of retail feed)",
          len(model["benchmarks"]) == 2, f"got {len(model['benchmarks'])}")

    # Contrast: BOTH keys truly ABSENT (never pulled/assembled) must still be a
    # hard load failure — only presence-with-empty-array is a valid empty universe.
    bad = {"meta": {}, "benchmarks": []}
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "no-source-keys.json"
        p.write_text(json.dumps(bad))
        try:
            core.load_findings(str(p))
            ok = False
        except core.FindingsError:
            ok = True
        check("both keys truly absent is still a hard load failure", ok)


def test_empty_retail_builder_bundle():
    """HM-599 acceptance: the CLI builder must exit 0 on a present-but-empty
    retail universe, emit all declared formats, state the root cause plainly in
    the md, and still render the campaign benchmark table (no-row-loss) in the
    xlsx — none of the 2 benchmark rows dropped."""
    print("test_empty_retail_builder_bundle")
    build_script = HERE.parent / "scripts" / "build_pmax_listing_filter.py"
    with tempfile.TemporaryDirectory() as td:
        result = subprocess.run(
            [sys.executable, str(build_script), "--input", str(EMPTY_RETAIL_FIXTURE),
             "--outdir", td, "--brand", "Lead Gen Co", "--formats", "md,html,xlsx"] + chart_args(),
            capture_output=True, text=True)
        check("builder exits 0 on present-but-empty retail universe",
              result.returncode == 0, f"rc={result.returncode} stderr={result.stderr}")
        out = Path(td)
        md_files = list(out.glob("*.md"))
        html_files = list(out.glob("*_explorer.html"))
        xlsx_files = list(out.glob("*.xlsx"))
        check("md emitted", len(md_files) == 1, f"got {len(md_files)}")
        check("html emitted", len(html_files) == 1, f"got {len(html_files)}")
        check("xlsx emitted", len(xlsx_files) == 1, f"got {len(xlsx_files)}")
        if not (md_files and html_files and xlsx_files):
            return

        md_text = md_files[0].read_text()
        check("md states the root cause plainly",
              "No retail listing groups returned" in md_text
              and "no Merchant Center feed" in md_text and "lead-gen PMax" in md_text)

        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_files[0]), data_only=False)
        ws = wb["Controls"]
        bench_rows = []
        r = 33  # aux "CAMPAIGN BENCHMARKS" start_row (pmax_listing_xlsx_spec.py)
        while ws.cell(row=r, column=1).value not in (None, ""):
            bench_rows.append(ws.cell(row=r, column=1).value)
            r += 1
        check("campaign benchmark table renders both campaigns (no-row-loss)",
              bench_rows == ["Lead Gen - Search", "Lead Gen - PMax"], bench_rows)


def test_dedupe_by_key():
    print("test_dedupe_by_key")
    bench = [{"campaign_id": 1, "campaign": "X", "clicks": 500, "cost": 1000.0, "conversions": 50.0}]
    dup = {"campaign_id": 1, "campaign": "X", "asset_group": "AG", "listing_group_id": "1~1",
           "listing_group": "L", "dimension": "", "impressions": 500, "clicks": 5, "cost": 20.0,
           "conversions": 0.0, "conversions_value": 0.0}
    m = core.compute_model({"meta": {}, "benchmarks": bench, "listing_groups": [dict(dup), dict(dup)]})
    rows = m["rows"]
    check("duplicate partition key merged to one row", len(rows) == 1, f"got {len(rows)}")
    check("merged cost summed (40.0)", bool(rows) and abs(rows[0]["cost"] - 40.0) < 1e-6,
          f"got {rows[0]['cost'] if rows else None}")
    check("merged ctr recomputed (10/1000=0.01)", bool(rows) and abs(rows[0]["ctr"] - 0.01) < 1e-9)


def test_fractional_conversions():
    print("test_fractional_conversions")
    # campaign cpa = 1000/40 = 25 ; F=1.5 -> bar 37.5. lg cost 80 / conv 2.0 = 40 > 37.5 -> Block 1
    bench = [{"campaign_id": 1, "campaign": "X", "clicks": 400, "cost": 1000.0, "conversions": 40.0}]
    lgs = [{"campaign_id": 1, "campaign": "X", "asset_group": "AG", "listing_group_id": "1~1",
            "listing_group": "L", "dimension": "", "impressions": 2000, "clicks": 30, "cost": 80.0,
            "conversions": 2.0, "conversions_value": 0.0}]
    s = core.compute_model({"meta": {}, "benchmarks": bench, "listing_groups": lgs})["summary"]
    check("fractional-benchmark converter flagged Block 1", s["block1"] == 1, f"got {s['block1']}")


def test_sensitivity_and_near_miss_shapes():
    print("test_sensitivity_and_near_miss_shapes")
    model = core.compute_model(core.load_findings(str(FIXTURE)))
    for key in ("sensitivity", "item_sensitivity"):
        sens = model[key]
        check(f"{key} has a row per ladder step", len(sens) == len(core.FACTOR_LADDER))
        check(f"{key} exactly one row flagged current", sum(1 for r in sens if r["is_current"]) == 1)
    for key in ("near_misses_block1", "near_misses_block2",
                "item_near_misses_block1", "item_near_misses_block2"):
        nm = model[key]
        check(f"{key} entries carry qualify_if_factor_le",
              all("qualify_if_factor_le" in r for r in nm))


def test_assemble_findings_from_raw():
    print("test_assemble_findings_from_raw")
    import assemble_findings as A
    rn = "customers/1/assetGroupListingGroupFilters/11~1"
    rn_ee = "customers/1/assetGroupListingGroupFilters/11~2"
    raw_lg = {"result": [
        {"campaign.id": 1, "campaign.name": "C", "asset_group.id": 11, "asset_group.name": "AG",
         "asset_group_product_group_view.asset_group_listing_group_filter": rn,
         "metrics.impressions": 600, "metrics.clicks": 6, "metrics.conversions": 1,
         "metrics.conversions_value": 100.0, "metrics.cost_micros": 30_000_000},
        # same partition split across raw rows (e.g. by a segment) -> must merge
        {"campaign.id": 1, "campaign.name": "C", "asset_group.id": 11, "asset_group.name": "AG",
         "asset_group_product_group_view.asset_group_listing_group_filter": rn,
         "metrics.impressions": 400, "metrics.clicks": 4, "metrics.conversions": 1,
         "metrics.conversions_value": 50.0, "metrics.cost_micros": 20_000_000},
        # the catch-all node: its label row has no case_value
        {"campaign.id": 1, "campaign.name": "C", "asset_group.id": 11, "asset_group.name": "AG",
         "asset_group_product_group_view.asset_group_listing_group_filter": rn_ee,
         "metrics.impressions": 100, "metrics.clicks": 1, "metrics.conversions": 0,
         "metrics.conversions_value": 0, "metrics.cost_micros": 5_000_000},
    ]}
    # the labels pull is structural — NO metrics; it only names each partition
    raw_labels = {"result": [
        {"asset_group_listing_group_filter.resource_name": rn,
         "asset_group_listing_group_filter.type": "UNIT_INCLUDED",
         "asset_group_listing_group_filter.case_value.product_brand.value": "Nike"},
        {"asset_group_listing_group_filter.resource_name": rn_ee,
         "asset_group_listing_group_filter.type": "UNIT_INCLUDED"},
    ]}
    raw_bench = {"result": [
        {"campaign.id": 1, "campaign.name": "C", "metrics.clicks": 500,
         "metrics.conversions": 50, "metrics.cost_micros": 1_000_000_000},
    ]}
    raw_prod = {"result": [
        {"campaign.id": 1, "campaign.name": "C", "segments.product_item_id": "SKU-1",
         "segments.product_title": "Anvil", "metrics.impressions": 300, "metrics.clicks": 3,
         "metrics.conversions": 1, "metrics.conversions_value": 80.0,
         "metrics.cost_micros": 12_000_000},
        # a legacy-Shopping campaign row -> filtered out by the PMax benchmark set
        {"campaign.id": 99, "campaign.name": "Shopping legacy", "segments.product_item_id": "SKU-9",
         "segments.product_title": "Old", "metrics.impressions": 50, "metrics.clicks": 1,
         "metrics.conversions": 0, "metrics.conversions_value": 0,
         "metrics.cost_micros": 1_000_000},
    ]}
    meta = {"client_name": "T", "account_id": "1", "currency": "CAD",
            "window_30d": "w30", "generated": "2026-07-06"}
    with tempfile.TemporaryDirectory() as td:
        p = {}
        for name, doc in (("lg", raw_lg), ("labels", raw_labels),
                          ("bench", raw_bench), ("prod", raw_prod)):
            p[name] = Path(td) / f"{name}.txt"
            p[name].write_text(json.dumps(doc))
        f = A.assemble(str(p["lg"]), str(p["labels"]), str(p["bench"]), str(p["prod"]),
                       dict(meta))
        lgs = f["listing_groups"]
        check("split partition merged (2 partitions)", len(lgs) == 2, f"{len(lgs)}")
        by_id = {r["listing_group_id"]: r for r in lgs}
        m = by_id.get("11~1")
        check("merged sums + micros converted",
              bool(m) and m["impressions"] == 1000 and m["clicks"] == 10
              and abs(m["cost"] - 50.0) < 1e-9 and m["conversions"] == 2
              and m["conversions_value"] == 150.0)
        check("brand label joined from the no-metrics labels pull",
              bool(m) and m["listing_group"] == "Brand: Nike" and m["dimension"] == "product_brand")
        ee = by_id.get("11~2")
        check("catch-all (no case_value) gets empty label",
              ee is not None and ee["listing_group"] == "" and ee["dimension"] == "")
        check("benchmark micros converted", abs(f["benchmarks"][0]["cost"] - 1000.0) < 1e-9)
        prods = f["products"]
        check("non-PMax product row filtered by the PMax campaign set",
              len(prods) == 1 and prods[0]["item_id"] == "SKU-1", f"{len(prods)}")
        check("product micros converted", bool(prods) and abs(prods[0]["cost"] - 12.0) < 1e-9)
        rec = f["meta"]["reconciliation"]
        check("reconciliation embedded with raw stamps",
              rec["listing_groups"]["rows"] == 2 and rec["products"]["rows"] == 1
              and rec["benchmarks"]["rows"] == 1 and len(rec.get("raw_files", [])) == 4)
        # the assembled output flows through the core: the catch-all renders
        model = core.compute_model(f)
        check("core renders the catch-all as 'Everything else'",
              any(r["label"] == "Everything else" for r in model["rows"]))
        # the assembled findings load clean through the core's verification...
        fp = Path(td) / "findings.json"
        fp.write_text(json.dumps(f))
        core.load_findings(str(fp))
        check("assembled findings pass core verification", True)
        # ...and a hand-edit is a hard load failure
        f["products"][0]["conversions"] += 2
        fp.write_text(json.dumps(f))
        try:
            core.load_findings(str(fp)); ok = False
        except core.FindingsError:
            ok = True
        check("hand-edited findings rejected by core", ok)


def test_bundle_md_html_parity_and_lazy():
    print("test_bundle_md_html_parity_and_lazy")
    import pmax_listing_spec as spec_mod
    from render import build_bundle
    from render import charts as C
    model = core.compute_model(core.load_findings(str(FIXTURE)))
    n_part = len(model["rows"])
    n_item = len(model["items"])
    with tempfile.TemporaryDirectory() as td:
        written = build_bundle(model, dict(spec_mod.SPEC), td, formats=("md", "html"))
        md = next(Path(td).glob("*.md")).read_text()
        html = next(Path(td).glob("*_explorer.html")).read_text()
        svgs = sorted(p.name for p in written if p.suffix == ".svg")

    def count_table(md_text, header):
        body = md_text.split("## " + header, 1)[1]
        body = re.split(r"\n## |\n---", body)[0]
        rows = [ln for ln in body.splitlines() if ln.startswith("| ")]
        return max(0, len(rows) - 1)  # minus the header row

    md_part = count_table(md, "All listing-group partitions (every row, with status)")
    md_item = count_table(md, "Products — every item (with status)")
    check("md partition table has every partition", md_part == n_part, f"{md_part} vs {n_part}")
    check("md products table has every product", md_item == n_item, f"{md_item} vs {n_item}")

    embedded = json.loads(re.search(r"^const MODEL = (.+);$", html, re.M).group(1))
    check("html embeds every partition", len(embedded["rows"]) == n_part,
          f"{len(embedded['rows'])} vs {n_part}")
    check("html embeds every product", len(embedded["items"]) == n_item,
          f"{len(embedded['items'])} vs {n_item}")
    # self-containment: the only opaque region allowed is the vendored chart
    # runtime, and only byte-equal to the committed, checksummed vendor files.
    blob = C.vendor_blob()
    check("explorer embeds the verified vendor runtime", blob in html)
    stripped = html.replace(blob, "")
    check("html is self-contained outside the verified vendor blob",
          len(re.findall(r"https?://|<link|src=|cdn", stripped)) == 0)
    # declared charts render as static SVGs and are referenced from the md
    check("both chart svgs written",
          svgs == ["flagged_spend_by_block.svg", "roas_spend_scatter.svg"], svgs)
    check("md has a Charts section", "## Charts" in md)
    check("md references charts relatively", "_charts/flagged_spend_by_block.svg)" in md)
    check("explorer embeds the chart specs", "const CHARTS = " in html)
    check("building md/html did not import openpyxl", "openpyxl" not in sys.modules)
    # HM-539 surfaces: tier concentration / tier signals / recommendations in
    # md, and the new row fields (cost_share, roas, tier_signal) in the html embed.
    check("md has a Prioritized recommendations section", "## Prioritized recommendations" in md)
    check("md has a Tier concentration section", "## Tier concentration (30d spend)" in md)
    check("md has a partition Tier signals section",
          "## Tier signals — partitions (spend concentrated in a weak-ROAS tier)" in md)
    check("md has a product Tier signals section",
          "## Tier signals — products (spend concentrated in a weak-ROAS tier)" in md)
    check("html rows embed cost_share/roas/tier_signal",
          all(k in embedded["rows"][0] for k in ("cost_share", "roas", "tier_signal", "signal_flags")))
    check("html embeds recommendations", "recommendations" in embedded)


def test_tier_concentration_and_signal_strict():
    print("test_tier_concentration_and_signal_strict")
    bench = [{"campaign_id": 1, "campaign": "C", "clicks": 1000, "cost": 5000.0, "conversions": 100.0}]

    def lg(code, cost, conv, value):
        return {"campaign_id": 1, "campaign": "C", "asset_group": "AG", "listing_group_id": code,
                "listing_group": code, "dimension": "", "impressions": 1000, "clicks": 10,
                "cost": cost, "conversions": conv, "conversions_value": value}

    # total cost 1000.0 across 3 partitions (default concentration_share_min=0.30,
    # weak_roas_max=1.00): A is over-concentrated AND weak ROAS -> tier_signal;
    # B sits exactly at the concentration bar (share 0.30, strict > required) and
    # is weak on ROAS alone (one flag only, not both) -> no tier_signal; C fires
    # neither.
    lgs = [
        lg("A", 400.0, 5.0, 200.0),   # share 0.40 (>0.30 fires); roas 0.50 (<1.00 fires)
        lg("B", 300.0, 5.0, 100.0),   # share 0.30 (at bar, NOT fired); roas 0.333 (fires alone)
        lg("C", 300.0, 5.0, 600.0),   # share 0.30 (not fired); roas 2.00 (not fired)
    ]
    m = core.compute_model({"meta": {}, "benchmarks": bench, "listing_groups": lgs})
    rows = {r["code"]: r for r in m["rows"]}
    check("row A cost_share == 0.40", abs(rows["A"]["cost_share"] - 0.4) < 1e-9)
    check("row A over_concentrated fires (0.40 > 0.30)",
          "over_concentrated" in rows["A"]["signal_flags"])
    check("row A weak_roas fires (0.50 < 1.00)", "weak_roas" in rows["A"]["signal_flags"])
    check("row A tier_signal True (both fire)", rows["A"]["tier_signal"] is True)
    check("row B at-bar cost_share (0.30) does NOT fire over_concentrated (strict >)",
          abs(rows["B"]["cost_share"] - 0.3) < 1e-9
          and "over_concentrated" not in rows["B"]["signal_flags"])
    check("row B weak_roas fires alone but tier_signal False (needs both)",
          "weak_roas" in rows["B"]["signal_flags"] and rows["B"]["tier_signal"] is False)
    check("row C fires neither signal",
          rows["C"]["signal_flags"] == [] and rows["C"]["tier_signal"] is False)
    s = m["summary"]
    check("summary tier_signals == 1", s["tier_signals"] == 1, f"got {s['tier_signals']}")
    check("summary signal_spend == 400.0", abs(s["signal_spend"] - 400.0) < 1e-6, f"got {s['signal_spend']}")
    conc = s["concentration"]
    check("concentration total == 1000.0", abs(conc["total"] - 1000.0) < 1e-6, f"got {conc['total']}")
    check("concentration top_share (top-3 of 3) == 1.0", abs(conc["top_share"] - 1.0) < 1e-6)
    check("concentration n == 3", conc["n"] == 3, f"got {conc['n']}")


def test_tier_signal_roas_boundary_strict():
    print("test_tier_signal_roas_boundary_strict")
    bench = [{"campaign_id": 1, "campaign": "C", "clicks": 100, "cost": 1000.0, "conversions": 10.0}]

    def lg(code, cost, value):
        return {"campaign_id": 1, "campaign": "C", "asset_group": "AG", "listing_group_id": code,
                "listing_group": code, "dimension": "", "impressions": 100, "clicks": 5,
                "cost": cost, "conversions": 1.0, "conversions_value": value}

    # both rows hold a 0.50 cost_share each (over_concentrated fires for both);
    # AT_BAR sits exactly on the weak-ROAS bar (roas == 1.00, strict < required).
    lgs = [lg("AT_BAR", 500.0, 500.0), lg("BELOW_BAR", 500.0, 499.0)]
    m = core.compute_model({"meta": {}, "benchmarks": bench, "listing_groups": lgs})
    rows = {r["code"]: r for r in m["rows"]}
    check("roas == 1.00 exactly at bar does NOT fire weak_roas (strict <)",
          abs(rows["AT_BAR"]["roas"] - 1.0) < 1e-9
          and "weak_roas" not in rows["AT_BAR"]["signal_flags"])
    check("roas < 1.00 fires weak_roas",
          rows["BELOW_BAR"]["roas"] < 1.0 and "weak_roas" in rows["BELOW_BAR"]["signal_flags"])
    check("AT_BAR tier_signal False (only over_concentrated fires)", rows["AT_BAR"]["tier_signal"] is False)
    check("BELOW_BAR tier_signal True (both fire)", rows["BELOW_BAR"]["tier_signal"] is True)


def test_tier_signal_zero_cost_no_roas_signal():
    print("test_tier_signal_zero_cost_no_roas_signal")
    bench = [{"campaign_id": 1, "campaign": "C", "clicks": 100, "cost": 1000.0, "conversions": 10.0}]
    lgs = [{"campaign_id": 1, "campaign": "C", "asset_group": "AG", "listing_group_id": "Z",
            "listing_group": "Z", "dimension": "", "impressions": 100, "clicks": 0,
            "cost": 0.0, "conversions": 0.0, "conversions_value": 0.0}]
    m = core.compute_model({"meta": {}, "benchmarks": bench, "listing_groups": lgs})
    r = m["rows"][0]
    check("zero-cost row has roas None (undefined, not 0)", r["roas"] is None)
    check("zero-cost row does not fire weak_roas (missing operand = no signal, never 0)",
          "weak_roas" not in r["signal_flags"])
    check("zero-cost row tier_signal False", r["tier_signal"] is False)


def test_recommendations_grounded_in_model():
    print("test_recommendations_grounded_in_model")
    model = core.compute_model(core.load_findings(str(FIXTURE)))
    recs = model["recommendations"]
    sevs = [r["severity"] for r in recs]
    check("recommendations present for the fixture (Block1+Block2 exist)", len(recs) > 0)
    check("severities are a subset of Critical/High/Medium", set(sevs) <= {"Critical", "High", "Medium"})
    check("a Critical rec cites Block 2",
          any(r["severity"] == "Critical" and "Block 2" in r["text"] for r in recs))
    check("a High rec cites Block 1",
          any(r["severity"] == "High" and "Block 1" in r["text"] for r in recs))
    check("every recommendation names the manual apply path (no Editor apply-CSV)",
          all("manual" in r["artifact"] for r in recs))
    # a genuinely empty universe -> zero recommendations (the "0/0 is clean" honesty
    # extends to the advisor layer: nothing is fabricated when there's nothing to flag)
    empty = core.compute_model({"meta": {}, "benchmarks": []})
    check("empty universe -> zero recommendations (honest, nothing fabricated)",
          empty["recommendations"] == [])


def test_mcp_vs_csv_identical_model():
    print("test_mcp_vs_csv_identical_model")
    import assemble_findings as AF
    import assemble_from_csv as AC

    # --- the same underlying data, assembled via the MCP raw-pull path ---
    raw_bench = {"result": [{"campaign.id": "CampX", "campaign.name": "CampX",
                             "metrics.clicks": 500, "metrics.conversions": 50,
                             "metrics.cost_micros": 1_000_000_000}]}
    raw_lg = {"result": [
        {"campaign.id": "CampX", "campaign.name": "CampX", "asset_group.id": "AGY",
         "asset_group.name": "AGY",
         "asset_group_product_group_view.asset_group_listing_group_filter":
             "customers/1/assetGroupListingGroupFilters/AGY~1",
         "metrics.impressions": 1000, "metrics.clicks": 30, "metrics.conversions": 1,
         "metrics.conversions_value": 30.0, "metrics.cost_micros": 61_000_000},
    ]}
    raw_labels = {"result": [
        {"asset_group_listing_group_filter.resource_name":
             "customers/1/assetGroupListingGroupFilters/AGY~1",
         "asset_group_listing_group_filter.type": "UNIT_INCLUDED",
         "asset_group_listing_group_filter.case_value.product_brand.value": "Nike"},
    ]}
    raw_prod = {"result": [
        {"campaign.id": "CampX", "campaign.name": "CampX", "segments.product_item_id": "SKU-1",
         "segments.product_title": "Anvil", "metrics.impressions": 500, "metrics.clicks": 10,
         "metrics.conversions": 0, "metrics.conversions_value": 0, "metrics.cost_micros": 40_000_000},
    ]}
    meta = {"client_name": "T", "account_id": "1", "currency": "CAD",
            "window_30d": "w30", "generated": "2026-07-12"}

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        paths = {}
        for name, doc in (("lg", raw_lg), ("labels", raw_labels), ("bench", raw_bench),
                          ("prod", raw_prod)):
            paths[name] = td / f"{name}.txt"
            paths[name].write_text(json.dumps(doc))
        findings_mcp = AF.assemble(str(paths["lg"]), str(paths["labels"]), str(paths["bench"]),
                                   str(paths["prod"]), dict(meta))

        # --- the SAME data, as Google Ads UI CSV exports ---
        bench_csv = td / "bench.csv"
        bench_csv.write_text("Campaign,Clicks,Cost,Conversions\nCampX,500,1000.00,50\n")
        lg_csv = td / "lg.csv"
        lg_csv.write_text(
            "Campaign,Asset group,Listing group,Impr.,Clicks,Cost,Conversions,Conv. value\n"
            'CampX,AGY,"Brand: Nike",1000,30,61.00,1,30.00\n')
        prod_csv = td / "prod.csv"
        prod_csv.write_text(
            "Campaign,Item ID,Item title,Impr.,Clicks,Cost,Conversions,Conv. value\n"
            "CampX,SKU-1,Anvil,500,10,40.00,0,0\n")
        findings_csv = AC.assemble(str(bench_csv), str(lg_csv), str(prod_csv), dict(meta))

    check("MCP path stamps meta.source=mcp", findings_mcp["meta"]["source"] == "mcp")
    check("CSV path stamps meta.source=user_csv", findings_csv["meta"]["source"] == "user_csv")

    model_mcp = core.compute_model(findings_mcp)
    model_csv = core.compute_model(findings_csv)

    check("summary identical MCP vs CSV (modulo provenance/source)",
          model_mcp["summary"] == model_csv["summary"], (model_mcp["summary"], model_csv["summary"]))
    check("products identical MCP vs CSV (no API-only join ids on this universe)",
          model_mcp["items"] == model_csv["items"])

    # Partitions: identical except the two fields only the API's structural
    # case_value pull can supply — `code` (the real filter-resource join id,
    # sourced from listing_group_id) and `dimension` (derived from
    # case_value.*). The flat UI export has neither, by design (documented in
    # assemble_from_csv.py).
    API_ONLY = {"code", "dimension"}
    r_mcp, r_csv = model_mcp["rows"][0], model_csv["rows"][0]
    check("partition row identical modulo API-only join ids",
          {k: v for k, v in r_mcp.items() if k not in API_ONLY}
          == {k: v for k, v in r_csv.items() if k not in API_ONLY},
          (r_mcp, r_csv))
    check("MCP partition carries the real listing-group filter id", r_mcp["code"] == "AGY~1")
    check("CSV partition has no filter id (flat export has no filter-resource id)",
          r_csv["code"] == "")
    check("MCP partition carries the case_value dimension", r_mcp["dimension"] == "product_brand")
    check("CSV partition has no dimension (flat export has no case_value)", r_csv["dimension"] == "")
    check("recommendations identical (grounded in the identical model, not raw data)",
          model_mcp["recommendations"] == model_csv["recommendations"])


def main():
    for t in (test_fixture_counts, test_no_benchmark_rows_present, test_block_boundaries_strict,
              test_no_benchmark_campaign_blocks_both, test_empty_universes,
              test_empty_retail_universe_present_but_empty,
              test_dedupe_by_key,
              test_fractional_conversions, test_sensitivity_and_near_miss_shapes,
              test_assemble_findings_from_raw, test_bundle_md_html_parity_and_lazy,
              test_tier_concentration_and_signal_strict, test_tier_signal_roas_boundary_strict,
              test_tier_signal_zero_cost_no_roas_signal, test_recommendations_grounded_in_model,
              test_mcp_vs_csv_identical_model,
              # runs last: reads an xlsx back with openpyxl in THIS process, which
              # would otherwise poison test_bundle_md_html_parity_and_lazy's
              # "did not import openpyxl" check for md/html-only builds
              test_empty_retail_builder_bundle):
        t()
    print()
    if _failures:
        print(f"FAILED ({len(_failures)}): {', '.join(_failures)}")
        return 1
    print("ALL TESTS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())

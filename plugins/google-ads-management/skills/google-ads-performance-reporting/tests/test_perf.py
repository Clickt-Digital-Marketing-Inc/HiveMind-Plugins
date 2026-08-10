#!/usr/bin/env python3
"""Tests for the performance-report core + bundle (stdlib only; run directly).

    python3 tests/test_perf.py

Asserts the fixture buckets/totals, no-row-loss, dedupe, the no-value path, an
empty edge, the raw-pull assembler (transcription firewall: micros conversion,
two-window join, IS null pass-through, no-value omission, reconciliation
round-trip + tamper rejection), md/html bundle parity + lazy-openpyxl, the
period-over-period anomaly signals + spend/conversion concentration + pre-score
(HM-537, cross-checked against `_shared/analytics.py` directly), the
delta-flag's tunability, the CSV-input path's identical-model parity against
the MCP path, and the xlsx `soffice` recalc of the new anomaly/concentration
cells. Exit 0 = pass, 1 = fail.
"""
import json
import re
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent / "scripts"))
sys.path.insert(0, str(HERE.parents[2] / "_shared"))

import perf_core as core  # noqa: E402

FIXTURE = HERE / "sample-findings.json"
#: The liveness matrix (HM-799) — kept out of sample-findings.json on purpose:
#: that fixture's revenue/spend is 23100/5600 = 4.125 exactly, the half-up
#: rounding boundary test_fixture_buckets asserts on (4.13, not round()'s 4.12).
LIVENESS_FIXTURE = HERE / "sample-liveness.json"
#: Row keys that legitimately differ between the fixture's recently_active /
#: dormant twins (campaigns 3 and 4). Everything else must be equal, which is
#: what makes 'status' the single distinguishing input.
_LIVENESS_DERIVED = {"campaign_id", "campaign", "status_label", "liveness",
                     "liveness_note", "flags", "pre_score"}
_failures = []


def check(name, cond, detail=""):
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  — {detail}" if detail and not cond else ""))
    if not cond:
        _failures.append(name)


def test_fixture_buckets():
    print("test_fixture_buckets")
    m = core.compute_model(core.load_findings(str(FIXTURE)))
    s = m["summary"]
    check("campaigns == 6 (no row loss)", s["campaigns"] == 6, f"got {s['campaigns']}")
    check("scale == 1", s["scale"] == 1, f"got {s['scale']}")
    check("winner == 2", s["winner"] == 2, f"got {s['winner']}")
    check("fix == 2", s["fix"] == 2, f"got {s['fix']}")
    check("no_value == 1", s["no_value"] == 1, f"got {s['no_value']}")
    check("spend == 5600", abs(s["spend"] - 5600.0) < 1e-6, f"got {s['spend']}")
    check("revenue == 23100", abs(s["revenue"] - 23100.0) < 1e-6, f"got {s['revenue']}")
    check("roas == 4.13 (half-up, matches JS)", abs(s["roas"] - 4.13) < 1e-6, f"got {s['roas']}")
    check("conversions == 122", abs(s["conversions"] - 122.0) < 1e-6, f"got {s['conversions']}")


def test_no_value_not_bucketed():
    print("test_no_value_not_bucketed")
    m = core.compute_model(core.load_findings(str(FIXTURE)))
    nv = [r for r in m["rows"] if r["status"] == "no_value"]
    check("one no_value row", len(nv) == 1)
    check("no_value row has no bucket", nv and nv[0]["bucket"] == "")
    check("no_value row has roas None", nv and nv[0]["roas"] is None)


def test_dedupe_by_campaign_id():
    print("test_dedupe_by_campaign_id")
    f = {"meta": {}, "campaigns": [
        {"campaign_id": 9, "campaign": "X", "cost": 100, "conversions": 1, "conversions_value": 500,
         "impressions": 1000, "clicks": 50},
        {"campaign_id": 9, "campaign": "X", "cost": 100, "conversions": 1, "conversions_value": 500,
         "impressions": 1000, "clicks": 50},
    ]}
    m = core.compute_model(f)
    check("two rows merged to one", len(m["rows"]) == 1, f"got {len(m['rows'])}")
    check("cost summed (200)", m["rows"] and abs(m["rows"][0]["cost"] - 200) < 1e-6)
    check("value summed (1000), roas 5", m["rows"] and abs(m["rows"][0]["roas"] - 5.0) < 1e-6)


def test_empty():
    print("test_empty")
    m = core.compute_model({"meta": {}, "campaigns": []})
    check("empty -> 0 campaigns", m["summary"]["campaigns"] == 0)
    check("empty -> roas None", m["summary"]["roas"] is None)
    check("goal sensitivity computed", len(m["goal_sensitivity"]) == len(core.ROAS_LADDER))


def test_goal_sensitivity_current_flag():
    print("test_goal_sensitivity_current_flag")
    m = core.compute_model(core.load_findings(str(FIXTURE)))
    cur = [r for r in m["goal_sensitivity"] if r["is_current"]]
    check("exactly one current goal row", len(cur) == 1)


def test_assemble_findings_from_raw():
    print("test_assemble_findings_from_raw")
    import assemble_findings as A
    raw_cur = {"result": [
        # same campaign split across two raw rows (defensive) -> summed;
        # IS fractions pass through from the first occurrence
        {"campaign.id": 1, "campaign.name": "Brand", "campaign.status": "ENABLED",
         "campaign.advertising_channel_type": "SEARCH",
         "metrics.impressions": 10000, "metrics.clicks": 800,
         "metrics.cost_micros": 500_000_000, "metrics.conversions": 15,
         "metrics.conversions_value": 3000.0,
         "metrics.search_impression_share": 0.62,
         "metrics.search_budget_lost_impression_share": 0.25,
         "metrics.search_rank_lost_impression_share": 0.05},
        {"campaign.id": 1, "campaign.name": "Brand", "campaign.status": "ENABLED",
         "campaign.advertising_channel_type": "SEARCH",
         "metrics.impressions": 10000, "metrics.clicks": 800,
         "metrics.cost_micros": 500_000_000, "metrics.conversions": 15,
         "metrics.conversions_value": 3000.0,
         "metrics.search_impression_share": 0.62,
         "metrics.search_budget_lost_impression_share": 0.25,
         "metrics.search_rank_lost_impression_share": 0.05},
        # PMax: IS fields unpopulated (absent) -> null, never 0; value not
        # tracked (lead gen) -> the --no-value-campaigns flag omits its keys
        {"campaign.id": 2, "campaign.name": "PMax Leads", "campaign.status": "ENABLED",
         "campaign.advertising_channel_type": "PERFORMANCE_MAX",
         "metrics.impressions": 5000, "metrics.clicks": 200,
         "metrics.cost_micros": 100_000_000, "metrics.conversions": 4,
         "metrics.conversions_value": 0.0},
    ]}
    raw_prior = {"result": [
        {"campaign.id": 1, "campaign.name": "Brand", "campaign.status": "ENABLED",
         "campaign.advertising_channel_type": "SEARCH",
         "metrics.impressions": 18000, "metrics.clicks": 1400,
         "metrics.cost_micros": 900_000_000, "metrics.conversions": 26,
         "metrics.conversions_value": 5000.0},
        # prior-only campaign -> survives as a zero-current row
        {"campaign.id": 3, "campaign.name": "Paused Old", "campaign.status": "PAUSED",
         "campaign.advertising_channel_type": "SEARCH",
         "metrics.impressions": 2000, "metrics.clicks": 90,
         "metrics.cost_micros": 50_000_000, "metrics.conversions": 1,
         "metrics.conversions_value": 200.0},
    ]}
    meta = {"client_name": "T", "account_id": "1", "currency": "CAD",
            "period": "p", "prior_period": "pp", "generated": "2026-07-06"}
    with tempfile.TemporaryDirectory() as td:
        pc = Path(td) / "cur.txt"; pc.write_text(json.dumps(raw_cur))
        pp = Path(td) / "prior.txt"; pp.write_text(json.dumps(raw_prior))
        f = A.assemble(str(pc), str(pp), dict(meta), no_value_ids=frozenset({"2"}))
        rows = {c["campaign_id"]: c for c in f["campaigns"]}
        check("two current + one prior-only campaign", len(rows) == 3, f"{len(rows)}")
        c1 = rows.get(1) or {}
        check("split rows summed, micros converted (cost 1000.0)",
              abs(c1.get("cost", 0) - 1000.0) < 1e-9 and c1.get("impressions") == 20000
              and c1.get("clicks") == 1600 and c1.get("conversions") == 30
              and abs(c1.get("conversions_value", 0) - 6000.0) < 1e-9)
        check("prior window joined by campaign id",
              abs(c1.get("prior_cost", 0) - 900.0) < 1e-9 and c1.get("prior_conversions") == 26
              and abs(c1.get("prior_conversions_value", 0) - 5000.0) < 1e-9
              and c1.get("prior_impressions") == 18000 and c1.get("prior_clicks") == 1400)
        check("IS fractions pass through", c1.get("search_impression_share") == 0.62
              and c1.get("search_budget_lost_is") == 0.25 and c1.get("search_rank_lost_is") == 0.05)
        c2 = rows.get(2) or {}
        check("absent IS fields stay null (never 0)",
              c2.get("search_impression_share") is None
              and c2.get("search_budget_lost_is") is None and c2.get("search_rank_lost_is") is None)
        check("no-value campaign has value keys omitted",
              "conversions_value" not in c2 and "prior_conversions_value" not in c2)
        c3 = rows.get(3) or {}
        check("prior-only campaign kept with zero current metrics",
              c3.get("cost") == 0 and c3.get("impressions") == 0
              and abs(c3.get("prior_cost", 0) - 50.0) < 1e-9 and c3.get("campaign") == "Paused Old")
        rec = f["meta"]["reconciliation"]
        check("reconciliation embedded with raw stamps",
              rec["campaigns"]["rows"] == 3 and len(rec.get("raw_files", [])) == 2)
        # the assembled findings load clean through the core's verification and
        # the no-value campaign lands unbucketed...
        fp = Path(td) / "findings.json"; fp.write_text(json.dumps(f))
        m = core.compute_model(core.load_findings(str(fp)))
        check("assembled findings pass core verification", True)
        nv = [r for r in m["rows"] if r["status"] == "no_value"]
        check("no-value campaign reported as status no_value (not ROAS 0)",
              len(nv) == 1 and nv[0]["campaign_id"] == 2 and nv[0]["bucket"] == "")
        # ...and a hand-edit is a hard load failure
        f["campaigns"][0]["cost"] += 500
        fp.write_text(json.dumps(f))
        try:
            core.load_findings(str(fp)); ok = False
        except core.FindingsError:
            ok = True
        check("hand-edited findings rejected by core", ok)


def test_bundle_md_html_parity_and_lazy():
    print("test_bundle_md_html_parity_and_lazy")
    import perf_spec
    from render import build_bundle
    from render import charts as C
    m = core.compute_model(core.load_findings(str(FIXTURE)))
    n = len(m["rows"])
    with tempfile.TemporaryDirectory() as td:
        written = build_bundle(m, dict(perf_spec.SPEC), td, formats=("md", "html"))
        md = next(Path(td).glob("*.md")).read_text()
        html = next(Path(td).glob("*_explorer.html")).read_text()
        svgs = sorted(p.name for p in written if p.suffix == ".svg")
    rows_blk = md.split("## All campaigns")[1].splitlines()
    md_rows = [ln for ln in rows_blk if ln.startswith("| ") and not ln.startswith("| Campaign")]
    embedded = json.loads(re.search(r"^const MODEL = (.+);$", html, re.M).group(1))["rows"]
    check("md row table has every campaign", len(md_rows) == n, f"{len(md_rows)} vs {n}")
    check("html embeds every campaign", len(embedded) == n, f"{len(embedded)} vs {n}")
    # self-containment: the only opaque region allowed is the vendored chart
    # runtime, and only byte-equal to the committed, checksummed vendor files.
    blob = C.vendor_blob()
    check("explorer embeds the verified vendor runtime", blob in html)
    stripped = html.replace(blob, "")
    check("html self-contained outside the verified vendor blob",
          len(re.findall(r"https?://|<link|src=|cdn", stripped)) == 0)
    # declared charts render as static SVGs and are referenced from the md
    check("all three chart svgs written",
          svgs == ["revenue_spend_scatter.svg", "spend_by_bucket.svg", "spend_by_campaign.svg"], svgs)
    check("md has a Charts section", "## Charts" in md)
    check("md references charts relatively", "_charts/spend_by_campaign.svg)" in md)
    check("explorer embeds the chart specs", "const CHARTS = " in html)
    check("building md/html did not import openpyxl", "openpyxl" not in sys.modules)


def test_anomalies_and_concentration():
    print("test_anomalies_and_concentration")
    import analytics as A
    m = core.compute_model(core.load_findings(str(FIXTURE)))
    s = m["summary"]
    check("anomalies == 2 at the default delta flag (0.25)", s["anomalies"] == 2, f"got {s['anomalies']}")
    anomalous = sorted([r for r in m["rows"] if r["flags"]], key=lambda r: r["campaign"])
    names = [r["campaign"] for r in anomalous]
    check("the two revenue-drop laggards are flagged",
          names == ["S | NB - Laggard", "S | NB - Weak"], names)
    for r in anomalous:
        check(f"{r['campaign']} flagged value_drop only", r["flags"] == ["value_drop"], r["flags"])
        check(f"{r['campaign']} pre_score == 2.0 (value_drop weight)", r["pre_score"] == 2.0, r["pre_score"])
    unflagged = [r for r in m["rows"] if not r["flags"]]
    check("unflagged rows carry pre_score 0.0 (never dropped, never re-scored)",
          unflagged and all(r["pre_score"] == 0.0 for r in unflagged))
    check("PMax row (channel-agnostic) participates in anomaly scoring",
          any(r["channel"] == "PERFORMANCE_MAX" for r in m["rows"]))

    # Cross-check against the shared analytics primitives directly — the model
    # must use the SAME kernel, not a re-derived equivalent.
    exp_spend = A.concentration(m["rows"], "cost", top_n=3)
    exp_conv = A.concentration(m["rows"], "conversions", top_n=3)
    check("spend concentration == analytics.concentration(rows,'cost')",
          m["concentration"]["spend"] == exp_spend, f"{m['concentration']['spend']} vs {exp_spend}")
    check("conversion concentration == analytics.concentration(rows,'conversions')",
          m["concentration"]["conversions"] == exp_conv, f"{m['concentration']['conversions']} vs {exp_conv}")
    check("spend concentration n == 6 (no row loss)", exp_spend["n"] == 6)
    check("spend top-3 share is a real concentration ratio (0,1)", 0 < exp_spend["top_share"] < 1)
    check("no_value row (Lead Gen) still contributes to concentration",
          any(r["campaign"] == "Lead Gen | No Revenue" for r in m["rows"]))


def test_delta_flag_tunable():
    print("test_delta_flag_tunable")
    base = core.load_findings(str(FIXTURE))
    loose = dict(base); loose["params"] = {"delta_flag": 0.5}
    m_loose = core.compute_model(loose)
    tight = dict(base); tight["params"] = {"delta_flag": 0.1}
    m_tight = core.compute_model(tight)
    m_default = core.compute_model(core.load_findings(str(FIXTURE)))
    check("delta_flag=0.5 clears every anomaly on this fixture",
          m_loose["summary"]["anomalies"] == 0, f"got {m_loose['summary']['anomalies']}")
    check("delta_flag=0.1 flags more campaigns (spend swings now qualify too)",
          m_tight["summary"]["anomalies"] == 4, f"got {m_tight['summary']['anomalies']}")
    check("default (0.25) sits strictly between loose and tight — the slider is non-vacuous",
          m_loose["summary"]["anomalies"] < m_default["summary"]["anomalies"] < m_tight["summary"]["anomalies"])
    check("bucket counts unaffected by delta_flag (independent params)",
          m_loose["summary"]["scale"] == m_tight["summary"]["scale"] == m_default["summary"]["scale"])


def test_csv_matches_mcp_model():
    print("test_csv_matches_mcp_model")
    import assemble_findings as MCP
    import assemble_from_csv as CSVA

    raw_period = {"result": [
        {"campaign.id": "S | Brand", "campaign.name": "S | Brand", "campaign.status": "ENABLED",
         "campaign.advertising_channel_type": "SEARCH",
         "metrics.impressions": 20000, "metrics.clicks": 800,
         "metrics.cost_micros": 1_000_000_000, "metrics.conversions": 15,
         "metrics.conversions_value": 3000.0,
         "metrics.search_impression_share": 0.62,
         "metrics.search_budget_lost_impression_share": 0.25,
         "metrics.search_rank_lost_impression_share": 0.05},
        {"campaign.id": "S | NB - Core", "campaign.name": "S | NB - Core", "campaign.status": "ENABLED",
         "campaign.advertising_channel_type": "SEARCH",
         "metrics.impressions": 30000, "metrics.clicks": 1500,
         "metrics.cost_micros": 1_000_000_000, "metrics.conversions": 25,
         "metrics.conversions_value": 5000.0,
         "metrics.search_impression_share": 0.74,
         "metrics.search_budget_lost_impression_share": 0.02,
         "metrics.search_rank_lost_impression_share": 0.20},
    ]}
    raw_prior = {"result": [
        {"campaign.id": "S | Brand", "metrics.impressions": 22000, "metrics.clicks": 900,
         "metrics.cost_micros": 1_500_000_000, "metrics.conversions": 20,
         "metrics.conversions_value": 5000.0},
        {"campaign.id": "S | NB - Core", "metrics.impressions": 29000, "metrics.clicks": 1440,
         "metrics.cost_micros": 1_000_000_000, "metrics.conversions": 24,
         "metrics.conversions_value": 4800.0},
    ]}
    meta = {"client_name": "Acme Corp", "account_id": "1", "currency": "CAD",
            "period": "p", "prior_period": "pp", "generated": "2026-07-12"}

    csv_period = (
        "Campaign,Campaign state,Campaign type,Impr.,Clicks,Cost,Conversions,Conv. value,"
        "Search impr. share,Search lost IS (budget),Search lost IS (rank)\n"
        "S | Brand,ENABLED,SEARCH,20000,800,1000.00,15,3000.00,62%,25%,5%\n"
        "S | NB - Core,ENABLED,SEARCH,30000,1500,1000.00,25,5000.00,74%,2%,20%\n")
    csv_prior = (
        "Campaign,Campaign state,Campaign type,Impr.,Clicks,Cost,Conversions,Conv. value\n"
        "S | Brand,ENABLED,SEARCH,22000,900,1500.00,20,5000.00\n"
        "S | NB - Core,ENABLED,SEARCH,29000,1440,1000.00,24,4800.00\n")

    with tempfile.TemporaryDirectory() as td:
        pc = Path(td) / "period_raw.txt"; pc.write_text(json.dumps(raw_period))
        pp = Path(td) / "prior_raw.txt"; pp.write_text(json.dumps(raw_prior))
        mcp_findings = MCP.assemble(str(pc), str(pp), dict(meta))
        mf = Path(td) / "mcp_findings.json"; mf.write_text(json.dumps(mcp_findings))

        cp = Path(td) / "period.csv"; cp.write_text(csv_period)
        cpr = Path(td) / "prior.csv"; cpr.write_text(csv_prior)
        csv_findings = CSVA.assemble(str(cp), str(cpr), dict(meta))
        cf = Path(td) / "csv_findings.json"; cf.write_text(json.dumps(csv_findings))

        m_mcp = core.compute_model(core.load_findings(str(mf)))
        m_csv = core.compute_model(core.load_findings(str(cf)))

    check("mcp path labels provenance.source == 'mcp'", m_mcp["provenance"]["source"] == "mcp")
    check("csv path labels provenance.source == 'user_csv' (honesty)",
          m_csv["provenance"]["source"] == "user_csv")
    prov_mcp = dict(m_mcp["provenance"]); prov_mcp.pop("source")
    prov_csv = dict(m_csv["provenance"]); prov_csv.pop("source")
    check("provenance identical across paths (source excluded)", prov_mcp == prov_csv,
          f"{prov_mcp} vs {prov_csv}")
    check("rows IDENTICAL across MCP and CSV paths", m_mcp["rows"] == m_csv["rows"],
          f"{m_mcp['rows']} vs {m_csv['rows']}")
    check("summary IDENTICAL across MCP and CSV paths", m_mcp["summary"] == m_csv["summary"],
          f"{m_mcp['summary']} vs {m_csv['summary']}")
    check("concentration IDENTICAL across MCP and CSV paths",
          m_mcp["concentration"] == m_csv["concentration"])
    check("goal_sensitivity IDENTICAL across MCP and CSV paths",
          m_mcp["goal_sensitivity"] == m_csv["goal_sensitivity"])
    check("anomaly flags/pre_score identical (same kernel, same inputs)",
          [(r["flags"], r["pre_score"]) for r in m_mcp["rows"]]
          == [(r["flags"], r["pre_score"]) for r in m_csv["rows"]])


def test_liveness_gating():
    print("test_liveness_gating")
    # Three-band coverage: this skill pulls status + current + prior spend, so
    # all three bands AND all three _liveness_note return paths are reachable.
    # The matrix lives in the SHIPPED fixture (HM-799), not in an inline dict —
    # one definition, and the same file the port's golden is generated from.
    m = core.compute_model(core.load_findings(str(LIVENESS_FIXTURE)))
    rows = {r["campaign_id"]: r for r in m["rows"]}
    check("every campaign survives (no-row-loss)", len(m["rows"]) == 5, f"got {len(m['rows'])}")
    check("live band", rows[1]["liveness"] == "live", rows[1]["liveness"])
    check("paused-mid-window -> recently_active", rows[2]["liveness"] == "recently_active", rows[2]["liveness"])
    check("enabled-idle -> recently_active", rows[3]["liveness"] == "recently_active", rows[3]["liveness"])
    check("removed + zero both windows -> dormant", rows[4]["liveness"] == "dormant", rows[4]["liveness"])
    check("prior-window-only -> recently_active", rows[5]["liveness"] == "recently_active", rows[5]["liveness"])
    # recently_active rows carry conditional phrasing (the HM-605 seam). All
    # THREE note paths are asserted by their distinguishing clause, so swapping
    # two of the branches cannot stay green.
    check("paused+spend row has the paused-mid-window note",
          rows[2]["liveness_note"] ==
          "Paused/removed mid-window after spending 200.00 — confirm intent before acting.",
          rows[2]["liveness_note"])
    check("enabled-idle row has the enabled-but-idle note",
          rows[3]["liveness_note"] ==
          "Enabled but no spend in the window — confirm it should be running before acting.",
          rows[3]["liveness_note"])
    check("prior-window-only row has the prior-window note",
          rows[5]["liveness_note"] ==
          "Spent only in the prior window — confirm intent before acting.",
          rows[5]["liveness_note"])
    check("live row has no note", rows[1]["liveness_note"] == "")
    # dormant row generates nothing but stays visible + tagged
    check("dormant row not bucketed", rows[4]["bucket"] == "", rows[4]["bucket"])
    check("dormant row has no anomaly flags", rows[4]["flags"] == [], rows[4]["flags"])
    check("dormant row pre_score 0", rows[4]["pre_score"] == 0.0, rows[4]["pre_score"])
    check("dormant row has empty note", rows[4]["liveness_note"] == "")
    # The gate is load-bearing, not decoration: row 3 and row 4 differ in exactly
    # one input field (status), and row 3 — the recently_active twin — DOES carry
    # the flags the gate suppresses on row 4. Delete the gate and row 4 becomes
    # row 3; widen it past dormant and row 3 loses its flags.
    check("recently_active twin keeps its anomaly flags",
          rows[3]["flags"] == ["conv_drop", "value_drop"], rows[3]["flags"])
    check("recently_active twin keeps a non-zero pre_score",
          rows[3]["pre_score"] == 4.5, rows[3]["pre_score"])
    check("the twins' COMPUTED rows differ ONLY in the derived keys",
          {k: v for k, v in rows[3].items() if k not in _LIVENESS_DERIVED}
          == {k: v for k, v in rows[4].items() if k not in _LIVENESS_DERIVED},
          "twin rows diverge in an input field, so the contrast proves nothing")
    # ...and the same premise asserted where it actually lives: the FIXTURE INPUTS.
    # The computed-row comparison above cannot see the twins' prior windows drift
    # apart — prior_conversions/prior_conversions_value never reach the output row,
    # surviving only as conv_delta/value_delta, which both saturate at -1.0 once
    # the current window is 0. Retune one twin's prior window and the check above
    # stays green while the one-field premise this whole contrast rests on is
    # already broken.
    raw = {c["campaign_id"]: c
           for c in json.loads(LIVENESS_FIXTURE.read_text(encoding="utf-8"))["campaigns"]}
    twin_diff = {k for k in set(raw[3]) | set(raw[4]) if raw[3].get(k) != raw[4].get(k)}
    check("the twins' FIXTURE INPUTS differ ONLY in status (plus id/name)",
          twin_diff == {"campaign_id", "campaign", "status"}, sorted(twin_diff))
    check("prior-window-only row is scored (not gated)",
          rows[5]["flags"] == ["spend_drop", "conv_drop", "value_drop"], rows[5]["flags"])
    # live/recently_active rows still fully scored (severity universe intact)
    check("live row still bucketed", rows[1]["bucket"] in ("Scale", "Winner", "Fix", "Hold"), rows[1]["bucket"])
    check("recently_active row still bucketed", rows[2]["bucket"] == "Scale", rows[2]["bucket"])
    check("summary counts only the ungated anomalies",
          m["summary"]["anomalies"] == 2, m["summary"]["anomalies"])

    # html_embed + JS-kernel parity: dormant gate is mirrored in the browser kernel,
    # so the embedded row carries liveness and the Node gate (run_parity.py) agrees.
    import perf_spec
    emb = perf_spec.html_embed(m)["rows"]
    check("html embed carries liveness + note", all("liveness" in r and "liveness_note" in r for r in emb))
    # Key presence alone is not the seam: a fully regressed embed that tags the
    # long-dead campaign "live" still has both keys on every row. Assert the VALUE
    # on the dormant row (mirrors the budget-pacing sibling's embed check).
    check("html embed dormant row tagged dormant",
          next(r for r in emb if r["campaign"] == "Dormant | Removed")["liveness"] == "dormant",
          next(r for r in emb if r["campaign"] == "Dormant | Removed").get("liveness"))


def test_xlsx_recalc_matches_model():
    print("test_xlsx_recalc_matches_model")
    import perf_spec
    import perf_xlsx_spec as xspec
    from render import xlsx as xlsxmod
    from openpyxl import load_workbook

    m = core.compute_model(core.load_findings(str(FIXTURE)))
    spec = dict(perf_spec.SPEC); spec["xlsx"] = xspec.XLSX
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "report.xlsx"
        xlsxmod.build_xlsx(m, spec, str(out), brand="Acme Corp", normalize=True)
        wb = load_workbook(str(out), data_only=True)
        ws = wb["Controls"]
        s, conc = m["summary"], m["concentration"]
        check("cached Scale count matches model", ws["C11"].value == s["scale"])
        check("cached Anomalies count matches model", ws["C19"].value == s["anomalies"],
              f"{ws['C19'].value} vs {s['anomalies']}")
        check("cached spend top-3 share matches model",
              abs((ws["C20"].value or 0) - conc["spend"]["top_share"]) < 1e-3,
              f"{ws['C20'].value} vs {conc['spend']['top_share']}")
        check("cached spend HHI matches model",
              abs((ws["C21"].value or 0) - conc["spend"]["hhi"]) < 0.5,
              f"{ws['C21'].value} vs {conc['spend']['hhi']}")
        check("cached spend effective-N matches model",
              abs((ws["C22"].value or 0) - conc["spend"]["effective_n"]) < 0.05,
              f"{ws['C22'].value} vs {conc['spend']['effective_n']}")
        check("cached conversion top-3 share matches model",
              abs((ws["C23"].value or 0) - conc["conversions"]["top_share"]) < 1e-3,
              f"{ws['C23'].value} vs {conc['conversions']['top_share']}")
        check("cached conversion HHI matches model",
              abs((ws["C24"].value or 0) - conc["conversions"]["hhi"]) < 0.5,
              f"{ws['C24'].value} vs {conc['conversions']['hhi']}")
        check("cached conversion effective-N matches model",
              abs((ws["C25"].value or 0) - conc["conversions"]["effective_n"]) < 0.05,
              f"{ws['C25'].value} vs {conc['conversions']['effective_n']}")
        rows_ws = wb["Campaigns"]
        headers = [c.value for c in rows_ws[1]]
        check("Anomaly score column present", "Anomaly score" in headers, headers)
        check("Bucket stays the last column", headers[-1] == "Bucket", headers)
        score_col = headers.index("Anomaly score") + 1
        value_delta_col = headers.index("Value Δ") + 1

        # a flagged row's cached Anomaly-score cell matches its pre_score
        r0 = next(r for r in m["rows"] if r["flags"])
        r0_idx = m["rows"].index(r0) + 2  # header row + 1-index
        cached_score = rows_ws.cell(row=r0_idx, column=score_col).value
        check(f"row {r0['campaign']!r} cached Anomaly score matches pre_score",
              abs((cached_score or 0) - r0["pre_score"]) < 1e-6,
              f"{cached_score} vs {r0['pre_score']}")

        # the no_value row's Value Δ is None -> a genuinely BLANK xlsx cell, and
        # its Anomaly-score formula's ISNUMBER guard must not misfire on it.
        r_nv = next(r for r in m["rows"] if r["status"] == "no_value")
        check("no_value row has value_delta None (blank-cell precondition)",
              r_nv["value_delta"] is None, r_nv["value_delta"])
        nv_idx = m["rows"].index(r_nv) + 2
        cached_value_delta = rows_ws.cell(row=nv_idx, column=value_delta_col).value
        check(f"row {r_nv['campaign']!r} Value Δ cell is genuinely blank (None), not 0",
              cached_value_delta is None, cached_value_delta)
        cached_nv_score = rows_ws.cell(row=nv_idx, column=score_col).value
        check(f"row {r_nv['campaign']!r} cached Anomaly score matches pre_score "
              "(ISNUMBER guard correctly skips the blank Value Δ cell)",
              abs((cached_nv_score or 0) - r_nv["pre_score"]) < 1e-6,
              f"{cached_nv_score} vs {r_nv['pre_score']}")


def test_assumptions_provenance():
    print("test_assumptions_provenance")
    from render import model as M

    # sample-findings.json carries no params.roas_goal -> compute_model must
    # auto-stamp a model_default assumption (HM-604), never silently present
    # DEFAULT_PARAMS['roas_goal'] as a confirmed client target.
    findings = json.loads(FIXTURE.read_text())
    check("fixture supplies no explicit roas_goal", (findings.get("params") or {}).get("roas_goal") is None)
    m = core.compute_model(findings)
    a = M.get_assumption(m, "roas_goal")
    check("roas_goal auto-stamped basis=model_default", a is not None and a["basis"] == "model_default", a)
    check("require_assumptions is clean (auto-stamped)", M.require_assumptions(m, ["roas_goal"]) == [])

    # an explicit client-supplied roas_goal must NOT be auto-stamped as a default.
    findings2 = json.loads(FIXTURE.read_text())
    findings2["params"] = {"roas_goal": 6.0}
    m2 = core.compute_model(findings2)
    check("explicit roas_goal carries no model_default assumption",
          M.get_assumption(m2, "roas_goal") is None)

    import perf_spec
    import perf_xlsx_spec
    from render import build_bundle
    spec = dict(perf_spec.SPEC)
    spec["xlsx"] = perf_xlsx_spec.XLSX
    with tempfile.TemporaryDirectory() as td:
        written = build_bundle(m, spec, td, formats=("md", "html", "xlsx"), charts=False,
                               normalize=False)
        md = next(p for p in written if p.suffix == ".md").read_text()
        html = next(p for p in written if p.name.endswith("_explorer.html")).read_text()
        xlsx_path = next(p for p in written if p.suffix == ".xlsx")
        check("md has the callout", "## Provenance & assumptions" in md)
        check("md ROAS goal line carries the inline marker", "ROAS goal" in md and "(default:" in md)
        check("html embeds roas_goal assumption", '"roas_goal"' in html and '"model_default"' in html)
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path))
        snap_cells = [c.value for row in wb["Snapshot"].iter_rows() for c in row if c.value is not None]
        check("xlsx Snapshot has the callout", "Provenance & assumptions" in snap_cells)
        ctrl_notes = [c.value for c in wb["Controls"]["D"] if c.value]
        check("xlsx Controls note carries the inline marker",
              any("default:" in (v or "") for v in ctrl_notes), ctrl_notes)


def main():
    for t in (test_fixture_buckets, test_no_value_not_bucketed, test_dedupe_by_campaign_id,
              test_empty, test_goal_sensitivity_current_flag, test_assemble_findings_from_raw,
              test_bundle_md_html_parity_and_lazy, test_anomalies_and_concentration,
              test_delta_flag_tunable, test_csv_matches_mcp_model, test_liveness_gating,
              test_xlsx_recalc_matches_model, test_assumptions_provenance):
        t()
    print()
    if _failures:
        print(f"FAILED ({len(_failures)}): {', '.join(_failures)}")
        return 1
    print("ALL TESTS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())

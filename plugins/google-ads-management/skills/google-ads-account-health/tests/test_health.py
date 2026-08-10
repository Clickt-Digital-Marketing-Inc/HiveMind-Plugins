#!/usr/bin/env python3
"""Tests for the account-health core + builder (stdlib only; run directly).

    python3 tests/test_health.py

Asserts the documented fixture scoring, no-row-loss, nullable-column
handling (per-check fields null on rows the check doesn't apply to),
MCP-vs-CSV model parity, bundle emit (md + xlsx, no HTML), and the xlsx
recalc matching the Python model. Exit 0 = all pass, 1 = a failure.
"""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent / "scripts"
SHARED = HERE.parents[2] / "_shared"
sys.path.insert(0, str(SCRIPTS))
sys.path.insert(0, str(SHARED))

import health_core as core          # noqa: E402
import health_artifacts as artifacts  # noqa: E402
import assemble_from_csv as csv_assembler  # noqa: E402
import assemble_findings as mcp_assembler  # noqa: E402

sys.path.insert(0, str(SHARED))
import csv_input as csv_input_mod   # noqa: E402

FIXTURE_MCP = HERE / "sample-findings.json"
FIXTURE_CSV = HERE / "sample-findings-csv.json"
_failures = []


def check(name, cond, detail=""):
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  — {detail}" if detail and not cond else ""))
    if not cond:
        _failures.append(name)


def chart_args():
    """`[]` when the pinned chart renderer is importable, `["--no-charts"]` when it is not.

    This skill's spec DECLARES charts, and `_shared/render/bundle.py` deliberately
    hard-fails (exit 2) when charts are declared while `vl-convert-python` is absent,
    rather than silently shipping a chartless report. That is correct product
    behavior — but the builder tests below assert on report/artifact EMISSION and on
    xlsx recalc, never on chart SVGs, so a missing optional native wheel must not
    turn them red (HM-803). Opting out only when the renderer is genuinely
    unavailable keeps the full chart path under test on any machine that has the
    dependency `requirements.txt` pins.

    The probe mirrors the guard it compensates for (`_shared/render/charts.py`'s
    `render_chart_svg`, and `_has_vl_convert()` in the shared toolkit's own tests):
    a REAL import, never `importlib.util.find_spec`. A dist that is locatable but
    not importable (arch-mismatched wheel, half-extracted native extension)
    resolves under `find_spec` while the builder still exits 2 — the exact red this
    helper exists to remove. The selected mode is printed so a green log can be
    told apart from a chart-blind one.
    """
    try:
        import vl_convert  # noqa: F401
        mode = []
    except Exception:      # absent, or present-but-unimportable
        mode = ["--no-charts"]
    print("    chart mode: " + ("charts rendered" if not mode
                                else "charts skipped (vl_convert unimportable)"))
    return mode


def test_fixture_scoring():
    print("test_fixture_scoring")
    model = core.compute_model(core.load_findings(str(FIXTURE_MCP)))
    s = model["summary"]
    check("universe == 12", s["universe"] == 12, f"got {s['universe']}")
    check("total_flagged == 5", s["total_flagged"] == 5, f"got {s['total_flagged']}")
    check("by_severity Critical=1/High=3/Medium=1",
          s["by_severity"] == {"Critical": 1, "High": 3, "Medium": 1}, f"got {s['by_severity']}")
    check("structural_score == 31.5", abs(s["structural_score"] - 31.5) < 1e-6, f"got {s['structural_score']}")
    by_check = s["by_check"]
    check("sprawl 1/3", by_check["sprawl"] == {"universe": 3, "flagged": 1})
    check("no_negatives 1/2 (Search-only)", by_check["no_negatives"] == {"universe": 2, "flagged": 1})
    check("automation_no_data 1/3", by_check["automation_no_data"] == {"universe": 3, "flagged": 1})
    check("naming 1/3", by_check["naming"] == {"universe": 3, "flagged": 1})
    check("pmax_cannibalization 1/1", by_check["pmax_cannibalization"] == {"universe": 1, "flagged": 1})


def test_top_fixes_ranking():
    print("test_top_fixes_ranking")
    model = core.compute_model(core.load_findings(str(FIXTURE_MCP)))
    tf = model["top_fixes"]
    check("5 flagged rows in top_fixes", len(tf) == 5, f"got {len(tf)}")
    got_order = [(r["check"], r["pre_score"]) for r in tf]
    want_order = [("automation_no_data", 9.0), ("no_negatives", 7.0),
                  ("pmax_cannibalization", 6.5), ("sprawl", 6.0), ("naming", 3.0)]
    check("ranked by pre_score desc", got_order == want_order, f"got {got_order}")


def test_no_row_loss():
    print("test_no_row_loss")
    findings = core.load_findings(str(FIXTURE_MCP))
    model = core.compute_model(findings)
    # 3 ad_group rows (sprawl) + 3 campaigns * 3 checks (no_negatives/automation/naming,
    # but no_negatives only for Search) + 1 pmax row for the PMax campaign
    n_ag = len(findings["ad_groups"])
    n_camp = len(findings["campaigns"])
    n_search = sum(1 for c in findings["campaigns"] if c["channel_type"] == "SEARCH")
    n_pmax = sum(1 for c in findings["campaigns"] if c["channel_type"] == "PERFORMANCE_MAX")
    expected = n_ag + (n_camp * 2) + n_search + n_pmax   # automation+naming for every campaign
    check("row count matches every (check, entity) pair", len(model["rows"]) == expected,
          f"got {len(model['rows'])}, expected {expected}")
    for i, r in enumerate(model["rows"]):
        check(f"row {i} has a status", "status" in r and r["status"] in ("scored", "config", "manual"))
        check(f"row {i} has pre_score", "pre_score" in r and isinstance(r["pre_score"], float))
        check(f"row {i} has is_flagged", "is_flagged" in r and isinstance(r["is_flagged"], bool))


def test_nullable_columns():
    print("test_nullable_columns")
    model = core.compute_model(core.load_findings(str(FIXTURE_MCP)))
    sprawl_rows = [r for r in model["rows"] if r["check"] == "sprawl"]
    other_fields = ("negative_count", "bidding_strategy_type", "conversions_30d",
                    "name_pattern_ok", "pmax_present", "brand_present", "has_brand_exclusion")
    for r in sprawl_rows:
        check("sprawl row has keyword_count", r["keyword_count"] is not None)
        check("sprawl row has ad_group_ctr", r["ad_group_ctr"] is not None)
        for f in other_fields:
            check(f"sprawl row's '{f}' is null (doesn't apply)", r[f] is None)

    pmax_rows = [r for r in model["rows"] if r["check"] == "pmax_cannibalization"]
    for r in pmax_rows:
        check("pmax row has pmax_present", r["pmax_present"] is not None)
        check("pmax row's has_brand_exclusion is ALWAYS None (never API-confirmable)",
              r["has_brand_exclusion"] is None)
        check("pmax row status == 'manual'", r["status"] == "manual")

    naming_rows = [r for r in model["rows"] if r["check"] == "naming"]
    for r in naming_rows:
        check("naming row status == 'config'", r["status"] == "config")

    scored_checks = ("sprawl", "no_negatives", "automation_no_data")
    for r in model["rows"]:
        if r["check"] in scored_checks:
            check(f"{r['check']} row status == 'scored'", r["status"] == "scored")


def test_sprawl_requires_both_conditions():
    print("test_sprawl_requires_both_conditions")
    model = core.compute_model(core.load_findings(str(FIXTURE_MCP)))
    core_row = next(r for r in model["rows"] if r["check"] == "sprawl" and r["entity_name"] == "Core")
    check("Core ad group (25 kw, 10% CTR) is NOT flagged — healthy CTR", not core_row["is_flagged"])
    check("Core ad group has the sprawl_size sub-signal (partial, not enough alone)",
          "sprawl_size" in core_row["flags"] and "sprawl_low_ctr" not in core_row["flags"])
    check("Core ad group pre_score is 0 when not flagged", core_row["pre_score"] == 0.0)


def test_mcp_csv_parity():
    print("test_mcp_csv_parity")
    m1 = core.compute_model(core.load_findings(str(FIXTURE_MCP)))
    m2 = core.compute_model(core.load_findings(str(FIXTURE_CSV)))
    check("MCP and CSV summaries are identical", m1["summary"] == m2["summary"],
          f"mcp={m1['summary']} csv={m2['summary']}")
    check("MCP meta.source == 'mcp'", m1["provenance"]["source"] == "mcp")
    check("CSV meta.source == 'user_csv'", m2["provenance"]["source"] == "user_csv")
    mcp_scores = sorted(r["pre_score"] for r in m1["rows"])
    csv_scores = sorted(r["pre_score"] for r in m2["rows"])
    check("MCP and CSV produce the same pre_score multiset", mcp_scores == csv_scores)


def test_empty_universe():
    print("test_empty_universe")
    f = {"meta": {}, "ad_groups": [], "campaigns": []}
    model = core.compute_model(f)
    s = model["summary"]
    check("empty -> universe 0", s["universe"] == 0)
    check("empty -> total_flagged 0", s["total_flagged"] == 0)
    check("empty -> top_fixes empty", model["top_fixes"] == [])


def test_orphan_negatives_regression():
    print("test_orphan_negatives_regression")
    import subprocess
    tmp = Path(tempfile.mkdtemp())
    try:
        out = tmp / "findings.json"
        r = subprocess.run([sys.executable, str(SCRIPTS / "assemble_findings.py"),
                            "--keywords", str(HERE / "raw" / "keywords.txt"),
                            "--adgroup-perf", str(HERE / "raw" / "adgroup_perf.txt"),
                            "--campaigns", str(HERE / "raw" / "campaigns.txt"),
                            "--negatives", str(HERE / "raw" / "negatives_orphan.txt"),
                            "--client-name", "Acme Corp", "--account-id", "123-456-7890",
                            "--currency", "CAD", "--window-30d", "2026-06-06 to 2026-07-05",
                            "--generated", "2026-07-05", "-o", str(out)],
                           capture_output=True, text=True)
        check("assembler exits 0", r.returncode == 0, r.stderr)
        check("assembler prints the absent-id NOTE naming campaign 9999",
              "NOTE:" in r.stderr and "9999" in r.stderr, r.stderr)

        findings = json.loads(out.read_text())
        # 3 of the 8 raw negative rows (campaign.id 9999) reference a campaign
        # absent from campaigns.txt (which only has 1001/1002/1003) — the raw
        # universe is 8, not the 5 that land on campaign 1002 in the array.
        rec = findings["meta"]["reconciliation"]
        check("reconciliation.raw_totals.negatives == 8 (the raw universe)",
              rec["raw_totals"]["negatives"] == 8.0, f"got {rec.get('raw_totals')}")
        check("campaigns array negative_count sums to 5 (post-join, excludes the orphan)",
              rec["campaigns"]["sums"]["negative_count"] == 5.0)
        check("orphan_negatives.count == 3", findings["orphan_negatives"]["count"] == 3,
              f"got {findings['orphan_negatives']}")
        check("orphan_negatives.campaign_ids == ['9999']",
              findings["orphan_negatives"]["campaign_ids"] == ["9999"])
        check("orphan_negatives.status == 'out_of_scope'",
              findings["orphan_negatives"]["status"] == "out_of_scope")

        # totals reconcile to the raw count: load_findings must NOT raise —
        # campaigns' negative_count (5) + orphan_negatives.count (3) == 8.
        loaded = core.load_findings(str(out))
        model = core.compute_model(loaded)
        check("model exposes orphan_negatives", model["orphan_negatives"]["count"] == 3)

        # A findings file that drops orphan_negatives (as the pre-fix code did,
        # implicitly) must hard-fail reconciliation — proving the raw-count
        # gate actually catches the silent-loss defect this test guards.
        tampered = json.loads(out.read_text())
        tampered["orphan_negatives"] = {"count": 0, "campaign_ids": [], "status": "out_of_scope"}
        tampered_path = tmp / "tampered.json"
        tampered_path.write_text(json.dumps(tampered))
        raised = False
        try:
            core.load_findings(str(tampered_path))
        except core.FindingsError:
            raised = True
        check("dropping orphan_negatives without updating the raw total hard-fails reconciliation",
              raised)

        outdir = tmp / "artifacts"
        r = subprocess.run([sys.executable, str(SCRIPTS / "build_health_report.py"),
                            "--input", str(out), "--outdir", str(outdir),
                            "--brand", "Acme Corp", "--formats", "md"] + chart_args(),
                           capture_output=True, text=True)
        check("builder exits 0 on the orphan fixture", r.returncode == 0, r.stderr)
        md_files = [p for p in outdir.glob("*.md")
                    if "account-health" in p.name and "_action_plan" not in p.name
                    and "_renaming" not in p.name]
        check("md report emitted", len(md_files) == 1, f"got {[p.name for p in md_files]}")
        if md_files:
            text = md_files[0].read_text()
            check("md report surfaces one orphan-negatives line",
                  "3 negative(s) belong to removed/out-of-scope campaigns" in text, text)
        action_plan = list(outdir.glob("*_action_plan.md"))
        check("action_plan.md emitted", len(action_plan) == 1)
        if action_plan:
            text = action_plan[0].read_text()
            check("action_plan.md surfaces one orphan-negatives line",
                  "3 negatives belong to removed/out-of-scope campaigns" in text, text)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_reconciliation_hard_fails_on_tamper():
    print("test_reconciliation_hard_fails_on_tamper")
    data = json.loads(FIXTURE_MCP.read_text())
    data["campaigns"][0]["negative_count"] = 999   # tamper after the fact
    tmp = Path(tempfile.mkstemp(suffix=".json")[1])
    tmp.write_text(json.dumps(data))
    try:
        raised = False
        try:
            core.load_findings(str(tmp))
        except core.FindingsError:
            raised = True
        check("tampered findings raise FindingsError on load", raised)
    finally:
        tmp.unlink()


def test_bundle_emit_and_xlsx_recalc():
    print("test_bundle_emit_and_xlsx_recalc")
    import subprocess
    tmp = Path(tempfile.mkdtemp())
    try:
        r = subprocess.run([sys.executable, str(SCRIPTS / "build_health_report.py"),
                            "--input", str(FIXTURE_MCP), "--outdir", str(tmp),
                            "--brand", "Acme Corp", "--formats", "md,xlsx"] + chart_args(),
                           capture_output=True, text=True)
        check("builder exits 0", r.returncode == 0, r.stderr)
        md_files = list(tmp.glob("*.md"))
        xlsx_files = list(tmp.glob("*.xlsx"))
        html_files = list(tmp.glob("*.html"))
        check("md report emitted", any("account-health" in p.name and "_action_plan" not in p.name
                                       and "_renaming" not in p.name for p in md_files))
        check("xlsx emitted", len(xlsx_files) == 1, f"got {[p.name for p in xlsx_files]}")
        check("NO html explorer emitted (reduced bundle)", len(html_files) == 0, f"got {html_files}")
        check("action_plan.md emitted", any(p.name.endswith("_action_plan.md") for p in md_files))
        check("renaming.md emitted", any(p.name.endswith("_renaming.md") for p in md_files))
        check("pause_list.csv emitted", any(p.name.endswith("_pause_list.csv") for p in tmp.glob("*.csv")))

        if xlsx_files:
            import health_xlsx_spec as xspec
            from render.xlsx import check_workbook
            rc = check_workbook(str(xlsx_files[0]), {"xlsx": xspec.XLSX})
            check("xlsx passes structural check_workbook", rc == 0)

            from openpyxl import load_workbook
            wb = load_workbook(str(xlsx_files[0]), data_only=True)
            c = wb["Controls"]
            model = core.compute_model(core.load_findings(str(FIXTURE_MCP)))
            s = model["summary"]
            check("xlsx C39 (total flagged) matches model", c["C39"].value == s["total_flagged"],
                  f"xlsx={c['C39'].value} model={s['total_flagged']}")
            check("xlsx C40 (universe) matches model", c["C40"].value == s["universe"],
                  f"xlsx={c['C40'].value} model={s['universe']}")
            check("xlsx C34 (sprawl flagged) matches model",
                  c["C34"].value == s["by_check"]["sprawl"]["flagged"])
            check("xlsx C38 (pmax flagged) matches model",
                  c["C38"].value == s["by_check"]["pmax_cannibalization"]["flagged"])
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_reject_html_format():
    print("test_reject_html_format")
    import subprocess
    tmp = Path(tempfile.mkdtemp())
    try:
        # chart_args() here too: --no-charts cannot change a format rejection, and
        # without it this invocation's green rests on build_health_report.py
        # validating --formats BEFORE the chart path — an ordering nothing pins
        # (HM-803 gate).
        r = subprocess.run([sys.executable, str(SCRIPTS / "build_health_report.py"),
                            "--input", str(FIXTURE_MCP), "--outdir", str(tmp),
                            "--formats", "html"] + chart_args(), capture_output=True, text=True)
        check("builder rejects --formats html (reduced bundle has no HTML spec)", r.returncode == 1)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_action_plan_artifacts():
    print("test_action_plan_artifacts")
    model = core.compute_model(core.load_findings(str(FIXTURE_MCP)))
    plan = artifacts.action_plan_md(model)
    check("action plan mentions Critical section", "## Critical" in plan)
    check("action plan mentions the automation fix", "Automation without data" in plan)
    rows = artifacts.pause_list_rows(model)
    check("pause_list has exactly the 1 flagged sprawl ad group", len(rows) == 1, f"got {len(rows)}")
    check("pause_list row names the right ad group", rows[0]["Ad Group"] == "Generic - CRM")
    renaming = artifacts.renaming_md(model)
    check("renaming worklist names the PMax campaign", "Shopping PMax Everything" in renaming)
    check("renaming worklist does NOT invent a compliant name",
          "confirm" in renaming.lower())


def test_csv_missing_column_raises():
    print("test_csv_missing_column_raises")
    tmp = Path(tempfile.mkdtemp())
    try:
        bad_csv = tmp / "adgroups_missing_keywords.csv"
        bad_csv.write_text("Campaign,Ad group,Clicks,Impr.\n"
                           "NonBrand_US_Search_CRM_2026,Generic - CRM,40,4000\n")  # no keyword-count column
        raised = False
        try:
            csv_assembler.assemble(str(bad_csv), str(HERE / "campaigns-export.csv"),
                                   {"client_name": "X", "account_id": "1", "currency": "USD",
                                    "window_30d": "a to b"})
        except csv_input_mod.CsvInputError as e:
            raised = True
            check("error names the missing column", "keyword" in str(e).lower(), str(e))
        check("missing 'Ad group keywords (enabled)' column raises CsvInputError", raised)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_assemble_findings_missing_perf_name_fallback():
    """An ad group that appears in the keywords pull but not the ad-group-perf
    pull (e.g. zero impressions in the window) must still get a usable
    campaign/ad-group name in the sprawl table, not a blank cell (HM-607 P9)."""
    print("test_assemble_findings_missing_perf_name_fallback")
    tmp = Path(tempfile.mkdtemp())
    try:
        keywords = tmp / "keywords.txt"
        keywords.write_text(json.dumps({"result": [
            {"campaign.id": "1001", "ad_group.id": "2001"},
            {"campaign.id": "1001", "ad_group.id": "2001"},
            # 2099 shows up only here — never in the perf pull below.
            {"campaign.id": "1001", "ad_group.id": "2099"},
        ]}))
        adgroup_perf = tmp / "adgroup_perf.txt"
        adgroup_perf.write_text(json.dumps({"result": [
            {"campaign.id": "1001", "campaign.name": "NonBrand_US_Search_CRM_2026",
             "ad_group.id": "2001", "ad_group.name": "Generic - CRM",
             "metrics.clicks": "40", "metrics.impressions": "4000"},
        ]}))
        campaigns = tmp / "campaigns.txt"
        campaigns.write_text(json.dumps({"result": [
            {"campaign.id": "1001", "campaign.name": "NonBrand_US_Search_CRM_2026",
             "campaign.status": "ENABLED", "campaign.advertising_channel_type": "SEARCH",
             "campaign.bidding_strategy_type": "MAXIMIZE_CONVERSIONS", "metrics.conversions": "1",
             "metrics.cost_micros": "500000000"},
        ]}))
        negatives = tmp / "negatives.txt"
        negatives.write_text(json.dumps({"result": []}))

        findings = mcp_assembler.assemble(
            str(keywords), str(adgroup_perf), str(campaigns), str(negatives),
            {"client_name": "X", "account_id": "1", "currency": "USD", "window_30d": "a to b"})

        by_id = {ag["ad_group_id"]: ag for ag in findings["ad_groups"]}
        check("perf-covered ad group keeps its real name", by_id["2001"]["ad_group"] == "Generic - CRM")
        missing = by_id["2099"]
        check("perf-missing ad group is not dropped (no-row-loss)", "2099" in by_id)
        check("perf-missing ad group gets a fallback ad_group label, not blank",
              missing["ad_group"] == "(name unavailable — id 2099)", missing["ad_group"])
        check("perf-missing ad group gets a fallback campaign label, not blank",
              missing["campaign"] == "(name unavailable — id 1001)", missing["campaign"])
        check("perf-missing ad group still carries its keyword_count", missing["keyword_count"] == 1)

        model = core.compute_model(findings)
        sprawl_names = {r["entity_name"] for r in model["rows"] if r["check"] == "sprawl"}
        check("fallback label reaches the sprawl rows the report renders",
              "(name unavailable — id 2099)" in sprawl_names, sprawl_names)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_liveness_gating():
    print("test_liveness_gating")
    # A dormant campaign (PAUSED, zero 30d spend) that WOULD trip automation_no_data
    # (automated bidding + 0 conversions) and no_negatives — plus a paused-but-spent
    # campaign that stays in scope as recently_active, and a live one.
    f = {"meta": {"currency": "CAD"}, "ad_groups": [
            {"campaign_id": "D", "campaign": "Dead", "ad_group_id": "d1", "ad_group": "dead-ag",
             "keyword_count": 40, "clicks": 0, "impressions": 5000},  # would trip sprawl
         ],
         "campaigns": [
            {"campaign_id": "L", "campaign": "Live_ok", "status": "ENABLED", "channel_type": "SEARCH",
             "bidding_strategy_type": "MAXIMIZE_CONVERSIONS", "conversions_30d": 0, "cost": 1200.0,
             "negative_count": 0},
            {"campaign_id": "P", "campaign": "Paused_spent", "status": "PAUSED", "channel_type": "SEARCH",
             "bidding_strategy_type": "MAXIMIZE_CONVERSIONS", "conversions_30d": 0, "cost": 300.0,
             "negative_count": 0},
            {"campaign_id": "D", "campaign": "Dead", "status": "PAUSED", "channel_type": "SEARCH",
             "bidding_strategy_type": "MAXIMIZE_CONVERSIONS", "conversions_30d": 0, "cost": 0.0,
             "negative_count": 0},
         ]}
    m = core.compute_model(f)
    rows = m["rows"]
    live_by_check = {(r["check"], r["campaign_id"] if r["entity_type"] == "campaign" else r["campaign_id"]): r
                     for r in rows}
    # every row present (no-row-loss) and tagged
    check("all rows carry a liveness tag", all(r.get("liveness") in
          ("live", "recently_active", "dormant") for r in rows), [r.get("liveness") for r in rows])
    live_rows = [r for r in rows if r["campaign_id"] == "L"]
    paused_rows = [r for r in rows if r["campaign_id"] == "P"]
    dead_rows = [r for r in rows if r["campaign_id"] == "D"]
    check("live campaign rows tagged live", all(r["liveness"] == "live" for r in live_rows))
    check("paused-but-spent campaign rows tagged recently_active",
          all(r["liveness"] == "recently_active" for r in paused_rows),
          [r["liveness"] for r in paused_rows])
    check("dead campaign (+ its ad group) tagged dormant",
          all(r["liveness"] == "dormant" for r in dead_rows), [(r["check"], r["liveness"]) for r in dead_rows])
    # the dead campaign would have tripped automation_no_data + no_negatives + sprawl,
    # but liveness gating produces ZERO flags on every dormant row (still present)
    check("dormant rows produce zero flags/severity",
          all((not r["is_flagged"]) and r["pre_score"] == 0.0 and r["severity"] is None
              for r in dead_rows),
          [(r["check"], r["is_flagged"]) for r in dead_rows])
    check("dormant sprawl ad-group row present but not flagged",
          any(r["check"] == "sprawl" and r["campaign_id"] == "D" and not r["is_flagged"] for r in rows))
    # the LIVE campaign's automation_no_data DOES fire (gate only removes the dead)
    check("live automated-no-data campaign still flags",
          any(r["check"] == "automation_no_data" and r["campaign_id"] == "L" and r["is_flagged"]
              for r in rows))
    # recently_active rows carry the conditional-phrasing seam
    paused_note = next(r["liveness_note"] for r in paused_rows)
    check("recently_active rows carry a conditional liveness_note",
          "confirm intent" in paused_note and "spending" in paused_note, paused_note)
    check("live rows have empty liveness_note", all(r["liveness_note"] == "" for r in live_rows))
    check("dormant rows have empty liveness_note", all(r["liveness_note"] == "" for r in dead_rows))
    # summary surfaces the liveness split
    bl = m["summary"]["by_liveness"]
    check("summary.by_liveness counts every row",
          bl["live"] + bl["recently_active"] + bl["dormant"] == len(rows), bl)
    check("summary.by_liveness has dormant rows", bl["dormant"] >= 2, bl)


def test_invalid_naming_regex_raises():
    print("test_invalid_naming_regex_raises")
    findings = json.loads(FIXTURE_MCP.read_text())
    findings["params"] = {"naming_regex": "("}   # unbalanced paren — invalid regex
    raised = False
    try:
        core.compute_model(findings)
    except core.FindingsError as e:
        raised = True
        check("error names the bad regex", "naming_regex" in str(e), str(e))
    check("invalid naming_regex raises a clear FindingsError (not a bare re.error)", raised)


def test_assumptions_provenance():
    print("test_assumptions_provenance")
    from render import model as M

    findings = json.loads(FIXTURE_MCP.read_text())
    check("fixture supplies no explicit naming_regex", (findings.get("params") or {}).get("naming_regex") is None)
    m = core.compute_model(findings)
    a = M.get_assumption(m, "naming_regex")
    check("naming_regex auto-stamped basis=model_default", a is not None and a["basis"] == "model_default", a)

    findings2 = json.loads(FIXTURE_MCP.read_text())
    findings2["params"] = {"naming_regex": r"^Custom_.*$"}
    m2 = core.compute_model(findings2)
    check("explicit naming_regex carries no assumption", M.get_assumption(m2, "naming_regex") is None)

    import health_spec
    import health_xlsx_spec
    from render import build_bundle
    spec = dict(health_spec.SPEC)
    spec["xlsx"] = health_xlsx_spec.XLSX
    with tempfile.TemporaryDirectory() as td:
        written = build_bundle(m, spec, td, formats=("md", "xlsx"), charts=False, normalize=False)
        md = next(p for p in written if p.suffix == ".md" and "action_plan" not in p.name
                 and "renaming" not in p.name).read_text()
        xlsx_path = next(p for p in written if p.suffix == ".xlsx")
        check("md has the callout", "## Provenance & assumptions" in md)
        check("md naming-convention line carries the marker", "(default:" in md)
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path))
        snap_cells = [c.value for row in wb["Checks snapshot"].iter_rows() for c in row if c.value is not None]
        check("xlsx Snapshot has the callout", "Provenance & assumptions" in snap_cells)
        check("xlsx Snapshot lists naming_regex", "naming_regex" in snap_cells)


def main() -> int:
    test_fixture_scoring()
    test_top_fixes_ranking()
    test_no_row_loss()
    test_nullable_columns()
    test_sprawl_requires_both_conditions()
    test_mcp_csv_parity()
    test_empty_universe()
    test_orphan_negatives_regression()
    test_reconciliation_hard_fails_on_tamper()
    test_csv_missing_column_raises()
    test_assemble_findings_missing_perf_name_fallback()
    test_liveness_gating()
    test_invalid_naming_regex_raises()
    test_action_plan_artifacts()
    test_bundle_emit_and_xlsx_recalc()
    test_reject_html_format()
    test_assumptions_provenance()
    print()
    if _failures:
        print(f"{len(_failures)} FAILURE(S): " + ", ".join(_failures))
        return 1
    print("All tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

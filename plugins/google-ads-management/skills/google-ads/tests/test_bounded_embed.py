#!/usr/bin/env python3
"""HM-339 bounded-embed guards (keywords waste filter + products).

Proves the two halves of the bounded-embed contract for every skill that opts into
a trimmed widget embed:
  1. The in-Claude tuner widget embeds only the trimmed in-play envelope and the
     assembled widget stays under a size ceiling, while `total_rows` still reports
     the true universe.
  2. The SAVED deliverables lose nothing: the md full-table and the xlsx rows sheet
     each carry the FULL universe (built by build_bundle over the full model, never
     the trimmed embed).

Plus a keywords-only guard that the envelope stays correct even when `ctr_factor`
starts above the slider max (1.0) — the regression guard for the hardened in_play cap.

Needs openpyxl. Run: python3 tests/test_bounded_embed.py
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent          # .../skills/google-ads/tests
HUB = HERE.parent                               # .../skills/google-ads
SKILLS = HUB.parent                             # .../skills
BUILD_WIDGET = HUB / "references" / "build_widget.py"
PY = sys.executable

# Assembled trimmed widgets are ~110 KB; an untrimmed embed (all rows) would blow
# well past this, so the ceiling fails loudly if the trim ever regresses.
SIZE_CEILING = 160_000

# Per-skill cases. md_row_re / xlsx_token match one cell per universe row in this
# skill's full table (the synthetic fixtures name rows "term NNNN" / "Product NNNN").
CASES = [
    {"id": "keywords", "skill": "google-ads-keywords-search-terms", "builder": "build_waste_filter.py",
     "fixture": "tests/sample-findings-large.json", "array": "search_terms",
     "md_section": "All loose-match terms", "md_row_re": r"(?m)^\|\s*term \d{4}",
     "xlsx_sheet": "Live filter", "xlsx_token": "term "},
    {"id": "products", "skill": "google-ads-products", "builder": "build_product_report.py",
     "fixture": "tests/product-sample-findings-large.json", "array": "products",
     "md_section": "All products", "md_row_re": r"(?m)^\|\s*Product \d{4}",
     "xlsx_sheet": "Live products", "xlsx_token": "Product "},
]

fails: list[str] = []


def check(cond, msg):
    if not cond:
        fails.append(msg)


def run_case(case) -> int:
    sdir = SKILLS / case["skill"]
    build = sdir / "scripts" / case["builder"]
    fixture = sdir / case["fixture"]
    universe = len(json.loads(fixture.read_text())[case["array"]])
    tag = case["id"]
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        wjson, whtml = td / "w.json", td / "w.html"

        # 1) widget: trimmed embed, honest total, under ceiling
        r = subprocess.run([PY, str(build), "--input", str(fixture), "--formats", "",
                            "--brand", "X", "--emit-widget", str(wjson)], capture_output=True, text=True)
        check(r.returncode == 0, f"[{tag}] emit-widget failed: {r.stderr.strip()}")
        emb = json.loads(wjson.read_text())["embed"]
        check(emb.get("total_rows") == universe,
              f"[{tag}] embed.total_rows {emb.get('total_rows')} == universe {universe}")
        check(len(emb["rows"]) < universe,
              f"[{tag}] embed trimmed below universe: {len(emb['rows'])} < {universe}")
        r = subprocess.run([PY, str(BUILD_WIDGET), "--data", str(wjson), "--out", str(whtml)],
                           capture_output=True, text=True)
        check(r.returncode == 0, f"[{tag}] assemble failed: {r.stderr.strip()}")
        size = whtml.stat().st_size if whtml.exists() else 1 << 30
        check(size < SIZE_CEILING, f"[{tag}] assembled widget {size}B < ceiling {SIZE_CEILING}B")
        head = whtml.read_bytes()[:200]
        check(head.lstrip().lower().startswith(b"<!doctype html>"),
              f"[{tag}] assembled widget starts with <!DOCTYPE html> (HM-607 H1): {head[:40]!r}")
        check(b'<meta charset="utf-8">' in head,
              f"[{tag}] assembled widget declares UTF-8 charset (HM-607 H1): {head[:80]!r}")

        # 2) no data loss: md + xlsx carry the FULL universe
        r = subprocess.run([PY, str(build), "--input", str(fixture), "--outdir", str(td),
                            "--formats", "md,xlsx", "--brand", "X"], capture_output=True, text=True)
        check(r.returncode == 0, f"[{tag}] md/xlsx build failed: {r.stderr.strip()}")
        md = next(td.glob("*.md")).read_text(encoding="utf-8")
        sec = md.split(case["md_section"], 1)[1] if case["md_section"] in md else ""
        sec = re.split(r"\n## ", sec)[0]
        check(len(re.findall(case["md_row_re"], sec)) == universe,
              f"[{tag}] md full table carries every row == {universe}")
        from openpyxl import load_workbook
        wb = load_workbook(next(td.glob("*.xlsx")), data_only=True)
        ws = wb[case["xlsx_sheet"]]
        cnt = sum(1 for row in ws.iter_rows(values_only=True)
                  for v in row if isinstance(v, str) and v.startswith(case["xlsx_token"]))
        check(cnt == universe, f"[{tag}] xlsx '{case['xlsx_sheet']}' carries every row: {cnt} == {universe}")
    return universe


def check_ctr_factor_envelope():
    """keywords-only: the envelope holds every flaggable row even when ctr_factor
    starts ABOVE the slider max (1.0). Guards the hardened in_play cap; has teeth
    (the old hardcoded `ctr < camp_ctr` undercounts at ctr_factor=2.0)."""
    kw = SKILLS / "google-ads-keywords-search-terms"
    sys.path.insert(0, str(kw / "scripts"))
    sys.path.insert(0, str(kw.parent.parent / "_shared"))
    import waste_filter_core as core            # noqa: E402
    import waste_filter_spec as spec            # noqa: E402
    base = json.loads((kw / "tests" / "sample-findings-large.json").read_text())
    for cf in (1.0, 2.0):
        f = dict(base)
        f["params"] = {"ctr_factor": cf}
        m = core.compute_model(f)
        env = [r for r in m["rows"] if spec.in_play(r, m["params"])]
        fb1, fb2 = m["summary"]["block1"], m["summary"]["block2"]
        eb1 = sum(1 for r in env if r["block"] == "Block 1")
        eb2 = sum(1 for r in env if r["block"] == "Block 2")
        check(eb1 == fb1 and eb2 == fb2,
              f"[keywords] ctr_factor={cf}: envelope holds every flag (env {eb1}/{eb2} == full {fb1}/{fb2})")
        check(len(env) < len(m["rows"]),
              f"[keywords] ctr_factor={cf}: envelope still trims ({len(env)} < {len(m['rows'])})")


def main() -> int:
    sizes = {c["id"]: run_case(c) for c in CASES}
    check_ctr_factor_envelope()
    if fails:
        print(f"FAIL — {len(fails)} bounded-embed problem(s):")
        for f in fails:
            print("  - " + f)
        return 1
    print("OK — bounded-embed: widgets trimmed + under ceiling; md + xlsx carry the full universe ("
          + ", ".join(f"{k}={v}" for k, v in sizes.items())
          + "); keywords ctr_factor>1 envelope holds. No data loss.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

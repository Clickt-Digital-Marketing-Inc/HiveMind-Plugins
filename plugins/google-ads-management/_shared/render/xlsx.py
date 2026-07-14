#!/usr/bin/env python3
"""Generic formula-driven xlsx renderer for the analytical bundle.

This is the ONLY module in the toolkit allowed to import openpyxl, and it is
imported lazily (only by bundle.build_bundle when 'xlsx' is requested), so the
md/html path stays dependency-free.

Driven by spec['xlsx'] (pure data — a skill's xlsx spec never imports openpyxl):

  controls_sheet / rows_sheet / snapshot_sheet : sheet names
  title(pr,brand) / subtitle(pr) / intro        : header text
  params  : [{row,label,key,fmt,note,dropdown}]  yellow input cells at column C
  toggles : {section_title,start_row,options,param_key,dropdown,note}  (optional)
  logic   : {title_row,title,blocks:[{head_row,head,rows:[(row,formula)]}]}  (optional)
  results : {title_row,title,items:[{row,label,cell,formula,fmt}]}
  aux     : [{title_row,title,header_row,start_row,source,columns:[(hdr,key,fmt)]}]
  rows_columns : ordered [{header,kind:'data'|'formula',key?,fmt?,scored?,formula?}]
  snapshot_sections(model) -> [section]    (defaults to spec['md_sections'])
  check   : {param_cells, cached_cell, status_header, qualifies_header}

Formula templates use tokens resolved against the built layout:
  {row}            current rows-sheet row number
  {C:Header}       column letter of that rows-sheet column
  {ctrl:key}       absolute Controls cell for that param (e.g. Controls!$C$5)
  {MT_RANGE}       absolute enum→toggle range for VLOOKUP scope
  {QR} / {COSTR}   absolute qualifies-column / cost-column ranges (results block)

Excel-open honesty: openpyxl output can fail to open in Excel-for-Mac, so the
file is normalized through LibreOffice (soffice) by default — preserves formulas,
caches values. If soffice is missing and normalize is requested, the build FAILS
(SystemExit 2): a shipped file that may not open is worse than a hard error.
"""
from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from . import model as M

# styles (Clickt-consistent; same palette as the reference workbook)
HDR = Font(bold=True, color="FFFFFF", size=11)
HDRFILL = PatternFill("solid", fgColor="1F4E78")
TITLE = Font(bold=True, size=14)
SECT = Font(bold=True, size=12, color="1F4E78")
BOLD = Font(bold=True)
ITAL = Font(italic=True)
GREY = Font(italic=True, color="808080")
INPUTF = Font(bold=True, size=12, color="1F4E78")
INPUTFILL = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="E2EFDA")
NBFILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MID = Side(style="medium", color="BF8F00")
INBORDER = Border(left=MID, right=MID, top=MID, bottom=MID)
CEN = Alignment(horizontal="center")
PCT, MONEY, NUM = "0.00%", "#,##0.00", "#,##0.00"
_FMT = {"PCT": PCT, "MONEY": MONEY, "NUM": NUM, "0.00": "0.00", "0": "0"}


def _fmt(name):
    return _FMT.get(name, name) if name else None


# --------------------------------------------------------------------------
def build_workbook(model: dict, spec: dict, brand: str, out_path: str) -> dict:
    x = spec["xlsx"]
    pr = model["provenance"]
    cur = pr.get("currency", "")
    rows = model["rows"]
    N = len(rows)
    LAST = N + 1

    cols = x["rows_columns"]
    headers = [c["header"] for c in cols]
    cl = {c["header"]: get_column_letter(i + 1) for i, c in enumerate(cols)}
    status_hdr = x["check"]["status_header"]
    qual_hdr = x["check"]["qualifies_header"]

    # control-cell map: param_key -> Controls!$C$<row>
    ctrl_cell = {p["key"]: f"{x['controls_sheet']}!$C${p['row']}" for p in x.get("params", [])}
    tog = x.get("toggles")
    mt_range = ""
    if tog:
        first = tog["start_row"]
        last = tog["start_row"] + len(tog["options"]) - 1
        mt_range = f"{x['controls_sheet']}!$B${first}:$C${last}"

    qcol, costcol = cl[qual_hdr], cl.get("Cost", "A")
    QR = f"'{x['rows_sheet']}'!${qcol}$2:${qcol}${LAST}"
    COSTR = f"'{x['rows_sheet']}'!${costcol}$2:${costcol}${LAST}"

    def resolve(tmpl: str, r0=None) -> str:
        s = tmpl
        if r0 is not None:
            s = s.replace("{row}", str(r0))
        for hdr, letter in cl.items():
            s = s.replace("{C:" + hdr + "}", letter)
        for key, ref in ctrl_cell.items():
            s = s.replace("{ctrl:" + key + "}", ref)
        s = s.replace("{MT_RANGE}", mt_range).replace("{QR}", QR).replace("{COSTR}", COSTR)
        # generic per-column range token: {R:Header} -> 'RowsSheet'!$<col>$2:$<col>$LAST
        for hdr, letter in cl.items():
            s = s.replace("{R:" + hdr + "}", f"'{x['rows_sheet']}'!${letter}$2:${letter}${LAST}")
        return s

    wb = Workbook()

    # ===== Controls sheet =====
    ws = wb.active
    ws.title = x["controls_sheet"]

    def put(cell, val, font=None, fmt=None, fill=None, border=None, align=None):
        c = ws[cell]
        c.value = val
        if font: c.font = font
        if fmt: c.number_format = _fmt(fmt)
        if fill: c.fill = fill
        if border: c.border = border
        if align: c.alignment = align

    client = pr.get("client_name") or brand or "Account"
    put("A1", x["title"](pr, brand) if callable(x.get("title")) else f"{spec['title']} — {client}", TITLE)
    if x.get("subtitle"):
        put("A2", x["subtitle"](pr), ITAL)
    if x.get("intro"):
        put("A3", x["intro"], ITAL)

    if x.get("params"):
        put(f"A{x.get('params_title_row', 4)}", x.get("params_title", "1 · FILTER PARAMETERS"), SECT)
        for p in x["params"]:
            r = p["row"]
            put(f"A{r}", p["label"], BOLD)
            put(f"C{r}", model["params"].get(p["key"], p.get("default")), INPUTF, p.get("fmt"),
                INPUTFILL, INBORDER, CEN)
            if p.get("note"):
                put(f"D{r}", p["note"])

    if tog:
        put(f"A{tog['title_row']}", tog["section_title"], SECT)
        scope = set(model["params"].get(tog["param_key"], []))
        for i, (label, enum) in enumerate(tog["options"]):
            r = tog["start_row"] + i
            put(f"A{r}", label)
            put(f"B{r}", enum)
            put(f"C{r}", "yes" if enum in scope else "no", INPUTF, None, INPUTFILL, INBORDER, CEN)
        if tog.get("note"):
            put(tog["note"][0], tog["note"][1], GREY)

    if x.get("logic"):
        lg = x["logic"]
        put(f"A{lg['title_row']}", lg["title"], SECT)
        for blk in lg["blocks"]:
            put(f"A{blk['head_row']}", blk["head"], BOLD)
            for r, formula in blk["rows"]:
                ws.merge_cells(f"A{r}:H{r}")
                ws.cell(row=r, column=1, value=resolve(formula))

    if x.get("results"):
        rs = x["results"]
        put(f"A{rs['title_row']}", rs["title"], SECT)
        for it in rs["items"]:
            put(f"A{it['row']}", it["label"], BOLD if not it.get("muted") else GREY)
            if "value_key" in it:
                val = model["summary"].get(it["value_key"])
            else:
                f = it["formula"]
                val = resolve(f) if str(f).startswith(("=", "{")) else f
            put(it["cell"], val, INPUTF if not it.get("muted") else GREY,
                it.get("fmt", "0"), align=CEN)

    for aux in x.get("aux", []):
        put(f"A{aux['title_row']}", aux["title"], SECT)
        for j, (hdr, _k, _f) in enumerate(aux["columns"], start=1):
            c = ws.cell(row=aux["header_row"], column=j, value=hdr)
            c.font = HDR; c.fill = HDRFILL; c.border = BORDER; c.alignment = CEN
        br = aux["start_row"]
        source = model.get(aux["source"], [])
        if aux.get("sort_key"):
            source = sorted(source, key=lambda b: b.get(aux["sort_key"], 0), reverse=True)
        for item in source:
            for j, (_hdr, key, fmt) in enumerate(aux["columns"], start=1):
                v = item.get(key)
                if isinstance(v, float):
                    v = round(v, 4)
                c = ws.cell(row=br, column=j, value=(v if v is not None else "n/a"))
                if fmt:
                    c.number_format = _fmt(fmt)
            br += 1

    # data validations
    for p in x.get("params", []):
        if p.get("dropdown"):
            dv = DataValidation(type="list", formula1=f'"{p["dropdown"]}"', allow_blank=False)
            ws.add_data_validation(dv); dv.add(ws[f"C{p['row']}"])
    if tog and tog.get("dropdown"):
        dv = DataValidation(type="list", formula1=f'"{tog["dropdown"]}"', allow_blank=False)
        ws.add_data_validation(dv)
        for i in range(len(tog["options"])):
            dv.add(ws[f"C{tog['start_row'] + i}"])
    for col, w in x.get("controls_widths", {"A": 36, "B": 13, "C": 14, "D": 22}).items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False

    # ===== Rows sheet (every row + Status) =====
    ws = wb.create_sheet(x["rows_sheet"])
    ws.append(headers)
    scored_status = x.get("scored_status", "scored")
    for i, r in enumerate(rows):
        r0 = i + 2
        scored = r["status"] == scored_status
        for c in cols:
            col_letter = cl[c["header"]]
            cell = ws[f"{col_letter}{r0}"]
            if c["kind"] == "data":
                if c["key"] == "__status__":
                    cell.value = r["status"].replace("_", " ")
                    if not scored:
                        cell.fill = NBFILL
                else:
                    cell.value = r.get(c["key"])
                    if c.get("fmt"):
                        cell.number_format = _fmt(c["fmt"])
            else:  # formula
                if c.get("scored") and not scored:
                    continue  # no-benchmark rows never classified
                cell.value = resolve(c["formula"], r0)
                if c.get("fmt"):
                    cell.number_format = _fmt(c["fmt"])
    # header style + freeze + autofilter + qualifying highlight
    for cnum in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=cnum)
        c.font = HDR; c.fill = HDRFILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.freeze_panes = x.get("rows_freeze", "C2")
    if N:
        ws.auto_filter.ref = f"A1:{cl[headers[-1]]}{LAST}"
        ws.conditional_formatting.add(f"A2:{cl[headers[-1]]}{LAST}",
                                      FormulaRule(formula=[f'${qcol}2<>""'], fill=GREEN))
    for c in cols:
        if c.get("width"):
            ws.column_dimensions[cl[c["header"]]].width = c["width"]

    # ===== Snapshot sheet (static sections) =====
    sect_fn = x.get("snapshot_sections") or spec.get("md_sections")
    if sect_fn:
        ws = wb.create_sheet(x["snapshot_sheet"])
        ws["A1"] = x.get("snapshot_title", "Snapshot"); ws["A1"].font = TITLE
        if x.get("snapshot_intro"):
            ws["A2"] = x["snapshot_intro"]; ws["A2"].font = ITAL
        r = 4
        for sec in sect_fn(model):
            ws.cell(row=r, column=1, value=sec["title"]).font = SECT
            r += 1
            if sec.get("note"):
                ws.cell(row=r, column=1, value=sec["note"]).font = GREY
                r += 1
            if not sec.get("rows"):
                ws.cell(row=r, column=1, value=sec.get("empty", "None.")).font = ITAL
                r += 2
                continue
            for j, h in enumerate(sec["headers"], start=1):
                c = ws.cell(row=r, column=j, value=h)
                c.font = HDR; c.fill = HDRFILL; c.border = BORDER; c.alignment = CEN
            r += 1
            for row in sec["rows"]:
                for j, val in enumerate(row, start=1):
                    ws.cell(row=r, column=j, value=val)
                r += 1
            r += 1
        for col, w in x.get("snapshot_widths", {"A": 44, "B": 30, "C": 14, "D": 16, "E": 22, "F": 8}).items():
            ws.column_dimensions[col].width = w

    try:
        wb.save(out_path)
    except Exception as e:  # pragma: no cover
        sys.stderr.write(f"ERROR: failed to write workbook: {e}\n"); sys.exit(2)
    return model["summary"]


# --------------------------------------------------------------------------
def find_soffice():
    for c in [shutil.which("soffice"), "/Applications/LibreOffice.app/Contents/MacOS/soffice",
              "/opt/homebrew/bin/soffice", "/usr/bin/soffice"]:
        if c and os.path.exists(c):
            return c
    return None


def normalize_with_libreoffice(path: str) -> None:
    soffice = find_soffice()
    if not soffice:
        sys.stderr.write(
            "ERROR: normalize requested but LibreOffice (soffice) was not found.\n"
            "Install LibreOffice (macOS: brew install --cask libreoffice) so the .xlsx opens\n"
            "reliably in Excel, or rebuild with normalize disabled to ship the un-normalized file.\n")
        sys.exit(2)
    with tempfile.TemporaryDirectory() as td:
        try:
            subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", td, path],
                           check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            sys.stderr.write(f"ERROR: LibreOffice normalization failed: {e}\n"); sys.exit(2)
        produced = Path(td) / (Path(path).stem + ".xlsx")
        if not produced.exists():
            sys.stderr.write("ERROR: LibreOffice produced no output.\n"); sys.exit(2)
        shutil.copyfile(produced, path)


def build_xlsx(model: dict, spec: dict, out_path: str, brand: str = "", normalize: bool = True) -> dict:
    M.require_model(model)
    if not spec.get("xlsx"):
        raise ValueError("spec declares no 'xlsx' layout")
    summary = build_workbook(model, spec, brand, out_path)
    if normalize:
        normalize_with_libreoffice(out_path)
    return summary


# --------------------------------------------------------------------------
def check_workbook(path: str, spec: dict) -> int:
    x = spec["xlsx"]
    chk = x["check"]
    try:
        wb = load_workbook(path)
        wb_v = load_workbook(path, data_only=True)
    except Exception as e:
        sys.stderr.write(f"ERROR: could not open workbook: {e}\n"); return 1
    problems = []
    for name in x["sheets"]:
        if name not in wb.sheetnames:
            problems.append(f"missing '{name}' sheet")
    cs = x["controls_sheet"]
    if cs in wb.sheetnames:
        c = wb[cs]
        for cell in chk.get("param_cells", []):
            if c[cell].value is None:
                problems.append(f"{cs}!{cell} (parameter input) is empty")
        cached = chk.get("cached_cell")
        if cached:
            if not str(c[cached].value or "").startswith("="):
                problems.append(f"{cs}!{cached} is not a formula")
            if wb_v[cs][cached].value is None:
                problems.append(f"{cs}!{cached} has no cached value — the file was not normalized "
                                "(it may not open in Excel). Rebuild with normalization enabled.")
    rs = x["rows_sheet"]
    if rs in wb.sheetnames:
        hdr = [v.value for v in wb[rs][1]]
        if hdr and hdr[-1] != chk["qualifies_header"]:
            problems.append(f"{rs} last column is not '{chk['qualifies_header']}'")
        if chk["status_header"] not in hdr:
            problems.append(f"{rs} is missing the '{chk['status_header']}' column (no-row-loss contract)")
    if problems:
        for pb in problems:
            sys.stderr.write(f"  - {pb}\n")
        sys.stderr.write(f"ERROR: workbook failed structural check ({len(problems)} issue(s))\n")
        return 1
    print(f"OK: '{path}' — {', '.join(x['sheets'])}; parameters present, formulas cached (normalized).")
    return 0

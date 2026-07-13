"""
Clickt brand system — shared openpyxl helpers for both calculator plugins.

Single source of truth for colour tokens, font choices, and the set of named
styles each .xlsx export uses. Keep this in sync with
clickt-theme/style.css :root (brand v1.0).

Usage:

    from openpyxl import Workbook
    from _brand.brand import register_named_styles, BRAND, fmt_currency_usd, ...

    wb = Workbook()
    register_named_styles(wb)
    ws = wb.active
    ws["A1"] = "Heading"
    ws["A1"].style = "clickt-h1"

The named styles registered here are referenced by name across both plugins.
"""

from __future__ import annotations


def _load_styles() -> None:
    """Lazily import openpyxl.styles into module globals.

    Keeps `import brand` dependency-free (the BRAND token dict and color helpers
    are stdlib) so the md/html renderers can import this module without openpyxl.
    Called by register_named_styles before any NamedStyle is built.
    """
    global Alignment, Border, Font, NamedStyle, PatternFill, Side
    from openpyxl.styles import (  # noqa: PLC0415
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
    )


# ─── Colour tokens (brand v1.0) ───────────────────────────────────────────────
# Hex strings without the leading "#"; openpyxl expects ARGB or RGB hex with
# no leading hash. Keep the alpha channel = FF (fully opaque).
BRAND = {
    # Primary
    "ink":          "FF0B0F0E",
    "bone":         "FFF4EFE6",
    "yellow":       "FFF3B61C",
    "teal":         "FF1F7A82",
    # Extended
    "yellow_soft":  "FFF8C84A",
    "yellow_deep":  "FFD99A0A",
    "teal_deep":    "FF0F4A52",
    "teal_pale":    "FF7FB8B0",
    "sage":         "FF5BA89A",
    # Mode-mapped neutrals (paper / dark)
    "paper":        "FFFFFFFF",          # card surface on light ground
    "paper_warm":   "FFECE6D9",          # alt warm cream (bg-2 token)
    "fg_dim":       "FF5C6361",          # muted body text on light ground
    "line":         "FFD6D2C4",          # warm bone-tinted divider
    "muted_line":   "FFE8E1D2",          # softer divider
    # States
    "red":          "FFB33A28",
    "green":        "FF2D7A4A",
    "amber":        "FFB8861B",
}


# Fonts. Workbooks ship with platform-default fallbacks if the chosen face is
# missing on the viewer's machine — openpyxl simply writes the name into the
# .xlsx; rendering is the spreadsheet app's job.
FONT_SERIF = "Fraunces"      # display: H1, hero values
FONT_SANS = "Inter"          # body
FONT_MONO = "JetBrains Mono"  # numbers, eyebrows, code-like labels


# ─── Number-format strings ────────────────────────────────────────────────────
# We store the raw float in the cell and apply display formatting via these
# format codes (per OPEN QUESTION #3 — store raw, format on top).
FMT_CURRENCY_USD = '"$"#,##0;[Red]"−$"#,##0'                # whole-dollar
FMT_CURRENCY_USD_2 = '"$"#,##0.00;[Red]"−$"#,##0.00'        # 2-decimal
FMT_PERCENT = "0.0%"                                          # 1-decimal pct
FMT_PERCENT_0 = "0%"                                          # 0-decimal pct
FMT_MULTIPLE = '0.00"×"'                                      # ROAS, ordersAt
FMT_INT = "#,##0"                                             # counts
FMT_MONTHS = '0" mo"'                                         # months


# ─── Borders ──────────────────────────────────────────────────────────────────
def _side(color: str = BRAND["line"], style: str = "thin") -> Side:
    return Side(border_style=style, color=color)


def _border(
    top: str | None = None,
    bottom: str | None = None,
    left: str | None = None,
    right: str | None = None,
) -> Border:
    return Border(
        top=_side(top) if top else Side(border_style=None),
        bottom=_side(bottom) if bottom else Side(border_style=None),
        left=_side(left) if left else Side(border_style=None),
        right=_side(right) if right else Side(border_style=None),
    )


# ─── Named styles ─────────────────────────────────────────────────────────────
def _styles() -> list[NamedStyle]:
    """Build the full set of NamedStyles. Called by register_named_styles."""

    styles: list[NamedStyle] = []

    # ─── Section titles ───
    # H1 — yellow band, ink Fraunces, used as the sheet title row.
    h1 = NamedStyle(name="clickt-h1")
    h1.font = Font(name=FONT_SERIF, size=20, bold=False, color=BRAND["ink"])
    h1.fill = PatternFill("solid", fgColor=BRAND["yellow"])
    h1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h1.border = _border(bottom=BRAND["ink"])
    styles.append(h1)

    # H2 — section heading, ink Fraunces 700, bone band.
    h2 = NamedStyle(name="clickt-h2")
    h2.font = Font(name=FONT_SERIF, size=14, bold=True, color=BRAND["ink"])
    h2.fill = PatternFill("solid", fgColor=BRAND["bone"])
    h2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h2.border = _border(bottom=BRAND["ink"])
    styles.append(h2)

    # Eyebrow — mono, dim, uppercase letter-tracked.
    eyebrow = NamedStyle(name="clickt-eyebrow")
    eyebrow.font = Font(name=FONT_MONO, size=9, color=BRAND["fg_dim"])
    eyebrow.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    styles.append(eyebrow)

    # ─── Table headers ───
    # Column header on a data table.
    th = NamedStyle(name="clickt-th")
    th.font = Font(name=FONT_MONO, size=9, bold=True, color=BRAND["fg_dim"])
    th.fill = PatternFill("solid", fgColor=BRAND["paper_warm"])
    th.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    th.border = _border(bottom=BRAND["ink"])
    styles.append(th)

    th_right = NamedStyle(name="clickt-th-right")
    th_right.font = Font(name=FONT_MONO, size=9, bold=True, color=BRAND["fg_dim"])
    th_right.fill = PatternFill("solid", fgColor=BRAND["paper_warm"])
    th_right.alignment = Alignment(horizontal="right", vertical="center", indent=1, wrap_text=True)
    th_right.border = _border(bottom=BRAND["ink"])
    styles.append(th_right)

    # ─── Body cells ───
    body = NamedStyle(name="clickt-body")
    body.font = Font(name=FONT_SANS, size=11, color=BRAND["ink"])
    body.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    body.border = _border(bottom=BRAND["muted_line"])
    styles.append(body)

    body_mono = NamedStyle(name="clickt-body-mono")
    body_mono.font = Font(name=FONT_MONO, size=11, color=BRAND["ink"])
    body_mono.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    body_mono.border = _border(bottom=BRAND["muted_line"])
    styles.append(body_mono)

    body_dim = NamedStyle(name="clickt-body-dim")
    body_dim.font = Font(name=FONT_SANS, size=10, color=BRAND["fg_dim"])
    body_dim.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    body_dim.border = _border(bottom=BRAND["muted_line"])
    styles.append(body_dim)

    # ─── Number cells (right-aligned, mono, with format codes applied at write time) ───
    num_currency = NamedStyle(name="clickt-num-currency")
    num_currency.font = Font(name=FONT_MONO, size=11, color=BRAND["ink"])
    num_currency.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    num_currency.number_format = FMT_CURRENCY_USD
    num_currency.border = _border(bottom=BRAND["muted_line"])
    styles.append(num_currency)

    num_currency_2 = NamedStyle(name="clickt-num-currency-2")
    num_currency_2.font = Font(name=FONT_MONO, size=11, color=BRAND["ink"])
    num_currency_2.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    num_currency_2.number_format = FMT_CURRENCY_USD_2
    num_currency_2.border = _border(bottom=BRAND["muted_line"])
    styles.append(num_currency_2)

    num_pct = NamedStyle(name="clickt-num-pct")
    num_pct.font = Font(name=FONT_MONO, size=11, color=BRAND["ink"])
    num_pct.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    num_pct.number_format = FMT_PERCENT
    num_pct.border = _border(bottom=BRAND["muted_line"])
    styles.append(num_pct)

    num_multiple = NamedStyle(name="clickt-num-multiple")
    num_multiple.font = Font(name=FONT_MONO, size=11, color=BRAND["ink"])
    num_multiple.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    num_multiple.number_format = FMT_MULTIPLE
    num_multiple.border = _border(bottom=BRAND["muted_line"])
    styles.append(num_multiple)

    num_int = NamedStyle(name="clickt-num-int")
    num_int.font = Font(name=FONT_MONO, size=11, color=BRAND["ink"])
    num_int.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    num_int.number_format = FMT_INT
    num_int.border = _border(bottom=BRAND["muted_line"])
    styles.append(num_int)

    num_months = NamedStyle(name="clickt-num-months")
    num_months.font = Font(name=FONT_MONO, size=11, color=BRAND["ink"])
    num_months.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    num_months.number_format = FMT_MONTHS
    num_months.border = _border(bottom=BRAND["muted_line"])
    styles.append(num_months)

    # ─── Hero values — bigger, Fraunscean ───
    hero_currency = NamedStyle(name="clickt-hero-currency")
    hero_currency.font = Font(name=FONT_SERIF, size=24, bold=False, color=BRAND["ink"])
    hero_currency.fill = PatternFill("solid", fgColor=BRAND["bone"])
    hero_currency.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    hero_currency.number_format = FMT_CURRENCY_USD
    hero_currency.border = _border(top=BRAND["ink"], bottom=BRAND["ink"])
    styles.append(hero_currency)

    hero_label = NamedStyle(name="clickt-hero-label")
    hero_label.font = Font(name=FONT_MONO, size=10, bold=True, color=BRAND["fg_dim"])
    hero_label.fill = PatternFill("solid", fgColor=BRAND["bone"])
    hero_label.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    hero_label.border = _border(top=BRAND["ink"], bottom=BRAND["ink"])
    styles.append(hero_label)

    # ─── Benchmark-band pill cells (label-only — colour comes from accent fills) ───
    band_strong = NamedStyle(name="clickt-band-strong")
    band_strong.font = Font(name=FONT_MONO, size=10, bold=True, color="FF1F4F2F")
    band_strong.fill = PatternFill("solid", fgColor="FFD9EBDF")
    band_strong.alignment = Alignment(horizontal="center", vertical="center")
    band_strong.border = _border(bottom=BRAND["muted_line"])
    styles.append(band_strong)

    band_healthy = NamedStyle(name="clickt-band-healthy")
    band_healthy.font = Font(name=FONT_MONO, size=10, bold=True, color="FF1F4F2F")
    band_healthy.fill = PatternFill("solid", fgColor="FFE8F2EC")
    band_healthy.alignment = Alignment(horizontal="center", vertical="center")
    band_healthy.border = _border(bottom=BRAND["muted_line"])
    styles.append(band_healthy)

    band_amber = NamedStyle(name="clickt-band-amber")
    band_amber.font = Font(name=FONT_MONO, size=10, bold=True, color="FF6B4A0E")
    band_amber.fill = PatternFill("solid", fgColor="FFFBE9C4")
    band_amber.alignment = Alignment(horizontal="center", vertical="center")
    band_amber.border = _border(bottom=BRAND["muted_line"])
    styles.append(band_amber)

    band_red = NamedStyle(name="clickt-band-red")
    band_red.font = Font(name=FONT_MONO, size=10, bold=True, color="FF7C2718")
    band_red.fill = PatternFill("solid", fgColor="FFF3D9D2")
    band_red.alignment = Alignment(horizontal="center", vertical="center")
    band_red.border = _border(bottom=BRAND["muted_line"])
    styles.append(band_red)

    band_none = NamedStyle(name="clickt-band-none")
    band_none.font = Font(name=FONT_MONO, size=10, color=BRAND["fg_dim"])
    band_none.alignment = Alignment(horizontal="center", vertical="center")
    band_none.border = _border(bottom=BRAND["muted_line"])
    styles.append(band_none)

    # ─── Footer / methodology body ───
    note = NamedStyle(name="clickt-note")
    note.font = Font(name=FONT_SANS, size=10, italic=False, color=BRAND["fg_dim"])
    note.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    styles.append(note)

    # ─── Brand byline ───
    byline = NamedStyle(name="clickt-byline")
    byline.font = Font(name=FONT_MONO, size=9, color=BRAND["fg_dim"])
    byline.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    styles.append(byline)

    return styles


def register_named_styles(wb) -> None:
    """Register the full Clickt named-style set on a workbook.

    Idempotent — safely skips styles that have already been registered (which
    happens when the same workbook is touched twice in a test run).
    """
    _load_styles()
    existing = {s.name for s in wb._named_styles}  # noqa: SLF001
    for style in _styles():
        if style.name in existing:
            continue
        wb.add_named_style(style)


# ─── Benchmark band → style name mapping ──────────────────────────────────────
# The classify() functions in the JSX return colour tokens; we map them to
# named styles registered above.
BAND_STYLE = {
    "Strong":   "clickt-band-strong",
    "Healthy":  "clickt-band-healthy",
    "Weak":     "clickt-band-amber",
    "Thin":     "clickt-band-amber",
    "Watch":    "clickt-band-amber",
    "Tight":    "clickt-band-amber",
    "High":     "clickt-band-red",
    "Bleeding": "clickt-band-red",
    "Loss":     "clickt-band-red",
    "Negative": "clickt-band-red",
    "—":        "clickt-band-none",
    "":         "clickt-band-none",
}


def band_style_for(label: str) -> str:
    """Return the named-style name for a classify() label string."""
    return BAND_STYLE.get(label, "clickt-band-none")

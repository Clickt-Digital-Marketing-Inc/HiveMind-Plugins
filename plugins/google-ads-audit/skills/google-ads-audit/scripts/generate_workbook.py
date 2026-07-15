#!/usr/bin/env python3
"""Generate a Google Ads audit workbook (.xlsx) from a findings JSON file.

The workbook follows a comprehensive 9-step audit framework plus modern
Google-specific deep checks. It is formula-driven: the agent supplies
PASS/FLAG/FAIL/N/A results and severities per check, and the workbook
computes the weighted Health Score, ICE scores, and the client summary
itself. This keeps the deliverable auditable and editable in Excel/Sheets.

Usage:
    # Build a workbook from findings JSON
    python3 generate_workbook.py --input findings.json \\
        --output ads-audit-acme-2026-06-24.xlsx --brand "Acme Corp"

    # Structurally validate an existing workbook (CI / quality gate)
    python3 generate_workbook.py --check --input ads-audit-acme-2026-06-24.xlsx

Findings JSON schema (see SKILL.md for the authoritative copy):
{
  "meta": {"client_name","account_id","currency","timezone",
           "business_model" ("Lead Gen"|"Ecommerce"),
           "date_range","search_terms_range","auditor","audit_date"},
  "data_inventory": [{"pull","resource","rows","status","notes"}],
  "kpis":           [{"metric","value","unit","benchmark","flag","notes"}],
  "sections":       [{"tab","title","checks":[
                        {"id","name","verify","applies_to",
                         "severity"(Critical|High|Medium|Low),
                         "result"(PASS|FLAG|FAIL|N/A),
                         "observed","recommendation"}]}],
  "findings":       [{"id","section","title","severity",
                      "recommendation","effort","horizon"(30|60|90),
                      "owner"}]
}

Exit codes: 0 success, 1 usage/validation error, 2 build error.
"""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover - environment guard
    sys.stderr.write(
        "ERROR: openpyxl is required. Install with:\n"
        "    python3 -m pip install --user openpyxl\n"
    )
    sys.exit(2)

# Scoring constants + the analysis-tab list are single-sourced in audit_model so the
# xlsx formulas, the markdown, and the HTML explorer can never disagree. The formula
# builders below INTERPOLATE these values rather than restating them, which is what
# makes that claim structurally true instead of merely intended.
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from audit_model import (  # noqa: E402
    SEVERITY_WEIGHTS, FLAG_SCORES, SEVERITY_IMPACT, GRADE_CUTOFFS, ANALYSIS_TABS,
    DEFAULT_SEVERITY, DEFAULT_IMPACT, normalize_findings,
)

# --------------------------------------------------------------------------
# Palette — Clickt org-wide design system (teal / lime / purple + ember signal)
# --------------------------------------------------------------------------
TAB_COLORS = {
    "setup": "FF1F7A82",        # HiveMind teal
    "analysis": "FFB4E01F",     # Performance lime
    "post_audit": "FF0F4A52",   # teal deep
    "deliverable": "FFF86B3C",  # ember (signal)
    "reference": "FF5C6470",    # slate
}
FILL_HEADER = PatternFill("solid", fgColor="FF07262B")   # abyss (teal ramp)
FILL_SECTION = PatternFill("solid", fgColor="FFF3F4F6")  # cloud
FILL_HILITE = PatternFill("solid", fgColor="FFEEF7D2")   # pale lime (input cells)

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFFFF")
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="FF0B0F0E")
FONT_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
FONT_BODY = Font(name="Calibri", size=10)
FONT_MUTED = Font(name="Calibri", size=9, italic=True, color="FF5C6470")
FMT_MONEY = "#,##0.00"
FMT_PCT = "0.0%"

THIN = Side(style="thin", color="FFCBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

# Scoring constants and ANALYSIS_TABS are imported from audit_model above.
ALL_TABS = (
    ["00_README", "01_Audit_Scope", "02_Data_Inventory"]
    + [name for name, _ in ANALYSIS_TABS]
    + ["12_Findings_Log", "13_ICE_Prioritization", "14_Action_Roadmap",
       "15_Client_Report", "16_Benchmarks", "17_Health_Score"]
)

# Check-table column layout (1-indexed). A..L
CHK_COLS = ["Check", "What to verify", "Applies to", "Severity", "Result",
            "Observed / evidence", "Recommendation",
            "_sev_w", "_flag", "_earned", "_possible"]
# Generated from the kernel's vocabularies, not restated: a dropdown that offers a
# different set than the (generated) formulas score is the same drift the formula
# builders were fixed for. N/A is a valid result but carries no FLAG_SCORES weight.
RESULTS_DV = '"' + ",".join(list(FLAG_SCORES) + ["N/A"]) + '"'
SEVERITY_DV = '"' + ",".join(SEVERITY_WEIGHTS) + '"'
MODEL_DV = '"Lead Gen,Ecommerce"'  # no kernel constant — business models are not scored
HORIZON_DV = '"30,60,90"'


# --------------------------------------------------------------------------
# Small styling helpers
# --------------------------------------------------------------------------
def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = FONT_TITLE
    c.fill = FILL_HEADER
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 26


def section(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_SECTION
    c.fill = FILL_SECTION
    return row + 1


def header_row(ws, row, labels, start_col=1):
    for j, label in enumerate(labels, start=start_col):
        c = ws.cell(row=row, column=j, value=label)
        c.font = FONT_HEAD
        c.fill = FILL_HEADER
        c.alignment = WRAP_TOP
        c.border = BORDER
    return row + 1


def kv(ws, row, label, value, value_is_formula=False):
    a = ws.cell(row=row, column=1, value=label)
    a.font = FONT_SECTION
    b = ws.cell(row=row, column=2, value=value)
    b.font = FONT_BODY
    b.alignment = WRAP_TOP
    return row + 1


# --------------------------------------------------------------------------
# Tab builders
# --------------------------------------------------------------------------
def build_readme(wb, meta):
    ws = wb.create_sheet("00_README")
    ws.sheet_properties.tabColor = TAB_COLORS["setup"]
    set_widths(ws, [26, 96])
    title(ws, f"Google Ads Audit — {meta.get('client_name', 'Client')}", 2)
    r = 3
    r = section(ws, r, "Purpose", 2)
    ws.cell(row=r, column=1, value="This workbook")
    ws.cell(row=r, column=2, value=(
        "A full Google Ads account audit built on a comprehensive 9-step framework "
        "plus modern Google-specific checks (PMax signals, Consent Mode v2, "
        "Enhanced Conversions, bidding-strategy fit, Demand Gen migration). "
        "Results are scored automatically on the 17_Health_Score tab.")
    ).alignment = WRAP_TOP
    r += 2
    r = section(ws, r, "How to read it", 2)
    for label, desc in [
        ("Setup tabs (blue)", "00-02: scope, the business-model switch, and the data that was pulled."),
        ("Analysis tabs (yellow)", "03-11: one tab per audit area. Set each Result to PASS / FLAG / FAIL / N/A."),
        ("Post-audit tabs (green)", "12-14: Findings Log, ICE prioritization, and the 30/60/90 roadmap."),
        ("Deliverable (red)", "15: the client-facing summary, driven by formulas."),
        ("Reference (grey)", "16-17: benchmarks/glossary and the weighted Health Score."),
    ]:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = FONT_SECTION
        ws.cell(row=r, column=2, value=desc).alignment = WRAP_TOP
        r += 1
    r += 1
    r = section(ws, r, "Workflow", 2)
    ws.cell(row=r, column=2, value=(
        "Set 01_Audit_Scope -> Business model. Fill Result + Observed on each "
        "analysis tab. Add issues to 12_Findings_Log, then fill Confidence and "
        "Ease on 13_ICE_Prioritization to rank them. Share 15_Client_Report.")
    ).alignment = WRAP_TOP
    ws.cell(row=r + 2, column=2, value="Generated by google-ads-audit skill.").font = FONT_MUTED


def build_scope(wb, meta):
    ws = wb.create_sheet("01_Audit_Scope")
    ws.sheet_properties.tabColor = TAB_COLORS["setup"]
    set_widths(ws, [26, 60])
    title(ws, "Audit Scope", 2)
    r = 3
    kv(ws, r, "Client", meta.get("client_name", ""))
    kv(ws, r + 1, "Account ID", meta.get("account_id", ""))
    # Row 5 holds the business-model master switch (named range below).
    model_cell = ws.cell(row=5, column=1, value="Business model")
    model_cell.font = FONT_SECTION
    bm = ws.cell(row=5, column=2, value=meta.get("business_model", "Lead Gen"))
    bm.font = FONT_BODY
    bm.fill = FILL_HILITE
    bm.border = BORDER
    dv = DataValidation(type="list", formula1=MODEL_DV, allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("B5")
    r = 6
    for label, key in [
        ("Currency", "currency"), ("Time zone", "timezone"),
        ("Date range (trend/structure)", "date_range"),
        ("Search-terms range", "search_terms_range"),
        ("Auditor", "auditor"), ("Audit date", "audit_date"),
    ]:
        kv(ws, r, label, meta.get(key, ""))
        r += 1
    ws.cell(row=r + 1, column=1, value="Named range 'business_model' -> B5; "
            "referenced by benchmark and report tabs.").font = FONT_MUTED
    # Define the named range used by other tabs.
    dn = DefinedName("business_model", attr_text="'01_Audit_Scope'!$B$5")
    wb.defined_names.add(dn)


def build_data_inventory(wb, inventory):
    ws = wb.create_sheet("02_Data_Inventory")
    ws.sheet_properties.tabColor = TAB_COLORS["setup"]
    set_widths(ws, [34, 26, 10, 12, 50])
    title(ws, "Data Inventory — GAQL pulls", 5)
    r = header_row(ws, 3, ["Pull", "GAQL resource", "Rows", "Status", "Notes"])
    if not inventory:
        ws.cell(row=r, column=1, value="(no pulls recorded)").font = FONT_MUTED
    for item in inventory:
        ws.cell(row=r, column=1, value=item.get("pull", "")).border = BORDER
        ws.cell(row=r, column=2, value=item.get("resource", "")).border = BORDER
        ws.cell(row=r, column=3, value=item.get("rows", "")).border = BORDER
        ws.cell(row=r, column=4, value=item.get("status", "")).border = BORDER
        nc = ws.cell(row=r, column=5, value=item.get("notes", ""))
        nc.border = BORDER
        nc.alignment = WRAP_TOP
        r += 1
    ws.freeze_panes = "A4"


def _num(v) -> str:
    """5.0 -> '5', 1.5 -> '1.5'. Keeps interpolated formulas free of float noise."""
    return str(int(v)) if float(v) == int(v) else str(v)


def _lookup_formula(cell: str, pairs, fallback: str) -> str:
    """=IF(cell="k1",v1,IF(cell="k2",v2,...,fallback)) generated from `pairs`.

    Generated, not restated: change a weight in audit_model and the workbook
    formula changes with it. Editing one side only — the drift these builders
    exist to prevent — is no longer expressible.
    """
    expr = fallback
    for key, val in reversed(list(pairs)):
        expr = f'IF({cell}="{key}",{val},{expr})'
    return f"={expr}"


# Check-table columns: A=Check B=verify C=applies D=Severity E=Result
# F=Observed G=Recommendation H=SevWeight I=Flag J=Earned K=Possible
def _severity_weight_formula(row):
    return _lookup_formula(f"D{row}",
                           [(s, _num(w)) for s, w in SEVERITY_WEIGHTS.items()], "0")


def _flag_formula(row):
    return _lookup_formula(f"E{row}",
                           [(r, _num(v)) for r, v in FLAG_SCORES.items()], '""')


def _grade_formula(cell: str) -> str:
    """=IF(NOT(ISNUMBER(B3)),"",IF(B3>=90,"A",...,"F")) from GRADE_CUTOFFS.

    Last GRADE_CUTOFFS pair is the catch-all. The ISNUMBER guard is load-bearing:
    when nothing is scoreable B3 is blank, and Excel ranks text above every number,
    so a bare `B3>=90` would grade an unscored audit "A".
    """
    *bands, (_, fallback) = GRADE_CUTOFFS
    expr = f'"{fallback}"'
    for cutoff, letter in reversed(bands):
        expr = f'IF({cell}>={_num(cutoff)},"{letter}",{expr})'
    return f'=IF(NOT(ISNUMBER({cell})),"",{expr})'


def _earned_formula(row):
    return f'=IF(OR(E{row}="N/A",E{row}=""),"",H{row}*I{row})'


def _possible_formula(row):
    return f'=IF(OR(E{row}="N/A",E{row}=""),"",H{row})'


def build_check_tab(wb, tab_name, tab_title, section_data, kpis=None):
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = TAB_COLORS["analysis"]
    # A Check | B verify | C applies | D (unused split) -> keep 7 visible + 4 helper
    set_widths(ws, [30, 40, 12, 11, 9, 38, 40, 8, 8, 9, 9])
    title(ws, tab_title, 7)
    r = 3

    # Optional KPI scorecard (Performance Review tab only).
    if kpis:
        r = section(ws, r, "KPI scorecard (informational)", 7)
        r = header_row(ws, r, ["Metric", "Value", "Unit", "Benchmark",
                               "Flag", "Notes", ""])
        for k in kpis:
            ws.cell(row=r, column=1, value=k.get("metric", "")).border = BORDER
            ws.cell(row=r, column=2, value=k.get("value", "")).border = BORDER
            ws.cell(row=r, column=3, value=k.get("unit", "")).border = BORDER
            ws.cell(row=r, column=4, value=k.get("benchmark", "")).border = BORDER
            fc = ws.cell(row=r, column=5, value=k.get("flag", ""))
            fc.border = BORDER
            fc.alignment = CENTER
            nc = ws.cell(row=r, column=6, value=k.get("notes", ""))
            nc.border = BORDER
            nc.alignment = WRAP_TOP
            r += 1
        r += 1

    r = section(ws, r, "Audit checks", 7)
    # Helper-column headers (hidden) included so the score math is transparent.
    header_row(ws, r, ["Check", "What to verify", "Applies to", "Severity",
                       "Result", "Observed / evidence", "Recommendation",
                       "SevW", "Flag", "Earn", "Poss"])
    head_row = r
    r += 1
    first = r
    checks = (section_data or {}).get("checks", [])
    if not checks:
        ws.cell(row=r, column=1, value="(no checks recorded for this area)").font = FONT_MUTED
        r += 1
    for chk in checks:
        name = chk.get("id", "")
        if chk.get("name"):
            name = f"{name} — {chk['name']}" if name else chk["name"]
        ws.cell(row=r, column=1, value=name).alignment = WRAP_TOP
        ws.cell(row=r, column=2, value=chk.get("verify", "")).alignment = WRAP_TOP
        ws.cell(row=r, column=3, value=chk.get("applies_to", "Both")).alignment = CENTER
        ws.cell(row=r, column=4, value=chk.get("severity", DEFAULT_SEVERITY)).alignment = CENTER
        rc = ws.cell(row=r, column=5, value=chk.get("result", ""))
        rc.alignment = CENTER
        rc.fill = FILL_HILITE
        ws.cell(row=r, column=6, value=chk.get("observed", "")).alignment = WRAP_TOP
        ws.cell(row=r, column=7, value=chk.get("recommendation", "")).alignment = WRAP_TOP
        ws.cell(row=r, column=8, value=_severity_weight_formula(r))
        ws.cell(row=r, column=9, value=_flag_formula(r))
        ws.cell(row=r, column=10, value=_earned_formula(r))
        ws.cell(row=r, column=11, value=_possible_formula(r))
        for col in range(1, 12):
            ws.cell(row=r, column=col).border = BORDER
        r += 1
    last = r - 1

    # Per-tab subtotals (used by 17_Health_Score). Placed at a stable label.
    r += 1
    ws.cell(row=r, column=1, value="Tab earned / possible:").font = FONT_SECTION
    if last >= first:
        ws.cell(row=r, column=10, value=f"=SUM(J{first}:J{last})")  # J = earned
        ws.cell(row=r, column=11, value=f"=SUM(K{first}:K{last})")  # K = possible
    else:
        ws.cell(row=r, column=10, value=0)
        ws.cell(row=r, column=11, value=0)
    ws._subtotal_row = r  # stash for health-score wiring

    # Data validation dropdowns on Result + Severity.
    if last >= first:
        dv_r = DataValidation(type="list", formula1=RESULTS_DV, allow_blank=True)
        dv_s = DataValidation(type="list", formula1=SEVERITY_DV, allow_blank=True)
        ws.add_data_validation(dv_r)
        ws.add_data_validation(dv_s)
        dv_r.add(f"E{first}:E{last}")
        dv_s.add(f"D{first}:D{last}")

    # Hide helper columns H..K.
    for col in ("H", "I", "J", "K"):
        ws.column_dimensions[col].hidden = True
    ws.freeze_panes = f"A{head_row + 1}"
    return ws


def build_findings_log(wb, findings):
    ws = wb.create_sheet("12_Findings_Log")
    ws.sheet_properties.tabColor = TAB_COLORS["post_audit"]
    set_widths(ws, [9, 24, 44, 12, 44, 10, 10, 16, 12])
    title(ws, "Findings Log — single source of truth", 9)
    r = header_row(ws, 3, ["ID", "Section", "Finding", "Severity",
                           "Recommendation", "Effort", "Horizon",
                           "Owner", "Status"])
    first = r
    if not findings:
        ws.cell(row=r, column=1, value="(no findings recorded)").font = FONT_MUTED
        r += 1
    for f in findings:
        ws.cell(row=r, column=1, value=f.get("id", "")).border = BORDER
        ws.cell(row=r, column=2, value=f.get("section", "")).border = BORDER
        ws.cell(row=r, column=3, value=f.get("title", "")).alignment = WRAP_TOP
        # DEFAULT_SEVERITY, not blank: 15_Client_Report COUNTIFs this column and
        # COUNTAs it for the total, so a blank here deletes the finding from the
        # client's report while the model still counts it.
        sc = ws.cell(row=r, column=4, value=f.get("severity", DEFAULT_SEVERITY))
        sc.alignment = CENTER
        ws.cell(row=r, column=5, value=f.get("recommendation", "")).alignment = WRAP_TOP
        ws.cell(row=r, column=6, value=f.get("effort", "")).alignment = CENTER
        ws.cell(row=r, column=7, value=f.get("horizon", "")).alignment = CENTER
        ws.cell(row=r, column=8, value=f.get("owner", "")).border = BORDER
        ws.cell(row=r, column=9, value=f.get("status", "Open")).border = BORDER
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = BORDER
        r += 1
    last = r - 1
    if last >= first:
        dv_sev = DataValidation(type="list", formula1=SEVERITY_DV, allow_blank=True)
        dv_h = DataValidation(type="list", formula1=HORIZON_DV, allow_blank=True)
        ws.add_data_validation(dv_sev)
        ws.add_data_validation(dv_h)
        dv_sev.add(f"D{first}:D{last}")
        dv_h.add(f"G{first}:G{last}")
    ws.freeze_panes = "A4"
    return first, last


def build_ice(wb, findings, log_first):
    ws = wb.create_sheet("13_ICE_Prioritization")
    ws.sheet_properties.tabColor = TAB_COLORS["post_audit"]
    set_widths(ws, [9, 46, 12, 10, 12, 9, 10, 10])
    title(ws, "ICE Prioritization (Impact x Confidence x Ease)", 8)
    ws.cell(row=2, column=1, value=("Impact is seeded from severity. Fill "
            "Confidence and Ease (1-10); ICE computes automatically. "
            "Sort by ICE descending to rank quick wins.")).font = FONT_MUTED
    r = header_row(ws, 4, ["ID", "Finding", "Severity", "Impact",
                           "Confidence", "Ease", "ICE", "Rank"])
    first = r
    for i, f in enumerate(findings):
        log_row = log_first + i
        ws.cell(row=r, column=1, value=f"='12_Findings_Log'!A{log_row}").border = BORDER
        ws.cell(row=r, column=2, value=f"='12_Findings_Log'!C{log_row}").alignment = WRAP_TOP
        ws.cell(row=r, column=3, value=f"='12_Findings_Log'!D{log_row}").alignment = CENTER
        # Same default as audit_model's impact seed — a mismatch here put one finding
        # at the bottom of the HTML's ICE list and mid-table in this one.
        imp = ws.cell(row=r, column=4, value=SEVERITY_IMPACT.get(
            f.get("severity", DEFAULT_SEVERITY), DEFAULT_IMPACT))
        imp.alignment = CENTER
        cc = ws.cell(row=r, column=5)
        cc.fill = FILL_HILITE
        ec = ws.cell(row=r, column=6)
        ec.fill = FILL_HILITE
        cc.alignment = CENTER
        ec.alignment = CENTER
        ws.cell(row=r, column=7, value=f'=IF(OR(E{r}="",F{r}=""),"",D{r}*E{r}*F{r})').alignment = CENTER
        ws.cell(row=r, column=8, value=f'=IF(G{r}="","",RANK(G{r},$G${first}:$G${first + max(len(findings) - 1, 0)}))').alignment = CENTER
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
        r += 1
    if not findings:
        ws.cell(row=r, column=1, value="(no findings)").font = FONT_MUTED
        r += 1
    last = r - 1
    if last >= first:
        dv = DataValidation(type="whole", operator="between", formula1="1",
                            formula2="10", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"E{first}:F{last}")
    ws.freeze_panes = "A5"


def build_roadmap(wb, findings):
    ws = wb.create_sheet("14_Action_Roadmap")
    ws.sheet_properties.tabColor = TAB_COLORS["post_audit"]
    set_widths(ws, [9, 50, 12, 18, 18, 14])
    title(ws, "Action Roadmap — 30 / 60 / 90", 6)
    r = 3
    buckets = [("30", "Next 30 days — quick wins & critical fixes"),
               ("60", "31-60 days — structural improvements"),
               ("90", "61-90 days — strategic / testing")]
    for horizon, label in buckets:
        r = section(ws, r, label, 6)
        r = header_row(ws, r, ["ID", "Action", "Severity", "Owner",
                               "Deadline", "Milestone"])
        rows = [f for f in findings if str(f.get("horizon", "")) == horizon]
        if not rows:
            ws.cell(row=r, column=1, value="(none assigned)").font = FONT_MUTED
            r += 1
        for f in rows:
            ws.cell(row=r, column=1, value=f.get("id", "")).border = BORDER
            ws.cell(row=r, column=2, value=f.get("title", "")).alignment = WRAP_TOP
            ws.cell(row=r, column=3, value=f.get("severity", "")).alignment = CENTER
            ws.cell(row=r, column=4, value=f.get("owner", "")).border = BORDER
            ws.cell(row=r, column=5, value="").border = BORDER
            ws.cell(row=r, column=6, value="").border = BORDER
            for col in range(1, 7):
                ws.cell(row=r, column=col).border = BORDER
            r += 1
        r += 1


def build_client_report(wb, meta, log_first, log_last):
    ws = wb.create_sheet("15_Client_Report")
    ws.sheet_properties.tabColor = TAB_COLORS["deliverable"]
    set_widths(ws, [30, 60])
    title(ws, f"Audit Summary — {meta.get('client_name', 'Client')}", 2)
    r = 3
    kv(ws, r, "Account", meta.get("account_id", ""))
    kv(ws, r + 1, "Period", meta.get("date_range", ""))
    kv(ws, r + 2, "Business model", "=business_model")
    r += 4
    r = section(ws, r, "Overall health", 2)
    ws.cell(row=r, column=1, value="Health Score").font = FONT_SECTION
    ws.cell(row=r, column=2, value="='17_Health_Score'!B3")
    r += 1
    ws.cell(row=r, column=1, value="Grade").font = FONT_SECTION
    ws.cell(row=r, column=2, value="='17_Health_Score'!B4")
    r += 2
    r = section(ws, r, "Findings by severity", 2)
    rng = f"'12_Findings_Log'!$D${log_first}:$D${log_last}"
    # Rows generated from SEVERITY_WEIGHTS — a restated list would silently omit a
    # tier the (generated) scoring formula prices, and the COUNTIFs would stop
    # summing to the total with nothing to show for it.
    for sev in SEVERITY_WEIGHTS:
        ws.cell(row=r, column=1, value=sev).font = FONT_BODY
        ws.cell(row=r, column=2, value=f'=COUNTIF({rng},"{sev}")')
        r += 1
    ws.cell(row=r, column=1, value="Total findings").font = FONT_SECTION
    # Count the findings by their ID, not by a field that happens to be populated.
    ws.cell(row=r, column=2,
            value=f"=COUNTA('12_Findings_Log'!$A${log_first}:$A${log_last})")
    r += 2
    r = section(ws, r, "Top priorities", 2)
    ws.cell(row=r, column=1, value=("See 13_ICE_Prioritization sorted by ICE "
            "for the ranked action list.")).font = FONT_MUTED


def build_benchmarks(wb, benchmarks):
    ws = wb.create_sheet("16_Benchmarks")
    ws.sheet_properties.tabColor = TAB_COLORS["reference"]
    set_widths(ws, [28, 22, 22, 26, 40])
    title(ws, "Reference Benchmarks & Glossary", 5)
    r = 3
    default_rows = [
        ["Metric", "DTC Ecommerce", "B2C Lead Gen", "B2B / SaaS", "Notes"],
        ["Search CTR", "4-6%", "4-7%", "2-5%", "Brand >> non-brand"],
        ["Avg CPC", "$0.50-$2", "$2-$8", "$4-$15", "Varies by vertical"],
        ["Conversion rate", "1.5-3%", "3-8%", "2-5%", "Landing-page dependent"],
        ["Search Impr. Share", ">65%", ">65%", ">60%", "Below = budget/rank loss"],
        ["Lost IS (Budget)", "<10%", "<10%", "<10%", "High = under-funded"],
        ["Lost IS (Rank)", "<20%", "<20%", "<25%", "High = QS/bid problem"],
        ["Quality Score", ">=7", ">=7", ">=6", "Weighted by cost"],
        ["Wasted spend", "<5%", "<5%", "<10%", ">$10 & 0 conv terms"],
    ]
    rows = benchmarks if benchmarks else default_rows
    r = header_row(ws, r, rows[0])
    for row in rows[1:]:
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            c.alignment = WRAP_TOP
        r += 1
    r += 1
    r = section(ws, r, "Glossary", 5)
    glossary = [
        ("IS", "Impression Share — impressions / eligible impressions."),
        ("Lost IS (Budget)", "Share of impressions missed due to limited budget."),
        ("Lost IS (Rank)", "Share missed due to low Ad Rank (bid/Quality Score)."),
        ("ICE", "Impact x Confidence x Ease prioritization score."),
        ("RSA", "Responsive Search Ad (replaces deprecated ETAs)."),
        ("PMax", "Performance Max — automated cross-network campaign type."),
    ]
    for term, desc in glossary:
        ws.cell(row=r, column=1, value=term).font = FONT_SECTION
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(row=r, column=2, value=desc).alignment = WRAP_TOP
        r += 1


def build_health_score(wb, analysis_sheets):
    ws = wb.create_sheet("17_Health_Score")
    ws.sheet_properties.tabColor = TAB_COLORS["reference"]
    set_widths(ws, [40, 16, 16, 40])
    title(ws, "Health Score (weighted, auto-computed)", 4)
    # Top-line score + grade.
    ws.cell(row=2, column=1, value="The workbook computes this from the Result "
            "+ Severity on each analysis tab. N/A is excluded.").font = FONT_MUTED
    ws.cell(row=3, column=1, value="Health Score (0-100)").font = FONT_SECTION
    ws.cell(row=4, column=1, value="Grade").font = FONT_SECTION

    r = 6
    r = header_row(ws, r, ["Analysis tab", "Earned", "Possible", "Score %"])
    first = r
    for name, sub_row in analysis_sheets:
        ws.cell(row=r, column=1, value=name).border = BORDER
        ws.cell(row=r, column=2, value=f"='{name}'!J{sub_row}").border = BORDER  # earned
        ws.cell(row=r, column=3, value=f"='{name}'!K{sub_row}").border = BORDER  # possible
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}=0,"",ROUND(B{r}/C{r}*100,1))').border = BORDER
        r += 1
    last = r - 1
    ws.cell(row=r, column=1, value="TOTAL").font = FONT_SECTION
    ws.cell(row=r, column=2, value=f"=SUM(B{first}:B{last})")
    ws.cell(row=r, column=3, value=f"=SUM(C{first}:C{last})")
    total_row = r
    # Wire the top-line cells to the totals.
    # Blank, not 0, when nothing is scoreable — mirrors the model's score=None.
    ws.cell(row=3, column=2,
            value=f'=IF(C{total_row}=0,"",ROUND(B{total_row}/C{total_row}*100,1))')
    ws.cell(row=4, column=2, value=_grade_formula("B3"))
    ws.cell(row=3, column=2).font = Font(name="Calibri", size=14, bold=True)
    ws.cell(row=4, column=2).font = Font(name="Calibri", size=14, bold=True)


def build_concentration_tab(wb, block):
    """Values-only concentration report (HHI / Effective-N / Gini / Pareto-ABC).

    Informational — never scored, never in ALL_TABS (check() must stay green
    for workbooks built without raw pull files)."""
    ws = wb.create_sheet("18_Concentration")
    ws.sheet_properties.tabColor = TAB_COLORS["reference"]
    set_widths(ws, [44, 14, 14, 12, 12, 8])
    title(ws, "Concentration — spend vs conversions (HHI)", 6)
    r = 3
    ws.cell(row=r, column=1, value=(
        "HHI bands (merger-guideline cutoffs): <1,500 unconcentrated · 1,500-2,500 "
        "moderate · >2,500 high. Small dimensions: read Effective-N instead.")
    ).alignment = WRAP_TOP
    r += 2
    for dim in block.get("dimensions", []):
        label = dim["label"] + (f" ({dim['window']})" if dim.get("window") else "")
        r = section(ws, r, label, 6)
        r = kv(ws, r, "Verdict", dim.get("verdict", ""))
        for side in ("spend", "conv"):
            m = dim.get(side)
            side_label = "Spend" if side == "spend" else "Conversions"
            if m:
                r = kv(ws, r, side_label,
                       f"HHI {m['hhi']:,.1f} ({m['band']}) · Effective-N {m['eff_n']} · "
                       f"Gini {m['gini']}")
            else:
                r = kv(ws, r, side_label, "no signal in this window")
        r = kv(ws, r, "Entities",
               f"{dim.get('n_entities', 0)} (from {dim.get('n_rows_raw', 0)} raw rows)")
        if dim.get("caveat"):
            r = kv(ws, r, "Caveat", dim["caveat"])
        r += 1
        r = header_row(ws, r, ["Entity", "Spend", "Conv", "Spend %", "Conv %", "ABC"])
        for t in dim.get("top", []):
            ws.cell(row=r, column=1, value=t["name"]).alignment = WRAP_TOP
            ws.cell(row=r, column=2, value=t["spend"]).number_format = FMT_MONEY
            ws.cell(row=r, column=3, value=t["conv"])
            ws.cell(row=r, column=4, value=t["spend_share"]).number_format = FMT_PCT
            ws.cell(row=r, column=5, value=t["conv_share"]).number_format = FMT_PCT
            ws.cell(row=r, column=6, value=t["abc"])
            r += 1
        tail = dim.get("tail")
        if tail:
            c = ws.cell(row=r, column=1, value=f"… plus {tail['n']} more (tail)")
            c.font = FONT_MUTED
            ws.cell(row=r, column=2, value=tail["spend"]).number_format = FMT_MONEY
            ws.cell(row=r, column=3, value=tail["conv"])
            ws.cell(row=r, column=4, value=tail["spend_share"]).number_format = FMT_PCT
            r += 1
        r += 1
    for note in block.get("notes", []):
        c = ws.cell(row=r, column=1, value=f"Note: {note}")
        c.font = FONT_MUTED
        c.alignment = WRAP_TOP
        r += 1
    return ws


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------
def build(findings_path: Path, output_path: Path, brand: str | None, *,
          concentration=None, findings_data: dict | None = None) -> int:
    # findings_data lets the orchestrator pass prescore-merged findings so the
    # xlsx agrees with the html/md; standalone CLI use still reads the file.
    data = findings_data if findings_data is not None else json.loads(findings_path.read_text())
    # Canonicalize before any result reaches a cell: this tab's formulas compare with
    # Excel's case-INSENSITIVE `=`, so "Fail" would score here but not in the model.
    data, _ = normalize_findings(data)
    meta = dict(data.get("meta", {}))
    if brand:
        meta["client_name"] = brand
    sections = {s.get("tab"): s for s in data.get("sections", [])}
    findings = data.get("findings", [])

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    build_readme(wb, meta)
    build_scope(wb, meta)
    build_data_inventory(wb, data.get("data_inventory", []))

    analysis_sheets = []
    for tab_name, tab_title in ANALYSIS_TABS:
        kpis = data.get("kpis") if tab_name == "04_Performance_Review" else None
        ws = build_check_tab(wb, tab_name, tab_title, sections.get(tab_name), kpis)
        analysis_sheets.append((tab_name, ws._subtotal_row))

    log_first, log_last = build_findings_log(wb, findings)
    build_ice(wb, findings, log_first)
    build_roadmap(wb, findings)
    build_client_report(wb, meta, log_first, log_last)
    build_benchmarks(wb, data.get("benchmarks"))
    build_health_score(wb, analysis_sheets)
    if concentration:
        build_concentration_tab(wb, concentration)

    # Order sheets per ALL_TABS.
    wb._sheets.sort(key=lambda s: ALL_TABS.index(s.title) if s.title in ALL_TABS else 99)
    wb.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return 0


XL_ERRORS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NUM!", "#NULL!")


def _find_soffice() -> str | None:
    return (shutil.which("soffice") or shutil.which("libreoffice")
            or next((p for p in
                     ("/Applications/LibreOffice.app/Contents/MacOS/soffice",)
                     if Path(p).is_file()), None))


def _recalculated(path: Path, soffice: str):
    """LibreOffice-recalculated copy of `path`, loaded data_only. None on failure.

    openpyxl writes no cached values, so a data_only load of our own output reads
    every cell as None — the arithmetic is invisible until something recalculates
    it. That is why the structural gate alone cannot see a #VALUE!.
    """
    with tempfile.TemporaryDirectory() as td:
        proc = subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir", td, str(path)],
            capture_output=True, timeout=300)
        out = Path(td) / path.name
        if proc.returncode != 0 or not out.is_file():
            return None
        return load_workbook(out, data_only=True)


def check(path: Path, *, recalc: bool = False, expect_score: float | None = None) -> int:
    """Quality gate for an existing workbook.

    Structural by default. `recalc=True` additionally recalculates via LibreOffice
    and inspects the COMPUTED values — the only way to catch a formula that builds
    fine and evaluates to an error. Pass `expect_score` to assert the recalculated
    Health Score equals the Python model's, which closes the xlsx parity loop
    end-to-end. Missing LibreOffice is reported and fails; it is never skipped
    silently, because a gate that quietly does nothing is worse than no gate.
    """
    problems = []
    wb = load_workbook(path)
    missing = [t for t in ALL_TABS if t not in wb.sheetnames]
    if missing:
        problems.append(f"missing tabs: {missing}")
    if "business_model" not in wb.defined_names:
        problems.append("named range 'business_model' not defined")
    if "17_Health_Score" in wb.sheetnames:
        hs = wb["17_Health_Score"]["B3"].value
        if not (isinstance(hs, str) and hs.startswith("=")):
            problems.append("17_Health_Score!B3 is not a formula")
    # Scan for literal Excel error text accidentally written into cells.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(e in v for e in XL_ERRORS):
                    problems.append(f"{ws.title}!{cell.coordinate} contains {v!r}")

    if recalc:
        soffice = _find_soffice()
        if not soffice:
            problems.append("--recalc needs LibreOffice (`soffice`) on PATH; not found")
        else:
            rwb = _recalculated(path, soffice)
            if rwb is None:
                problems.append("LibreOffice failed to recalculate the workbook")
            else:
                for ws in rwb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str) and cell.value in XL_ERRORS:
                                problems.append(
                                    f"{ws.title}!{cell.coordinate} evaluates to {cell.value}")
                # Only assert the value when the caller says what to expect. A model
                # score of None is a legitimately unscored audit, where B3 is blank.
                if expect_score is not None:
                    b3 = (rwb["17_Health_Score"]["B3"].value
                          if "17_Health_Score" in rwb.sheetnames else None)
                    if not isinstance(b3, (int, float)):
                        problems.append(f"17_Health_Score!B3 recalculates to {b3!r}, not a number")
                    elif abs(float(b3) - expect_score) > 0.05:
                        problems.append(f"17_Health_Score!B3 recalculates to {b3} but the "
                                        f"model says {expect_score} — xlsx/Python parity break")

    if problems:
        sys.stderr.write("CHECK FAILED:\n  - " + "\n  - ".join(problems) + "\n")
        return 1
    how = "structure valid" + (", recalculated" if recalc else "")
    print(f"CHECK OK: {path.name} — {len(wb.sheetnames)} tabs, {how}.")
    return 0


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Build/validate a Google Ads audit workbook.")
    ap.add_argument("--input", required=True, help="findings JSON (build) or .xlsx (--check)")
    ap.add_argument("--output", help="output .xlsx path (build mode)")
    ap.add_argument("--brand", help="client/brand name override")
    ap.add_argument("--check", action="store_true", help="validate an existing workbook")
    ap.add_argument("--recalc", action="store_true",
                    help="with --check: recalculate via LibreOffice and inspect the "
                         "computed values (needs `soffice` on PATH)")
    args = ap.parse_args(argv)

    in_path = Path(args.input)
    if not in_path.exists():
        sys.stderr.write(f"ERROR: input not found: {in_path}\n")
        return 1

    if args.check:
        return check(in_path, recalc=args.recalc)

    if not args.output:
        sys.stderr.write("ERROR: --output is required in build mode.\n")
        return 1
    try:
        return build(in_path, Path(args.output), args.brand)
    except (ValueError, KeyError, json.JSONDecodeError) as exc:
        sys.stderr.write(f"ERROR building workbook: {exc}\n")
        return 2


if __name__ == "__main__":
    sys.exit(main())

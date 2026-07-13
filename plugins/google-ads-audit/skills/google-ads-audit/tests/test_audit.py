#!/usr/bin/env python3
# Copyright (c) 2026 Clickt Digital Marketing Inc. All rights reserved.
"""Conformance tests for the google-ads-audit renderers. Stdlib only.

Run: python3 tests/test_audit.py
Guards: score correctness, no-loss, N/A exclusion, JS<->Python<->Excel constant parity,
self-containment (GSAP sentinel + checksum), and determinism.
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
SCRIPTS = HERE.parent / "scripts"
sys.path.insert(0, str(SCRIPTS))

import audit_model
import audit_html
import audit_md

EXAMPLE = SCRIPTS / "example_findings.json"
FINDINGS = json.loads(EXAMPLE.read_text())

FAILS = []


def check(name, cond, detail=""):
    print(("  ok   " if cond else "FAIL   ") + name + ("" if cond else "   :: " + detail))
    if not cond:
        FAILS.append(name)


# --- 1-3: model correctness, no-loss, N/A exclusion --------------------------
model = audit_model.compute_model(FINDINGS, generated="2026-06-24T00:00:00")
check("health score == 43.7 (hand oracle)", model["health"]["score"] == 43.7,
      str(model["health"]))
check("grade == D", model["health"]["grade"] == "D")
check("earned/possible == 19.0 / 43.5", model["health"]["earned"] == 19.0 and model["health"]["possible"] == 43.5,
      str(model["health"]))
check("no-loss: 24 checks", model["provenance"]["n_checks"] == 24, str(model["provenance"]["n_checks"]))
check("no-loss: 5 findings", model["provenance"]["n_findings"] == 5)
check("counts 6/9/5/4", (model["summary"]["n_pass"], model["summary"]["n_flag"],
      model["summary"]["n_fail"], model["summary"]["n_na"]) == (6, 9, 5, 4), str(model["summary"]))

na = [c for s in model["sections"] for c in s["checks"] if c["result"] == "N/A"]
check("4 N/A checks excluded (earned & possible None)",
      len(na) == 4 and all(c["earned"] is None and c["possible"] is None for c in na),
      f"{len(na)} n/a rows")

# a section that is entirely FAIL + N/A scores 0 with a real denominator (not div-by-zero)
kw = [s for s in model["sections"] if s["tab"] == "05_Keyword_Strategy"][0]
check("all-FAIL section score 0.0 (not None)", kw["score_pct"] == 0.0, str(kw["score_pct"]))


# --- 4: JS <-> Python constant parity (the render kernel mirrors the model) ---
tpl = audit_html._TEMPLATE


def js_map(name):
    m = re.search(name + r"\s*=\s*\{([^}]*)\}", tpl)
    out = {}
    for part in m.group(1).split(","):
        k, v = part.split(":")
        out[k.strip()] = float(v.strip())
    return out


check("JS SEV_W == audit_model.SEVERITY_WEIGHTS", js_map("SEV_W") == audit_model.SEVERITY_WEIGHTS)
check("JS FLAG == audit_model.FLAG_SCORES", js_map("FLAG") == audit_model.FLAG_SCORES)
check("JS IMPACT == audit_model.SEVERITY_IMPACT", js_map("IMPACT") == audit_model.SEVERITY_IMPACT)
js_grades = [(int(n), g) for n, g in re.findall(r"\[(\d+),'([A-F])'\]", tpl)]
check("JS GRADES == audit_model.GRADE_CUTOFFS", js_grades == [tuple(x) for x in audit_model.GRADE_CUTOFFS],
      str(js_grades))


# --- 5: Excel <-> Python parity (workbook single-sources from audit_model) ----
try:
    import generate_workbook as gw
    check("xlsx SEVERITY_WEIGHTS is audit_model's", gw.SEVERITY_WEIGHTS == audit_model.SEVERITY_WEIGHTS)
    check("xlsx SEVERITY_IMPACT is audit_model's", gw.SEVERITY_IMPACT == audit_model.SEVERITY_IMPACT)
    check("xlsx ANALYSIS_TABS is audit_model's", gw.ANALYSIS_TABS == audit_model.ANALYSIS_TABS)
    with tempfile.TemporaryDirectory() as td:
        xlsx = Path(td) / "t.xlsx"
        rc_build = gw.build(EXAMPLE, xlsx, "Acme Corp")
        rc_check = gw.check(xlsx)
        check("xlsx build == 0", rc_build == 0)
        check("xlsx --check gate == 0", rc_check == 0)
except SystemExit:
    check("openpyxl available for xlsx tests", False, "openpyxl not installed")


# --- 6: self-containment -----------------------------------------------------
html = audit_html.render_html(model, animate=True)
blob = audit_html.gsap_blob()
check("GSAP blob embedded", blob in html)
stripped = html.replace(blob, "")
hits = re.findall(r"https?://|<link|src=|cdn", stripped)
check("no external refs outside GSAP sentinels", not hits, str(hits[:4]))

gsap_bytes = (SCRIPTS / "vendor" / "gsap.min.js").read_bytes()
want = (SCRIPTS / "vendor" / "SHA256SUMS").read_text().split()[0]
check("vendored GSAP matches SHA256SUMS", hashlib.sha256(gsap_bytes).hexdigest() == want)

h_noanim = audit_html.render_html(model, animate=False)
check("animate=False carries zero GSAP bytes",
      audit_html.GSAP_BEGIN not in h_noanim and "GreenSock" not in h_noanim)
check("animate=False still self-contained", not re.findall(r"https?://|<link|src=|cdn", h_noanim))


# --- 7: determinism (pure fn of findings, except provenance.generated) -------
a = audit_html.render_html(audit_model.compute_model(FINDINGS, generated="TS1"))
b = audit_html.render_html(audit_model.compute_model(FINDINGS, generated="TS2"))
check("HTML deterministic modulo generated", a.replace("TS1", "TS") == b.replace("TS2", "TS"))
ma = audit_md.render_md(audit_model.compute_model(FINDINGS, generated="TS1"))
mb = audit_md.render_md(audit_model.compute_model(FINDINGS, generated="TS2"))
check("markdown deterministic modulo generated", ma.replace("TS1", "TS") == mb.replace("TS2", "TS"))


# --- 8: concentration — metric hand-oracles ----------------------------------
import concentration as conc

FIX = HERE / "fixtures"
spend6 = [50, 20, 10, 10, 5, 5]
conv6 = [30, 0, 5, 0, 1, 0]
check("hhi(spend6) == 3150.0", abs(conc.hhi(spend6) - 3150.0) < 1e-9, str(conc.hhi(spend6)))
check("effective_n(spend6) == 3.17", round(conc.effective_n(spend6), 2) == 3.17)
check("gini(spend6) == 0.45", round(conc.gini(spend6), 3) == 0.45, str(conc.gini(spend6)))
check("abc(spend6) == AAABBC", conc.pareto_abc(spend6) == ["A", "A", "A", "B", "B", "C"],
      str(conc.pareto_abc(spend6)))
check("hhi(conv6) == 7145.1", round(conc.hhi(conv6), 1) == 7145.1, str(conc.hhi(conv6)))
check("effective_n(conv6) == 1.40", round(conc.effective_n(conv6), 2) == 1.40)
check("gini(conv6) == 0.769", round(conc.gini(conv6), 3) == 0.769, str(conc.gini(conv6)))
check("abc(conv6) crossing-inclusive == ACBCCC",
      conc.pareto_abc(conv6) == ["A", "C", "B", "C", "C", "C"], str(conc.pareto_abc(conv6)))

check("verdict fragility", conc.verdict(3000, 3000)[0] == "fragility")
check("verdict consolidate", conc.verdict(1000, 3000)[0] == "consolidate")
check("verdict review_bidding", conc.verdict(3000, 1000)[0] == "review_bidding")
check("verdict diversified", conc.verdict(1000, 1000)[0] == "diversified")
check("verdict no_conv_signal", conc.verdict(3000, None)[0] == "no_conv_signal")
check("verdict insufficient", conc.verdict(None, None)[0] == "insufficient")

check("equal k=4: hhi 2500 / band moderate / eff_n 4 / gini 0",
      conc.hhi([1, 1, 1, 1]) == 2500.0 and conc.hhi_band(2500.0) == "moderate"
      and conc.effective_n([1, 1, 1, 1]) == 4.0 and conc.gini([1, 1, 1, 1]) == 0.0)
check("band boundary 1500 -> moderate", conc.hhi_band(1500.0) == "moderate")
check("single entity: hhi 10000 / eff_n 1 / gini 0 / lorenz / abc A",
      conc.hhi([7]) == 10000.0 and conc.effective_n([7]) == 1.0 and conc.gini([7]) == 0.0
      and conc.lorenz_points([7]) == [[0.0, 0.0], [1.0, 1.0]]
      and conc.pareto_abc([7]) == ["A"])


# --- 9: concentration — end-to-end from fixtures ------------------------------
camp_rows = conc.load_rows(str(FIX / "conc_campaigns.json"),
                           require_fields=["campaign.name", "metrics.cost_micros",
                                           "metrics.conversions"])
kw_rows = conc.load_rows(str(FIX / "conc_keywords.json"))
st_rows = conc.load_rows(str(FIX / "conc_search_terms_31.json"))
block = conc.compute_concentration(campaign_rows=camp_rows, keyword_rows=kw_rows,
                                   search_term_rows=st_rows,
                                   windows={"90d": "LAST_90_DAYS", "30d": "LAST_30_DAYS"})
dims = {d["key"]: d for d in block["dimensions"]}
check("four dimensions present",
      set(dims) == {"search_terms", "keywords", "campaigns", "campaign_types"}, str(set(dims)))
dc = dims["campaigns"]
check("campaigns spend hhi 3150.0 / band high",
      dc["spend"]["hhi"] == 3150.0 and dc["spend"]["band"] == "high", str(dc["spend"]))
check("campaigns conv hhi 7145.1", dc["conv"]["hhi"] == 7145.1, str(dc["conv"]))
check("campaigns verdict fragility", dc["verdict_key"] == "fragility")
check("campaigns top sorted spend desc, name asc tie-break",
      [t["name"] for t in dc["top"]][:4] == ["camp-a", "camp-b", "camp-c", "camp-d"])
check("campaigns small-N caveat fires", "Effective-N" in (dc["caveat"] or ""))
check("keywords aggregate across ad groups: 3 entities from 4 rows",
      dims["keywords"]["n_entities"] == 3 and dims["keywords"]["n_rows_raw"] == 4)
check("keywords top is blue widgets [EXACT] 50.0",
      dims["keywords"]["top"][0]["name"] == "blue widgets [EXACT]"
      and dims["keywords"]["top"][0]["spend"] == 50.0, str(dims["keywords"]["top"][0]))
dst = dims["search_terms"]
vals31 = [float(32 - i) for i in range(1, 32)]
check("embed cap: top 25 of 31, tail n 6",
      len(dst["top"]) == 25 and dst["tail"]["n"] == 6, str(dst["tail"]))
check("cap does not distort math: hhi on full universe",
      dst["spend"]["hhi"] == round(conc.hhi(vals31), 1), str(dst["spend"]["hhi"]))
check("lorenz downsampled <= 101 pts", len(dst["lorenz"]["spend"]) <= 101)
check("windows labeled per dimension",
      dst["window"] == "LAST_30_DAYS" and dc["window"] == "LAST_90_DAYS")

zc_rows = conc.load_rows(str(FIX / "conc_campaigns_zeroconv.json"))
zc = conc.compute_concentration(campaign_rows=zc_rows)
check("zero-conv: conv None + no_conv_signal",
      zc["dimensions"][0]["conv"] is None
      and zc["dimensions"][0]["verdict_key"] == "no_conv_signal")
check("absent inputs -> None", conc.compute_concentration() is None)

try:
    conc.load_rows(str(FIX / "conc_keywords.json"),
                   require_fields=["search_term_view.search_term"])
    check("require_fields catches wrong file", False)
except conc.RawResultError:
    check("require_fields catches wrong file", True)


# --- 10: concentration — through the renderers --------------------------------
model_c = audit_model.compute_model(FINDINGS, generated="2026-06-24T00:00:00",
                                    concentration=block)
html_c = audit_html.render_html(model_c)
check("html embeds concentration data", '"verdict_key"' in html_c and '"dimensions"' in html_c)
check("html without concentration embeds null", '"concentration":null' in html)
stripped_c = html_c.replace(audit_html.gsap_blob(), "")
check("html with concentration still self-contained",
      not re.findall(r"https?://|<link|src=|cdn", stripped_c))
js_grades_still = re.findall(r"\[(\d+),'([A-F])'\]", audit_html._TEMPLATE)
check("GRADES parity regex unpolluted by concentration JS",
      [(int(n), g) for n, g in js_grades_still] == audit_model.GRADE_CUTOFFS,
      str(js_grades_still))

md_c = audit_md.render_md(model_c)
check("md has Concentration section", "## Concentration" in md_c)
check("md without concentration omits section",
      "## Concentration" not in audit_md.render_md(model))
c1 = audit_html.render_html(audit_model.compute_model(FINDINGS, generated="TS1", concentration=block))
c2 = audit_html.render_html(audit_model.compute_model(FINDINGS, generated="TS2", concentration=block))
check("HTML with concentration deterministic modulo generated",
      c1.replace("TS1", "TS") == c2.replace("TS2", "TS"))

try:
    import generate_workbook as gw2
    with tempfile.TemporaryDirectory() as td:
        x1 = Path(td) / "with.xlsx"
        rc = gw2.build(EXAMPLE, x1, "Acme Corp", concentration=block)
        check("xlsx with concentration builds", rc == 0)
        from openpyxl import load_workbook as _lw
        check("18_Concentration present", "18_Concentration" in _lw(x1).sheetnames)
        check("check() green with concentration tab", gw2.check(x1) == 0)
        x2 = Path(td) / "without.xlsx"
        gw2.build(EXAMPLE, x2, "Acme Corp")
        check("18_Concentration absent when not provided",
              "18_Concentration" not in _lw(x2).sheetnames)
        check("check() green without concentration tab", gw2.check(x2) == 0)
except SystemExit:
    print("  skip  xlsx concentration checks (openpyxl unavailable)")


# --- 11: manual UI-CSV path (the no-MCP adapter) ------------------------------
import manual_csv as mc

check("ui_num coercions",
      mc.ui_num("--") == 0.0 and mc.ui_num(" --") == 0.0 and mc.ui_num("") == 0.0
      and mc.ui_num("2,791.40") == 2791.40 and mc.ui_num("52.76%") == 52.76
      and mc.ui_num("CA$0.00") == 0.0 and mc.ui_num('"?"') == 0.0
      and mc.ui_num("1,000.00") == 1000.0)

uc_rows, uc_meta = mc.campaigns_rows(str(FIX / "ui_campaigns.csv"))
check("ui campaigns: 3 rows (Total rows dropped, multiline cell survived)",
      len(uc_rows) == 3, str(len(uc_rows)))
check("ui campaigns: date range extracted",
      uc_meta["date_range"] == "April 1, 2026 - July 11, 2026", uc_meta["date_range"])
check("ui campaigns: micros synthesized + ' --' conversions -> 0",
      uc_rows[0]["metrics.cost_micros"] == 1_000_000_000
      and uc_rows[1]["metrics.conversions"] == 0.0
      and uc_rows[2]["metrics.conversions"] == 5.5)
check("ui campaigns: channel types verbatim",
      [r["campaign.advertising_channel_type"] for r in uc_rows]
      == ["Search", "Performance Max", "Search"])

uk_rows, _ = mc.keywords_rows(str(FIX / "ui_keywords.csv"))
us_rows, us_meta = mc.search_terms_rows(str(FIX / "ui_search_terms.csv"))
ui_block = conc.compute_concentration(
    campaign_rows=uc_rows, keyword_rows=uk_rows, search_term_rows=us_rows,
    windows={"campaigns": uc_meta["date_range"], "search_terms": "ST-WINDOW"})
ui_dims = {d["key"]: d for d in ui_block["dimensions"]}
check("ui keywords aggregate across ad groups: 2 entities from 3 rows",
      ui_dims["keywords"]["n_entities"] == 2 and ui_dims["keywords"]["n_rows_raw"] == 3)
check("ui keywords top: quoted phrase text kept verbatim, spend summed",
      ui_dims["keywords"]["top"][0]["name"] == '"blue widgets" [Phrase match]'
      and ui_dims["keywords"]["top"][0]["spend"] == 500.0,
      str(ui_dims["keywords"]["top"][0]))
check("ui search terms: duplicate term aggregated (2 entities from 3 rows)",
      ui_dims["search_terms"]["n_entities"] == 2
      and ui_dims["search_terms"]["top"][0]["name"] == "term one"
      and ui_dims["search_terms"]["top"][0]["spend"] == 1300.0,
      str(ui_dims["search_terms"]["top"][:1]))
check("ui per-dimension windows from files",
      ui_dims["campaigns"]["window"] == "April 1, 2026 - July 11, 2026"
      and ui_dims["search_terms"]["window"] == "ST-WINDOW")
check("ui campaign_types present from UI 'Campaign type' column",
      "campaign_types" in ui_dims and ui_dims["campaign_types"]["n_entities"] == 2)

try:
    mc.load_ui_csv(str(FIX / "ui_keywords.csv"), kind="search_terms")
    check("ui wrong-report guard", False)
except mc.ManualCsvError:
    check("ui wrong-report guard", True)


# --- 12: deterministic pre-scorer ---------------------------------------------
import prescore as ps

check("ui_frac coercions",
      mc.ui_frac("52.76%") == (0.5276, None) and mc.ui_frac("< 10%") == (0.10, "<")
      and mc.ui_frac("> 90%") == (0.90, ">") and mc.ui_frac(" --") == (None, None)
      and mc.ui_frac("") == (None, None))
check("ui_int_or_none", mc.ui_int_or_none("7") == 7 and mc.ui_int_or_none(" --") is None)
check("canonicalizers",
      ps.canon_match("Phrase match (close variant)") == "PHRASE"
      and ps.canon_match("BROAD") == "BROAD"
      and ps.canon_kw_text('"Blue  Widgets"') == "blue widgets"
      and ps.canon_kw_text("[red widgets]") == "red widgets"
      and ps.canon_strategy("Manual CPC (enhanced)") == "ENHANCED_CPC"
      and ps.canon_enum("Below average") == ps.canon_enum("BELOW_AVERAGE"))

# raw-path prescore over the extended fixtures.
# IS oracle arithmetic (Search rows, cost weights 50/10/5, total 65):
#   PR-04 (50*60+10*80+5*40)/65 = 4000/65 = 61.5 FLAG
#   PR-05 (50*25+10*5 +5*10)/65 = 1350/65 = 20.8 FAIL
#   PR-06 (50*10+10*15+5*45)/65 =  875/65 = 13.5 PASS
#   PR-01 clicks 520 / impr 13000 = 4.0% PASS
camp_rows2 = conc.load_rows(str(FIX / "conc_campaigns.json"))
st_rows2 = conc.load_rows(str(FIX / "prescore_search_terms.json"))
kw_rows2 = conc.load_rows(str(FIX / "prescore_keywords.json"))
pres = ps.compute_prescore(campaign_rows=camp_rows2, keyword_rows=kw_rows2,
                           search_term_rows=st_rows2, business_model="Lead Gen")
C = pres["checks"]
check("PR-04 61.5 FLAG", "61.5%" in C["PR-04"]["observed"] and C["PR-04"]["result"] == "FLAG",
      str(C.get("PR-04")))
check("PR-05 20.8 FAIL", "20.8%" in C["PR-05"]["observed"] and C["PR-05"]["result"] == "FAIL",
      str(C.get("PR-05")))
check("PR-06 13.5 PASS", "13.5%" in C["PR-06"]["observed"] and C["PR-06"]["result"] == "PASS",
      str(C.get("PR-06")))
check("PR-01 4.0 PASS", "4.0%" in C["PR-01"]["observed"] and C["PR-01"]["result"] == "PASS",
      str(C.get("PR-01")))
check("BB-02 PASS (no eCPC)", C["BB-02"]["result"] == "PASS")
# KW-02: wasted = 12 + (6+6 split term) = 24 of 103 total = 23.3% FAIL;
# 'small waste' $9 <= $10 excluded; 'already excluded' removed from numerator.
check("KW-02 23.3% FAIL (split-term agg, EXCLUDED removed)",
      "23.3%" in C["KW-02"]["observed"] and C["KW-02"]["result"] == "FAIL",
      str(C.get("KW-02")))
check("KW-05 2 dupes FLAG", C["KW-05"]["result"] == "FLAG"
      and C["KW-05"]["observed"].startswith("2 "), str(C.get("KW-05")))
check("AS-03 1 cross-campaign dupe FLAG", C["AS-03"]["result"] == "FLAG"
      and C["AS-03"]["observed"].startswith("1 "), str(C.get("AS-03")))
check("KW-03 BROAD-in-ManualCPC FLAG", C["KW-03"]["result"] == "FLAG",
      str(C.get("KW-03")))
qs_kpi = next((k for k in pres["kpis"] if k["metric"] == "Quality Score (cost-wtd)"), None)
check("QS cost-wtd 6.8 FLAG", qs_kpi and qs_kpi["value"] == 6.8 and qs_kpi["flag"] == "FLAG",
      str(qs_kpi))
check("severities enforced from framework",
      C["PR-05"]["severity"] == "High" and C["KW-05"]["severity"] == "Medium")
check("evidence blocks computed",
      "Blended CPA" in pres["evidence"]["PR-02"]["observed"]
      and "match type" in pres["evidence"]["KW-04"]["observed"])

# skipped-when-missing: minimal rows without IS/strategy/ad-group fields
bare = [{"campaign.name": "x", "campaign.advertising_channel_type": "SEARCH",
         "metrics.cost_micros": 1_000_000, "metrics.conversions": 1}]
pres_bare = ps.compute_prescore(campaign_rows=bare)
skipped_ids = {s["id"] for s in pres_bare["skipped"]}
check("graceful skips on old-style input",
      {"PR-01", "PR-04", "PR-05", "PR-06", "BB-02"} <= skipped_ids
      and not pres_bare["checks"], str(sorted(skipped_ids)))
check("prescore None on no input", ps.compute_prescore() is None)

# UI-path prescore: (1000*60+250.5*80)/1250.5 = 64.0 FLAG; lost-budget boundary
# '< 10%' -> (1000*10+250.5*5)/1250.5 = 9.0 PASS with a bounds note.
uc2, _ = mc.campaigns_rows(str(FIX / "ui_campaigns.csv"))
pres_ui = ps.compute_prescore(campaign_rows=uc2)
check("ui PR-04 64.0 FLAG", "64.0%" in pres_ui["checks"]["PR-04"]["observed"]
      and pres_ui["checks"]["PR-04"]["result"] == "FLAG", str(pres_ui["checks"].get("PR-04")))
check("ui PR-05 9.0 PASS with bounds note",
      "9.0%" in pres_ui["checks"]["PR-05"]["observed"]
      and pres_ui["checks"]["PR-05"]["result"] == "PASS"
      and any("bounds" in n for n in pres_ui["notes"]), str(pres_ui["notes"]))
check("ui BB-02 PASS from Bid strategy type", pres_ui["checks"]["BB-02"]["result"] == "PASS")

# merge: corrections, injection, kpi replacement, Health flip.
# After merge: earned = KW-02 FAIL 0 + KW-06 1.5 + PR-01 1.5 + PR-04 .75 + PR-06 1.5
#              + PR-05 0 + BB-02 1.5 = 6.75; possible = 13.5 -> 50.0 grade D.
pf = json.loads((FIX / "prescore_findings.json").read_text())
pres_m = ps.compute_prescore(campaign_rows=camp_rows2, search_term_rows=st_rows2,
                             business_model="Lead Gen")
merged, block, plog = ps.merge_into_findings(pf, pres_m)
check("merge is pure (input untouched)",
      pf["sections"][0]["checks"][0]["result"] == "PASS")
check("KW-02 corrected PASS->FAIL",
      block["corrected"] == [{"id": "KW-02", "from": "PASS", "to": "FAIL"}]
      and any("KW-02 PASS->FAIL" in l for l in plog), str(block["corrected"]))
check("recommendation preserved on corrected check",
      next(c for s in merged["sections"] for c in s["checks"]
           if c["id"] == "KW-02")["recommendation"] == "keep monitoring")
check("injection creates 04_Performance_Review in order",
      [s["tab"] for s in merged["sections"]]
      == ["04_Performance_Review", "05_Keyword_Strategy", "08_Budget_Bidding"],
      str([s["tab"] for s in merged["sections"]]))
check("injected section uses canonical title",
      merged["sections"][0]["title"] == dict(audit_model.ANALYSIS_TABS)["04_Performance_Review"])
m_merged = audit_model.compute_model(merged, generated="2026-06-24T00:00:00", prescore=block)
check("merged Health 50.0 / D (hand oracle: 6.75/13.5)",
      m_merged["health"]["score"] == 50.0 and m_merged["health"]["grade"] == "D",
      str(m_merged["health"]))
kpi_metrics = [k["metric"] for k in merged["kpis"]]
ctr_row = next(k for k in merged["kpis"] if k["metric"] == "CTR")
check("kpi CTR replaced in place, Conv rate survives",
      kpi_metrics[0] == "CTR" and ctr_row["value"] == 4.0 and "Conv rate" in kpi_metrics,
      str(kpi_metrics))
check("merge with None is a no-op", ps.merge_into_findings(pf, None) == (pf, None, []))

# renderers + guards with prescore present
html_p = audit_html.render_html(m_merged)
check("html shows machine-scored overview line", "machine-scored" in html_p)
check("html without prescore embeds null", '"prescore":null' in html)
check("concentration-null assertion still intact", '"concentration":null' in html)
stripped_p = html_p.replace(audit_html.gsap_blob(), "")
check("html with prescore self-contained",
      not re.findall(r"https?://|<link|src=|cdn", stripped_p))
js_grades_p = re.findall(r"\[(\d+),'([A-F])'\]", audit_html._TEMPLATE)
check("GRADES parity unpolluted by prescore",
      [(int(n), g) for n, g in js_grades_p] == audit_model.GRADE_CUTOFFS)
md_p = audit_md.render_md(m_merged)
check("md prescore footer with correction ids",
      "checks machine-scored" in md_p
      and "KW-02" in md_p.rsplit("checks machine-scored", 1)[1][:250],
      repr(md_p.rsplit("checks machine-scored", 1)[-1][:120]))
p1 = ps.compute_prescore(campaign_rows=camp_rows2, search_term_rows=st_rows2,
                         business_model="Lead Gen")
check("prescore deterministic", json.dumps(pres_m, sort_keys=True) == json.dumps(p1, sort_keys=True))

# workbook threading
try:
    import generate_workbook as gw3
    with tempfile.TemporaryDirectory() as td:
        x = Path(td) / "pres.xlsx"
        check("xlsx builds from merged findings_data",
              gw3.build(EXAMPLE, x, "Acme Corp", findings_data=merged) == 0)
        from openpyxl import load_workbook as _lw2
        ws = _lw2(x)["05_Keyword_Strategy"]
        kw02 = [tuple(str(c) for c in row) for row in ws.iter_rows(values_only=True)
                if any("KW-02" in str(c) for c in row)]
        check("corrected FAIL lands in the xlsx",
              kw02 and any("FAIL" in cell for cell in kw02[0]), str(kw02[:1]))
        check("check() green on prescored workbook", gw3.check(x) == 0)
        x2 = Path(td) / "disk.xlsx"
        check("no-kwarg build still reads disk", gw3.build(EXAMPLE, x2, "Acme Corp") == 0)
except SystemExit:
    print("  skip  xlsx prescore checks (openpyxl unavailable)")


print()
if FAILS:
    print(f"FAILED ({len(FAILS)}): " + ", ".join(FAILS))
    sys.exit(1)
print("All audit conformance tests passed.")
sys.exit(0)
